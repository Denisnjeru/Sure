import datetime
from email import message
from multiprocessing import context
import os
import tempfile
from pathlib import Path
from unicodedata import category
from django import apps
from django.conf import settings
from django.template.loader import render_to_string
from django.core.mail import EmailMultiAlternatives

from celery import Task, shared_task
from celery_progress.backend import ProgressRecorder
from backend.celery import app

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from sentry_sdk import capture_exception
from apps.common.utils import get_local_filepath
from apps.core.utils import format_excel, insert_image, get_file_path
from .utils import convertfiletobase64, convertbase64tofile 
from .models import (
    AuctionInvitee,
    AuctionItems,
    AuctionItemResponses,
    Auction
)

def invite_suppliers_auction(auction_id):
    auction = Auction.objects.filter(id=auction_id).first()
    for invitee in AuctionInvitee.objects.filter(auction=auction):
        if invitee.supplier is None:
            supplier = "Esteemed Vendor"
        else:
            supplier = invitee.supplier
        
        email_subject = "Tendersure Auction Invitation"
        to_email = invitee.email

        # send default from Tendersure
        message = render_to_string(
            "emails/auction_inivitation.html",
            {
                "supplier": supplier,
                "auction": auction,
                "buyer_name": "Tendersure Team",
                "buyer_logo": "tendersure_logo"
            }
        )
        email = EmailMultiAlternatives(email_subject, message, to=[to_email])
        email.attach_alternative(message, "text/html")
        email.send(fail_silently=True)
    
    context = {"message": "Invitations to participate sent to suppliers."}
    return context

@shared_task(bind=True)
def create_auction_items(self, auction_id):
    auction = Auction.objects.filter(id=auction_id).first()
    if auction is not None:
        """
        Self Evaluate, evaluated by system
        """
        filepath = get_local_filepath(category.items_template.url)

        workbook = load_workbook(filepath, data_only=True)
        try:
            worksheet = workbook.get_sheet_by_name("AUCTION ITEMS")

            for i in range(9, worksheet.max_row + 1):
                try:
                    if len(worksheet.cell(row=i, column=2).value) > 1:
                        if worksheet.cell(row=8, column=3).value != "UoM":
                            unit_of_measure = "N/A"
                        else:
                            unit_of_measure = worksheet.cell(row=i, column=3).value

                        if (
                            worksheet.cell(row=8, column=5).value
                            != "Specification 1"
                        ):
                            specification_1 = "N/A"
                        else:
                            specification_1 = worksheet.cell(row=i, column=5).value
                        if (
                            worksheet.cell(row=8, column=6).value
                            != "Specification 2"
                        ):
                            specification_2 = "N/A"
                        else:
                            specification_2 = worksheet.cell(row=i, column=6).value
                        if worksheet.cell(row=8, column=3).value != "Qty":
                            qty = 1
                        else:
                            qty = worksheet.cell(row=i, column=3).value
                        data = {
                            "auction_id": auction.id,
                            "description": worksheet.cell(
                                row=i, column=2
                            ).value,
                            "unit_of_measure": unit_of_measure,
                            "specification_1": specification_1,
                            "specification_2": specification_2,
                            "item_number": worksheet.cell(row=i, column=1).value,
                            "quantity": qty,
                            "current_price": 0,
                            "price_validity_months": 12,
                        }
                        new__auction_item = AuctionItems(**data)
                        new__auction_item.save()
                        print("Auction item created here")
                    else:
                        message = "Column 2 must not be empty"
                        return {"result": "error", "error_message": message}

                except Exception as e:
                    print(e)
                    message = str(e)
                    return {"result": "error", "error_message": message}

            message = "Auction item created successfully"
            return {
                "result": "success",
                "success_message": message,
                "Item": new__auction_item.id,
            }
        except Exception as e:
            message = "Column 2 must not be empty"
            return {"result": "error", "error_message": message}

    message = "Auction not found!"
    return {"result": "error", "error_message": message}

@shared_task(bind=True)
def download_auction_items_template_import(self, auction_id, auction_type):
    """
    Download template to import bids
    """
    time = datetime.datetime.now()
    dir_name = Path(
        "media/Auction/templates/item_template/%s/%s/%s" % (time.year, time.month, time.day)
    )  # folder structure
    dir_name.mkdir(parents=True, exist_ok=True)
    time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
    auction_items = "Auction_Items"
    filepath = "{}/{}_{}.xlsx".format(dir_name, auction_items, time_only)

    workbook = Workbook()
    items_sheet = workbook.active
    items_sheet.title = "ITEMs"
    buyer_logo_url = "static/core/img/Tendersure_Logo.png"
    buyer_logo = Image("static/core/img/Tendersure_Logo.png")
    buyer_logo_height, buyer_logo_width = buyer_logo.height, buyer_logo.width
    max_height = 75
    ratio = buyer_logo_height / max_height
    buyer_logo.height, buyer_logo.width = buyer_logo_height, buyer_logo_width

    buyer_logo_anchor = "B1"

    items_sheet.title = "ITEMs"
    logo_worksheet = items_sheet.title
    items_sheet.append(("", "", ""))
    items_sheet.append(("", "", ""))
    items_sheet.append(("", "", ""))
    items_sheet.append(("", "", ""))
    if auction_type == 'Reverse Auction':
        items_sheet.append(
            (
                "",
                "h_Name",
                "h_Item Description",
                "h_Reserve Price",
                "h_Minimum Decrement"
            )
        )
    else:
        items_sheet.append(
            (
                "",
                "h_Name",
                "h_Item Description",
                "h_Reserve Price",
                "h_Minimum Increment"
            )
        )

    data = {}
    # data["ITEM PRICES_B"] = 8
    data["ITEM PRICES_C"] = 5
    data["ITEM PRICES_D"] = 39
    data["ITEM PRICES_E"] = 20
    data["ITEM PRICES_F"] = 15
    data["ITEM PRICES_G"] = 15

    auction_items = (
        AuctionItems.objects.filter(auction_id=auction_id)
        .order_by("id")
        .only("id")
    )

    try:
        result = 0
        progress_recorder = ProgressRecorder(self)
        total_for_progress = auction_items.count()

        for count, item in enumerate(auction_items):
            result += 1
            progress_recorder.set_progress(result, total_for_progress)
            if auction_type == 'Reverse Auction': 
                items_sheet.append(
                    (
                        "",
                        item.name,
                        item.description,
                        item.reserve_price,
                        item.minimum_decrement,
                    ),
                )
            else:
                items_sheet.append(
                    (
                        "",
                        item.name,
                        item.description,
                        item.reserve_price,
                        item.minimum_increment,
                    ),
                )
        workbook.save(filepath)
        format_excel(filepath, data=data)
        filepath = insert_image(
            filepath, logo_worksheet, buyer_logo_anchor, image_url=buyer_logo_url
        )
        context = {
            "result": "success",
            "response_message": "Template generated successfully",
            # "filepath": filepath.split("/", 1)[1],
            "filepath": filepath,
        }

        return context

    except Exception as e:
        # capture_exception(e)
        print(e)
        context = {
            "result": "error",
            "response_message": "Error generating template",
        }
        return context

@shared_task(bind=True)
def download_auction_bids_import(self, auction_id, auction_type):
    """
    Download import bid template
    """
    time = datetime.datetime.now()
    dir_name = Path(
        "media/Auction/templates/bid_template/%s/%s/%s" % (time.year, time.month, time.day)
    )  # folder structure
    dir_name.mkdir(parents=True, exist_ok=True)
    time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
    auction_items = "Auction_Items"
    filepath = "{}/{}_{}.xlsx".format(dir_name, auction_items, time_only)

    workbook = Workbook()
    items_sheet = workbook.active
    items_sheet.title = "ITEMs"
    buyer_logo_url = "static/core/img/Tendersure_Logo.png"
    buyer_logo = Image("static/core/img/Tendersure_Logo.png")
    buyer_logo_height, buyer_logo_width = buyer_logo.height, buyer_logo.width
    max_height = 75
    ratio = buyer_logo_height / max_height
    buyer_logo.height, buyer_logo.width = buyer_logo_height, buyer_logo_width

    buyer_logo_anchor = "B1"

    items_sheet.title = "ITEMs"
    logo_worksheet = items_sheet.title
    items_sheet.append(("", "", ""))
    items_sheet.append(("", "", ""))
    items_sheet.append(("", "", ""))
    items_sheet.append(("", "", ""))
    if auction_type == 'Reverse Auction':
        items_sheet.append(
            (
                "",
                "h_Name",
                "h_Item Description",
                "h_Reserve Price",
                "h_Minimum Decrement",
                "h_Bid Price",
            )
        )
    else:
        items_sheet.append(
            (
                "",
                "h_Name",
                "h_Item Description",
                "h_Reserve Price",
                "h_Minimum Increment",
                "h_Bid Price",
            )
        )

    data = {}
    # data["ITEM PRICES_B"] = 8
    data["ITEM PRICES_C"] = 5
    data["ITEM PRICES_D"] = 39
    data["ITEM PRICES_E"] = 20
    data["ITEM PRICES_F"] = 15
    data["ITEM PRICES_G"] = 15

    auction_items = (
        AuctionItems.objects.filter(auction_id=auction_id)
        .order_by("id")
        .only("id")
    )

    try:
        result = 0
        progress_recorder = ProgressRecorder(self)
        total_for_progress = auction_items.count()

        for count, item in enumerate(auction_items):
            result += 1
            progress_recorder.set_progress(result, total_for_progress)
            if auction_type == 'Reverse Auction': 
                items_sheet.append(
                    (
                        "",
                        item.name,
                        item.description,
                        item.reserve_price,
                        item.minimum_decrement,
                    ),
                )
            else:
                items_sheet.append(
                    (
                        "",
                        item.name,
                        item.description,
                        item.reserve_price,
                        item.minimum_increment,
                    ),
                )
        workbook.save(filepath)
        format_excel(filepath, data=data)
        filepath = insert_image(
            filepath, logo_worksheet, buyer_logo_anchor, image_url=buyer_logo_url
        )
        context = {
            "result": "success",
            "response_message": "Template generated successfully",
            # "filepath": filepath.split("/", 1)[1],
            "filepath": filepath,
        }

        return context

    except Exception as e:
        # capture_exception(e)
        print(e)
        context = {
            "result": "error",
            "response_message": "Error generating template",
        }
        return context

@shared_task(bind=True)
def submit_auction(self, auction_id, supplier_id, bid_template):
    messages = []
    auction = Auction.objects.filter(id=auction_id).first()
    supplier = apps.apps.get_model("suppliers", "Supplier").objects.get(id==supplier_id)

    filepath = get_local_filepath(bid_template)
    print(filepath)

    workbook = load_workbook(filepath, data_only=True)
    if len(workbook.worksheets) > 2:
        messages.append(
            "Please make sure the document you are uploading has the same format as the downloaded template"
        )
        context = {
            "result": "error",
            "response_message": messages,
        }
        return context
    
    if not "ITEMs" in workbook.sheetnames:
        messages.append(
            'Please make sure the document you are uploading has a worksheet with the name "ITEMs"'
        )
        context = {
            "result": "error",
            "response_message": messages,
        }
        return context
    
    try:
        worksheet = workbook["ITEMs"]
        old_response = None
        updated_items = False
        # capture submissions for the rest of the items
        for i in range(5, worksheet.max_row + 1):
            if (
                worksheet.cell(row=i, column=2).value != None
                and worksheet.cell(row=i, column=1).value != "Item ID"
            ):
                print(worksheet.cell(row=i, column=2).value)
                column_data = ""
                cell_data = ""
                for k in range(5, worksheet.max_column):
                    column_data += "%s," % str(
                        worksheet.cell(row=4, column=k).value
                    )
                    cell_data += "%s," % str(worksheet.cell(row=i, column=k).value)

                auction_item = AuctionItems.objects.filter(
                    auction=auction,
                    description=worksheet.cell(row=i, column=2).value,
                    pk=worksheet.cell(row=i, column=1).value,
                ).first()
                # if rfq_item is None:
                #     raise ValueError("Item mismatch")
                if auction_item is not None:
                    total = worksheet.cell(row=i, column=worksheet.max_column).value
                    item_number = worksheet.cell(row=i, column=1).value

                    # check for old response on the item and update or create a new one.
                    old_responses = AuctionItemResponses.objects.filter(
                        auction_item=auction_item, supplier=supplier
                    )
                    if old_responses.count() > 0:
                        old_response = old_responses.first()
                        if total is not None:
                            old_response.total = hash(total)
                        else:
                            old_response.total = hash(0)
                        old_response.cell_data = cell_data
                        old_response.column_data = column_data
                        old_response.save()
                        updated_items = True
                    else:
                        new_total = ""
                        if total is not None:
                            new_total = hash(total)
                        else:
                            new_total = hash(0)
                        item_response = AuctionItemResponses.objects.filter(
                            rfq_item=auction_item, supplier=supplier
                        ).first()
                        if item_response is not None:
                            item_response.bid_price = new_total
                            item_response.save()
                        else:
                            AuctionItemResponses.objects.create(
                                supplier=supplier,
                                auction_item=auction_item,
                                bid_price=new_total,
                            )
        # create supplier response
        messages.append(
            "Auction Item Responses created successfuly"
        )
        context = {
            "result": "success",
            "response_message": messages,
        }
        return context
    except Exception as e:
        print(e)
        # todo delete that excel from the server
        # "Rfq / 2019 / 08 / 24 / RFQ_Template_for_Ouma_Company_Buyers.xlsx"

        # dir_name = "media/%s" % excel_url.rsplit('/', 1)[0]
        # match_string = excel_url.rsplit('/', 1)[-1]
        # delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

        messages.append(
            "There is a problem with your uploaded Excel. Verify and retry."
        )
        context = {
            "result": "error",
            "response_message": messages,
        }
        return context

@shared_task(bind=True)
def import_auction_items(self, auction_id, auction_type, file_url):
    messages = []
    file = convertbase64tofile(file_url)
    # Write the file to a temporary location, deletion is guaranteed
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_file = os.path.join(tmp_dir, 'temp.xlsx')
        with open(tmp_file, 'wb') as f:
            f.write(file)
        # Process the file
        # filepath = f"{settings.BASE_DIR}{file_url}"
        workbook = load_workbook(tmp_file, data_only=True)
        worksheets = workbook.sheetnames[0]
        if len(worksheets) < 1:
            messages.append("There are no sheets to upload from the excel.")
            context = {"result": "error", "response_message": messages}
            return context

        items_count = 0

        try:

            category_sheet = workbook["ITEMs"]

            for i in range(7, category_sheet.max_row + 1):
                items_count += 1
                item_name = category_sheet["B{}".format(i)].value
                item_description = category_sheet["C{}".format(i)].value
                reserve_price = category_sheet["D{}".format(i)].value
                minimum_decrement_increment = category_sheet["E{}".format(i)].value

                print(f'item name: {item_name}')
                print(f'item_id: {auction_id}')
                # if AuctionItems.objects.filter(
                #     auction_id=auction_id, name=item_name, description=item_description
                # ).exists():
                try:
                    if auction_type == 'Reverse Auction':
                        auction_item, created = AuctionItems.objects.update_or_create(
                            auction_id=auction_id,
                            name=item_name,
                            defaults={
                                "reserve_price": reserve_price,
                                "minimum_decrement": minimum_decrement_increment,
                                "description":item_description
                            }
                        )
                        print(f'object: {auction_item}')
                        print(f'its created: {created}')
                    else:
                        auction_item, created = AuctionItems.objects.update_or_create(
                            auction_id=auction_id,
                            name=item_name,
                            defaults={
                                "reserve_price": reserve_price,
                                "minimum_increment": minimum_decrement_increment,
                                "description":item_description
                            }
                        )
                        print(f'object 2: {auction_item}')
                        print(f'its created 2: {created}')
                except Exception as e:
                    # capture_exception(e)
                    print(e)
                # else:
                #     messages.append({"error": "The item does not exists"})

            # os.remove(filepath)
            messages.append(
                "Imports succeeded. {} items added to auction.".format(
                    items_count
                )
            )
            context = {"result": "success", "response_message": messages}
            return context

        except Exception as e:
            # capture_exception(e)
            print(e)
            messages.append("Failed to auction items")
            context = {"result": "error", "response_message": messages}
            return context

@shared_task(bind=True)
def import_auction_bids(self, auction_type, file_url):
    # messages = []

    # filepath = f"{settings.BASE_DIR}{file_url}"
    # workbook = load_workbook(filepath, data_only=True)

    # worksheets = workbook.sheetnames[0]
    # if len(worksheets) < 1:
    #     messages.append("There are no sheets to upload from the excel.")
    #     context = {"result": "error", "response_message": messages}
    #     return context

    # items_count = 0

    # try:

    #     category_sheet = workbook["ITEMs"]

    #     for i in range(7, category_sheet.max_row + 1):
    #         items_count += 1
    #         item_id = category_sheet["C{}".format(i)].value
    #         item_name = category_sheet["D{}".format(i)].value
    #         item_description = category_sheet["E{}".format(i)].value
    #         reserve_price = category_sheet["F{}".format(i)].value
    #         minimum_decrement_increment = category_sheet["G{}".format(i)].value

    #         if AuctionItems.objects.filter(
    #             auction_id=item_id, name=item_name, item_description=item_description
    #         ).exists():
    #             try:
    #                 if auction_type == 'Reverse Auction':
    #                     rfq_item = AuctionItems.objects.update_or_create(
    #                         auction_id=item_id,
    #                         item_name=item_name,
    #                         item_description=item_description,
    #                         defaults={
    #                             "reserve_price": reserve_price,
    #                             "minimum_decrement": minimum_decrement_increment,
    #                         }
    #                     )
    #                 else:
    #                     rfq_item = AuctionItems.objects.update_or_create(
    #                         auction_id=item_id,
    #                         item_name=item_name,
    #                         item_description=item_description,
    #                         defaults={
    #                             "reserve_price": reserve_price,
    #                             "minimum_increment": minimum_decrement_increment,
    #                         }
    #                     )
    #             except Exception as e:
    #                 # capture_exception(e)
    #                 print(e)
    #         else:
    #             messages.append({"error": "The item does not exists"})

    #     os.remove(filepath)
    #     messages.append(
    #         "Imports succeeded. {} items added to auction.".format(
    #             items_count
    #         )
    #     )
    #     context = {"result": "success", "response_message": messages}
    #     return context

    # except Exception as e:
    #     # capture_exception(e)
    #     print(e)
    #     messages.append("Failed to auction items")
    #     context = {"result": "error", "response_message": messages}
    #     return context
    pass