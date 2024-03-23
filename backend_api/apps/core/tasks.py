import os
import time
from celery import shared_task
from celery.utils.log import get_task_logger
from contextlib import contextmanager

from celery_progress.backend import ProgressRecorder
from django import apps
from django.core.cache import cache
from hashlib import md5

from openpyxl import load_workbook
from sentry_sdk import capture_exception

from apps.core.models import CategoryType, CategoryTypeSupplier
from backend import settings

logger = get_task_logger(__name__)

LOCK_EXPIRE = 60 * 10  # Lock expires in 10 minutes

@contextmanager
def memcache_lock(lock_id, oid):
    timeout_at = time.monotonic() + LOCK_EXPIRE - 3
    # cache.add fails if the key already exists
    status = cache.add(lock_id, oid, LOCK_EXPIRE)
    try:
        yield status
    finally:
        # memcache delete is very slow, but we have to use it to take
        # advantage of using add() for atomic locking
        if time.monotonic() < timeout_at and status:
            # don't release the lock if we exceeded the timeout
            # to lessen the chance of releasing an expired lock
            # owned by someone else
            # also don't release the lock if we didn't acquire it
            cache.delete(lock_id)


@shared_task(bind=True)
def import_category_suppliers(self, file_url):
    messages = []

    filepath = f"{settings.BASE_DIR}{file_url}"
    workbook = load_workbook(filepath, data_only=True)

    worksheets = workbook.sheetnames[1:]
    if len(worksheets) < 1:

        context = {"messages": messages}
        messages.append("There are no sheets to upload from the excel.")
        return context

    categories_count = 0
    suppliers = 0

    progress_recorder = ProgressRecorder(self)
    result = 0

    for sheet in worksheets:
        result += 1
        progress_recorder.set_progress(result, len(worksheets))

        sheet_name_list = sheet.split("_")
        category_sheet = workbook[sheet]

        if CategoryType.objects.filter(innitials=sheet_name_list[-1]).count() == 1:
            category_type = CategoryType.objects.filter(
                innitials=sheet_name_list[-1]
            ).first()
            categories_count += 1

            for i in range(7, category_sheet.max_row + 1):
                # for each of the rows create a question after checking for existing section
                company_name = category_sheet["C{}".format(i)].value
                primary_email = category_sheet["E{}".format(i)].value
                alternative_email = category_sheet["F{}".format(i)].value
                primary_phone = category_sheet["G{}".format(i)].value
                alternative_phone = category_sheet["H{}".format(i)].value
                country = category_sheet["I{}".format(i)].value

                supplier = None
                supplier1 = None
                supplier2 = None

                if primary_email != None:
                    supplier1 = apps.apps.get_model('suppliers', 'Supplier').objects.filter(username=primary_email).first()
                elif alternative_email != None:
                    supplier2 = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                        username=alternative_email
                    ).first()
                else:
                    pass

                if supplier1 != None:
                    supplier = supplier1
                elif supplier2 != None:
                    supplier = supplier2
                else:
                    pass

                if supplier == None:
                    suppliers += 1
                    try:
                        category_supplier = (
                            CategoryTypeSupplier.objects.update_or_create(
                                primary_email=primary_email,
                                alternative_email=alternative_email,
                                category_type=category_type,
                                defaults={
                                    "primary_email": primary_email,
                                    "alternative_email": alternative_email,
                                    "primary_phone": primary_phone,
                                    "alternative_phone": alternative_phone,
                                    "category_type": category_type,
                                    "company_name": company_name,
                                    "country": country,
                                },
                            )
                        )
                    except Exception as e:
                        capture_exception(e)
                else:
                    suppliers += 1
                    try:
                        category_supplier = (
                            CategoryTypeSupplier.objects.update_or_create(
                                supplier=supplier,
                                category_type=category_type,
                                defaults={
                                    "supplier": supplier,
                                    "primary_email": primary_email,
                                    "alternative_email": alternative_email,
                                    "primary_phone": primary_phone,
                                    "alternative_phone": alternative_phone,
                                    "category_type": category_type,
                                    "company_name": company_name,
                                    "country": country,
                                },
                            )
                        )
                    except Exception as e:
                        capture_exception(e)

    os.remove(filepath)
    messages.append(
        "Imports succeeded. {} suppliers added to {} categories.".format(
            suppliers, categories_count
        )
    )
    context = {"messages": messages}
    return context


#Exmpale of using memcache_lock -- Ensures task is executed one at a time
# @task(bind=True)
# def import_feed(self, feed_url):
#     # The cache key consists of the task name and the MD5 digest
#     # of the feed URL.
#     feed_url_hexdigest = md5(feed_url).hexdigest()
#     lock_id = '{0}-lock-{1}'.format(self.name, feed_url_hexdigest)
#     logger.debug('Importing feed: %s', feed_url)
#     with memcache_lock(lock_id, self.app.oid) as acquired:
#         if acquired:
#             return Feed.objects.import_feed(feed_url).url
#     logger.debug(
#         'Feed %s is already being imported by another worker', feed_url)