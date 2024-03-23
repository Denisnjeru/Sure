import datetime
import fnmatch
import os
from pathlib import Path
from random import randint as j
from string import ascii_uppercase as a

import requests
from django.core.mail import get_connection
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.styles import Side, Alignment, Border, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter, column_index_from_string
from io import BytesIO
from django.http import HttpResponse
from django.template.loader import get_template, render_to_string
import xhtml2pdf.pisa as pisa
from backend.storage_backends import PrivateMediaStorage
from django.template.loader import render_to_string
from weasyprint import HTML, CSS
from django.conf import settings


SENDBLUE_API_KEY = "HDBUScRdLF5znwav"
default_from_email_sendinblue = "info@tendersure.co.ke"
my_host = "smtp-relay.sendinblue.com"
my_username = "njerudenis@qedsolutions.co.ke"  # this is exactly the value 'apikey'
my_password = SENDBLUE_API_KEY
my_port = 587
my_use_tls = True
default_from_email_sendinblue = "info@tendersure.co.ke"

connection = get_connection(
    host=my_host,
    port=my_port,
    username=my_username,
    password=my_password,
    use_tls=my_use_tls,
)


def ns(b):
    c = ''
    for d in range(b + 1): c += str(d)
    return c


def dm():
    e = ns(9);
    e += '.!#'
    return {0: a[:13], 1: a[:12:-1], 2: e}


def hash(g):
    k = len(str(g));
    i = str(j(0, 2));
    m = dm();
    n = '';
    p = 0
    for l in range(k - 1): i += str(j(0, 2))
    for o in str(i): n += m[int(o)][m[2].index((str(g)[p]))]; p += 1
    i += n
    return i


def show(s):
    t = s[:int(len(s) / 2)];
    u = s[int(len(s) / 2):];
    f = dm();
    k = '';
    q = 0
    for h in str(t): k += f[2][f[int(h)].index(str(u[int(q)]))]; q += 1
    return float(k)

def weasy_pdf(template_src, context_dict, file_path, file_name):
    Path(file_path).mkdir(parents=True, exist_ok=True)
    output_filename = "%s/%s" % (file_path, file_name)
    result_file = open(output_filename, "w+b")

    pdf_file = HTML(string=render_to_string(template_name=template_src, context=context_dict)).write_pdf(stylesheets=[CSS(string='body { font-family: Roboto !important; } @page {margin: 0.5cm;}')])
    result_file.write(pdf_file)
    result_file.close()
    return output_filename

class Render:

    @staticmethod
    def render(path: str, params: dict):
        html = get_template(path).render(params)
        response = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), response, encoding='UTF-8')
        if not pdf.err:
            return HttpResponse(response.getvalue(), content_type='application/pdf')
        else:
            return HttpResponse("Error Rendering PDF", status=400)


def delete_matching_files_in_directory(dir_name, match_string, extension=None):
    """delete matching files in a directory"""
    try:
        for file in os.listdir(dir_name):
            if fnmatch.fnmatch(file, "*" + match_string + "*" + extension):
                os.remove("{}/{}".format(dir_name, file))
    except:
        pass
    return True


def get_file_path(path):
    A = PrivateMediaStorage()
    headers = {"ResponseContentDisposition": f"attachment;"}
    time = datetime.datetime.now()
    file_url = A.url(f"{path}", expire=300, parameters=headers, http_method="GET")
    dir_name = Path(
        "media/temp/{}/{}/{}".format(time.year, time.month, time.day)
    )  # folder structure
    dir_name.mkdir(parents=True, exist_ok=True)
    file_name = os.path.basename(f"{path}")
    filepath = "{}/{}".format(dir_name, file_name)
    r = requests.get(file_url)
    with open("{}".format(filepath), "wb") as f:
        f.write(r.content)
    return filepath


def format_excel(excel_url, data=None):
    """Formats excel as per qed's specs. Remember to append these strings to the start of your cell values in order
    to apply the specified formats. After formatting, it calls resize cells. So if you intended to attach images, attach
    after this function"""
    """
    Key values:
    Nb_     No borders for this cell
    qT_     Quantitative total for this column
    h_      Cell in header row. Should apply heading format
    O_      Outlier value
    T_      Just a Total or any number that needs number formatting
    Tc_      Just a Total or any number that needs number formatting and center
    Tcnd_   Just a Total or any number that needs number formatting, center with no decimals
    Ac_     Align center
    B_      Bold
    Bnb_    Bold no borders
    Bac_    Bold align center
    Bar_    Bold align right
    Ar_     Align right
    hc_     Header anc center
    hl_     Header Left
    Qac_    Quality Assurance Comment
    hhc_    Header anc center vertical
    Pc_     Position value center
    Tw_     Totals or any number that needs number formatting but with warning fill
    """

    bd = Side(style="thin", color="000000")
    bd_double = Side(style="double", color="ff0000")
    align_right = Alignment(horizontal="right", vertical="top", wrapText=True)
    align_center = Alignment(horizontal="center", vertical="top", wrapText=True)
    align_left = Alignment(horizontal="left", wrapText=True)
    align_left_wrap_text = Alignment(horizontal="left", vertical="top", wrapText=True)
    center_bottom = Alignment(horizontal="center", vertical="bottom", wrapText=True)
    center_top = Alignment(horizontal="center", vertical="top", wrapText=True)
    left_top = Alignment(horizontal="left", vertical="top", wrapText=True)
    left_bottom = Alignment(horizontal="left", vertical="bottom", wrapText=True)
    workbook = load_workbook(excel_url)
    sheets = workbook.worksheets

    for worksheet in sheets:
        sheet_rows = worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
            min_col=2,
            max_col=worksheet.max_column,
        )
        for rows in sheet_rows:
            for cell in rows:
                cell_value = "{}".format(cell.value)
                if cell_value != "" and cell_value != "None":
                    if cell_value[:3] == "Nb_":
                        cell.value = cell_value[3:]
                    else:
                        cell.border = Border(top=bd, left=bd, right=bd, bottom=bd)
                        alignment = Alignment(vertical="top", wrapText=True)
                        cell.alignment = alignment
                if cell_value[:3] == "Ac_":
                    cell.value = cell_value[3:]
                    cell.alignment = align_center

                if cell_value[:4] == "Bnb_":
                    cell.value = cell_value[4:]
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(vertical="top", wrapText=True)
                    no_borders = Side(border_style=None)
                    no_border = Border(
                        left=no_borders,
                        right=no_borders,
                        top=no_borders,
                        bottom=no_borders,
                    )
                    cell.border = no_border

                if cell_value[:3] == "Pr_":
                    cell.value = cell_value[3:]
                    cell.alignment = align_right
                    cell.border = Border(
                        left=Side(border_style=None, color="FF000000"),
                        top=Side(border_style=None, color="FF000000"),
                        right=Side(border_style=None, color="FF000000"),
                        bottom=Side(border_style=None, color="FF000000"),
                    )

                if cell_value[:4] == "Bac_":
                    cell.value = cell_value[4:]
                    cell.alignment = align_center
                    cell.font = Font(bold=True)

                if cell_value[:2] == "B_":
                    cell.value = cell_value[2:]
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(vertical="top", wrapText=True)

                if cell_value[:4] == "Bar_":
                    cell.value = cell_value[4:]
                    cell.alignment = align_right
                    cell.font = Font(bold=True)

                if cell_value[:3] == "Ar_":
                    cell.value = cell_value[3:]
                    cell.alignment = align_right

                if cell_value[:3] == "Al_":
                    cell.value = cell_value[3:]
                    cell.alignment = left_top

                if cell_value[:3] == "qT_":
                    # cell.value = cell_value[3:]
                    try:
                        cell.value = float(cell_value[3:])
                    except:
                        cell.value = cell_value[3:]

                    cell.number_format = "#,##0.00_-"
                    cell.alignment = align_right
                    cell.font = Font(bold=True)

                if cell_value[:2] == "h_":
                    cell.value = cell_value[2:]
                    cell.fill = PatternFill(
                        start_color="DEEBF7", end_color="DEEBF7", fill_type="solid"
                    )
                    cell.alignment = left_bottom
                    cell.font = Font(color="8F1E22")
                    cell.font = Font(bold=True)

                if cell_value[:3] == "hc_":
                    cell.value = cell_value[3:]
                    cell.fill = PatternFill(
                        start_color="DEEBF7", end_color="DEEBF7", fill_type="solid"
                    )
                    cell.font = Font(color="8F1E22")
                    cell.font = Font(bold=True)
                    cell.alignment = center_bottom
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

                if cell_value[:3] == "hl_":
                    cell.value = cell_value[3:]
                    cell.fill = PatternFill(
                        start_color="DEEBF7", end_color="DEEBF7", fill_type="solid"
                    )
                    cell.font = Font(color="8F1E22")
                    cell.font = Font(bold=True)
                    cell.alignment = left_bottom
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

                if cell_value[:4] == "hhc_":
                    cell.value = cell_value[4:]
                    cell.fill = PatternFill(
                        start_color="DEEBF7", end_color="DEEBF7", fill_type="solid"
                    )

                    cell.font = Font(color="8F1E22")
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="bottom",
                        textRotation=90,
                        wrapText=True,
                    )
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

                if cell_value[:2] == "O_":
                    # cell.value = float(cell_value[2:])
                    try:
                        cell.value = float(cell_value[2:])
                    except:
                        cell.value = cell_value[2:]
                    cell.number_format = "#,##0.00_-"
                    cell.fill = PatternFill(
                        start_color="fcd5b5", end_color="fcd5b5", fill_type="solid"
                    )
                    cell.alignment = Alignment(vertical="top", wrapText=True)

                if cell_value[:2] == "T_":
                    try:
                        cell.value = float(cell_value[2:])
                    except:
                        cell.value = cell_value[2:]
                    cell.alignment = left_top
                    cell.number_format = "#,##0.00_-"
                    cell.alignment = Alignment(vertical="top", wrapText=True)

                if cell_value[:3] == "Tc_":
                    try:
                        cell.value = float(cell_value[3:])
                    except:
                        cell.value = cell_value[3:]
                    cell.number_format = "#,##0.00_-"
                    cell.alignment = center_top

                if cell_value[:4] == "Tcb_":
                    try:
                        cell.value = float(cell_value[4:])
                    except:
                        cell.value = cell_value[4:]
                    cell.number_format = "#,##0.00_-"
                    cell.alignment = center_top
                    cell.font = Font(bold=True)

                if cell_value[:3] == "Nc_":
                    try:
                        cell.value = float(cell_value[3:])
                    except:
                        cell.value = cell_value[3:]
                    cell.number_format = numbers.BUILTIN_FORMATS[3]
                    cell.alignment = center_top

                if cell_value[:3] == "Tr_":
                    try:
                        cell.value = float(cell_value[3:])
                    except:
                        cell.value = cell_value[3:]
                    cell.number_format = "#,##0.00_-"
                    cell.alignment = align_right

                if cell_value[:3] == "Tw_":
                    try:
                        cell.value = float(cell_value[3:])
                    except:
                        cell.value = cell_value[3:]
                    cell.number_format = "#,##0.00_-"
                    cell.fill = PatternFill(
                        start_color="FF8F00", end_color="FF8F00", fill_type="solid"
                    )
                    cell.alignment = Alignment(vertical="top", wrapText=True)
                if cell_value[:5] == "Tcnd_":
                    try:
                        cell.value = float(cell_value[5:])
                    except:
                        cell.value = cell_value[5:]
                    cell.number_format = "#,##0"
                    cell.alignment = center_top
                if cell_value[:4] == "Hmr_":
                    c = cell_value.split("_", 2)
                    cell.value = c[2]
                    cell.fill = PatternFill(
                        start_color="DEEBF7", end_color="DEEBF7", fill_type="solid"
                    )

                    cell.font = Font(color="8F1E22")
                    cell.font = Font(bold=True)
                    cell.alignment = center_top
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
                    try:
                        start = cell.coordinate
                        end = (
                            f"{get_column_letter(column_index_from_string(cell.column) + int(c[1]))}"
                            + f"{int(cell.row)}"
                        )
                        worksheet.merge_cells(f"{start}:{end}")
                    except Exception as e:
                        # capture_exception(e)
                        pass

                if cell_value[:5] == "Nmrl_":
                    c = cell_value.split("_", 2)
                    cell.value = c[2]
                    cell.font = Font(color="8F1E22")
                    cell.font = Font(bold=True)
                    cell.alignment = align_left_wrap_text
                    try:
                        start = cell.coordinate
                        end = (
                            f"{get_column_letter(column_index_from_string(cell.column) + int(c[1]))}"
                            + f"{int(cell.row)}"
                        )
                        worksheet.merge_cells(f"{start}:{end}")
                        cell.border = Border(top=bd, left=bd, right=bd, bottom=bd)
                    except Exception as e:
                        # capture_exception(e)
                        pass

                if cell_value[:4] == "Qac_":
                    c = cell_value.split("_", 1)
                    cell_value = c[1]
                    if cell_value.__contains__("("):
                        try:
                            c = cell_value.split("(", 1)
                            cell.value = c[0]
                            comment = f"{c[1].replace(')', '')}"
                            cell.comment = Comment("{}".format(comment), "Tendersure")
                        except Exception as e:
                            # capture_exception(e)
                            pass

    workbook.save(excel_url)
    try:
        resize_cells(excel_url=excel_url, data=data)
    except Exception as e:
        print(e)
    return excel_url


def resize_cells(excel_url, data=None):
    workbook = load_workbook(excel_url)
    worksheets = workbook.sheetnames
    if len(worksheets) > 0:
        for sheet in worksheets:
            current_worksheet = workbook[sheet]

            for col in current_worksheet.columns:
                length = max(len(str(cell.value)) for cell in col)
                if length > 62:
                    if data:
                        if f"{sheet}_{col[0].column_letter}" in data.keys():
                            current_worksheet.column_dimensions[
                                col[0].column_letter
                            ].width = data[f"{sheet}_{col[0].column_letter}"]
                        else:
                            current_worksheet.column_dimensions[
                                col[0].column_letter
                            ].width = 62
                    else:
                        current_worksheet.column_dimensions[col[0].column_letter].width = 62
                else:
                    if data:
                        if f"{sheet}_{col[0].column_letter}" in data.keys():
                            current_worksheet.column_dimensions[
                                col[0].column_letter
                            ].width = data[f"{sheet}_{col[0].column_letter}"]
                        elif f"{sheet}_OTHER" in data.keys():
                            current_worksheet.column_dimensions[
                                col[0].column_letter
                            ].width = data[f"{sheet}_OTHER"]
                        else:
                            current_worksheet.column_dimensions[
                                col[0].column_letter
                            ].width = length
                    else:
                        current_worksheet.column_dimensions[
                            col[0].column_letter
                        ].width = length
    workbook.save(excel_url)
    return excel_url


def insert_image(excel_url, worksheet_name, anchor, image_url=None, max_height=75):
    """Inserts an image into the excel at the specified anchor"""
    workbook = load_workbook(excel_url)
    worksheet = workbook[worksheet_name]
    if image_url is None:
        image_url = "static/core/img/Tendersure_Logo.png"
    image = Image(image_url)
    image_height, image_width = image.height, image.width
    ratio = image_height / max_height
    image.height, image.width = image_height / ratio, image_width / ratio
    image.anchor = worksheet[anchor]

    qed_logo = Image("static/core/img/Tendersure_Logo.png")
    q_logo_height, q_logo_width = qed_logo.height, qed_logo.width
    max_height = 75
    ratio = q_logo_height / max_height
    qed_logo.height, qed_logo.width = q_logo_height / ratio, q_logo_width / ratio
    qed_logo.anchor = worksheet["F1"]

    worksheet.add_image(image, anchor)
    worksheet.add_image(qed_logo, "F1")
    workbook.save(excel_url)

    return excel_url


def write_pdf(template_src, context_dict, file_path, file_name):
    Path(file_path).mkdir(parents=True, exist_ok=True)
    output_filename = "%s/%s" % (file_path, file_name)
    result_file = open(output_filename, "w+b")
    pisa_status = pisa.CreatePDF(
        render_to_string(template_src, context_dict), dest=result_file
    )
    result_file.close()
    return output_filename


def criteria_files(instance, filename):
    year = datetime.datetime.now().year
    category_type = instance.category_type
    return "%s/%s/%s" % (
        category_type.name.replace(" ", "_"),
        year,
        filename,
    )

def get_document_url(document_key):
    A = PrivateMediaStorage()
    headers = {"ResponseContentDisposition": f"attachment;"}
    response = A.url(
        f"{document_key}",
        expire=300,
        parameters=headers,
        http_method="GET",
    )

    return response


def get_local_filepath(file_url):
    media_root = settings.MEDIA_ROOT
    nw_media_root = media_root.split("/")[:-1]
    nw_url = "/".join(nw_media_root)
    return nw_url + file_url