import datetime
from pathlib import Path

from celery import shared_task
from django import apps
from openpyxl import Workbook
from sentry_sdk import capture_exception

from apps.core.utils import get_file_path, delete_matching_files_in_directory, format_excel, insert_image


def category_bidder_payments_report(category_id):
    category = apps.apps.get_model('prequal', 'Category').objects.filter(
        id=category_id).prefetch_related('prequalification').first()
    try:
        job = category.prequalification
        company = job.company
        time = datetime.datetime.now()

        dir_name = Path(
            "media/bidder_payments/%s/%s/%s"
            % (company.company_name, time.year, job.job_title)
        )  # folder structure
        dir_name.mkdir(parents=True, exist_ok=True)
        time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
        filepath = "{}/{}_{}_{}.xlsx".format(
            dir_name, job.job_code, job.job_title, time_only
        )

        # delete similar reports run this month
        match_string = "{}_{}".format(job.id, company.company_name.replace(" ", "_"))
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

        workbook = Workbook()

        category_order = apps.apps.get_model('prequal', 'CategoryOrder')
        orders = category_order.objects.filter(
            payment_status=category_order.PAID, category_id=category.id, target__model='prequalification'
        )
        paid_bidders = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
            id__in=orders.only('supplier_id').values('supplier_id')
        )
        suppliers = paid_bidders

        bidder_payments_worksheet = workbook.active
        bidder_payments_worksheet.title = "Bidder Payments"
        if company.company_logo_url is not None and company.company_logo_url != "":
            buyer_logo_url = get_file_path(company.company_logo_url)
        else:
            buyer_logo_url = "qed/static/img/no-company-image128-128.png"

        qed_logo_anchor = "K1"
        qed_logo_worksheet = bidder_payments_worksheet.title
        buyer_logo_anchor = "C1"
        buyer_logo_worksheet = bidder_payments_worksheet.title
        bidder_payments_worksheet["C5"] = "Nb_{}".format(category.name)
        bidder_payments_worksheet["C6"] = "Nb_{}".format(category.unique_reference)
        bidder_payments_worksheet["C7"] = "Nb_Bidder Payments"
        bidder_payments_worksheet["C8"] = "Nb_"

        bidder_payments_worksheet.append(
            ("", "hc_No", "h_Company Name", "h_Contact Person", "h_Email Address", "h_Phone Number")
        )

        for count, supplier in enumerate(suppliers, start=1):
            bidder_payments_worksheet.append(
                (
                    "",
                    "Ac_%d" % (count),
                    supplier.company_name,
                    supplier.contact_name,
                    supplier.email_address,
                    supplier.phone_number,
                )
            )

        workbook.save(filepath)
        format_excel(filepath)
        insert_image(
            excel_url=filepath,
            worksheet_name=qed_logo_worksheet,
            anchor=qed_logo_anchor,
        )
        insert_image(
            filepath, buyer_logo_worksheet, buyer_logo_anchor, image_url=buyer_logo_url
        )
        context = {
            "filepath": filepath,
            "response_message": "Report generated successfully",
        }
        return context
    except Exception as e:
        capture_exception(e)


@shared_task(bind=True)
def category_responsive_bidders_report(self, category_id):
    category = apps.apps.get_model('prequal', 'Category').objects.filter(id=category_id).first()
    time = datetime.datetime.now()
    try:
        suppliers = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                id__in=apps.apps.get_model('prequal', 'SupplierResponse').objects.filter(
                    question__section__category_id=category.id).only('supplier_id').values('supplier_id')
            )
        job = category.prequalification
        company = job.company

        dir_name = Path(
            "media/responsive_bidders/%s/%s/%s"
            % (company.company_name, time.year, job.job_title)
        )  # folder structure
        dir_name.mkdir(parents=True, exist_ok=True)
        time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
        filepath = "{}/{}_{}_{}.xlsx".format(
            dir_name, job.job_code, job.job_title, time_only
        )

        # delete similar reports run this month
        match_string = "{}_{}".format(job.id, company.company_name.replace(" ", "_"))
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

        workbook = Workbook()
        # create and save a workbook to later populate
        responsive_bidders_worksheet = workbook.active
        responsive_bidders_worksheet.title = "Responsive Bidders"
        if company.company_logo_url is not None and company.company_logo_url != "":
            buyer_logo_url = get_file_path(company.company_logo_url)
        else:
            buyer_logo_url = "qed/static/img/no-company-image128-128.png"

        qed_logo_anchor = "K1"
        qed_logo_worksheet = responsive_bidders_worksheet.title
        buyer_logo_anchor = "C1"
        buyer_logo_worksheet = responsive_bidders_worksheet.title
        responsive_bidders_worksheet["C5"] = "Nb_{}".format(category.name)
        responsive_bidders_worksheet["C6"] = "Nb_{}".format(category.unique_reference)
        responsive_bidders_worksheet["C7"] = "Nb_Responsive Bidders"
        responsive_bidders_worksheet["C8"] = "Nb_"

        responsive_bidders_worksheet.append(
            ("", "hc_No", "h_Company Name", "h_Contact Person",
             "h_Email Address", "h_Phone Number")
        )

        for count, supplier in enumerate(suppliers, start=1):
            responsive_bidders_worksheet.append(
                ("", "Ac_%d" % (count), supplier.company_name, supplier.contact_name,
                 supplier.email_address, supplier.phone_number)
            )

        workbook.save(filepath)
        format_excel(filepath)
        insert_image(
            excel_url=filepath,
            worksheet_name=qed_logo_worksheet,
            anchor=qed_logo_anchor,
        )
        insert_image(
            filepath, buyer_logo_worksheet, buyer_logo_anchor, image_url=buyer_logo_url
        )
    except Exception as e:
        capture_exception(e)


def category_non_responsive_bidders_report(self, category_id):
    try:
        category = apps.apps.get_model('prequal', 'Category').objects.filter(id=category_id).first()
        time = datetime.datetime.now()
        if category is not None:
            job = category.prequalification
            company = job.company

            supplier_response = apps.apps.get_model('prequal', 'SupplierResponse')
            category_order = apps.apps.get_model('core', 'CategoryOrder')

            responsive_bidders = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                id__in=supplier_response.objects.filter(
                    question__section__category_id=category.id).only('supplier_id').values('supplier_id')
            )
            paid_bidders = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                id__in=category_order.objects.filter(
                    payment_status=category_order.PAID, category_id=category.id, target__model='prequalification'
                ).only('supplier_id').values('supplier_id')
            )
            suppliers = paid_bidders.difference(responsive_bidders)

            dir_name = Path(
                "media/non_responsive_bidders/%s/%s/%s"
                % (company.company_name, time.year, job.job_title)
            )  # folder structure
            dir_name.mkdir(parents=True, exist_ok=True)
            time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
            filepath = "{}/{}_{}_{}.xlsx".format(
                dir_name, job.job_code, job.job_title, time_only
            )

            # delete similar reports run this month
            match_string = "{}_{}".format(job.id, company.company_name.replace(" ", "_"))
            delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

            # create and save a workbook to later populate
            workbook = Workbook()

            # create and save a workbook to later populate
            responsive_bidders_worksheet = workbook.active
            responsive_bidders_worksheet.title = "Non-Responsive Bidders"
            if company.company_logo_url is not None and company.company_logo_url != "":
                buyer_logo_url = get_file_path(company.company_logo_url)
            else:
                buyer_logo_url = "qed/static/img/no-company-image128-128.png"

            qed_logo_anchor = "K1"
            qed_logo_worksheet = responsive_bidders_worksheet.title
            buyer_logo_anchor = "C1"
            buyer_logo_worksheet = responsive_bidders_worksheet.title
            responsive_bidders_worksheet["C5"] = "Nb_{}".format(category.name)
            responsive_bidders_worksheet["C6"] = "Nb_{}".format(category.unique_reference)
            responsive_bidders_worksheet["C7"] = "Nb_Non-Responsive Bidders"
            responsive_bidders_worksheet["C8"] = "Nb_"

            responsive_bidders_worksheet.append(
                (
                    "",
                    "hc_No",
                    "h_Company Name",
                    "h_Contact Person",
                    "h_Email Address",
                    "h_Phone Number",
                )
            )

            for count, supplier in enumerate(suppliers, start=1):
                responsive_bidders_worksheet.append(
                    (
                        "",
                        "Ac_%d" % (count),
                        supplier.company_name,
                        supplier.contact_name,
                        supplier.email_address,
                        supplier.phone_number,
                    )
                )

            workbook.save(filepath)
            format_excel(filepath)
            insert_image(
                excel_url=filepath,
                worksheet_name=qed_logo_worksheet,
                anchor=qed_logo_anchor,
            )
            insert_image(
                filepath, buyer_logo_worksheet, buyer_logo_anchor, image_url=buyer_logo_url
            )

            context = {
                "filepath": filepath,
                "response_message": "Report generated successfully",
            }
            return context
    except Exception as e:
        capture_exception(e)