import datetime
import os
from pathlib import Path

from celery import Task, shared_task
from celery_progress.backend import ProgressRecorder
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.drawing.image import Image
from sentry_sdk import capture_exception
from django.utils.html import escape
from django.contrib.contenttypes.models import ContentType
from apps.core.models import CategoryOrder, CategoryTypeSupplier
from apps.core.utils import get_file_path, format_excel, insert_image, delete_matching_files_in_directory
from apps.prequal.models import *
from apps.prequal.utils import (
    get_locations, c_suppliers, non_responsive_bidders, get_supplier_data, get_directors, 
    get_cat_supplier_locations, ratio_scores_before_after_qa, ratios_before_after_qa, variance_ratios
)

class QARankingReport(Task):
    name = "QARankingReport"

    def run(self, *args, **kwargs):
        context = self.qa_ranking_report(kwargs["job_id"])
        return context

    def supplier_total_section_score(self, parent_section, section_scores):
        t = 0
        if parent_section.is_scored:
            score = section_scores.filter(section_id=parent_section.id).first()
            if score is not None:
                t += score.score_after_qa
            if parent_section.has_child_sections:
                for child_section in parent_section.child_sections:
                    t += self.supplier_total_section_score(child_section, section_scores)
        return t

    def questions_short_name(self):
        data = {
            "Tax compliance": "TCC", "PIN certificate": "PIN", "Certificate of Incorporation": "COI",
            "CR12 Certificate or equivalent": "CR12", "NCA certificate": "NCA-C", "NCA license": "NCA-L",
            "Institute of Certified Public Accountants of Kenya Certificate": "ICPAK",
            "Institute of  Certified Public Secretaries of Kenya": "ICPSK", "Registration certificate": "COI",
            "Energy Petroleum Regulatory Authority License": "EPRA", "County government business permit": "BP",
            "Pharmacy and poisons board License": "PPB", "Insurance and Regulatory Authority License": "IRA",
            "National identification number": "ID", "Police Clearance Certificate": "PCC",
            "Water Services Regulatory Board License": "WASREB",
            "National Industrial Training Authority Certificate": "NITA",
            "Certified Public Accountants of Kenya Certificate": "CPAK",
            "Association of Chartered Certified Accountants Ceritificate": "ACCA",
            "Law Society of Kenya Registration Certificate": "LSK",
            "National Transport and Safety Authority Registration Ceritificate": "NTSA",
            "Marketing Society of Kenya Certificate": "MSK",
            "Association of Practitioners in Advertising Certificate": "APA",
            "International Air Transport Association License": "IATA-L",
            "International Air Transport Association Certificate": "IATA-C",
            "Kenya Association of Travel Agents Certificate": "KATA", "Engineers Board of Kenya License": "EBK-L",
            "Engineers Board of Kenya certificate": "EBK-C", "Institution of Engineers of Kenya Certificate": "IEK",
            "Kenya Engineering Technology Registration Board License": "KETRB",
            "Board of Registration of Architects and Quantity Surveyors Certificate": "BORAQS-C",
            "Board of Registration of Architects and Quantity Surveyors License": "BORAQS-L",
            "Architectural Association of Kenya Certificate": "AAK",
            "Institute of Quantity Surveyors of Kenya Certificate": "IQSK",
            "Institute of Human Resources Management Certificate": "IHRM",
            "Directorate Of Occupational Safety And Health Services Certificate": "DOSH",
            "Marketing and Social Research Association Certificate": "MSRA",
            "Work Injury Benefits Act": "WIBA", "National Environmental Management Authority Certificate": "NEMA-C",
            "National Environmental Management Authority License": "NEMA-L", "Environmental Impact Assessment": "EIA",
            "Kenya Bureau of Standards": "KEBS", "Energy Regulatory Commission License": "ERC",
            "Institution of Surveyors of Kenya": "ISK", "Kenya Film Classification Board": "KFCB",
            "Estate Agents Registration Board": "EARB", "Valuers Registration Board": "VRB",
            "Authorised Economic Operator": "AEO",
            "Institute of Certified Safety and Security Proffesionals of Kenya": "ICSSPK",
            "Kenya Security Industry Association": "KSIA", "Protective Security Industry Association": "PSIA",
            "Private Security Regulatory Authority": "PSRA", "Kenya Animal Genetic Resources Centre": "KAGRC",
            "Kenya Veterinary Board": "KVB"

        }
        return data

    def qa_ranking_report(self, job_id):
        try:
            progress_recorder = ProgressRecorder(self)
            result = 0

            job = Prequalification.objects.get(id=job_id)
            company = job.company

            categories = Category.objects.filter(
                prequalification_id=job.id).only("id", "unique_reference", "name").order_by("unique_reference")

            time = datetime.datetime.now()
            dir_name = Path("media/prequal/job/reports/{}/{}/{}".format(time.year, time.month, time.day))

            dir_name.mkdir(parents=True, exist_ok=True)
            filepath = f"{dir_name}/{job_id}_QA_Ranking_Report_{time.hour}{time.minute}{time.second}.xlsx"

            workbook = Workbook()
            summary_sheet = workbook["Sheet"]
            summary_sheet.title = "Summary"

            buyer_logo_anchor = "B1"

            if company.company_logo_url is not None and company.company_logo_url != "":
                buyer_logo_url = get_file_path(company.company_logo_url)
            else:
                buyer_logo_url = "static/core/img/no-company-image128-128.png"

            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(
                (
                    " ", "hc_#", "hc_Category Code", "hl_Category Title", "hhc_Responsive",
                    "hhc_Non-Responsive", "hhc_Total Bidders", "hhc_Prequalified Bidders", "hhc_Letters",
                )
            )

            data = {
                "Summary_B": 5, "Summary_C": 21, "Summary_D": 41, "Summary_E": 6, "Summary_F": 6,
                "Summary_G": 6, "Summary_H": 6, "Summary_I": 6, "QA_OTHER_B": 41, "QA_OTHER_C": 8,
                "QA_OTHER_D": 8, "QA_OTHER_E": 8, "QA_OTHER_F": 8,
            }

            # totals
            total_responsive = 0
            total_non_responsive = 0
            total_bidders = 0
            total_prequalified = 0
            total_letters = 0
            total_category = categories.count()
            category_count = 1
            for category in categories:
                data[f"{category.unique_reference.replace('/', '_')}_C"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_D"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_E"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_F"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_G"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_H"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_I"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_J"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_K"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_L"] = 5
                # data[f"{category.unique_reference.replace('/', '_')}_M"] = 8
                data[f'{category.unique_reference.replace("/", "_")}_B'] = 31

                result += 1
                progress_recorder.set_progress(result, total_category)
                resp = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                    id__in=SupplierResponse.objects.filter(
                    question__section__category_id=category.id).only('supplier_id').values('supplier_id').distinct()
                ).count()

                all = CategoryOrder.objects.filter(
                    category_id=category.id, payment_status=CategoryOrder.PAID, target__model='category', 
                    target__app_label='prequalification').count()
                non_resp = all - resp

                if non_resp < 0:
                    non_resp = 0

                audited_accounts_question = Question.objects.filter(
                    description="For limited liability companies, attach audited accounts for the last two years, for "
                                "sole proprietors and partnerships, attach your most recent management accounts",
                    section__category_id=category.id).first()

                summary_sheet.append(
                    (
                        " ", f"Ac_{category_count}", f"Al_{category.unique_reference}",
                        f"Al_{category.name}",
                        f"Nc_{resp}",
                        f"Nc_{non_resp}",
                        f"Nc_{all}",
                        f"Nc_{category.suppliers_in_qa}",
                        f"Nc_{AwardLetter.objects.filter(category_id=category.id).count()}",
                    )
                )
                try:
                    total_responsive += resp
                    total_non_responsive += non_resp
                    total_bidders += all
                    total_prequalified += category.suppliers_in_qa
                    total_letters += AwardLetter.objects.filter(
                        category_id=category.id
                    ).count()
                except:
                    total_responsive = 0
                    total_non_responsive = 0
                    total_bidders = 0
                    total_prequalified = 0
                    total_letters = 0

                category_type = category.category_type
                category_count += 1

                category_worksheet = workbook.create_sheet(category.unique_reference.replace("/", "_"))
                category_worksheet.append(("", ""))
                category_worksheet["B2"] = f"Bnb_Category: {category.name}"
                category_worksheet["B3"] = ""
                category_worksheet["B4"] = "Number of bidders"
                category_worksheet["C4"] = f"Nc_{all}"
                category_worksheet["B5"] = "Number of pre-qualified bidders"
                category_worksheet["B6"] = "Number of responsive bidders"
                category_worksheet["C6"] = f"Nc_{resp}"
                category_worksheet["B7"] = "Pass mark"
                category_worksheet["C7"] = f"Tcnd_{category.pass_score}"
                category_worksheet.append(("", ""))

                category_headers = [
                    " ", "hl_Bidder", "hhc_Rank After QA", "hhc_Score After QA",
                    "hhc_Rank Before QA", "hhc_Score Before QA",
                ]

                category_worksheet.merge_cells("B2:H2")
                prequalified_bidders = 0

                short_descriptions_to_exclude = [
                    "Debt/equity ratio", "NP margin", "GP margin", "Cash ratio", "Current ratio"]

                questions = Question.objects.filter(
                    section__category_id=category.id, is_qa=True).exclude(
                    short_description__in=short_descriptions_to_exclude).only(
                    "id", "short_description", "description")

                sections = Section.objects.filter(
                    category_id=category.id, parent_section__isnull=True).order_by("id")

                self.append_category_headers(sections, category_headers, questions, category_worksheet)
                responsive_bidders = Supplier.objects.filter(
                    id__in=SupplierCategoryScore.objects.filter(
                    category_id=category.id).order_by("rank_after_qa").only("supplier_id").values('supplier_id')
                )

                for supplier in responsive_bidders:
                    supplier_prequal_response = SupplierCategoryScore.objects.filter(
                            supplier_id=supplier.id, category_id=category.id).first()
                    section_scores = SupplierSectionScore.objects.filter(
                        supplier_id=supplier.id, section__category_id=category.id)

                    prequal_score = supplier_prequal_response.score

                    if prequal_score < category.pass_score:
                        score_after_qa = "Ac_N/A"
                    else:
                        prequalified_bidders += 1
                        score_after_qa = (
                            f"Tcnd_{supplier_prequal_response.score_after_qa}"
                            if supplier_prequal_response.score_after_qa is not None
                            else "Ac_N/A"
                        )

                    c_data = {
                        "supplier": supplier, "supplier_prequal_response": supplier_prequal_response,
                        "sections": sections, "prequal_score": prequal_score,
                        "score_after_qa": score_after_qa, "section_scores": section_scores,
                        "category": category
                    }


                    if prequal_score >= category.pass_score:
                        # get section score for supplier over pass score
                        supplier_info = self.section_score_for_supplier_over_pass_score(data=c_data)
                    else:
                        # get section score for supplier under pass score
                        supplier_info = self.section_score_for_supplier_under_pass_score(data=c_data)

                    for question in questions:
                        qa_response = QualityAssuranceResponse.objects.filter(
                                supplier_id=supplier.id, quality_assurance_question__question_id=question.id,
                            ).only("id", "outcome", "comment").first()

                        if qa_response is not None:
                            supplier_info.extend([f"Qac_{qa_response.outcome} ({qa_response.comment})"])
                        else:
                            supplier_info.extend([" "])

                    try:
                        location = "N/A"
                        pass
                    except:
                        location = "N/A"

                    supplier_info.extend(
                        [
                            f"Al_{supplier.contact_name.title()}", f"Ar_{supplier.phone_number}", f"Al_{supplier.email}",
                            f"{ILLEGAL_CHARACTERS_RE.sub(r'', str(location))}",
                        ]
                    )


                    rank_comment = ""
                    if not prequal_score < float(category.pass_score):
                        section_score_instances = SupplierSectionScore.objects.filter(
                            supplier_id=supplier.id, section__category_id=category.id).prefetch_related('section')
                        for instance in section_score_instances:
                            if instance.score_after_qa != instance.score and instance.score_after_qa is not None:
                                difference = instance.score_after_qa - instance.score
                                if difference > 0:
                                    difference = f"+{difference}"
                                rank_comment = rank_comment + f"{instance.section.short_name}: {difference}, "

                    if audited_accounts_question:
                        ac_supplier_response = SupplierResponse.objects.filter(
                            question_id=audited_accounts_question.id, supplier_id=supplier.id).first()
                        if ac_supplier_response is None:
                            rank_comment = rank_comment + "Audited Accounts Not Attached, "
                        elif ac_supplier_response is not None:
                            if ac_supplier_response.options is None and ac_supplier_response.document_url is None:
                                rank_comment = rank_comment + "Audited Accounts Not Attached, "

                    if rank_comment != "":
                        rank_data = supplier_info[3]
                        supplier_info[3] = f"{rank_data}_{rank_comment}"

                    category_worksheet.append(supplier_info)
                # Append number of prequalified bidders
                category_worksheet["C5"] = f"Nc_{prequalified_bidders}"

                # non responsive bidders
                non_responsive_bidders = Supplier.objects.filter(
                    id__in=CategoryOrder.objects.filter(
                        category_id=category.id, payment_status=CategoryOrder.PAID,
                        target__model='category',  target__app_label='prequalification').exclude(
                        supplier_id__in=[r.id for r in responsive_bidders]).only("supplier_id")
                )

                for supplier in non_responsive_bidders:
                    supplier_info = [
                        "",
                        supplier.company_name.title(),
                        "Ac_N/R",
                        "Ac_N/R",
                        "Ac_N/R",
                        "Ac_N/R",
                    ]

                    for section in sections:
                        if section.has_child_sections:
                            if section.is_scored:
                                supplier_info.extend(["Ac_N/R"])
                            else:
                                evaluate_this = False
                                for s in section.child_sections:
                                    if s.is_scored:
                                        evaluate_this = True
                                        break
                                if evaluate_this == True:
                                    supplier_info.extend(["Ac_N/R"])
                        else:
                            if section.is_scored:
                                supplier_info.extend(["Ac_N/R"])

                    for question in questions:
                        qa_response = (
                            QualityAssuranceResponse.objects.filter(
                                supplier_id=supplier.id,quality_assurance_question__question_id=question.id,
                            ).only("id", "outcome", "comment").first()
                        )
                        if qa_response is not None:
                            supplier_info.extend(["Ac_N/R"])
                        else:
                            supplier_info.extend(["Ac_N/R"])
                    supplier_info.extend(["Ac_N/R", "Ac_N/R", "Ac_N/R", "Ac_N/R"])
                    category_worksheet.append(supplier_info)

                # append notes
                category_worksheet.append(("", ""))
                category_worksheet.append(("", "Notes"))
                category_worksheet.append(("", "N/A - Not Applicable"))
                category_worksheet.append(("", "N/R - None responsive"))
                category_worksheet.freeze_panes = category_worksheet["D10"]

            summary_sheet.append(
                (
                    " ", "", "", "qT_Total", "Ac_{}".format(total_responsive),
                    "Ac_{}".format(total_non_responsive), "Ac_{}".format(total_bidders),
                    "Ac_{}".format(total_prequalified), "Ac_{}".format(total_letters),
                )
            )

            summary_sheet.freeze_panes = summary_sheet["D10"]
            workbook.save(filepath)
            format_excel(filepath, data=data)

            # insert_image(excel_url=filepath, worksheet_name="Summary", anchor=qed_logo_anchor)
            try:
                insert_image(
                    excel_url=filepath, worksheet_name="Summary", anchor=buyer_logo_anchor, image_url=buyer_logo_url)
            except:
                insert_image(
                    excel_url=filepath, worksheet_name="Summary", anchor=buyer_logo_anchor, image_url="",
                )

            try:
                time = datetime.datetime.now()
                with open(filepath, "rb") as l:
                    storage = PrivateMediaStorage()
                    url = f"prequal/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
                    storage.save(url, l)

                    job_report, created = JobReport.objects.update_or_create(
                        prequalification_id=job_id, defaults={"qa_ranking_report": url}
                    )
                    path = job_report.qa_ranking_report.url
                    context = {
                        "filepath": path,
                        "response_message": "Report generated successfully",
                    }
                    return context
            except Exception as e:
                print(e)
                capture_exception(e)
                context = {"response_message": "Error generating qa ranking report"}
                return context
        except Exception as e:
            print(str(e))
            # capture_exception(e)
            context = {"response_message": "Error generating qa ranking report", "errors": f"{str(e)}"}
            return context

    def append_category_headers(self, prequal_sections, category_headers, questions, category_worksheet):
        question_short_names = self.questions_short_name()
        for section in prequal_sections:
            if section.has_child_sections:
                if section.is_scored:
                    s_total = section.section_score
                    for s in section.child_sections:
                        if s.has_child_sections:
                            for x in s.child_sections:
                                s_total += x.section_score
                        else:
                            s_total += s.section_score
                    category_headers.extend(
                        [f"hhc_{section.short_name}({int(s_total)})"]
                    )
                else:
                    evaluate_this = False
                    for s in section.child_sections:
                        if s.has_child_sections:
                            for x in s.child_sections:
                                if x.is_scored:
                                    evaluate_this = True
                                    break
                        else:
                            if s.is_scored:
                                evaluate_this = True
                                break
                    if evaluate_this == True:
                        s_total = 0
                        for s in section.child_sections:
                            if s.has_child_sections:
                                for x in s.child_sections:
                                    s_total += x.section_score
                            else:
                                s_total += s.section_score
                        category_headers.extend(
                            [f"hhc_{section.short_name}({int(s_total)})"]
                        )
            elif section.is_scored:
                category_headers.extend(
                    [
                        f"hhc_{section.short_name}({int(section.section_score)})"
                    ]
                )

        for question in questions:
            if question.short_description is not None:
                if question.short_description in question_short_names:
                    category_headers.extend(
                        [f"hl_{question_short_names[question.short_description]}"]
                    )
                else:
                    category_headers.extend(
                        [f"hl_{question.short_description}"]
                    )
            else:
                category_headers.extend([f"hl_{question.description}"])

        category_headers.extend(
            ["hl_Contact Person", "hl_Phone Number", "hl_Email Address", "hl_Main Location", ]
        )
        category_worksheet.append(category_headers)
        category_worksheet.merge_cells("B2:H2")

        return

    def section_score_for_supplier_over_pass_score(self, data):
        print("got here")
        sections = data["sections"]
        supplier = data["supplier"]
        supplier_prequal_response = data["supplier_prequal_response"]
        prequal_score = data["prequal_score"]
        score_after_qa = data["score_after_qa"]
        category = data["category"]
        section_scores = data["section_scores"]

        supplier_info = [
            "",
            supplier.short_name if supplier.short_name is not None else supplier.company_name, 
            f"Ac_{supplier_prequal_response.rank_after_qa}"
            if supplier_prequal_response.rank_after_qa and not prequal_score < category.pass_score
            else "Ac_N/A",
            score_after_qa, f"Nc_{supplier_prequal_response.rank}", f"Tcnd_{prequal_score}",
        ]
        for section in sections:
            if section.is_parent_section_scored:
                section_total = self.supplier_total_section_score(section, section_scores)
                supplier_info.extend([f"Tcnd_{section_total}"])
        print(supplier_info)
        return supplier_info

    def section_score_for_supplier_under_pass_score(self, data):
        sections = data["sections"]
        supplier = data["supplier"]
        supplier_prequal_response = data["supplier_prequal_response"]
        prequal_score = data["prequal_score"]
        score_after_qa = data["score_after_qa"]
        category = data["category"]
        section_scores = data["section_scores"]

        supplier_info = [
            "",
            supplier.short_name if supplier.short_name is not None else supplier.company_name,
            f"Ac_{supplier_prequal_response.rank_after_qa}"
            if supplier_prequal_response.rank_after_qa and not prequal_score < category.pass_score
            else "Ac_N/A",
            score_after_qa, 
            f"Nc_{supplier_prequal_response.rank}", f"Tcnd_{prequal_score}",
        ]
        print(supplier_info)
        for section in sections:
            if section.is_parent_section_scored:
                section_total = self.supplier_total_section_score(section, section_scores)
                supplier_info.extend([f"Tcnd_{section_total}"])
        print(supplier_info)
        return supplier_info


class QAReport(Task):
    name = "QAReport"
    progress_recorder = None
    result = 0

    def run(self, *args, **kwargs):
        self.progress_recorder = ProgressRecorder(self)
        context = self.report(kwargs["job_id"])
        return context

    def report(self, job_id):
        try:
            job = Prequalification.objects.get(id=job_id)
            time = datetime.datetime.now()
            company = job.company
            dir_name = Path("media/qa_reports/{}/{}".format(time.year, time.month))
            dir_name.mkdir(parents=True, exist_ok=True)
            time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
            filepath = "{}/{}_{}_{}.xlsx".format(
                dir_name, job_id, job.company.company_name.replace(" ", "_"), time_only
            )

            # delete old files that were run in the same month
            match_string = "{}_{}".format(
                job_id, job.company.company_name.replace(" ", "_")
            )
            delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

            workbook = Workbook()
            summary_sheet = workbook["Sheet"]
            summary_sheet.title = "Summary"
            categories = Category.objects.filter(job=job)

            summary_sheet.append((" ", " ", ""))
            summary_sheet.append((" ", " ", ""))
            summary_sheet.append((" ", " ", ""))
            summary_sheet.append((" ", " ", ""))
            summary_sheet.append(
                (
                    " ",
                    "hc_#",
                    "hc_Category Code",
                    "hc_Category Name",
                    "hc_Responsive Bidders",
                    "hc_Non-Responsive Bidders",
                    "hc_Total Bidders",
                    "hc_QA Bidders",
                )
            )
            summary_sheet.freeze_panes = summary_sheet["A7"]

            number = 1
            # progress_recorder = ProgressRecorder(self)
            # result = 0

            qed_logo_anchor = "H1"
            buyer_logo_anchor = "B1"

            if company.company_logo_url is not None and company.company_logo_url != "":
                buyer_logo_url = get_file_path(company.company_logo_url)
            else:
                buyer_logo_url = "qed/static/img/no-company-image128-128.png"

            category_count = categories.count()
            for category in categories:
                self.result += 1
                self.progress_recorder.set_progress(self.result, category_count)

                resp = (
                    SupplierResponse.objects.filter(
                        question__section__name="Declaration",
                        question__section__category_id=category.id,
                    )
                    .values_list("supplier", flat=True)
                    .distinct()
                    .count()
                )

                all = CategoryOrder.objects.filter(
                    category_id=category.id, payment_status=CategoryOrder.PAID
                ).count()

                non_resp = all - resp
                category.update_status_open()
                summary_sheet.append(
                    (
                        " ",
                        number,
                        category.unique_reference,
                        category.name,
                        resp,
                        non_resp,
                        all,
                        category.suppliers_in_qa,
                    )
                )
                number += 1

                supplier_responses = SupplierResponse.objects.filter(
                    question__section__category=category, question__is_qa=True
                )
                list_suppliers = []
                for supplier_response in supplier_responses:
                    supplier = supplier_response.supplier
                    list_suppliers.append(supplier)
                list_suppliers = set(list_suppliers)

                worksheet = workbook.create_sheet(
                    category.unique_reference.replace("/", "_")
                )
                worksheet["B4"] = "Category: {}".format(category.name)
                supplier_headers = ["", "", "", ""]

                for supplier in list_suppliers:
                    supplier_headers.append(f"hc_{supplier.company_name.upper()}")
                    supplier_headers.append("")
                    supplier_headers.append("")
                    supplier_headers.append("")

                worksheet.append(supplier_headers)

                sections = Section.objects.filter(category=category)
                data_rows = []
                for section in sections:
                    if section.has_qa_questions:
                        section_headers = [
                            " ",
                            f"hc_{section.short_name}",
                            "hc_Max Score",
                            "hc_Verify Document",
                        ]
                        for supplier in list_suppliers:
                            section_headers.append("hc_Number")
                            section_headers.append("hc_Date")
                            section_headers.append("hc_Comment")
                            section_headers.append("hc_Score After QA")

                        data_rows.append(tuple(section_headers))

                        qa_questions = QualityAssuranceQuestion.objects.filter(
                            question__section=section
                        )
                        for qa_question in qa_questions:
                            if qa_question.question.max_score is not None:
                                question_row = [
                                    " ",
                                    qa_question.question.short_description,
                                    int(qa_question.question.max_score),
                                    qa_question.verification_instruction,
                                ]
                            else:
                                question_row = [
                                    " ",
                                    qa_question.question.short_description,
                                    "0",
                                    qa_question.verification_instruction,
                                ]

                            for supplier in list_suppliers:
                                supplier_qa_response = supplier.qa_question_response(
                                    question_id=qa_question.question_id
                                )
                                if supplier_qa_response is not None:
                                    question_row.append(supplier_qa_response.number)
                                    if supplier_qa_response.date is not None:
                                        question_row.append(
                                            supplier_qa_response.date.date()
                                        )
                                    else:
                                        question_row.append(" ")
                                    question_row.append(supplier_qa_response.outcome)
                                    if supplier_qa_response.score_after_qa is not None:
                                        question_row.append(
                                            int(supplier_qa_response.score_after_qa)
                                        )
                                    else:
                                        question_row.append(int("0"))
                                else:
                                    question_row.extend([" ", " ", " ", " "])

                            data_rows.append(question_row)

                for row in data_rows:
                    worksheet.append(row)

            workbook.save(filepath)
            format_excel(filepath)
            insert_image(
                excel_url=filepath, worksheet_name="Summary", anchor=qed_logo_anchor
            )
            try:
                insert_image(
                    excel_url=filepath,
                    worksheet_name="Summary",
                    anchor=buyer_logo_anchor,
                    image_url=buyer_logo_url,
                )
            except:
                insert_image(
                    excel_url=filepath,
                    worksheet_name="Summary",
                    anchor=buyer_logo_anchor,
                    image_url="",
                )

            try:
                time = datetime.datetime.now()
                with open(filepath, "rb") as l:
                    storage = PrivateMediaStorage()
                    url = f"prequal/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
                    storage.save(url, l)
                    print(url)
                    job_report, created = JobReport.objects.update_or_create(
                        job_id=job_id, defaults={"qa_summary_excel": url}
                    )
                    context = {
                        "filepath": job_report.qa_ranking_excel.url,
                        "job_id": job_id,
                        "response_message": "Report generated successfully",
                    }
                    return context
            except Exception as e:
                capture_exception(e)
                context = {"response_message": "Error generating qa ranking report"}
                return context
        except Exception as e:
            capture_exception(e)


class DueDiligenceRankingReport(Task):
    name = "DueDiligenceRankingReport"

    def run(self, *args, **kwargs):
        context = self.report(kwargs["job_id"])
        return context

    def report(self, job_id):
        try:
            progress_recorder = ProgressRecorder(self)
            result = 0

            data = {}
            job = Prequalification.objects.get(id=job_id)
            categories = Category.objects.filter(prequalification_id=job_id)
            total_for_progress = categories.count()
            company = job.company

            time = datetime.datetime.now()
            dir_name = Path("media/prequal/job/reports/{}/{}".format(time.year, time.month))
            dir_name.mkdir(parents=True, exist_ok=True)
            # time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
            filepath = "{}/{}_DD_Ranking_Report.xlsx".format(
                dir_name, job_id
            )

            workbook = Workbook()
            summary_sheet = workbook.get_sheet_by_name("Sheet")
            summary_sheet.title = "Summary"

            qed_logo_anchor = "G2"
            buyer_logo_anchor = "B2"

            if company.company_logo_url is not None and company.company_logo_url != "":
                buyer_logo_url = get_file_path(company.company_logo_url)
            else:
                buyer_logo_url = "static/core/img/no-company-image128-128.png"

            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))

            summary_sheet.append(
                (
                    " ", "hc_#", "hc_Category No", "hc_Category Title",
                    "hhc_Responsive", "hhc_Non-Responsive", "hhc_Total Bidders",
                    "hhc_Prequalified Bidders QA", "hhc_Due Diligence", "hhc_Success/Regret Letters",
                )
            )

            additional_data = {
                "Summary_C": 21, "Summary_D": 62, "Summary_E": 6.2, "Summary_F": 6.2,
                "Summary_G": 6.2, "Summary_H": 6.2, "Summary_I": 6.2, "Summary_J": 6.2
            }
            data.update(additional_data)
            # data["Summary_C"] = 21
            # data["Summary_D"] = 62
            # data["Summary_E"] = 6.2
            # data["Summary_F"] = 6.2
            # data["Summary_G"] = 6.2
            # data["Summary_H"] = 6.2
            # data["Summary_I"] = 6.2
            # data["Summary_J"] = 6.2

            responsive_bidders = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                id__in=SupplierResponse.objects.filter(
                    question__section__category__prequalification_id=job_id).only('supplier_id').values('supplier_id')
            )
            paid_bidders = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                id__in=CategoryOrder.objects.filter(
                    payment_status=CategoryOrder.PAID, job_id=job_id, target__model='prequalification'
                ).only('supplier_id').values('supplier_id')
            )
            non_responsive_bidders = paid_bidders.difference(responsive_bidders)

            responsive_bidders_count = responsive_bidders.count()
            non_responsive_bidders_count = paid_bidders.count() - responsive_bidders.count()
            total_bidders = paid_bidders.count()

            suppliers_in_qa = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                id__in=QualityAssuranceResponse.objects.filter(
                    quality_assurance_question__quality_assurance__category__prequalification_id=job_id
                ).only('supplier_id').values('supplier_id')
            ).count()

            dd_suppliers = DueDiligenceSupplier.objects.filter(
                due_diligence__category__prequalification_id=job_id).count()

            qa_and_dd_suppliers = (AwardLetter.objects.filter(category__prequalification_id=job_id).count()
                                   + RegretLetter.objects.filter(category__prequalification_id=job_id).count())

            section_scores = SupplierSectionScore.objects.filter(section__category__prequalification_id=job_id)
            category_count = 1
            for category in categories:
                result += 1
                progress_recorder.set_progress(result, total_for_progress)

                summary_sheet.append(
                    (
                        "", category_count, category.unique_reference, category.name,
                        responsive_bidders_count, non_responsive_bidders_count,
                        total_bidders, suppliers_in_qa, dd_suppliers, qa_and_dd_suppliers
                    )
                )
                category_count += 1

                category_worksheet = workbook.create_sheet(
                    category.unique_reference.replace("/", "_")
                )
                category_worksheet.append((" ", " "))
                category_worksheet["B2"] = f"Category: {category.name}"
                category_worksheet["B3"] = "Number of bidders"
                category_worksheet["C3"] = f"Ac_{category.total_bidders}"
                category_worksheet["B4"] = "Number of pre-qualified bidders"
                category_worksheet["C4"] = f"Ac_{category.count_qualified_bidders}"
                category_worksheet["B5"] = "Number of responsive bidders"
                category_worksheet["C5"] = f"Ac_{category.responsive_bidder_count}"
                category_worksheet["B6"] = "Pass mark"
                category_worksheet["C6"] = f"Tcnd_{category.pass_score}"
                category_worksheet.append(("", ""))

                main_headers = ["", "", "hc_Prequalification Results", "", ""]

                category_headers = [
                    " ", "hc_Bidder", "hhc_Rank", "hhc_Total", "hhc_Score After QA",
                ]

                data[f'{category.unique_reference.replace("/", "_")}_B'] = 36

                questions = Question.objects.filter(
                    section__category_id=category.id, is_qa=True).order_by("id")

                sections = category.sections
                no_of_sections = 0
                for section in sections:
                    if section.is_scored:
                        no_of_sections += 1
                        category_headers.append(f"hhc_{section.short_name}")
                        main_headers.append("")
                main_headers.append("hc_Quality Assurance")

                no_of_questions = 0
                for question in questions:
                    no_of_questions += 1
                    category_headers.append(f"hhc_{question.short_description}")
                    main_headers.append("")
                main_headers.append("hhc_Tender Committee")

                category_worksheet.append(main_headers)
                category_worksheet.append(category_headers)
                category_worksheet.merge_cells(
                    start_row=8,
                    start_column=3,
                    end_row=8,
                    end_column=5 + no_of_sections,
                )

                category_worksheet.merge_cells(
                    start_row=8,
                    start_column=6 + no_of_sections,
                    end_row=8,
                    end_column=6 + no_of_sections + no_of_questions,
                )

                sorted_participants = apps.apps.get_model('prequal', 'SupplierCategoryScore').objects.filter(
                    category_id=category.id).prefetch_related('supplier').order_by("rank")

                for sorted_participant in sorted_participants:
                    p_score = sorted_participant.score
                    s_qa = sorted_participant.score_after_qa
                    supplier = sorted_participant.supplier
                    supplier_info = [
                        " ",
                        supplier.company_name.upper(),
                        f"Ac_{sorted_participant.rank}",
                        f"Tc_{p_score}" if p_score is not None else " ",
                        f"Tc_{s_qa}" if s_qa is not None else " ",
                    ]

                    for section in sections:
                        section_score = section_scores.filter(id=section.id, supplier_id=supplier.id).first()
                        s_score = section_score.score if section_score is not None else 0
                        if section.is_scored:
                            supplier_info.extend(
                                [f"Tcnd_{s_score}" if s_score is not None else " "]
                            )

                    for question in questions:
                        qa_response = QualityAssuranceResponse.objects.filter(
                            supplier=supplier,
                            quality_assurance_question__question=question,
                        ).first()
                        if qa_response is not None:
                            supplier_info.extend(
                                [qa_response.outcome + "(" + qa_response.comment + ")"]
                            )
                        else:
                            supplier_info.extend([" "])
                    category_worksheet.append(supplier_info)
                category_worksheet.freeze_panes = category_worksheet["F10"]
            summary_sheet.freeze_panes = summary_sheet["E7"]
            workbook.save(filepath)
            format_excel(filepath, data=data)
            insert_image(
                excel_url=filepath, worksheet_name="Summary", anchor=qed_logo_anchor
            )

            try:
                insert_image(
                    excel_url=filepath,
                    worksheet_name="Summary",
                    anchor=buyer_logo_anchor,
                    image_url=buyer_logo_url,
                )
            except:
                insert_image(
                    excel_url=filepath,
                    worksheet_name="Summary",
                    anchor=buyer_logo_anchor,
                    image_url="",
                )

            time = datetime.datetime.now()
            with open(filepath, "rb") as l:
                storage = PrivateMediaStorage()
                url = f"prequal/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
                storage.save(url, l)

                job_report, created = JobReport.objects.update_or_create(
                    prequalification_id=job_id, defaults={"dd_ranking_report": url}
                )
                path = job_report.dd_ranking_report.url
                context = {
                    "filepath": path,
                    "response_message": "Report generated successfully",
                }
                return context
        except Exception as e:
            capture_exception(e)
            print(e)
            context = {
                "response_message": "Report generation error",
                "messages": [f"{e}",]
            }
            return context


# class KraComplianceReport(Task):
#     name = "KraComplianceReport"
#     progress_recorder = None
#     result = 0
#
#     def run(self, *args, **kwargs):
#         self.progress_recorder = ProgressRecorder(self)
#         context = self.report()
#         return context
#
#     def report(self):
#         try:
#             data = {}
#             time = datetime.datetime.now()
#             dir_name = Path(
#                 "media/document_update/reports/{}/{}".format(time.year, time.month)
#             )
#             dir_name.mkdir(parents=True, exist_ok=True)
#             time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
#             filepath = "{}/document_update_report_{}.xlsx".format(dir_name, time_only)
#             data["KRA Compliance Report_B"] = 6
#             data["KRA Compliance Report_C"] = 62
#             data["KRA Compliance Report_D"] = 15
#             data["KRA Compliance Report_E"] = 15
#             data["KRA Compliance Report_G"] = 40
#             data["KRA Compliance Report_H"] = 40
#             data["KRA Compliance Report_I"] = 40
#             data["KRA Compliance Report_J"] = 40
#
#             workbook = Workbook()
#             summary_sheet = workbook.get_sheet_by_name("Sheet")
#             summary_sheet.title = "KRA Compliance Report"
#
#             qed_logo_anchor = "H1"
#             buyer_logo_anchor = "B1"
#             today = datetime.date.today()
#             summary_sheet.append(("", ""))
#             summary_sheet.append(("", ""))
#             summary_sheet.append(("", ""))
#             summary_sheet.append(("", ""))
#             summary_sheet.append(("", ""))
#             summary_sheet.append(
#                 (
#                     "",
#                     "hc_#",
#                     "hl_Supplier",
#                     "hc_Expiry Date",
#                     "hl_Status",
#                     "hc_Renewal Date",
#                     "hl_Client 1",
#                     "hl_Client 2",
#                     "hl_Client 3",
#                     "hl_Client 4",
#                 )
#             )
#             suppliers = Supplier.objects.filter(
#                 id__in=SupplierResponse.objects.all()
#                 .only("supplier_id")
#                 .values("supplier_id")
#                 .distinct("supplier_id")
#             )
#             total_for_progress = suppliers.count()
#             count = 1
#             for supplier in suppliers:
#                 self.result += 1
#                 self.progress_recorder.set_progress(self.result, total_for_progress)
#                 current_profile = SupplierProfile.objects.filter(
#                     supplier_id=supplier.id
#                 ).last()
#                 # previous_profile = SupplierProfile.objects.filter(supplier_id=supplier.id).order_by('-id')[1]
#
#                 if current_profile:
#                     companies = Company.objects.filter(
#                         id__in=SupplierResponse.objects.filter(
#                             supplier_id=supplier.id
#                         ).values("question__section__category__job__company_id")
#                     ).order_by("-id")
#                     try:
#                         company_one = companies[0].company_name
#                     except:
#                         company_one = "N/A"
#
#                     try:
#                         company_two = companies[1].company_name
#                     except:
#                         company_two = "N/A"
#
#                     try:
#                         company_three = companies[2].company_name
#                     except:
#                         company_three = "N/A"
#
#                     try:
#                         company_four = companies[3].company_name
#                     except:
#                         company_four = "N/A"
#
#                     if current_profile.kra_compliance_expiry_date:
#                         if current_profile.kra_compliance_expiry_date < today:
#                             supplier_row = [
#                                 "",
#                                 count,
#                                 supplier.company_name,
#                                 str(current_profile.kra_compliance_expiry_date),
#                                 "Expired",
#                                 "N/A",
#                                 company_one,
#                                 company_two,
#                                 company_three,
#                                 company_four,
#                             ]
#                         elif current_profile.kra_compliance_expiry_date > today:
#                             supplier_row = [
#                                 "",
#                                 count,
#                                 supplier.company_name,
#                                 str(current_profile.kra_compliance_expiry_date),
#                                 "Upto Date",
#                                 current_profile.created_at.strftime("%Y-%m-%d"),
#                                 company_one,
#                                 company_two,
#                                 company_three,
#                                 company_four,
#                             ]
#                         elif current_profile.kra_compliance_expiry_date == today:
#                             supplier_row = [
#                                 "",
#                                 count,
#                                 supplier.company_name,
#                                 str(current_profile.kra_compliance_expiry_date),
#                                 "Expires Today",
#                                 "N/A",
#                                 company_one,
#                                 company_two,
#                                 company_three,
#                                 company_four,
#                             ]
#
#                         summary_sheet.append(supplier_row)
#                         count += 1
#                     else:
#                         supplier_row = [
#                             "",
#                             count,
#                             supplier.company_name,
#                             "N/A",
#                             "N/A",
#                             "N/A",
#                             "N/A",
#                             "N/A",
#                             "N/A",
#                             "N/A",
#                         ]
#                         summary_sheet.append(supplier_row)
#                         count += 1
#                 else:
#                     supplier_row = [
#                         "",
#                         count,
#                         supplier.company_name,
#                         "N/A",
#                         "N/A",
#                         "N/A",
#                         "N/A",
#                         "N/A",
#                         "N/A",
#                         "N/A",
#                     ]
#                     summary_sheet.append(supplier_row)
#                     count += 1
#
#             summary_sheet.freeze_panes = summary_sheet["G7"]
#             workbook.save(filepath)
#             format_excel(filepath, data=data)
#
#             try:
#                 insert_image(
#                     excel_url=filepath,
#                     worksheet_name="KRA Compliance Report",
#                     anchor=buyer_logo_anchor,
#                 )
#             except:
#                 insert_image(
#                     excel_url=filepath,
#                     worksheet_name="KRA Compliance Report",
#                     anchor=buyer_logo_anchor,
#                     image_url="",
#                 )
#             # insert_image(
#             #     excel_url=filepath,
#             #     worksheet_name="Summary",
#             #     anchor=buyer_logo_anchor,
#             #     image_url=buyer_logo_url,
#             # )
#
#             context = {
#                 "filepath": filepath.split("/", 1)[1],
#                 "response_message": "Report generated successfully",
#             }
#             return context
#         except Exception as e:
#             capture_exception(e)


# class TradingLicenseReport(Task):
#     name = "TradingComplianceReport"
#     progress_recorder = None
#     result = 0
#
#     def run(self, *args, **kwargs):
#         self.progress_recorder = ProgressRecorder(self)
#         context = self.report()
#         return context
#
#     def report(self):
#         try:
#             data = {}
#             time = datetime.datetime.now()
#             dir_name = Path(
#                 "media/document_update/reports/{}/{}".format(time.year, time.month)
#             )
#             dir_name.mkdir(parents=True, exist_ok=True)
#             time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
#             filepath = "{}/document_update_report_{}.xlsx".format(dir_name, time_only)
#             data["Trading License Report_B"] = 6
#             data["Trading License Report_C"] = 62
#             data["Trading License Report_D"] = 15
#             data["Trading License Report_E"] = 15
#             data["Trading License Report_G"] = 40
#             data["Trading License Report_H"] = 40
#             data["Trading License Report_I"] = 40
#             data["Trading License Report_J"] = 40
#
#             workbook = Workbook()
#             summary_sheet = workbook.get_sheet_by_name("Sheet")
#             summary_sheet.title = "Trading License Report"
#
#             qed_logo_anchor = "H1"
#             buyer_logo_anchor = "B1"
#             today = datetime.date.today()
#             summary_sheet.append(("", ""))
#             summary_sheet.append(("", ""))
#             summary_sheet.append(("", ""))
#             summary_sheet.append(("", ""))
#             summary_sheet.append(("", ""))
#             summary_sheet.append(
#                 (
#                     "",
#                     "hc_#",
#                     "hl_Supplier",
#                     "hc_Expiry Date",
#                     "hl_Status",
#                     "hc_Renewal Date",
#                     "hl_Client 1",
#                     "hl_Client 2",
#                     "hl_Client 3",
#                     "hl_Client 4",
#                 )
#             )
#
#             suppliers = Supplier.objects.all()
#
#             total_for_progress = suppliers.count()
#             count = 1
#             for supplier in suppliers:
#                 self.result += 1
#                 self.progress_recorder.set_progress(self.result, total_for_progress)
#                 current_profile = SupplierProfile.objects.filter(
#                     supplier_id=supplier.id
#                 ).last()
#                 # previous_profile = SupplierProfile.objects.filter(supplier_id=supplier.id).order_by('-id')[1]
#
#                 if current_profile:
#                     companies = Company.objects.filter(
#                         id__in=SupplierResponse.objects.filter(
#                             supplier_id=supplier.id
#                         ).values("question__section__category__job__company_id")
#                     ).order_by("-id")
#                     try:
#                         company_one = companies[0].company_name
#                     except:
#                         company_one = "N/A"
#
#                     try:
#                         company_two = companies[1].company_name
#                     except:
#                         company_two = "N/A"
#
#                     try:
#                         company_three = companies[2].company_name
#                     except:
#                         company_three = "N/A"
#
#                     try:
#                         company_four = companies[3].company_name
#                     except:
#                         company_four = "N/A"
#
#                     if current_profile.kra_trading_licence_expiry_date:
#                         if current_profile.kra_trading_licence_expiry_date < today:
#                             supplier_row = [
#                                 "",
#                                 count,
#                                 supplier.company_name,
#                                 str(current_profile.kra_trading_licence_expiry_date),
#                                 "Expired",
#                                 "N/A",
#                                 company_one,
#                                 company_two,
#                                 company_three,
#                                 company_four,
#                             ]
#                         elif current_profile.kra_trading_licence_expiry_date > today:
#                             supplier_row = [
#                                 "",
#                                 count,
#                                 supplier.company_name,
#                                 str(current_profile.kra_trading_licence_expiry_date),
#                                 "Upto Date",
#                                 current_profile.created_at.strftime("%Y-%m-%d"),
#                                 company_one,
#                                 company_two,
#                                 company_three,
#                                 company_four,
#                             ]
#                         elif current_profile.kra_trading_licence_expiry_date == today:
#                             supplier_row = [
#                                 "",
#                                 count,
#                                 supplier.company_name,
#                                 str(current_profile.kra_trading_licence_expiry_date),
#                                 "Expires Today",
#                                 "N/A",
#                                 company_one,
#                                 company_two,
#                                 company_three,
#                                 company_four,
#                             ]
#
#                         summary_sheet.append(supplier_row)
#                         count += 1
#                     else:
#                         supplier_row = [
#                             "",
#                             count,
#                             supplier.company_name,
#                             "N/A",
#                             "N/A",
#                             "N/A",
#                             "N/A",
#                             "N/A",
#                             "N/A",
#                             "N/A",
#                         ]
#                         summary_sheet.append(supplier_row)
#                         count += 1
#                 else:
#                     supplier_row = [
#                         "",
#                         count,
#                         supplier.company_name,
#                         "N/A",
#                         "N/A",
#                         "N/A",
#                         "N/A",
#                         "N/A",
#                         "N/A",
#                         "N/A",
#                     ]
#                     summary_sheet.append(supplier_row)
#                     count += 1
#
#             summary_sheet.freeze_panes = summary_sheet["G7"]
#             workbook.save(filepath)
#             format_excel(filepath, data=data)
#
#             try:
#                 insert_image(
#                     excel_url=filepath,
#                     worksheet_name="Trading License Report",
#                     anchor=buyer_logo_anchor,
#                 )
#             except:
#                 insert_image(
#                     excel_url=filepath,
#                     worksheet_name="Trading License Report",
#                     anchor=buyer_logo_anchor,
#                     image_url="",
#                 )
#             # insert_image(
#             #     excel_url=filepath,
#             #     worksheet_name="Summary",
#             #     anchor=buyer_logo_anchor,
#             #     image_url=buyer_logo_url,
#             # )
#
#             context = {
#                 "filepath": filepath.split("/", 1)[1],
#                 "response_message": "Report generated successfully",
#             }
#             return context
#         except Exception as e:
#             capture_exception(e)


def prequal_bidders_information_report(data):
    participants = data['participants']
    category = data['category']
    sections = data['sections']

    time = datetime.datetime.now()
    # bd = Side(style='thin', color="000000")
    dir_name = Path(
        "media/prequal_bidders_info/{}/{}".format(time.year, time.month)
    )  # folder structure
    dir_name.mkdir(parents=True, exist_ok=True)

    time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
    filepath = "{}/{}_Bidders_Information_{}_{}.xlsx".format(
        dir_name,
        category.id,
        category.name.replace(" ", "_")
        .replace(":", "")
        .replace(",", "")
        .replace("-", "_")
        .replace("/", "_"),
        time_only,
    )

    workbook = Workbook()

    for section in sections:
        workbook.create_sheet(
            section.short_name_version.replace(" ", "_").replace(":", "_")
            .replace("(", "_").replace(")", "_")
        )
        section_name = section.short_name
        section_questions = section.questions
        current_worksheet = workbook[
            section.short_name_version.replace(" ", "_").replace(":", "_")
            .replace("(", "_").replace(")", "_")
        ]

        current_worksheet["B2"] = "Nb_{}".format(category.name)
        current_worksheet["B3"] = "Nb_%s" % section_name
        current_worksheet["B4"] = ""
        headers = [" ", "h_%s" % section_name]
        for supplier in participants:
            headers.append("h_%s" % supplier.company_name)
        current_worksheet.append(tuple(headers))
        i = 1
        for question in section_questions:
            question_row = [" ", question.short_description_value]
            for supplier in participants:
                response = supplier.prequal_question_response(question)
                question_row.append(response)
            current_worksheet.append(tuple(question_row))
            i += 1

    workbook.remove_sheet(workbook["Sheet"])
    workbook.save(filepath)
    format_excel(filepath)
    return filepath


def prequal_evaluation_report_context(data):
    participants = data['participants']
    category = data['category']
    sections = data['sections']

    data = {
        "questions": category.questions,
        "category": category,
        "company": category.prequalification.company,
        "time": timezone.now(),
        "sections": sections,
        "suppliers": participants,
    }
    return data


class PrequalInterimReport(Task):
    """
    Interim prequal report before QA
    """

    name = "PrequalInterimReport"
    progress_recorder = None
    result = 0

    def run(self, *args, **kwargs):
        self.progress_recorder = ProgressRecorder(self)
        context = self.report(kwargs["job_id"])
        return context

    def report(self, job_id):
        try:
            job = Prequalification.objects.filter(id=job_id).first()
            company = job.company

            categories = Category.objects.filter(
                prequalification_id=job.id).only("id", "unique_reference", "name").order_by("unique_reference")

            time = datetime.datetime.now()
            dir_name = Path(
                "media/prequal/job/reports/{}/{}/{}".format(time.year, time.month, time.day)
            )
            dir_name.mkdir(parents=True, exist_ok=True)
            time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
            filepath = "{}/{}_Interim_Technical_Report.xlsx".format(
                dir_name, job_id
            )

            workbook = Workbook()
            summary_sheet = workbook["Sheet"]
            summary_sheet.title = "Summary"

            qed_logo_anchor = "H1"
            buyer_logo_anchor = "B1"

            if company.company_logo_url is not None and company.company_logo_url != "":
                buyer_logo_url = get_file_path(company.company_logo_url)
            else:
                buyer_logo_url = f"static/core/img/no-company-image128-128.png"

            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(
                (
                    " ", "hc_#", "hc_Category Code", "hl_Category Title", "hhc_Responsive",
                    "hhc_Non-Responsive", "hhc_Total Bidders",
                )
            )

            data = {
                "Summary_B": 5, "Summary_C": 21, "Summary_D": 41,
                "Summary_E": 6, "Summary_F": 6, "Summary_G": 6,
            }

            progress_recorder = ProgressRecorder(self)
            result = 0

            # totals
            total_responsive = 0
            total_non_responsive = 0
            total_bidders = 0
            total_prequalified = 0
            total_letters = 0

            category_count = 1
            for category in categories:
                data[f"{category.unique_reference.replace('/', '_')}_C"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_D"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_E"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_F"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_G"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_H"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_I"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_J"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_K"] = 18
                data[f"{category.unique_reference.replace('/', '_')}_L"] = 18
                data[f'{category.unique_reference.replace("/", "_")}_B'] = 31

                result += 1
                progress_recorder.set_progress(result, categories.count())
                resp = (
                    SupplierResponse.objects.filter(
                        question__section__category_id=category.id
                    )
                    .only('supplier_id').values('supplier_id').distinct()
                    .count()
                )
                category_order = apps.apps.get_model('core', 'CategoryOrder')
                all = category_order.objects.filter(
                    category_id=category.id, payment_status=category_order.PAID, target__model='prequalification'
                ).count()

                non_resp = all - resp
                if non_resp < 0:
                    non_resp = 0

                summary_sheet.append(
                    (
                        " ", f"Ac_{category_count}", f"Al_{category.unique_reference}",
                        f"Al_{category.name}", f"Nc_{resp}", f"Nc_{non_resp}", f"Nc_{all}",
                    )
                )
                try:
                    total_responsive += resp
                    total_non_responsive += non_resp
                    total_bidders += all
                except:
                    total_responsive = 0
                    total_non_responsive = 0
                    total_bidders = 0

                category_type = category.category_type

                category_count += 1
                category_worksheet = workbook.create_sheet(
                    category.unique_reference.replace("/", "_")
                )
                category_worksheet.append(("", ""))
                category_worksheet["B2"] = f"Bnb_Category: {category.name}"
                category_worksheet["B3"] = ""
                category_worksheet["B4"] = "Number of bidders"
                category_worksheet["C4"] = f"Nc_{all}"
                category_worksheet["B5"] = "Number of responsive bidders"
                category_worksheet["C5"] = f"Nc_{resp}"
                category_worksheet["B6"] = "Proposed pass mark"
                category_worksheet["C6"] = f"Tcnd_{category.pass_score}"
                category_worksheet.append(("", ""))

                category_headers = [
                    " ",
                    "hl_Bidder",
                    "hhc_Rank",
                    "hhc_Score Before QA",
                ]

                category_worksheet.merge_cells("B2:E2")
                prequalified_bidders = 0

                questions = Question.objects.filter(
                    section__category_id=category.id, is_qa=True).only("id", "short_description", "description")
                sections = Section.objects.filter(
                    category_id=category.id, parent_section__isnull=True).order_by("id")

                for section in sections:
                    if section.has_child_sections:
                        if section.is_scored:
                            s_total = section.section_score
                            for s in section.child_sections:
                                if s.has_child_sections:
                                    for x in s.child_sections:
                                        s_total += x.section_score
                                else:
                                    s_total += s.section_score
                            category_headers.extend(
                                [f"hhc_{section.short_name}({int(s_total)})"]
                            )
                        else:
                            evaluate_this = False
                            for s in section.child_sections:
                                if s.has_child_sections:
                                    for x in s.child_sections:
                                        if x.is_scored:
                                            evaluate_this = True
                                            break
                                else:
                                    if s.is_scored:
                                        evaluate_this = True
                                        break
                            if evaluate_this == True:
                                s_total = 0
                                for s in section.child_sections:
                                    if s.has_child_sections:
                                        for x in s.child_sections:
                                            s_total += x.section_score
                                    else:
                                        s_total += s.section_score
                                category_headers.extend(
                                    [f"hhc_{section.short_name}({int(s_total)})"]
                                )
                    elif section.is_scored:
                        category_headers.extend(
                            [
                                f"hhc_{section.short_name}({int(section.section_score)})"
                            ]
                        )

                category_headers.extend(
                    [
                        "hl_Contact Person",
                        "hl_Phone Number",
                        "hl_Email Address",
                        "hl_Main Location",
                    ]
                )
                category_worksheet.append(category_headers)

                responsive_bidders = c_suppliers(category, interim=True)

                for supplier in responsive_bidders:
                    supplier_prequal_response = (
                        SupplierCategoryScore.objects.filter(
                            supplier_id=supplier.id, category_id=category.id
                        )
                        .only("score_after_qa", "rank", "rank_after_qa")
                        .first()
                    )
                    prequal_score = supplier_prequal_response.score

                    supplier_info = [
                        "",
                        supplier.company_name,
                        f"Nc_{supplier_prequal_response.rank}",
                        f"Tcnd_{prequal_score}",
                    ]

                    for section in sections:
                        if section.has_child_sections:
                            if section.is_scored:
                                section_total = supplier.prequal_section_score(
                                    section
                                )
                                for s in section.child_sections:
                                    if s.has_child_sections:
                                        if s.is_scored:
                                            section_total += (
                                                supplier.prequal_section_score(s)
                                            )
                                            for x in s.child_sections:
                                                if x.is_scored:
                                                    section_total += supplier.prequal_section_score(
                                                        x
                                                    )
                                        else:
                                            evaluate_this = False
                                            for x in s.child_sections:
                                                if x.is_scored:
                                                    evaluate_this = True
                                                    break
                                            if evaluate_this == True:
                                                section_total += (
                                                    supplier.prequal_section_score(
                                                        s
                                                    )
                                                )
                                                for x in s.child_sections:
                                                    if x.is_scored:
                                                        section_total += supplier.prequal_section_score(
                                                            x
                                                        )

                                    else:
                                        if s.is_scored:
                                            section_total += (
                                                supplier.prequal_section_score(s)
                                            )
                                supplier_info.extend([f"Tcnd_{section_total}"])
                            else:
                                evaluate_this = False
                                for s in section.child_sections:
                                    if s.is_scored:
                                        evaluate_this = True
                                        break
                                if evaluate_this == True:
                                    section_total = supplier.prequal_section_score(
                                        section
                                    )
                                    for s in section.child_sections:
                                        if s.has_child_sections:
                                            if s.is_scored:
                                                section_total += (
                                                    supplier.prequal_section_score(
                                                        s
                                                    )
                                                )
                                                for x in s.child_sections:
                                                    if x.is_scored:
                                                        section_total += supplier.prequal_section_score(
                                                            x
                                                        )
                                            else:
                                                evaluate_this = False
                                                for x in s.child_sections:
                                                    if x.is_scored:
                                                        evaluate_this = True
                                                        break
                                                if evaluate_this == True:
                                                    section_total += supplier.prequal_section_score(
                                                        s
                                                    )
                                                    for x in s.child_sections:
                                                        if x.is_scored:
                                                            section_total += supplier.prequal_section_score(
                                                                x
                                                            )
                                        else:
                                            if s.is_scored:
                                                section_total += (
                                                    supplier.prequal_section_score(
                                                        s
                                                    )
                                                )
                                    supplier_info.extend([f"Tcnd_{section_total}"])
                        else:
                            if section.is_scored:
                                section_total = supplier.prequal_section_score(
                                    section
                                )
                                supplier_info.extend([f"Tcnd_{section_total}"])

                    try:
                        # supplier_locations = get_cat_supplier_locations(cat_type_supplier)
                        supplier_locations = get_locations(
                            category=category, supplier_id=supplier.id
                        )
                        location = supplier_locations.get("main_location", "Ar_N/A")
                    except:
                        location = "N/A"

                    try:
                        supplier_info.extend(
                            [
                                f"Al_{supplier.contact_name}",
                                f"Al_{supplier.phone_number}",
                                f"Al_{supplier.email}",
                                f"{location}",
                            ]
                        )
                    except Exception as e:
                        capture_exception(e)

                    category_worksheet.append(supplier_info)

                # Append number of prequalified bidders
                # category_worksheet["C4"] = f"Nc_{prequalified_bidders}"

                # non responsive bidders
                for supplier in non_responsive_bidders(category, responsive_bidders):
                    supplier_info = [
                        "",
                        supplier.company_name,
                        "Ac_N/R",
                        "Ac_N/R",
                    ]
                    for section in sections:
                        if section.has_child_sections:
                            if section.is_scored:
                                supplier_info.extend(["Ac_N/R"])
                            else:
                                evaluate_this = False
                                for s in section.child_sections:
                                    if s.is_scored:
                                        evaluate_this = True
                                        break
                                if evaluate_this == True:
                                    supplier_info.extend(["Ac_N/R"])
                        else:
                            if section.is_scored:
                                supplier_info.extend(["Ac_N/R"])

                    supplier_info.extend(["Ac_N/R", "Ac_N/R", "Ac_N/R", "Ac_N/R"])
                    category_worksheet.append(supplier_info)

                # append notes
                category_worksheet.append(("", ""))
                category_worksheet.append(("", "Notes"))
                category_worksheet.append(("", "N/A - Not Applicable"))
                category_worksheet.append(("", "N/R - None responsive"))
                category_worksheet.freeze_panes = category_worksheet["C9"]

            summary_sheet.append(
                (
                    " ", "", "", "qT_Total", "Ac_{}".format(total_responsive),
                    "Ac_{}".format(total_non_responsive), "Ac_{}".format(total_bidders),
                )
            )
            summary_sheet.freeze_panes = summary_sheet["E7"]
            workbook.save(filepath)
            format_excel(filepath, data=data)
            insert_image(
                excel_url=filepath, worksheet_name="Summary", anchor=qed_logo_anchor
            )
            insert_image(
                excel_url=filepath,
                worksheet_name="Summary",
                anchor=buyer_logo_anchor,
                image_url=buyer_logo_url,
            )

            time = datetime.datetime.now()
            with open(filepath, "rb") as l:
                storage = PrivateMediaStorage()
                url = f"prequal/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
                storage.save(url, l)
                job_report, created = JobReport.objects.update_or_create(
                    prequalification_id=job_id, defaults={"interim_report": url}
                )
                path = job_report.interim_report.url
                context = {
                    "filepath": path,
                    "response_message": "Report generated successfully",
                }
                return context
        #     try:
        #         time = datetime.datetime.now()
        #         with open(filepath, "rb") as l:
        #             storage = PrivateMediaStorage()
        #             url = f"prequal/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
        #             storage.save(url, l)
        #
        #             job_report, created = JobReport.objects.update_or_create(
        #                 job_id=job_id, defaults={"intermediate_prequal_excel": url}
        #             )
        #             context = {
        #                 "filepath": job_report.intermediate_prequal_excel.url,
        #                 "job_id": job_id,
        #                 "response_message": "Report generated successfully",
        #             }
        #             return context
        #     except Exception as e:
        #         capture_exception(e)
        #         print(e)
        #         context = {
        #             "response_message": "Error generating interim prequal report"
        #         }
        #         return context

        except Exception as e:
            print(e)
            capture_exception(e)
            context = {
                "response_message": "Report generation error",
                "messages": [f"{e}", ]
            }
            return context


@shared_task(bind=True)
def bidder_locations_report(self, job_id):
    """
    Prequalified suppliers location and contact details
    """
    try:
        job = Prequalification.objects.filter(id=job_id).first()
        company = job.company
        time = datetime.datetime.now()
        data = {}
        # dir_name = Path("media/rfq_reports/{}/{}".format(time.year, time.month))
        dir_name = Path(
            "media/bidder_locations/%s/%s/%s"
            % (company.company_name, time.year, job.title)
        )  # folder structure
        dir_name.mkdir(parents=True, exist_ok=True)
        # time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
        filepath = "{}/{}_Bidder_Locations_Report.xlsx".format(
            dir_name, job.unique_reference
        )

        # delete similar reports run this month
        match_string = "{}_{}".format(job.id, company.company_name.replace(" ", "_"))
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

        # create and save a workbook to later populate
        workbook = Workbook()

        # create and save a workbook to later populate
        bidder_location_details_worksheet = workbook.active
        bidder_location_details_worksheet.title = "Summary"
        if company.company_logo_url is not None and company.company_logo_url != "":
            buyer_logo_url = get_file_path(company.company_logo_url)
        else:
            buyer_logo_url = "static/core/img/no-company-image128-128.png"

        qed_logo_anchor = "K1"
        qed_logo_worksheet = bidder_location_details_worksheet.title
        buyer_logo_anchor = "C1"
        buyer_logo_worksheet = bidder_location_details_worksheet.title
        bidder_location_details_worksheet["C5"] = "Nb_{}".format(job.title)
        bidder_location_details_worksheet["C6"] = "Nb_{}".format(job.unique_reference)
        bidder_location_details_worksheet["C7"] = "Nb_Bidder Location Details"
        bidder_location_details_worksheet["C8"] = "Nb_"

        bidder_location_details_worksheet.append(
            ("", "hc_No", "h_Category Code", "h_Category Name")
        )
        bidder_location_details_worksheet.merge_cells("C5:D5")

        data["Summary_C"] = 25
        data["Summary_D"] = 62

        categories = Category.objects.filter(prequalification_id=job_id).only("id", "unique_reference", "name").order_by(
            "unique_reference"
        )
        total = categories.count()
        result = 0
        progress_recorder = ProgressRecorder(self)

        for count, category in enumerate(categories, start=1):
            result += 1
            progress_recorder.set_progress(result, total)
            bidder_location_details_worksheet.append(
                (
                    "",
                    "Ac_%d" % (count),
                    category.unique_reference,
                    category.name,
                )
            )

            workbook.create_sheet(category.unique_reference.replace("/", "_"))
            category_bidder_details_worksheet = workbook[
                category.unique_reference.replace("/", "_")
            ]
            empty_row = [""]
            empty_row = tuple(empty_row)
            category_bidder_details_worksheet.append(empty_row)
            category_bidder_details_worksheet["C2"] = "Nb_{}".format(category.name)
            category_bidder_details_worksheet["C3"] = "Nb_{}".format(
                category.unique_reference
            )
            category_bidder_details_worksheet["C4"] = "Nb_"

            data[f'category.unique_reference.replace("/", "_")_C'] = 25
            data[f'category.unique_reference.replace("/", "_")_D'] = 5
            data[f'category.unique_reference.replace("/", "_")_E'] = 14.6
            data[f'category.unique_reference.replace("/", "_")_F'] = 25
            data[f'category.unique_reference.replace("/", "_")_G'] = 14.6
            data[f'category.unique_reference.replace("/", "_")_H'] = 30
            data[f'category.unique_reference.replace("/", "_")_I'] = 35

            category_bidder_details_worksheet.merge_cells("C2:E2")
            category_bidder_details_worksheet.append(
                (
                    "",
                    "hc_No",
                    "h_Company Name",
                    "h_Rank",
                    "h_Marks After QA",
                    "h_Contact Person",
                    "h_Phone Number",
                    "h_Email Address",
                    "h_Physical Address",
                    "h_Main Location",
                    "h_Branches",
                    "h_Location 3",
                    "h_Location 4",
                    "h_Location 5",
                    "h_Location 6",
                    "h_Location 7",
                    "h_Location 8",
                    "h_Location 9",
                    "h_Location 10",
                )
            )

            supplier_totals = SupplierCategoryScore.objects.filter(category_id=category.id).order_by(
                "rank"
            ).prefetch_related('supplier')

            for count, supplier_total in enumerate(supplier_totals, start=1):
                supplier = supplier_total.supplier
                score_after_qa = supplier_total.score_after_qa
                rank_after_qa = supplier_total.rank_after_qa

                if score_after_qa is None:
                    score_after_qa = "Ar_N/A"
                    rank_after_qa = "Ar_N/A"

                if rank_after_qa is None:
                    rank_after_qa = "Ar_N/A"

                locations = get_locations(category=category, supplier_id=supplier_total.supplier_id)

                category_bidder_details_worksheet.append(
                    (
                        "", "Ac_%d" % (count), supplier.company_name,rank_after_qa, score_after_qa,
                        supplier.contact_name, supplier.phone_number, supplier.address,
                        " ".join(locations.get("main_location", "Ar_N/A").split()),
                        " ".join(locations.get("location_1", "Ar_N/A").split()),
                        " ".join(locations.get("location_2", "Ar_N/A").split()),
                        " ".join(locations.get("location_3", "Ar_N/A").split()),
                        " ".join(locations.get("location_4", "Ar_N/A").split()),
                        " ".join(locations.get("location_5", "Ar_N/A").split()),
                        " ".join(locations.get("location_6", "Ar_N/A").split()),
                        " ".join(locations.get("location_7", "Ar_N/A").split()),
                        " ".join(locations.get("location_8", "Ar_N/A").split()),
                        " ".join(locations.get("location_9", "Ar_N/A").split()),
                        " ".join(locations.get("location_10", "Ar_N/A").split()),
                    )
                )
                category_bidder_details_worksheet.freeze_panes = (
                    category_bidder_details_worksheet["D6"]
                )

        bidder_location_details_worksheet.freeze_panes = (
            bidder_location_details_worksheet["D10"]
        )
        workbook.save(filepath)
        format_excel(filepath, data=data)
        insert_image(
            excel_url=filepath,
            worksheet_name=qed_logo_worksheet,
            anchor=qed_logo_anchor,
        )
        insert_image(
            filepath, buyer_logo_worksheet, buyer_logo_anchor, image_url=buyer_logo_url
        )

        time = datetime.datetime.now()
        with open(filepath, "rb") as l:
            storage = PrivateMediaStorage()
            url = f"prequal/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
            storage.save(url, l)

            job_report, created = JobReport.objects.update_or_create(
                prequalification_id=job_id, defaults={"bidder_locations_report": url}
            )
            context = {
                "response_message": "Report generated successfully",
                "filepath": job_report.bidder_locations_report.url,
            }
            return context

    except Exception as e:
        capture_exception(e)
        print(e)
        context = {
            "response_message": "Report generation error",
        }
        return context


@shared_task(bind=True)
def prequalified_suppliers_report(self, job_id):
    try:
        job = Prequalification.objects.filter(id=job_id).first()
        company = job.company
        time = datetime.datetime.now()
        data = {}
        dir_name = Path(
            "media/suppliers_list_report/%s/%s/%s"
            % (company.company_name, time.year, job.title)
        )  # folder structure
        dir_name.mkdir(parents=True, exist_ok=True)
        time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
        filepath = "{}/{}_Prequalified_Suppliers_Report.xlsx".format(
            dir_name, job.unique_reference
        )

        # delete similar reports run this month
        match_string = "{}_{}".format(job.id, company.company_name.replace(" ", "_"))
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

        # create and save a workbook to later populate
        workbook = Workbook()

        # create and save a workbook to later populate
        suppliers_list_worsheet = workbook.active
        suppliers_list_worsheet.title = "SUPPLIERS"
        if company.company_logo_url is not None and company.company_logo_url != "":
            buyer_logo_url = get_file_path(company.company_logo_url)
        else:
            buyer_logo_url = "static/core/img/no-company-image128-128.png"

        qed_logo_anchor = "D1"
        qed_logo_worksheet = suppliers_list_worsheet.title
        buyer_logo_anchor = "B1"
        buyer_logo_worksheet = suppliers_list_worsheet.title
        suppliers_list_worsheet["B5"] = job.title
        suppliers_list_worsheet["B6"] = job.unique_reference
        suppliers_list_worsheet["B7"] = "Supplier Details"
        suppliers_list_worsheet["B8"] = "Nb_"

        data["SUPPLIERS_C"] = 35
        data["SUPPLIERS_D"] = 25
        data["SUPPLIERS_E"] = 30
        data["SUPPLIERS_F"] = 25
        data["SUPPLIERS_G"] = 20
        data["SUPPLIERS_H"] = 62

        suppliers_list_worsheet.merge_cells("B5:D5")

        categories = Category.objects.filter(prequalification_id=job_id).only("id", "unique_reference", "name").order_by(
            "unique_reference"
        )
        total = categories.count()
        result = 0
        progress_recorder = ProgressRecorder(self)

        for count, category in enumerate(categories, start=1):
            suppliers_list_worsheet.append(
                ("", f"Category: {category.name}", "", "", "", "", "", "",)
            )
            suppliers_list_worsheet.append(
                ("", "h_Bidder", "h_Contact Person", "h_Phone Number",
                 "h_Email Address", "h_Postal Address", "h_Tax PIN", "h_Physical Address",
                 "h_Main Location", "h_Branches")
            )

            result += 1
            progress_recorder.set_progress(result, total)

            suppliers = Supplier.objects.filter(
                id__in=AwardLetter.objects.filter(category_id=category.id).only("supplier_id").values_list(
                    "supplier_id", flat=True
                )
            )

            for count, supplier in enumerate(suppliers, start=1):
                supplier_data = get_supplier_data(
                    category_id=category.id, supplier_id=supplier.id
                )
                try:
                    suppliers_list_worsheet.append(
                        (
                            "", supplier.company_name, supplier.contact_name,
                            supplier.phone_number, supplier.address,
                            supplier_data["Section"].get("P.O Box address", "N/A"),
                            supplier_data.get("Tax_ID", "N/A"),
                            supplier_data["Section"].get("Office_Address", "N/A"),
                            supplier_data["Section"].get("Main_Office", "N/A"),
                            supplier_data["Section"].get("Branch_Office", "N/A"),
                        )
                    )
                except Exception as e:
                    capture_exception(e)

            suppliers_list_worsheet.append(
                ("", "", "", "", "", "", "", "")
            )

        suppliers_list_worsheet.freeze_panes = suppliers_list_worsheet["C8"]

        workbook.save(filepath)
        format_excel(filepath, data=data)
        insert_image(
            excel_url=filepath,
            worksheet_name=qed_logo_worksheet,
            anchor=qed_logo_anchor,
        )
        insert_image(
            filepath, buyer_logo_worksheet, buyer_logo_anchor, image_url=buyer_logo_url
        )

        time = datetime.datetime.now()
        with open(filepath, "rb") as l:
            storage = PrivateMediaStorage()
            url = f"prequal/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
            storage.save(url, l)

            job_report, created = JobReport.objects.update_or_create(
                prequalification_id=job_id, defaults={"prequalified_suppliers_report": url}
            )

            context = {
                "response_message": "Report generated successfully",
                "filepath": job_report.prequalified_suppliers_report.url,
            }
            return context

    except Exception as e:
        capture_exception(e)
        print(e)
        context = {
            "response_message": "Report generation error",
        }
        return context


@shared_task(bind=True)
def job_bidder_payments_report(self, job_id):
    try:
        result = 0
        progress_recorder = ProgressRecorder(self)

        job = apps.apps.get_model('prequal', 'Prequalification').objects.filter(id=job_id).first()
        company = job.company
        time = datetime.datetime.now()
        # dir_name = Path("media/rfq_reports/{}/{}".format(time.year, time.month))
        dir_name = Path(
            "media/bidder_payments/%s/%s/%s"
            % (company.company_name, time.year, job.title)
        )  # folder structure
        dir_name.mkdir(parents=True, exist_ok=True)
        # time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
        filepath = "{}/{}_Bidder_Payments_Report.xlsx".format(
            dir_name, job.unique_reference
        )

        # delete similar reports run this month
        match_string = "{}_{}".format(job.id, company.company_name.replace(" ", "_"))
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

        # create and save a workbook to later populate
        workbook = Workbook()

        bidder_payments_worksheet = workbook.active
        bidder_payments_worksheet.title = "Summary"

        if company.company_logo_url is not None and company.company_logo_url != "":
            buyer_logo_url = company.company_logo_url
        else:
            buyer_logo_url = "static/core/img/no-company-image128-128.png"

        qed_logo_anchor = "K1"
        qed_logo_worksheet = bidder_payments_worksheet.title
        buyer_logo_anchor = "C1"
        buyer_logo_worksheet = bidder_payments_worksheet.title
        bidder_payments_worksheet["C5"] = "Nb_{}".format(job.title)
        bidder_payments_worksheet["C6"] = "Nb_{}".format(job.unique_reference)
        bidder_payments_worksheet["C7"] = "Nb_Bidder Payments"
        bidder_payments_worksheet["C8"] = "Nb_"

        bidder_payments_worksheet.append(
            ("", "hc_No", "h_Category Code", "h_Category Name", "h_Bidder Payments",
             "h_Amount")
        )
        categories = Category.objects.filter(prequalification_id=job_id)
        total = categories.count()
        total_job_bid_charge = 0
        for count, category in enumerate(categories, start=1):
            result += 1
            progress_recorder.set_progress(result, total)

            orders = CategoryOrder.objects.filter(
                payment_status=CategoryOrder.PAID, category_id=category.id, target__model='prequalification'
            )
            paid_bidders = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                id__in=orders.only('supplier_id').values('supplier_id')
            )
            paid_bidders_count = orders.count()
            # orders = orders.aggregate(total=Sum("category__bid_charge"))
            total_bid_charge = float(category.bid_charge) * paid_bidders_count
            total_job_bid_charge += total_bid_charge

            bidder_payments_worksheet.append(
                ("", "Ac_%d" % (count), category.unique_reference,
                 category.name, paid_bidders_count, total_bid_charge
                )
            )

            workbook.create_sheet(category.unique_reference.replace("/", "_"))
            category_bidder_payments_worksheet = workbook[
                category.unique_reference.replace("/", "_")
            ]
            empty_row = [""]
            empty_row = tuple(empty_row)
            category_bidder_payments_worksheet.append(empty_row)

            category_bidder_payments_worksheet["C2"] = "Nb_{}".format(category.name)
            category_bidder_payments_worksheet["C3"] = "Nb_{}".format(
                category.unique_reference
            )
            category_bidder_payments_worksheet["C4"] = "Nb_"

            category_bidder_payments_worksheet.append(
                ("", "hc_No", "h_Company Name", "h_Contact Person", "h_Email Address", "h_Phone Number",
                )
            )

            for count, supplier in enumerate(paid_bidders, start=1):
                category_bidder_payments_worksheet.append(
                    (
                        "", "Ac_%d" % (count), supplier.company_name, supplier.contact_name,
                        supplier.address, supplier.phone_number,
                    )
                )
        bidder_payments_worksheet.append(
            ("", "", "Ac_Total", "", "", total_job_bid_charge)
        )
        bidder_payments_worksheet.merge_cells(
            start_row=bidder_payments_worksheet.max_row,
            start_column=3, end_row=bidder_payments_worksheet.max_row,
            end_column=5,
        )
        workbook.save(filepath)
        format_excel(filepath)
        insert_image(excel_url=filepath, worksheet_name=qed_logo_worksheet,
            anchor=qed_logo_anchor,
        )
        insert_image(
            filepath, buyer_logo_worksheet, buyer_logo_anchor, image_url=buyer_logo_url
        )

        time = datetime.datetime.now()
        with open(filepath, "rb") as l:
            storage = PrivateMediaStorage()
            url = f"prequal/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
            storage.save(url, l)
            job_report, created = JobReport.objects.update_or_create(
                prequalification_id=job_id, defaults={"bidder_payments_report": url}
            )
            context = {
                "response_message": "Report generated successfully",
                "filepath": job_report.bidder_payments_report.url,
            }
            return context
    except Exception as e:
        capture_exception(e)
        print(e)
        context = {
            "response_message": "Report generation error"
        }
        return context


@shared_task(bind=True)
def responsive_bidders_report(self, job_id):
    try:
        result = 0
        progress_recorder = ProgressRecorder(self)

        job = Prequalification.objects.filter(id=job_id).first()
        company = job.company
        time = datetime.datetime.now()
        # dir_name = Path("media/rfq_reports/{}/{}".format(time.year, time.month))
        dir_name = Path(
            "media/responsive_bidders/%s/%s/%s"
            % (company.company_name, time.year, job.title)
        )  # folder structure
        dir_name.mkdir(parents=True, exist_ok=True)
        # time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
        filepath = "{}/{}_Responsive_Bidders_Report.xlsx".format(
            dir_name, job.unique_reference
        )

        # delete similar reports run this month
        match_string = "{}_{}".format(job.id, company.company_name.replace(" ", "_"))
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

        # create and save a workbook to later populate
        workbook = Workbook()

        responsive_bidders_worksheet = workbook.active
        responsive_bidders_worksheet.title = "Summary"

        if company.company_logo_url is not None and company.company_logo_url != "":
            buyer_logo_url = company.company_logo_url
        else:
            buyer_logo_url = "static/core/img/no-company-image128-128.png"

        qed_logo_anchor = "K1"
        qed_logo_worksheet = responsive_bidders_worksheet.title
        buyer_logo_anchor = "C1"
        buyer_logo_worksheet = responsive_bidders_worksheet.title
        responsive_bidders_worksheet["C5"] = "Nb_{}".format(job.title)
        responsive_bidders_worksheet["C6"] = "Nb_{}".format(job.unique_reference)
        responsive_bidders_worksheet["C7"] = "Nb_Responsive Bidders"
        responsive_bidders_worksheet["C8"] = "Nb_"

        responsive_bidders_worksheet.append(
            ("", "hc_No", "h_Category Name", "h_Category Code", "h_Responsive Bidders")
        )
        categories = Category.objects.filter(id=job_id)
        total = categories.count()
        total_responsive_bidders = 0

        for count, category in enumerate(categories, start=1):
            result += 1
            progress_recorder.set_progress(result, total)

            responsive_bidders = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                id__in=SupplierResponse.objects.filter(
                    question__section__category_id=category.id).only('supplier_id').values('supplier_id')
            )
            responsive_bidders_worksheet.append(
                ("", "Ac_%d" % (count), category.name,
                 category.unique_reference, responsive_bidders.count(),
                 )
            )
            workbook.create_sheet(category.unique_reference.replace("/", "_"))
            category_responsive_bidders_worksheet = workbook[
                category.unique_reference.replace("/", "_")
            ]
            empty_row = [""]
            empty_row = tuple(empty_row)
            category_responsive_bidders_worksheet.append(empty_row)

            category_responsive_bidders_worksheet["C2"] = "Nb_{}".format(category.name)
            category_responsive_bidders_worksheet["C3"] = "Nb_{}".format(
                category.unique_reference
            )
            category_responsive_bidders_worksheet["C4"] = "Nb_"

            category_responsive_bidders_worksheet.append(
                (
                    "",
                    "hc_No",
                    "h_Company Name",
                    "h_Contact Person",
                    "h_Email Address",
                    "h_Phone Number",
                )
            )
            total_responsive_bidders += responsive_bidders.count()
            for count, supplier in enumerate(responsive_bidders, start=1):
                category_responsive_bidders_worksheet.append(
                    (
                        "",
                        "Ac_%d" % (count),
                        supplier.company_name,
                        supplier.contact_name,
                        supplier.address,
                        supplier.phone_number,
                    )
                )

        responsive_bidders_worksheet.append(
            ("", "", "", "h_Total", total_responsive_bidders)
        )
        workbook.save(filepath)
        format_excel(filepath)
        insert_image(
            excel_url=filepath,
            worksheet_name=qed_logo_worksheet,
            anchor=qed_logo_anchor,
        )
        insert_image(
            filepath, buyer_logo_worksheet, buyer_logo_anchor, image_url=buyer_logo_url
        )

        time = datetime.datetime.now()
        with open(filepath, "rb") as l:
            storage = PrivateMediaStorage()
            url = f"prequal/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
            storage.save(url, l)

            job_report, created = JobReport.objects.update_or_create(
                prequalification_id=job_id, defaults={"responsive_bidders_report": url}
            )
            context = {
                "response_message": "Report generated successfully",
                "filepath": job_report.responsive_bidders_report.url,
            }
            return context
    except Exception as e:
        capture_exception(e)
        print(e)
        context = {
            "response_message": "Report generation error"
        }
        return context


@shared_task(bind=True)
def non_responsive_bidders_report(self, job_id):
    try:
        result = 0
        progress_recorder = ProgressRecorder(self)
        job = Prequalification.objects.filter(id=job_id).first()
        company = job.company
        time = datetime.datetime.now()
        dir_name = Path(
            "media/non_responsive_bidders/%s/%s/%s"
            % (company.company_name, time.year, job.title)
        )  # folder structure
        dir_name.mkdir(parents=True, exist_ok=True)
        # time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
        filepath = "{}/{}_Non_Responsive_Bidders_Report.xlsx".format(
            dir_name, job.unique_reference
        )

        # delete similar reports run this month
        match_string = "{}_{}".format(job.id, company.company_name.replace(" ", "_"))
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

        # create and save a workbook to later populate
        workbook = Workbook()

        responsive_bidders_worksheet = workbook.active
        responsive_bidders_worksheet.title = "Summary"
        total_non_responsive = 0

        if company.company_logo_url is not None and company.company_logo_url != "":
            buyer_logo_url = company.company_logo_url
        else:
            buyer_logo_url = "static/core/img/no-company-image128-128.png"

        qed_logo_anchor = "K1"
        qed_logo_worksheet = responsive_bidders_worksheet.title
        buyer_logo_anchor = "C1"
        buyer_logo_worksheet = responsive_bidders_worksheet.title
        responsive_bidders_worksheet["C5"] = "Nb_{}".format(job.title)
        responsive_bidders_worksheet["C6"] = "Nb_{}".format(job.unique_reference)
        responsive_bidders_worksheet["C7"] = "Nb_Non-Responsive Bidders"
        responsive_bidders_worksheet["C8"] = "Nb_"

        responsive_bidders_worksheet.append(
            (
                "", "hc_No", "h_Category Name", "h_Category Code", "h_Non-Responsive Bidders",
            )
        )
        categories = Category.objects.filter(prequalification_id=job.id)
        total_for_progress = categories.count()
        for count, category in enumerate(categories, start=1):
            result += 1
            progress_recorder.set_progress(result, total_for_progress)

            responsive_bidders = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                id__in=SupplierResponse.objects.filter(
                    question__section__category_id=category.id).only('supplier_id').values('supplier_id')
            )
            paid_bidders = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                id__in=CategoryOrder.objects.filter(
                    payment_status=CategoryOrder.PAID, category_id=category.id, target__model='prequalification'
                ).only('supplier_id').values('supplier_id')
            )
            non_responsive_bidders_count = paid_bidders.count() - responsive_bidders.count()
            non_responsive_bidders = paid_bidders.difference(responsive_bidders)

            responsive_bidders_worksheet.append(
                ("", "Ac_%d" % (count), category.name, category.unique_reference,
                 non_responsive_bidders_count,
                 )
            )
            total_non_responsive += non_responsive_bidders_count

            workbook.create_sheet(category.unique_reference.replace("/", "_"))
            category_responsive_bidders_worksheet = workbook[
                category.unique_reference.replace("/", "_")
            ]
            empty_row = [""]
            empty_row = tuple(empty_row)
            category_responsive_bidders_worksheet.append(empty_row)

            category_responsive_bidders_worksheet["C2"] = "Nb_{}".format(category.name)
            category_responsive_bidders_worksheet["C3"] = "Nb_{}".format(
                category.unique_reference
            )
            category_responsive_bidders_worksheet["C4"] = "Nb_"

            category_responsive_bidders_worksheet.append(
                (
                    "", "hc_No", "h_Company Name", "h_Contact Person", "h_Email Address", "h_Phone Number",
                )
            )

            for count, supplier in enumerate(non_responsive_bidders, start=1):
                category_responsive_bidders_worksheet.append(
                    (
                        "", "Ac_%d" % (count), supplier.company_name, supplier.contact_name,
                        supplier.address, supplier.phone_number,
                    )
                )
        responsive_bidders_worksheet.append(
            (
                "", "", "", "hc_Total", total_non_responsive,
            )
        )
        workbook.save(filepath)
        format_excel(filepath)
        insert_image(
            excel_url=filepath,
            worksheet_name=qed_logo_worksheet,
            anchor=qed_logo_anchor,
        )
        insert_image(
            filepath, buyer_logo_worksheet, buyer_logo_anchor, image_url=buyer_logo_url
        )

        time = datetime.datetime.now()
        with open(filepath, "rb") as l:
            storage = PrivateMediaStorage()
            url = f"prequal/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
            storage.save(url, l)

            job_report, created = JobReport.objects.update_or_create(
                prequalification_id=job_id, defaults={"non_responsive_bidders_report": filepath}
            )
            context = {
                "response_message": "Report generated successfully",
                "filepath": job_report.non_responsive_bidders_report.url,
            }
            return context
    except Exception as e:
        capture_exception(e)
        print(e)
        context = {
            "response_message": "Report generation error"
        }
        return context


@shared_task(bind=True)
def directors_report(self, job_id):
    """
    Directors report
    """
    try:
        result = 0
        progress_recorder = ProgressRecorder(self)

        job = Prequalification.objects.filter(id=job_id).first()
        company = job.company
        time = datetime.datetime.now()
        data = {}
        # dir_name = Path("media/rfq_reports/{}/{}".format(time.year, time.month))
        dir_name = Path(
            "media/directors_report/%s/%s/%s"
            % (company.company_name, time.year, job.title)
        )  # folder structure
        dir_name.mkdir(parents=True, exist_ok=True)
        time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
        filepath = "{}/{}_Directors_Report.xlsx".format(
            dir_name, job.unique_reference
        )

        # delete similar reports run this month
        match_string = "{}_{}".format(job.id, company.company_name.replace(" ", "_"))
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

        # create and save a workbook to later populate
        workbook = Workbook()

        # create and save a workbook to later populate
        bidder_location_details_worksheet = workbook.active
        bidder_location_details_worksheet.title = "Summary"
        if company.company_logo_url is not None and company.company_logo_url != "":
            buyer_logo_url = get_file_path(company.company_logo_url)
        else:
            buyer_logo_url = "static/core/img/no-company-image128-128.png"

        qed_logo_anchor = "K1"
        qed_logo_worksheet = bidder_location_details_worksheet.title
        buyer_logo_anchor = "C1"
        buyer_logo_worksheet = bidder_location_details_worksheet.title
        bidder_location_details_worksheet["C5"] = "Nb_{}".format(job.title)
        bidder_location_details_worksheet["C6"] = "Nb_{}".format(job.unique_reference)
        bidder_location_details_worksheet["C7"] = "Nb_Bidder Location Details"
        bidder_location_details_worksheet["C8"] = "Nb_"

        bidder_location_details_worksheet.append(
            ("", "hc_No", "h_Category Code", "h_Category Name", "h_Number of Directors")
        )
        bidder_location_details_worksheet.merge_cells("C5:D5")
        data["Summary_C"] = 25
        data["Summary_D"] = 62

        categories = Category.objects.filter(prequalification_id=job_id).only("id", "unique_reference", "name").order_by(
            "unique_reference"
        )
        total = categories.count()

        for count, category in enumerate(categories, start=1):
            result += 1
            progress_recorder.set_progress(result, total)

            bidder_location_details_worksheet.append(
                ("", "Ac_%d" % (count), category.unique_reference, category.name, 1)
            )
            bidder_location_details_worksheet.freeze_panes = (
                bidder_location_details_worksheet["D10"]
            )
            workbook.create_sheet(category.unique_reference.replace("/", "_"))
            category_bidder_details_worksheet = workbook[
                category.unique_reference.replace("/", "_")
            ]
            empty_row = [""]
            empty_row = tuple(empty_row)
            category_bidder_details_worksheet.append(empty_row)
            category_bidder_details_worksheet["C2"] = "Nb_{}".format(category.name)
            category_bidder_details_worksheet["C3"] = "Nb_{}".format(
                category.unique_reference
            )
            category_bidder_details_worksheet["C4"] = "Nb_"
            category_bidder_details_worksheet.merge_cells("C2:E2")
            data[f'category.unique_reference.replace("/", "_")_C'] = 25
            data[f'category.unique_reference.replace("/", "_")_D'] = 5
            data[f'category.unique_reference.replace("/", "_")_E'] = 14.6
            data[f'category.unique_reference.replace("/", "_")_F'] = 25
            data[f'category.unique_reference.replace("/", "_")_G'] = 14.6
            data[f'category.unique_reference.replace("/", "_")_H'] = 30
            data[f'category.unique_reference.replace("/", "_")_I'] = 35

            category_bidder_details_worksheet.append(
                (
                    "", "hc_No", "h_Company Name", "h_Rank", "h_Marks After QA",
                    "h_Name 1", "h_Name 2", "h_Name 3", "h_Name 4", "h_Name 5",
                )
            )

            supplier_totals = SupplierCategoryScore.objects.filter(category_id=category.id).order_by(
                "rank"
            ).prefetch_related('supplier')

            for count, supplier_total in enumerate(supplier_totals, start=1):
                supplier = supplier_total.supplier
                try:
                    score_after_qa = supplier_total.score_after_qa
                    rank_after_qa = supplier_total.rank_after_qa
                except:
                    score_after_qa = "Ar_N/A"
                    rank_after_qa = "Ar_N/A"

                if score_after_qa is None:
                    score_after_qa = "Ar_N/A"
                    rank_after_qa = "Ar_N/A"

                if rank_after_qa is None:
                    rank_after_qa = "Ar_N/A"

                directors = get_directors(
                    category_id=category.id, supplier_id=supplier.id
                )

                try:
                    category_bidder_details_worksheet.append(
                        (
                            "", "Ac_%d" % (count), supplier.company_name, rank_after_qa,
                            score_after_qa,
                            " ".join(directors[0].get("director_1").get("name", "Ar_N/A").split()),
                            " ".join(directors[1].get("director_2").get("name", "Ar_N/A").split()),
                            " ".join(directors[2].get("director_3").get("name", "Ar_N/A").split()),
                            " ".join(directors[3].get("director_4").get("name", "Ar_N/A").split()),
                            " ".join(directors[4].get("director_5").get("name", "Ar_N/A").split()
                            ),
                        )
                    )
                except Exception as e:
                    capture_exception(e)

            category_bidder_details_worksheet.freeze_panes = (
                category_bidder_details_worksheet["D6"]
            )
        workbook.save(filepath)
        format_excel(filepath, data=data)
        insert_image(
            excel_url=filepath,
            worksheet_name=qed_logo_worksheet,
            anchor=qed_logo_anchor,
        )
        insert_image(
            filepath, buyer_logo_worksheet, buyer_logo_anchor, image_url=buyer_logo_url
        )

        time = datetime.datetime.now()
        with open(filepath, "rb") as l:
            storage = PrivateMediaStorage()
            url = f"prequal/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
            storage.save(url, l)
            job_report, created = JobReport.objects.update_or_create(
                prequalification_id=job_id, defaults={"directors_report": url}
            )
            context = {
                "response_message": "Report generated successfully",
                "filepath": job_report.directors_report.url,
            }
            return context

    except Exception as e:
        capture_exception(e)
        print(e)
        context = {
            "response_message": "Report generation error"
        }
        return context


@shared_task(bind=True)
def participation_status_report(self, job_id):
    """
    Checks status of Invited suppliers against participation status
    """
    try:
        job = Prequalification.objects.get(id=job_id)
        company = job.company
        data = {}
        time = datetime.datetime.now()

        categories = Category.objects.filter(invite_only=True, prequalification_id=job_id)
        # categories = job.categories
        if len(categories) < 1:
            return {"response_message": "Job has no invite only categories"}
        else:
            result = 0
            progress_recorder = ProgressRecorder(self)
            total_for_progress = categories.count()

            prequal_participants = []
            for cat in categories:
                result += 1
                progress_recorder.set_progress(result, total_for_progress)

                participants = Supplier.objects.filter(
                    id__in=SupplierResponse.objects.filter(
                        question__section__category_id=cat.id
                    ).only('supplier_id').values('supplier_id').distinct()
                )
                invited_suppliers = Supplier.objects.filter(
                    id__in=CategoryInvite.objects.filter(
                        category_id=cat.id).only('supplier_id').values('supplier_id').distinct()
                )

                prequal_data = {
                    "category": cat,
                    "participants": participants,
                    "invitees": invited_suppliers,
                }

                prequal_participants.append(prequal_data)

            dir_name = Path(
                "media/prequal_participants/%s/%s/%s"
                % (company.company_name, time.year, job.title)
            )  # folder structure
            dir_name.mkdir(parents=True, exist_ok=True)
            # time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
            filepath = "{}/{}_Participation_Status_Report.xlsx".format(
                dir_name, job.unique_reference
            )
            # delete similar reports run this month
            match_string = "{}_{}".format(
                job.id, company.company_name.replace(" ", "_")
            )
            delete_matching_files_in_directory(dir_name, match_string, ".xlsx")
            # create and save a workbook to later populate
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = f"Summary"

            if company.company_logo_url is not None and company.company_logo_url != "":
                buyer_logo_url = get_file_path(company.company_logo_url)
            else:
                buyer_logo_url = "static/core/img/no-company-image128-128.png"
            qed_logo_anchor = "D1"
            qed_logo_worksheet = worksheet.title
            buyer_logo_anchor = "C1"
            buyer_logo_worksheet = worksheet.title

            worksheet["C5"] = "Nb_{}".format(job.title)
            worksheet["C6"] = "Nb_{}".format(job.unique_reference)
            worksheet["C7"] = "Nb_Prequal Participation Report"
            worksheet["C8"] = "Nb_"

            worksheet.append(
                (
                    "",
                    "",
                    "",
                    "hc_Participants",
                )
            )
            worksheet.append(
                ("", "hc_No", "h_Category Name", "h_Invited", "h_Yes", "h_No")
            )
            worksheet.merge_cells("D9:F9")

            data["Summary_C"] = 62
            data["Summary_D"] = 10
            data["Summary_E"] = 10
            data["Summary_F"] = 10

            total_invited = 0
            total_participants = 0
            total_non_participants = 0

            total_for_progress += len(prequal_participants)
            for count, prequal_data in enumerate(prequal_participants, start=1):
                result += 1
                progress_recorder.set_progress(result, total_for_progress)

                category = prequal_data["category"]
                invitees = prequal_data["invitees"]
                participants = prequal_data["participants"]
                # percentage_participation = "{:.2f}".format(
                #     len(participants) / len(invitees)
                # )
                if len(invitees) > 0:
                    percentage_participation = "{}%".format(
                        round(((len(participants) / len(invitees)) * 100))
                    )
                else:
                    percentage_participation = 0

                non_participants = len(prequal_data["invitees"]) - len(
                    prequal_data["participants"]
                )
                worksheet.append(
                    (
                        "",
                        "Ac_%d" % (count),
                        prequal_data["category"].name,
                        "Ac_{}".format(len(invitees)),
                        "Ac_{}".format(len(participants)),
                        "Ac_{}".format(non_participants),
                    )
                )

                total_invited += len(invitees)
                total_participants += len(participants)
                total_non_participants += non_participants

                # create individual sheets
                workbook.create_sheet(category.unique_reference.replace("/", "_"))
                details_worksheet = workbook[
                    category.unique_reference.replace("/", "_")
                ]
                empty_row = [""]
                empty_row = tuple(empty_row)
                details_worksheet.append(empty_row)
                details_worksheet["C2"] = "Nb_{}".format(category.name)
                details_worksheet["C3"] = "Nb_"
                details_worksheet["C4"] = "Number of bidders invited"
                details_worksheet["C5"] = "Number of bidders participated"
                details_worksheet["C6"] = "% Participation"
                details_worksheet["D4"] = "Ac_{}".format(len(invitees))
                details_worksheet["D5"] = "Ac_{}".format(len(participants))
                details_worksheet["D6"] = "Ac_{}".format(percentage_participation)
                details_worksheet["C7"] = "Nb_"

                details_worksheet.append(
                    (
                        "",
                        "",
                        "",
                        "",
                        "hc_Participation Status",
                    )
                )

                details_worksheet.append(
                    (
                        "",
                        "hc_No",
                        "h_Supplier",
                        "h_Phone Number",
                        "h_Yes",
                        "h_No",
                    )
                )
                details_worksheet.merge_cells("E8:F8")
                details_worksheet.merge_cells("C2:E2")
                # formating
                data[f'{category.unique_reference.replace("/", "_")}_C'] = 35
                data[f'{category.unique_reference.replace("/", "_")}_D'] = 15
                data[f'{category.unique_reference.replace("/", "_")}_E'] = 10
                data[f'{category.unique_reference.replace("/", "_")}_F'] = 10

                status_yes = ""
                status_no = ""
                for count, supplier in enumerate(invitees, start=1):
                    if supplier is not None:
                        if isinstance(supplier, str):
                            sup = Supplier.objects.filter(email=supplier).first()
                            if sup is not None:
                                company = sup.company_name
                                phone = sup.phone_number
                            else:
                                company = supplier
                                phone = "N/A"
                        else:
                            company = supplier.company_name
                            phone = supplier.phone_number
                            sup = supplier

                        if sup in participants:
                            status_yes = "Yes"
                            status_no = " "
                        else:
                            status_no = "No"
                            status_yes = " "

                        details_worksheet.append(
                            (
                                "",
                                "Ac_%d" % (count),
                                company,
                                phone,
                                status_yes,
                                status_no,
                            )
                        )

            worksheet.append(
                (
                    "",
                    "",
                    "Total",
                    "Ac_{}".format(total_invited),
                    "Ac_{}".format(total_participants),
                    "Ac_{}".format(total_non_participants),
                )
            )
            workbook.save(filepath)
            format_excel(filepath, data=data)
            insert_image(
                excel_url=filepath,
                worksheet_name=qed_logo_worksheet,
                anchor=qed_logo_anchor,
            )
            insert_image(
                filepath,
                buyer_logo_worksheet,
                buyer_logo_anchor,
                image_url=buyer_logo_url,
            )

            time = datetime.datetime.now()
            with open(filepath, "rb") as l:
                storage = PrivateMediaStorage()
                url = f"prequal/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
                storage.save(url, l)

                job_report, created = JobReport.objects.update_or_create(
                    prequalification_id=job_id, defaults={"participation_status_report": url}
                )
                context = {
                    "response_message": "Report generated successfully",
                    "filepath": job_report.participation_status_report.url,
                }
                return context
    except Exception as e:
        print(e)
        capture_exception(e)
        context = {
            "response_message": "Report generation error"
        }
        return context

# Current Suppliers Report


def suppliers_by_category(category):
    """
    Get the current suppliers by specific category
    """
    if category is not None:
        company = category.prequalification.company
        current_suppliers = apps.apps.get_model('core', 'CurrentSupplier').objects.filter(company=company)
        current_supplier_by_categories = []
        category_current_supliers = current_suppliers.filter(category_id=category.id)

        for cs in category_current_supliers:
            current_supplier_by_categories.append(cs)

        return current_supplier_by_categories


def category_participation_count(category):
    """
    Get the participation status for every category
    """
    cat_list = suppliers_by_category(category)
    non_participants = []
    participants = []
    total = []
    for sup in cat_list:
        total.append(sup)
        if sup.participation_status == "Not Participated":
            non_participants.append(sup)
        else:
            participants.append(sup)

    return {
        "non_participants": len(non_participants),
        "participants": len(participants),
        "total": len(total),
    }


@shared_task(bind=True)
def download_current_suppliers(self, job_id):
    data = {}
    job = Prequalification.objects.filter(id=job_id).first()
    company = job.company
    time = datetime.datetime.now()
    dir_name = Path(
        "media/current_suppliers/%s/%s/%s"
        % (company.company_name, time.year, job.title)
    )  # folder structure
    dir_name.mkdir(parents=True, exist_ok=True)
    time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
    filepath = "{}/{}_{}_{}.xlsx".format(
        dir_name, job.unique_reference, job.title, time_only
    )

    # delete similar reports run this month
    match_string = "{}_{}".format(job.id, company.company_name.replace(" ", "_"))
    delete_matching_files_in_directory(dir_name, match_string, ".xlsx")
    # create and save a workbook to later populate
    current_suppliers_workbook = Workbook()

    # create and save a workbook to later populate
    current_suppliers_worksheet = current_suppliers_workbook.active
    current_suppliers_worksheet.title = f"Summary"

    if company.company_logo_url is not None and company.company_logo_url != "":
        buyer_logo_url = get_file_path(company.company_logo_url)
    else:
        buyer_logo_url = "qed/static/img/no-company-image128-128.png"
    qed_logo_anchor = "D1"
    qed_logo_worksheet = current_suppliers_worksheet.title
    buyer_logo_anchor = "B1"
    buyer_logo_worksheet = current_suppliers_worksheet.title

    # current_suppliers_worksheet["C2"] = f"{job.company.company_name}"
    # current_suppliers_worksheet["C3"] = "Current Suppliers"
    # current_suppliers_worksheet["C4"] = ""
    current_suppliers_worksheet["C5"] = "Nb_{}".format(job.title)
    current_suppliers_worksheet["C6"] = "Nb_{}".format(job.unique_reference)
    current_suppliers_worksheet["C7"] = "Nb_Current Suppliers Report"
    current_suppliers_worksheet["C8"] = "Nb_"

    current_suppliers_worksheet.append(
        (
            "",
            "",
            "",
            "",
            "hc_Participation",
            "",
            "",
        )
    )
    current_suppliers_worksheet.merge_cells("C5:D5")
    current_suppliers_worksheet.merge_cells("E9:G9")
    current_suppliers_worksheet.append(
        ("", "hc_No", "h_Category Code", "h_Category Name", "h_Yes", "h_No", "h_Total")
    )
    data["Summary_C"] = 18
    data["Summary_D"] = 62
    data["Summary_E"] = 7.7
    data["Summary_F"] = 7.7
    data["Summary_G"] = 7.7
    current_suppliers = (
        apps.apps.get_model('core', 'CurrentSupplier').objects.filter(company=company, job_id=job.id, target=ContentType.objects.get_for_model(apps.apps.get_model('prequal', 'Category')))
        .order_by("category_id")
    )
    try:
        current_supplier_categories = []
        for cs in current_suppliers:
            if cs.category is not None:
                current_supplier_categories.append(cs.category)

        cat_list = []
        for x in current_supplier_categories:
            if x not in cat_list:
                cat_list.append(x)

        total_partcipants = 0
        total_non_participants = 0
        total_suppliers = 0

        result = 0
        progress_recorder = ProgressRecorder(self)
        total_for_progress = len(cat_list)

        for count, category in enumerate(cat_list, start=1):
            result += 1
            progress_recorder.set_progress(result, total_for_progress)

            sups = category_participation_count(category)
            current_suppliers_worksheet.append(
                (
                    "",
                    "Ac_%d" % (count),
                    category.unique_reference,
                    category.name,
                    "Ac_{}".format(sups["participants"]),
                    "Ac_{}".format(sups["non_participants"]),
                    "Ac_{}".format(sups["total"]),
                )
            )
            total_partcipants += sups["participants"]
            total_non_participants += sups["non_participants"]
            total_suppliers += sups["total"]

            current_suppliers_workbook.create_sheet(
                category.unique_reference.replace("/", "_")
            )
            current_suppliers_details_worksheet = current_suppliers_workbook[
                category.unique_reference.replace("/", "_")
            ]
            percentage_participation = "{}%".format(
                round(((sups["participants"] / sups["total"]) * 100))
            )
            empty_row = [""]
            empty_row = tuple(empty_row)
            current_suppliers_details_worksheet.append(empty_row)
            current_suppliers_details_worksheet["C2"] = "Nb_{}".format(category.name)
            current_suppliers_details_worksheet["C3"] = "Nb_{}".format(
                category.unique_reference
            )
            current_suppliers_details_worksheet["C4"] = "Nb_"
            current_suppliers_details_worksheet["C5"] = "Number of bidders invited"
            current_suppliers_details_worksheet["C6"] = "Number of bidders participated"
            current_suppliers_details_worksheet["C7"] = "% Participation"
            current_suppliers_details_worksheet["D5"] = "Ac_{}".format(sups["total"])
            current_suppliers_details_worksheet["D6"] = "Ac_{}".format(
                sups["participants"]
            )
            current_suppliers_details_worksheet["D7"] = "Ac_{}".format(
                percentage_participation
            )
            current_suppliers_details_worksheet["C8"] = "Nb_"

            current_suppliers_details_worksheet.merge_cells("C2:E2")
            data[f'category.unique_reference.replace("/", "_")_C'] = 25
            data[f'category.unique_reference.replace("/", "_")_D'] = 15
            data[f'category.unique_reference.replace("/", "_")_E'] = 16
            data[f'category.unique_reference.replace("/", "_")_F'] = 28
            data[f'category.unique_reference.replace("/", "_")_G'] = 28
            data[f'category.unique_reference.replace("/", "_")_H'] = 25

            current_suppliers_details_worksheet.append(
                (
                    "",
                    "hc_No",
                    "h_Company Name",
                    "h_Status",
                    "h_Tax PIN Number",
                    "h_Email",
                    "h_Alternative Email",
                    "h_Phone",
                )
            )

            current_suppliers_by_category = current_suppliers.filter(category_id=category.id)

            total_for_progress += len(current_suppliers_by_category)

            for count, category_supplier in enumerate(
                current_suppliers_by_category, start=1
            ):
                result += 1
                progress_recorder.set_progress(result, total_for_progress)
                current_suppliers_details_worksheet.append(
                    (
                        "",
                        "Ac_%d" % (count),
                        category_supplier.supplier_name,
                        category_supplier.participation_status,
                        category_supplier.get_tax_pin(),
                        category_supplier.supplier_email,
                        category_supplier.alternative_email,
                        category_supplier.supplier_phone,
                    )
                )
        current_suppliers_worksheet.append(
            (
                "",
                "",
                "",
                "h_Total",
                total_partcipants,
                total_non_participants,
                total_suppliers,
            )
        )
        current_suppliers_worksheet.freeze_panes = current_suppliers_worksheet["D11"]
        current_suppliers_workbook.save(filepath)
        format_excel(filepath, data=data)
        insert_image(
            excel_url=filepath,
            worksheet_name=qed_logo_worksheet,
            anchor=qed_logo_anchor,
        )
        insert_image(
            filepath, buyer_logo_worksheet, buyer_logo_anchor, image_url=buyer_logo_url
        )
        context = {
            "response_message": "Report generated successfully",
            "filepath": filepath.split("/", 1)[1],
        }

        return context

    except Exception as e:
        capture_exception(e)
        return True


@shared_task(bind=True)
def job_qed_category_suppliers(self, job_id):
    data = {}
    job = Prequalification.objects.filter(id=job_id).first()
    company = job.company
    categories = Category.objects.filter(prequalification_id=job_id)

    time = datetime.datetime.now()
    dir_name = Path(
        "media/category_types/jobs/%d/suppliers/%s" % (job_id, time.year)
    )  # folder structure
    dir_name.mkdir(parents=True, exist_ok=True)
    time_only = "%d_%d_%d" % (time.month, time.day, time.hour)
    all_suppliers = "QED_Category_Suppliers"
    filepath = "{}/{}_{}.xlsx".format(dir_name, all_suppliers, time_only)

    # delete similar reports run this month
    match_string = "{}_{}".format(all_suppliers, time_only)
    delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

    # create and save a workbook to later populate
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    buyer_logo_url = "static/core/img/QED2.png"
    buyer_logo = Image("static/core/img/QED2.png")
    buyer_logo_height, buyer_logo_width = buyer_logo.height, buyer_logo.width
    max_height = 75
    ratio = buyer_logo_height / max_height
    buyer_logo.height, buyer_logo.width = buyer_logo_height, buyer_logo_width

    buyer_logo.anchor = summary_sheet["C1"]
    summary_sheet.add_image(buyer_logo, "C1")
    summary_sheet.title = "Summary"
    logo_worksheet = summary_sheet.title
    summary_sheet.append(("", "", ""))
    summary_sheet.append(("", "", ""))
    summary_sheet.append(("", "", ""))
    summary_sheet.append(("", "", ""))
    summary_sheet.append(
        (
            "",
            "hc_No",
            "h_QED Code",
            "h_Category Name",
            "h_Category Code",
            "h_Suppliers Count",
            "h_Paid Bidders",
        )
    )
    data["Summary_C"] = 18
    data["Summary_D"] = 62
    data["Summary_E"] = 7.7
    data["Summary_F"] = 7.7
    data["Summary_G"] = 7.7

    # try:
    result = 0
    progress_recorder = ProgressRecorder(self)
    total_for_progress = categories.count()

    total_suppliers_count = 0
    total_paid_bidders = 0
    for count, category in enumerate(categories, start=1):
        result += 1
        progress_recorder.set_progress(result, total_for_progress)

        suppliers = category.category_type.suppliers_list(job.company.country)
        category_total_bidders = category.total_bidders
        category_type = category.category_type
        summary_sheet.append(
            (
                "",
                "Ac_%d" % (count),
                category.category_type.innitials,
                category.name,
                category.unique_reference,
                suppliers["supplier_count"],
                category_total_bidders,
            )
        )
        total_paid_bidders += category_total_bidders
        total_suppliers_count += int(suppliers["supplier_count"])

        # create and save a workbook to later populate
        category_name = (
            f"{category.unique_reference}_{category_type.innitials}"
        )

        category_type_worksheet = workbook.create_sheet(category_name)
        category_type_worksheet.title = category_name
        category_type_worksheet["C2"] = "Nb_Category Name"
        category_type_worksheet["D2"] = "Nb_{}".format(category.name)
        category_type_worksheet["C3"] = "Nb_QED Category".format(
            category_type.name)

        category_type_worksheet["D3"] = "Nb_{}".format(category_type.name)
        category_type_worksheet["C4"] = "Nb_QED Code".format(
            category_type.innitials)
        category_type_worksheet["D4"] = "Nb_{}".format(
            category_type.innitials)
        category_type_worksheet["C5"] = "Nb_"

        category_type_worksheet.merge_cells("D2:F2")
        category_type_worksheet.merge_cells("D3:F3")

        data[f"category.unique_reference_category.category_type.innitials_C"] = 30
        data[f"category.unique_reference_category.category_type.innitials_D"] = 15
        data[f"category.unique_reference_category.category_type.innitials_E"] = 28
        data[f"category.unique_reference_category.category_type.innitials_F"] = 28
        data[f"category.unique_reference_category.category_type.innitials_G"] = 28
        data[f"category.unique_reference_category.category_type.innitials_H"] = 28
        data[f"category.unique_reference_category.category_type.innitials_I"] = 10

        category_type_worksheet.append(
            (
                "", "hc_No", "h_Company_Name", "h_Status (Paid/Not_Paid)",
                "h_Primary Email", "h_Alternative Email", "h_Primary Phone",
                "h_Alternative Phone", "h_Country", "h_Location 1",
                "h_Location 2", "h_Location 3", "h_Location 4", "h_Location 5",
                "h_Location 6", "h_Location 7", "h_Location 8", "h_Location 9",
                "h_Location 10", "h_Other Locations",
            )
        )

        total_for_progress += len(suppliers["registered_suppliers"])
        for count, supplier in enumerate(
            suppliers["registered_suppliers"], start=1
        ):
            result += 1
            progress_recorder.set_progress(result, total_for_progress)
            category_type = category.category_type

            supplier_paid_categories = Category.objects.filter(
                id__in=apps.apps.get_model('core', 'CategoryOrder').objects.filter(
                    target__model='prequalification', supplier_id=supplier.id
                ).only('category_id').values('category_id').distinct()
            )

            cat_type_supplier = CategoryTypeSupplier.objects.filter(
                supplier=supplier, category_type=category_type
            ).first()

            supplier_locations = get_cat_supplier_locations(cat_type_supplier)
            try:
                location_1 = supplier_locations[0]
            except:
                location_1 = "N/A"
            try:
                location_2 = supplier_locations[1]
            except:
                location_2 = "N/A"
            try:
                location_3 = supplier_locations[2]
            except:
                location_3 = "N/A"
            try:
                location_4 = supplier_locations[3]
            except:
                location_4 = "N/A"
            try:
                location_5 = supplier_locations[4]
            except:
                location_5 = "N/A"
            try:
                location_6 = supplier_locations[5]
            except:
                location_6 = "N/A"
            try:
                location_7 = supplier_locations[6]
            except:
                location_7 = "N/A"
            try:
                location_8 = supplier_locations[7]
            except:
                location_8 = "N/A"
            try:
                location_9 = supplier_locations[8]
            except:
                location_9 = "N/A"
            try:
                location_10 = supplier_locations[9]
            except:
                location_10 = "N/A"

            if len(supplier_locations) > 10:
                try:
                    others = supplier_locations[10:]
                except:
                    others = ["NA"]
                others = ",".join(others)
            others = "NA"

            if category in supplier_paid_categories:
                status = "Paid"
            else:
                status = "Not Paid"

            category_type_worksheet.append(
                (
                    "", "Ac_%d" % (count), supplier.company_name, status,
                    supplier.email, " ", supplier.phone_number, " ", supplier.country,
                    location_1, location_2, location_3, location_4, location_5,
                    location_6, location_7, location_8, location_9, location_10, others,
                )
            )

        for count, supplier in enumerate(
            suppliers["old_suppliers"],
            start=len(suppliers["registered_suppliers"]) + 1,
        ):
            status = "Not Paid"

            if supplier.alternative_email == "":
                alternative_email = " "
            else:
                alternative_email = supplier.alternative_email

            if supplier.primary_phone == "":
                primary_phone = " "
            else:
                primary_phone = supplier.primary_phone

            if supplier.alternative_phone == "":
                alternative_phone = " "
            else:
                alternative_phone = supplier.alternative_phone

            category_type_worksheet.append(
                (
                    "", "Ac_%d" % (count), supplier.company_name, status,
                    supplier.primary_email, alternative_email, primary_phone,
                    alternative_phone, supplier.country,
                )
            )
        category_type_worksheet.freeze_panes = category_type_worksheet["D7"]
    summary_sheet.append(
        ("", "", "", "", "Total", total_suppliers_count, total_paid_bidders)
    )
    summary_sheet.freeze_panes = summary_sheet["E7"]
    workbook.save(filepath)
    format_excel(filepath)
    filepath = insert_image(
        filepath, logo_worksheet, buyer_logo.anchor, image_url=buyer_logo_url
    )

    time = datetime.datetime.now()
    with open(filepath, "rb") as l:
        storage = PrivateMediaStorage()
        url = f"prequal/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
        storage.save(url, l)

        job_report, created = JobReport.objects.update_or_create(
            prequalification_id=job_id, defaults={"category_suppliers_report": url}
        )
        context = {
            "response_message": "Report generated successfully",
            "filepath": job_report.category_suppliers_report.url,
        }
        return context


    # except Exception as e:
    #     capture_exception(e)
    #     print(e)
    #     return False


class FinancialRatiosReport(Task):
    name = "TenderFinancialRatiosReport"
    progress_recorder = None
    result = 0

    def run(self, *args, **kwargs):
        self.progress_recorder = ProgressRecorder(self)
        context = self.report(kwargs["category_id"])
        return context

    def report(self, category_id):
        # try:
        print(category_id)
        category = Category.objects.get(id=category_id)
        company = category.prequalification.company

        section = category.sections.filter(name="Financial Ratios").first()

        time = datetime.datetime.now()
        dir_name = Path(
            "media/category/ratios/reports/{}/{}/{}".format(
                time.year, time.month, time.day
            )
        )
        dir_name.mkdir(parents=True, exist_ok=True)
        time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
        filepath = "{}/{}_{}_{}.xlsx".format(
            dir_name, category_id, company.company_name.replace(" ", "_"), time_only
        )

        workbook = Workbook()
        summary_sheet = workbook["Sheet"]
        summary_sheet.title = "Summary"

        qed_logo_anchor = "H1"
        buyer_logo_anchor = "B1"
        qed_logo = Image("static/core/img/Tendersure_Logo.png")

        if company.company_logo_url is not None and company.company_logo_url != "":
            buyer_logo = get_file_path(company.company_logo_url)
        else:
            buyer_logo = "static/core/img/no-company-image128-128.png"

        summary_sheet.append(("", "", ""))
        summary_sheet.append(("", "", ""))
        summary_sheet.append(("", "", ""))
        summary_sheet.append(("", "", ""))
        summary_sheet.append(("", "", ""))
        summary_sheet["B6"] = f"Bnb_{category.name.title()}"
        summary_sheet["B7"] = ""
        summary_sheet["B8"] = f"Bnb_{'Financial Ratio Summary Report'}"
        summary_sheet.append(("", "", ""))
        summary_sheet.merge_cells("B6:E6")
        summary_sheet.merge_cells("B8:E8")
        summary_sheet.append(
            (
                " ", "hc_#", "hl_Supplier Name", "hc_Score Before QA",
                "hc_Score After QA", "hc_Variance", "hl_Comments",
            )
        )

        data = {
            "Summary_B": 5, "Summary_C": 41, "Summary_D": 15,
            "Summary_E": 15, "Summary_F": 15, "Summary_G": 15,
        }

        progress_recorder = ProgressRecorder(self)
        result = 0

        responsive_bidders = category.responsive_bidders

        supplier_count = 1
        for supplier in responsive_bidders:
            if section is not None:
                score_before_qa = SupplierSectionScore.objects.filter(
                    supplier=supplier, section=section
                ).first()

                question_in_qa = QualityAssuranceResponse.objects.filter(
                    quality_assurance_question__question__section=section,
                    supplier__id=supplier.id,
                )
                if question_in_qa:
                    total_score = 0
                    for question in question_in_qa:
                        if question.score_after_qa is not None:
                            total_score += float(question.score_after_qa)
                    score_after_qa = total_score
                else:
                    score_after_qa = 0

            supplier_name = supplier.short_name.replace("/", "_") if supplier.short_name is not None else supplier.company_name.replace("/", "_")
            supplier_worksheet = workbook.create_sheet(
                supplier_name
            )
            data[f"{supplier_name}_B"] = 20
            data[f"{supplier_name}_C"] = 15
            data[f"{supplier_name}_D"] = 15
            data[f"{supplier_name}_E"] = 15
            data[f"{supplier_name}_F"] = 15
            data[f"{supplier_name}_G"] = 15
            data[f"{supplier_name}_H"] = 15

            result += 1
            progress_recorder.set_progress(result, len(responsive_bidders))

            if score_before_qa is not None:
                variance = float(score_after_qa) - float(score_before_qa.score)
            else:
                variance = 0

            summary_sheet.append(
                (
                    " ",
                    f"Ac_{supplier_count}",
                    f"Al_{supplier_name}",
                    f"Nc_{float(score_before_qa.score)}"
                    if score_before_qa is not None
                    else f"Nc_{0}",
                    f"Nc_{score_after_qa}",
                    f"Nc_{variance}",
                )
            )

            supplier_count += 1

            supplier_worksheet.append(("", ""))
            supplier_worksheet[
                "B2"
            ] = f"Bnb_Supplier: {supplier.company_name.title()}"
            supplier_worksheet["B3"] = ""
            supplier_worksheet["B4"] = f"Bnb_{'Latest Financial Statements'}"
            supplier_worksheet.append(("", ""))
            supplier_worksheet.append(("", ""))

            worksheet_headers = [
                " ",
                "hl_Description",
                "hc_Supplier (KES)",
                "hc_QA (KES)",
                "hc_Variance (KES)",
                "hc_Comment",
            ]

            supplier_worksheet.merge_cells("B2:E2")
            supplier_worksheet.merge_cells("B4:E4")

            supplier_worksheet.append(worksheet_headers)
            financial_ratio = FinancialRatio.objects.filter(
                supplier=supplier, section=section
            ).first()
            if financial_ratio is not None:
                equity = [
                    "",
                    "Equity",
                    f"Tr_{float(financial_ratio.equity)}"
                    if financial_ratio.equity is not None
                    else 0,
                    f"Tr_{float(financial_ratio.equity_after_qa)}"
                    if financial_ratio.equity_after_qa is not None
                    else 0,
                
                ]
                long_term_loans = [
                    "",
                    "Long Term Loans (Debt)",
                    f"Tr_{float(financial_ratio.debtors)}"
                    if financial_ratio.debtors is not None
                    else 0,
                    f"Tr_{float(financial_ratio.debtors_after_qa)}"
                    if financial_ratio.debtors_after_qa is not None
                    else 0,
                ]
                curr_liabilities = [
                    "",
                    "Current Liabilities",
                    f"Tr_{float(financial_ratio.curr_liabilities)}"
                    if financial_ratio.curr_liabilities is not None
                    else 0,
                    f"Tr_{float(financial_ratio.curr_liabilities_after_qa)}"
                    if financial_ratio.curr_liabilities_after_qa is not None
                    else 0,
                
                ]
                fixed_assets = [
                    "",
                    "Fixed Assets",
                    f"Tr_{float(financial_ratio.fixed_assets)}"
                    if financial_ratio.fixed_assets is not None
                    else 0,
                    f"Tr_{float(financial_ratio.fixed_assets_after_qa)}"
                    if financial_ratio.fixed_assets_after_qa is not None
                    else 0,
                    
                ]
                curr_assets = [
                    "",
                    "Curent Assets",
                    f"Tr_{float(financial_ratio.current_assets)}"
                    if financial_ratio.current_assets is not None
                    else 0,
                    f"Tr_{float(financial_ratio.current_assets_after_qa)}"
                    if financial_ratio.current_assets_after_qa is not None
                    else 0,
                ]
                cash = [
                    "",
                    "Cash",
                    f"Tr_{float(financial_ratio.cash)}"
                    if financial_ratio.cash is not None
                    else 0,
                    f"Tr_{float(financial_ratio.cash_after_qa)}"
                    if financial_ratio.cash_after_qa is not None
                    else 0,
                ]
                turnover = [
                    "",
                    "Turnover",
                    f"Tr_{float(financial_ratio.turnover)}"
                    if financial_ratio.turnover is not None
                    else 0,
                    f"Tr_{float(financial_ratio.turnover_after_qa)}"
                    if financial_ratio.turnover_after_qa is not None
                    else 0,
                ]
                gross_profit = [
                    "",
                    "Gross Profit",
                    f"Tr_{float(financial_ratio.gross_profit)}"
                    if financial_ratio.gross_profit is not None
                    else 0,
                    f"Tr_{float(financial_ratio.gross_profit_after_qa)}"
                    if financial_ratio.gross_profit_after_qa is not None
                    else 0,
                ]
                net_profit = [
                    "",
                    "Net Profit",
                    f"Tr_{float(financial_ratio.net_profit)}"
                    if financial_ratio.net_profit is not None
                    else 0,
                    f"Tr_{float(financial_ratio.net_profit_after_qa)}"
                    if financial_ratio.net_profit_after_qa is not None
                    else 0,
                ]

                variances = variance_ratios(financial_ratio.id)

                equity.extend([f'Tr_{variances["equity"]["variance"]}', variances["equity"]["comment"]])
                long_term_loans.extend([f'Tr_{variances["debtors"]["variance"]}', variances["debtors"]["comment"]])
                curr_liabilities.extend([f'Tr_{variances["curr_liabilities"]["variance"]}', variances["curr_liabilities"]["comment"]])
                fixed_assets.extend([f'Tr_{variances["f_assets"]["variance"]}', variances["f_assets"]["comment"]])
                curr_assets.extend([f'Tr_{variances["c_assets"]["variance"]}', variances["c_assets"]["comment"]])
                cash.extend([f'Tr_{variances["cash"]["variance"]}', variances["cash"]["comment"]])
                turnover.extend([f'Tr_{variances["turnover"]["variance"]}', variances["turnover"]["comment"]])
                gross_profit.extend([f'Tr_{variances["g_profit"]["variance"]}', variances["g_profit"]["comment"]])
                net_profit.extend([f'Tr_{variances["n_profit"]["variance"]}', variances["n_profit"]["comment"]])

                supplier_worksheet.append(equity)
                supplier_worksheet.append(long_term_loans)
                supplier_worksheet.append(curr_liabilities)
                supplier_worksheet.append(fixed_assets)
                supplier_worksheet.append(curr_assets)
                supplier_worksheet.append(cash)
                supplier_worksheet.append(turnover)
                supplier_worksheet.append(gross_profit)
                supplier_worksheet.append(net_profit)
                supplier_worksheet.append(("", "", ""))
                supplier_worksheet.append(("", "", ""))
                supplier_worksheet.append(("", "", ""))

                ratio_headers = [
                    " ",
                    "hl_Ratios",
                    "hc_Before QA",
                    "hc_After QA",
                    "hc_Variance",
                    "hc_Score Before QA",
                    "hc_Score After QA",
                    "hc_Variance",
                ]
                supplier_worksheet.append(ratio_headers)
                ratios = ratios_before_after_qa(instance_id=financial_ratio.id)
                debt_equity = [
                    "",
                    "Debt/equity ratio",
                    "Tc_{}".format(ratios["debt_equity"]["before_qa"]),
                    "Tc_{}".format(ratios["debt_equity"]["after_qa"]),
                    "Tc_{}".format(ratios["debt_equity"]["variance"]),
                ]
                current_ratio = [
                    "",
                    "Current ratio",
                    "Tc_{}".format(ratios["current_ratio"]["before_qa"]),
                    "Tc_{}".format(ratios["current_ratio"]["after_qa"]),
                    "Tc_{}".format(ratios["current_ratio"]["variance"]),
                ]
                cash_ratio = [
                    "",
                    "Cash ratio",
                    "Tc_{}".format(ratios["cash_ratio"]["before_qa"]),
                    "Tc_{}".format(ratios["cash_ratio"]["after_qa"]),
                    "Tc_{}".format(ratios["cash_ratio"]["variance"]),
                ]
                gp_margin = [
                    "",
                    "GP Margin",
                    "Tc_{}".format(ratios["gp_margin"]["before_qa"]),
                    "Tc_{}".format(ratios["gp_margin"]["after_qa"]),
                    "Tc_{}".format(ratios["gp_margin"]["variance"]),
                ]
                np_margin = [
                    "",
                    "NP Margin",
                    "Tc_{}".format(ratios["np_margin"]["before_qa"]),
                    "Tc_{}".format(ratios["np_margin"]["after_qa"]),
                    "Tc_{}".format(ratios["np_margin"]["variance"]),
                ]

                scores = ratio_scores_before_after_qa(
                    supplier_id=supplier.id, section_id=section.id
                )

                for score in scores:
                    if score["index"] == 0:
                        ratio_debt_equity = [
                            "Tc_{}".format(score["before_qa"]),
                            "Tc_{}".format(score["after_qa"]),
                            "Tc_{}".format(score["variance"]),
                        ]
                        debt_equity.extend(ratio_debt_equity)
                    if score["index"] == 1:
                        ratio_current = [
                            "Tc_{}".format(score["before_qa"]),
                            "Tc_{}".format(score["after_qa"]),
                            "Tc_{}".format(score["variance"]),
                        ]
                        current_ratio.extend(ratio_current)
                    if score["index"] == 2:
                        ratio_cash = [
                            "Tc_{}".format(score["before_qa"]),
                            "Tc_{}".format(score["after_qa"]),
                            "Tc_{}".format(score["variance"]),
                        ]
                        cash_ratio.extend(ratio_cash)
                    if score["index"] == 3:
                        ratio_gp = [
                            "Tc_{}".format(score["before_qa"]),
                            "Tc_{}".format(score["after_qa"]),
                            "Tc_{}".format(score["variance"]),
                        ]
                        gp_margin.extend(ratio_gp)
                    if score["index"] == 4:
                        ratio_np = [
                            "Tc_{}".format(score["before_qa"]),
                            "Tc_{}".format(score["after_qa"]),
                            "Tc_{}".format(score["variance"]),
                        ]
                        np_margin.extend(ratio_np)

                supplier_worksheet.append(debt_equity)
                supplier_worksheet.append(current_ratio)
                supplier_worksheet.append(cash_ratio)
                supplier_worksheet.append(gp_margin)
                supplier_worksheet.append(np_margin)

            formulas_row = []
            formulas_row.extend(["", "", "", "", "qT_Total"])

            formulas_row.extend(["qT_{}".format(f"= SUM(F21:F25)")])
            formulas_row.extend(["qT_{}".format(f"= SUM(G21:G25)")])
            formulas_row.extend(["qT_{}".format(f"= SUM(H21:H25)")])
            supplier_worksheet.append(tuple(formulas_row))

        workbook.save(filepath)
        format_excel(filepath, data=data)
        # insert_image(
        #     excel_url=filepath,
        #     worksheet_name="Summary",
        #     anchor=qed_logo_anchor,
        # )

        insert_image(
            excel_url=filepath,
            worksheet_name="Summary",
            anchor=buyer_logo_anchor,
            image_url=buyer_logo,
        )

        time = datetime.datetime.now()
        with open(filepath, "rb") as l:
            storage = PrivateMediaStorage()
            url = f"prequal/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
            storage.save(url, l)

            category_report, created = CategoryReport.objects.update_or_create(
                category_id=category_id, 
                defaults={"financial_ratios": url}
            )
            path = category_report.financial_ratios .url
            context = {
                "filepath": path,
                "response_message": "Report generated successfully",
            }
            return context

        # except Exception as e:
            # capture_exception(e)
            # context = {"response_message": "Error generating report", "errors": f"{e}"}
            # return context