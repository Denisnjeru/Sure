import datetime
import imp
import math
import os
from pathlib import Path
from zipfile import ZipFile

import pytz
import requests
from celery import shared_task, Task
from celery_progress.backend import ProgressRecorder
from django import apps
from django.contrib.auth.models import User
from django.core.files import File
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.timezone import make_aware
from django.contrib.contenttypes.models import ContentType
from openpyxl import load_workbook
from sentry_sdk import capture_exception
from backend.celery import app
from backend.settings.base import BASE_DIR
from backend.storage_backends import PrivateMediaStorage
from .models import Prequalification, Category, QaNcaaResponse, Section, Question, MarkingScheme, SupplierResponse, \
    SupplierSectionScore, SupplierCategoryScore, QualityAssuranceResponse, QualityAssurance, QualityAssuranceQuestion, \
    ClientDocument, AwardLetter, QaTccResponse, QaCr12Response, QaBusinessPermitResponse, QaPoisonsBoardResponse, \
    QaPinCertificateResponse, QaIncorporationCertificateResponse
from apps.ocr.ocr import verify_nca, verify_tcc, verify_cr12, verify_businesspermit, verify_poisonsboardcert, verify_pincert, \
    verify_incorporationcert
from .reports import QARankingReport, PrequalInterimReport, DueDiligenceRankingReport, FinancialRatiosReport
from .utils import get_file
from ..core.utils import weasy_pdf
from .email_notifications import SendEmailNotifications, SendSMSNotifications


def job_questions_upload(job_id):
    messages = []
    job = Prequalification.objects.get(id=job_id)
    categories = Category.objects.filter(prequalification_id=job_id)
    if categories.count() > 0:

        try:
            # A = PrivateMediaStorage()
            # headers = {"ResponseContentDisposition": f"attachment;"}
            # time = datetime.datetime.now()
            # file_url = A.url(
            #     f"{job.questions_template}", expire=300, parameters=headers, http_method="GET",
            # )
            # dir_name = Path("media/temp/{}/{}/{}".format(time.year, time.month, time.day))
            # dir_name.mkdir(parents=True, exist_ok=True)
            # file_name = os.path.basename(f"{job.questions_template}")
            # filepath = "{}/{}".format(dir_name, file_name)
            # r = requests.get(file_url)
            # with open("{}".format(filepath), "wb") as f:
            #     f.write(r.content)

            workbook = load_workbook(job.question_template, data_only=True)

            worksheets = workbook.sheetnames
            if len(worksheets) < 1:
                messages.append('There are no sections to upload from the excel')
                context = {
                    "messages": messages
                }
                return context

            data = upload_question_data(workbook=workbook, worksheets=worksheets, job=job)

            messages.append(
                "Imports succeeded. {} new sections and {} new questions added to {} categories.".format(
                    data['sections_count'], data['questions_count'], data['categories_count']
                ),
            )
            context = {
                "messages": messages
            }
            return context

        except Exception as e:
            # for category in categories:
            #     for section in category.sections:
            #         section.delete()
            print(e)
            messages.append(f"{str(e)}")
            messages.append("There is a problem with your upload excel. Check instructions carefully, correct, upload.")
            context = {
                "messages": messages
            }
            return context

    else:
        context = {
            "messages": messages
        }
        return context


@shared_task(bind=True)
def category_questions_upload(self, job_id, category_id):
    messages = []
    categories_count = 0
    questions_count = 0
    sections_count = 0
    job = Prequalification.objects.get(id=job_id)
    category = Category.objects.filter(id=category_id).first()

    progress_recorder = ProgressRecorder(self)
    result = 0

    if category is not None:

        try:
            # A = PrivateMediaStorage()
            # headers = {"ResponseContentDisposition": f"attachment;"}
            # time = datetime.datetime.now()
            # file_url = A.url(
            #     f"{category.questions_template}", expire=300, parameters=headers, http_method="GET",
            # )
            # dir_name = Path("media/temp/{}/{}/{}".format(time.year, time.month, time.day))
            # dir_name.mkdir(parents=True, exist_ok=True)
            # file_name = os.path.basename(f"{job.questions_template}")
            # filepath = "{}/{}".format(dir_name, file_name)
            # r = requests.get(file_url)
            # with open("{}".format(filepath), "wb") as f:
            #     f.write(r.content)

            workbook = load_workbook(category.question_template, data_only=True)

            worksheets = workbook.sheetnames[1:]
            if len(worksheets) < 1:
                messages.append('There are no sections to upload from the excel')
                context = {
                    "messages": messages
                }
                return context

            category_sheet = workbook.get_sheet_by_name('Questions')

            trans_category_name = category_sheet["D2"].value
            category_short_name = category_sheet["C3"].value
            category.category_short_name = category_short_name
            category.trans_name = trans_category_name
            category.save()
            categories_count += 1

            # delete all old questions and sections
            sections = Section.objects.filter(category_id=category_id)
            sections.delete()

            total_for_progress = category_sheet.max_row - 5
            for i in range(5, category_sheet.max_row + 1):
                # for each of the rows create a question after checking for existing section
                section_name = category_sheet["A{}".format(i)].value
                trans_name = category_sheet["B{}".format(i)].value
                short_name = category_sheet["C{}".format(i)].value
                old_sections = Section.objects.filter(
                    name=section_name, category=category
                )
                if old_sections.count() > 0:
                    section = old_sections.first()
                else:
                    section = Section.objects.create(
                        name=section_name,
                        trans_name=trans_name,
                        short_name=short_name,
                        description="Questions in {}".format(
                            section_name
                        ),
                        category=category,
                    )
                    sections_count += 1

                trans_description = category_sheet["F{}".format(i)].value
                if trans_description == "" or trans_description == "Question short description":
                    question_short_description = category_sheet["D{}".format(i)].value
                    trans_short_description = category_sheet["E{}".format(i)].value
                else:
                    question_short_description = category_sheet["G{}".format(i)].value
                    trans_short_description = category_sheet["H{}".format(i)].value

                q = Question.objects.create(
                    section=section,
                    description=category_sheet["E{}".format(i)].value,
                    trans_description=trans_description,
                    short_description=question_short_description,
                    trans_short_description=trans_short_description,
                    answer_type=category_sheet["I{}".format(i)].value,
                    is_required=category_sheet["J{}".format(i)].value,
                    is_scored=category_sheet["K{}".format(i)].value,
                    max_score=category_sheet["L{}".format(i)].value,
                    is_qa=category_sheet["M{}".format(i)].value,
                    is_dd=category_sheet["N{}".format(i)].value,
                )
                questions_count += 1

                if q.answer_type == Question.TYPE_CHECKBOX or q.answer_type == Question.TYPE_SELECT:
                    MarkingScheme.objects.create(
                        question_id=q.id,
                        options=category_sheet["O{}".format(i)].value,
                        score=category_sheet["P{}".format(i)].value,
                    )

                result += 1
                progress_recorder.set_progress(result, total_for_progress)

            messages.append(
                "Imports succeeded. {} new sections and {} new questions added to {} categories.".format(
                    sections_count, questions_count, categories_count
                ),
            )
            context = {
                "messages": messages,
                "response_message": "Template processed successfully",
                "categories_count": categories_count,
                "questions_count": questions_count,
                "sections_count": sections_count
            }
            return context

        except Exception as e:
            sections = Section.objects.filter(category_id=category_id)
            sections.delete()

            messages.append(f"{str(e)}")
            messages.append("There is a problem with your upload excel. Check instructions carefully, correct, upload.")
            context = {
                "messages": messages,
                "response_message": "Template error",
            }
            print(f"{e}")
            return context

    else:
        context = {
            "messages": messages,
            "response_message": "Template error",
        }
        return context


def upload_question_data(workbook, worksheets, job):
    categories_count = 0
    questions_count = 0
    sections_count = 0
    for sheet in worksheets:
        category_sheet = workbook.get_sheet_by_name(sheet)
        category = Category.objects.filter(prequalification_id=job.id, unique_reference=sheet).first()

        if category is not None:
            trans_category_name = category_sheet["D2"].value
            category_short_name = category_sheet["C3"].value
            category.category_short_name = category_short_name
            category.trans_name = trans_category_name
            category.save()
            categories_count += 1
            # delete all old questions and sections
            for section in category.sections:
                section.delete()

            for i in range(5, category_sheet.max_row + 1):
                # for each of the rows create a question after checking for existing section
                section_name = category_sheet["A{}".format(i)].value
                trans_name = category_sheet["B{}".format(i)].value
                short_name = category_sheet["C{}".format(i)].value
                old_sections = Section.objects.filter(
                    name=section_name, category_id=category.id
                )
                if old_sections.count() > 0:
                    section = old_sections.first()
                else:
                    print(section_name)
                    section = Section.objects.create(
                        name=section_name, trans_name=trans_name,
                        short_name=short_name, category_id=category.id,
                        description="Questions in {}".format(
                            section_name
                        ),
                    )
                    sections_count += 1

                trans_description = category_sheet["F{}".format(i)].value
                if trans_description == "" or trans_description == "Question short description":
                    question_short_description = category_sheet["D{}".format(i)].value
                    trans_short_description = category_sheet["E{}".format(i)].value
                else:
                    question_short_description = category_sheet["G{}".format(i)].value
                    trans_short_description = category_sheet["H{}".format(i)].value

                q = Question.objects.create(
                    section=section,
                    description=category_sheet["E{}".format(i)].value,
                    trans_description=trans_description,
                    short_description=question_short_description,
                    trans_short_description=trans_short_description,
                    answer_type=category_sheet["I{}".format(i)].value,
                    is_required=category_sheet["J{}".format(i)].value,
                    is_scored=category_sheet["K{}".format(i)].value,
                    max_score=category_sheet["L{}".format(i)].value,
                    is_qa=category_sheet["M{}".format(i)].value,
                    is_dd=category_sheet["N{}".format(i)].value,
                )
                questions_count += 1

                if q.answer_type == Question.TYPE_CHECKBOX or q.answer_type == Question.TYPE_SELECT:
                    MarkingScheme.objects.create(
                        question_id=q.id,
                        options=category_sheet["O{}".format(i)].value,
                        score=category_sheet["P{}".format(i)].value,
                    )

    context = {
        "categories_count": categories_count,
        "questions_count": questions_count,
        "sections_count": sections_count
    }
    return context


class EvaluatePrequal(Task):
    name = "EvaluatePrequal"

    def run(self, *args, **kwargs):
        context = self.calculate_category_scores(kwargs["category_id"])
        return context

    def calculate_category_scores(self, category_id):
        result = 0
        total_for_progress = 3
        progress_recorder = ProgressRecorder(self)

        sections = Section.objects.filter(category_id=category_id)
        supplier_responses = SupplierResponse.objects.filter(question__section__category_id=category_id)

        suppliers = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
            id__in=supplier_responses.only('supplier_id').values('supplier_id').distinct()
        )
        qa = QualityAssurance.objects.filter(category_id=category_id).first()
        qa_responses = QualityAssuranceResponse.objects.filter(
            quality_assurance_question__quality_assurance__category_id=category_id
        )
        questions = Question.objects.filter(section__category_id=category_id, is_scored=True)
        for supplier in suppliers:
            supplier_score = 0
            for section in sections:
                section_questions = questions.filter(section_id=section.id)
                section_score = 0
                for question in section_questions:
                    score = 0
                    if question.is_scored:
                        question_response = supplier_responses.filter(
                            supplier_id=supplier.id, question_id=question.id).first()
                        if (question_response is not None and question_response.options is not None
                                and len(question_response.options) > 0):
                            if question.answer_type == Question.TYPE_UPLOAD:
                                score = float(question.max_score)
                            elif question_response.options in question.options:
                                response_index = question.options.index(question_response.options)
                                score = float(question.scores[response_index])
                            elif question_response.options == "True":
                                score = float(question.max_score)
                    section_score += score

                SupplierSectionScore.objects.update_or_create(
                    section_id=section.id,
                    supplier_id=supplier.id,
                    defaults={"score": section_score}
                )
                supplier_score += section_score

            technically_qualified = False
            if supplier_score >= 50:
                technically_qualified = True

            meets_all_mandatory_requirements = True
            qa_responses = qa_responses.filter(supplier_id=supplier.id)
            if qa_responses.filter(outcome="Fail").count() > 0:
                meets_all_mandatory_requirements = False
            elif qa_responses.count() < 1:
                meets_all_mandatory_requirements = False

            # calculate score after qa
            if qa:
                supplier_responses = supplier_responses.filter(supplier_id=supplier.id)
                score_after_qa = self.calculate_qa_scores(category_id, supplier, supplier_responses, questions)
            else:
                score_after_qa = None

            SupplierCategoryScore.objects.update_or_create(
                category_id=category_id,
                supplier_id=supplier.id,
                defaults={
                    "score": supplier_score,
                    "score_after_qa": score_after_qa,
                    "meets_all_mandatory_requirements": meets_all_mandatory_requirements,
                    "technically_qualified": technically_qualified
                },
            )

        result += 1
        progress_recorder.set_progress(result, total_for_progress)

        # rank participants according to category score
        self.rank_participants(category_id=category_id)

        result += 1
        progress_recorder.set_progress(result, total_for_progress)

        # rank participants according to qa score
        if qa:
            self.qa_rank_participants(category_id=category_id)

        result += 1
        progress_recorder.set_progress(result, total_for_progress)

        context = {
            "response_message": "Score calculation complete"
        }
        return context

    def calculate_qa_scores(self, category_id, supplier, supplier_responses, questions):
        total_score = 0
        for section in Section.objects.filter(category_id=category_id):
            section_score = 0
            questions = questions.filter(section_id=section.id)
            for question in questions:
                if not question.is_qa and question.is_scored:
                    score = 0
                    question_response = supplier_responses.filter(
                        supplier_id=supplier.id, question_id=question.id).first()
                    if (question_response is not None and question_response.options is not None
                            and len(question_response.options) > 0):
                        if question.answer_type == Question.TYPE_UPLOAD:
                            score = float(question.max_score)
                        elif question_response.options in question.options:
                            response_index = question.options.index(question_response.options)
                            score = float(question.scores[response_index])
                        elif question_response.options == "True":
                            score = float(question.max_score)
                    total_score += score
                    section_score += score
                else:
                    question_in_qa = QualityAssuranceResponse.objects.filter(
                        quality_assurance_question__question_id=question.id,
                        supplier_id=supplier.id,
                    ).first()

                    if question_in_qa.score_after_qa is not None:
                        total_score += float(question_in_qa.score_after_qa)
                        section_score += float(question_in_qa.score_after_qa)

            SupplierSectionScore.objects.update_or_create(
                section_id=section.id, supplier_id=supplier.id,
                defaults={
                    "score_after_qa": section_score
                }
            )
        return total_score

    def rank_participants(self, category_id):
        scores = SupplierCategoryScore.objects.filter(category_id=category_id).order_by(
            "-score"
        )
        rank = 0
        first_score = scores.first()
        if first_score is not None:
            first_score.rank = 1
            first_score.save()
            prev_score = first_score
            for score in scores:
                if score.score == prev_score.score:
                    score.rank = prev_score.rank
                else:
                    score.rank = rank + 1
                rank += 1
                score.save()
                prev_score = score

    def qa_rank_participants(self, category_id):
        qa = QualityAssurance.objects.filter(category_id=category_id).first()
        if qa is not None:
            scores = SupplierCategoryScore.objects.filter(category_id=category_id).order_by(
                "-score_after_qa"
            )
            rank_after_qa = 0
            if scores.first().score_after_qa is not None:
                first_score = scores.first()
            else:
                first_score = scores[0]

            if first_score is not None:
                first_score.rank_after_qa = 1
                first_score.save()
                prev_score = first_score
                for score in scores:
                    if score.score_after_qa == None:
                        score.rank_after_qa = None
                        rank_after_qa += 0
                        score.save()
                        continue
                    elif score.score_after_qa == prev_score.score_after_qa:
                        score.rank_after_qa = prev_score.rank_after_qa
                    else:
                        score.rank_after_qa = rank_after_qa + 1
                    rank_after_qa += 1
                    score.save()
                    prev_score = score
        else:
            pass

    def calculate_financial_ratios_after_qa(self, category_id):
        """
        Calculate Ratios After QA
        debt/equity = long term loans/equity
        current ratio = current assets/current liabilities (current assets+ cash)/current liabilities
        cash ratio = cash/current liabilities
        GP margin = GP/turnover
        NP margin = NP/Turnover
        """
        ratio_instances = apps.apps.get_model('prequal', 'FinancialRatio').objects.filter(section__category_id=category_id)
        for ratio_instance in ratio_instances:
            section = ratio_instance.section

            all_questions = Question.objects.filter(
                section_id=ratio_instance.section_id).order_by("id")

            first_question = all_questions.first()
            other_questions = all_questions.exclude(id=first_question.id)
            supplier = ratio_instance.supplier

            section_score = 0
            try:
                debt_equity_ratio = float(ratio_instance.debtors_after_qa) / float(
                    ratio_instance.equity_after_qa)
            except:
                debt_equity_ratio = 0
            try:
                current_ratio = (
                    float(ratio_instance.current_assets_after_qa) + float(ratio_instance.cash_after_qa)
                ) / float(ratio_instance.curr_liabilities_after_qa)
            except:
                current_ratio = 0

            try:
                cash_ratio = float(ratio_instance.cash_after_qa) / float(
                    ratio_instance.curr_liabilities_after_qa)
            except:
                cash_ratio = 0

            try:
                gp_margin = (
                    float(ratio_instance.gross_profit_after_qa) / float(ratio_instance.turnover_after_qa)
                ) * 100
            except:
                gp_margin = 0

            try:
                np_margin = (float(ratio_instance.net_profit_after_qa)/ float(ratio_instance.turnover_after_qa)) * 100
            except:
                np_margin = 0

            c = 0
            ratio_list = [debt_equity_ratio, current_ratio, cash_ratio, gp_margin, np_margin]

            # create QAResponse
            for question in other_questions:
                ratio = ratio_list[c]
                question_options = question.options
                if ratio is not None:
                    for option in question_options:
                        start = int(float(option.split("-", 1)[0]) * 10)
                        end = int(float(option.split("-", 1)[1]) * 10)
                        q_ratio = int(ratio) * 10
                        if q_ratio in range(start, end):
                            my_index = question_options.index(option)
                            score_after_qa = float(question.scores[my_index])
                            data = {
                                "supplier": supplier,
                                "question": question,
                                "score_after_qa": score_after_qa,
                            }
                        else:
                            score_after_qa = 0
                            data = {
                                "supplier": supplier,
                                "question": question,
                                "score_after_qa": score_after_qa,
                            }

                qa_question = QualityAssuranceQuestion.objects.filter(question=question).first()

                qa_res, created = QualityAssuranceResponse.objects.update_or_create(
                    supplier=data["supplier"],
                    quality_assurance_question=qa_question,
                    defaults={
                        "score_after_qa": data["score_after_qa"],
                        "comment": "",
                    },
                )
        return


class ConductOcr(Task):
    name = "ConductOcr"
    messages = []
    responses = []

    def run(self, *args, **kwargs):
        qa_instance = QualityAssurance.objects.filter(category_id=kwargs["category_id"]).first()
        self.ocr_tcc(qa_instance, kwargs['category_id'])
        self.ocr_cr12(qa_instance, kwargs['category_id'])
        self.ocr_business_permit(qa_instance, kwargs['category_id'])
        self.ocr_poisons_board_license(qa_instance, kwargs['category_id'])
        self.ocr_tax_pin_cert(qa_instance, kwargs['category_id'])
        self.ocr_incorporation_certificate(qa_instance, kwargs['category_id'])
        self.ocr_nca_certificate(qa_instance, kwargs['category_id'])

        context = {
            "response_message": "OCR process complete",
            "messages": self.messages,
            "responses": self.responses,
        }
        print(context)
        return context

    def ocr_tcc(self, qa_instance, category_id):
        tcc = QualityAssuranceQuestion.objects.filter(
            question__short_description__icontains='Tax compliance',
            quality_assurance_id=qa_instance.id
        ).first()
        if tcc:
            try:
                supplier_responses = SupplierResponse.objects.filter(
                    question_id=tcc.question_id
                )
                q = Question.objects.filter(
                    section__category_id=category_id, short_description="PIN number").first()
                for supplier_response in supplier_responses:
                    r = verify_tcc(document=f"{supplier_response.document_url}")
                    self.responses.append(r)
                    # if r['kra_pin'] and r['expiry_date']:
                    date = make_aware(r['expiry_date'], timezone=pytz.timezone("Africa/Nairobi")) if r[
                        'expiry_date'] else None

                    pin_number_outcome = "Fail"
                    if q is not None:
                        s_r = SupplierResponse.objects.filter(
                            question_id=q.id, supplier_id=supplier_response.supplier_id).first()
                        if s_r is not None:
                            if s_r.options:
                                if s_r.options == r['kra_pin']:
                                    pin_number_outcome = "Pass"

                    expiry_date_outcome = "Fail"
                    if date is not None:
                        if datetime.datetime.now().date() <= date.date():
                            expiry_date_outcome = "Pass"
        
                    overall_outcome = "Pass"
                    if pin_number_outcome == "Fail" or expiry_date_outcome == "Fail":
                        overall_outcome = "Fail"
                    
                    instance, created = QualityAssuranceResponse.objects.update_or_create(
                        supplier_id=supplier_response.supplier_id,
                        quality_assurance_question_id=tcc.id,
                        defaults={
                            "number": r['kra_pin'], "ocr_response": r,
                            "outcome": overall_outcome
                        }
                    )
                    print("we created qa response")
                    QaTccResponse.objects.update_or_create(
                        qa_response_id=instance.id,
                        defaults={
                            # company name
                            "pin_number": r['kra_pin'], "pin_number_outcome": pin_number_outcome,
                            "expiry_date": r['expiry_date'], "expiry_date_outcome": expiry_date_outcome
                        }
                    )
                    print("we created tcc response")
                self.messages.append(f'Processing done for TCC {supplier_responses.count()}')
            except Exception as e:
                # capture_exception(e)
                print(f"TCC {e}")
        else:
            self.messages.append('NO TCC IN QA')
        return

    def ocr_cr12(self, qa_instance, category_id):
        cr12 = QualityAssuranceQuestion.objects.filter(
            question__short_description__icontains='CR12 Certificate or equivalent',
            quality_assurance_id=qa_instance.id
        ).first()
        if cr12:
            try:
                supplier_responses = SupplierResponse.objects.filter(
                    question_id=cr12.question_id
                )
                q = Question.objects.filter(
                    section__category_id=category_id,
                    short_description="Company registration certificate number").first()
                for supplier_response in supplier_responses:
                    r = verify_cr12(document=f"{supplier_response.document_url}")
                    self.responses.append(f'{r}')
                    # if r['cr12_ref_no']:

                    company_number_outcome = "Fail"
                    if q is not None:
                        s_r = SupplierResponse.objects.filter(
                            question_id=q.id, supplier_id=supplier_response.supplier_id).first()
                        if s_r is not None:
                            if s_r.options:
                                if s_r.options == r['cr12_ref_no']:
                                    company_number_outcome = "Pass"

                    instance, created = QualityAssuranceResponse.objects.update_or_create(
                        supplier_id=supplier_response.supplier_id,
                        quality_assurance_question_id=cr12.id,
                        defaults={
                            "number": r['cr12_ref_no'],
                            "date": make_aware(
                                datetime.datetime.now(), timezone=pytz.timezone("Africa/Nairobi")),
                            "ocr_response": r, "outcome": company_number_outcome
                        }
                    )

                    QaCr12Response.objects.update_or_create(
                        qa_response_id=instance.id,
                        defaults={
                            "company_number": r["cr12_ref_no"],
                            # "document_date": data["date"],
                            "company_number_outcome": company_number_outcome,
                        }
                    )
                self.messages.append(f'Processing done for CR12 {supplier_responses.count()}')
            except Exception as e:
                # capture_exception(e)
                print(f"CR12 {e}")
        else:
            self.messages.append('NO CR12 IN QA')

    def ocr_business_permit(self, qa_instance, category_id):
        business_permit = QualityAssuranceQuestion.objects.filter(
            question__short_description__icontains='Business permit',
            quality_assurance_id=qa_instance.id
        ).first()
        if business_permit:
            try:
                supplier_responses = SupplierResponse.objects.filter(
                    question_id=business_permit.question_id,
                )
                for supplier_response in supplier_responses:
                    r = verify_businesspermit(document=f"{supplier_response.document_url}")
                    self.responses.append(r)
                    # if r['business_id']:

                    business_name_outcome = "Fail"
                    supplier = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                        id=supplier_response.supplier_id).first()
                    if supplier is not None:
                        if supplier.company_name.lower() == r['business_name']:
                            business_name_outcome = "Pass"

                    instance, created = QualityAssuranceResponse.objects.update_or_create(
                        supplier_id=supplier_response.supplier_id,
                        quality_assurance_question_id=business_permit.id,
                        defaults={
                            "number": r['business_id'],
                            "date": make_aware(
                                datetime.datetime.now(), timezone=pytz.timezone("Africa/Nairobi")),
                            "ocr_response": r, "outcome": business_name_outcome
                        }
                    )

                    QaBusinessPermitResponse.objects.update_or_create(
                        qa_response_id=instance.id,
                        defaults={
                            "business_id": r['business_id'], "date": r["year"],
                            "business_name": r["business_name"],
                            "business_name_outcome": business_name_outcome
                        }
                    )
                    # QualityAssuranceResponse.objects.update_or_create()
                self.messages.append(f'Processing done for Business Permit {supplier_responses.count()}')
            except Exception as e:
                # capture_exception(e)
                print(f"Permit {e}")
        else:
            self.messages.append('NO BUSINESS PERMIT IN QA')

    def ocr_poisons_board_license(self, qa_instance, category_id):
        poisons_boardcert = QualityAssuranceQuestion.objects.filter(
            question__short_description__icontains='Pharmacy and poisons board License',
            quality_assurance_id=qa_instance.id
        ).first()
        if poisons_boardcert:
            try:
                supplier_responses = SupplierResponse.objects.filter(
                    question_id=poisons_boardcert.question_id
                )
                for supplier_response in supplier_responses:
                    r = verify_poisonsboardcert(document=f"{supplier_response.document_url}")
                    self.responses.append(r)
                    # if r['business_id'] and r['expiry_date']:
                    date = make_aware(r['expiry_date'], timezone=pytz.timezone("Africa/Nairobi")) if r[
                        'expiry_date'] else None

                    company_name_outcome = "Fail"
                    supplier = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                        id=supplier_response.supplier_id).first()
                    if supplier is not None:
                        if supplier.company_name.lower() == r['business_name']:
                            company_name_outcome = "Pass"

                    expiry_date_outcome = "Fail"
                    if datetime.datetime.now().date() <= date.date():
                        expiry_date_outcome = "Pass"

                    overall_outcome = "Pass"
                    if company_name_outcome == "Fail" or expiry_date_outcome == "Fail":
                        overall_outcome = "Fail"

                    instance, created = QualityAssuranceResponse.objects.update_or_create(
                        supplier_id=supplier_response.supplier_id,
                        quality_assurance_question_id=poisons_boardcert.id,
                        defaults={
                            "number": r['company_name'],
                            "date": date, "ocr_response": r, "outcome": overall_outcome
                        }
                    )

                    QaPoisonsBoardResponse.objects.update_or_create(
                        qa_response_id=instance.id,
                        defaults={
                            "expiry_date": date,
                            "company_name": r["company_name"],
                            "company_name_outcome": company_name_outcome,
                            "expiry_date_outcome": expiry_date_outcome
                        }
                    )
                self.messages.append(f'Processing done for POISONS {supplier_responses.count()}')
            except Exception as e:
                # capture_exception(e)
                print(f"Poisons {e}")
        else:
            self.messages.append('NO POISIONS BOARD QUESTION IN QA')

    def ocr_tax_pin_cert(self, qa_instance, category_id):
        pin_cert = QualityAssuranceQuestion.objects.filter(
            question__short_description__icontains='PIN certificate',
            quality_assurance_id=qa_instance.id
        ).first()
        if pin_cert:
            try:
                supplier_responses = SupplierResponse.objects.filter(
                    question_id=pin_cert.question_id
                )
                q = Question.objects.filter(
                    section__category_id=category_id, short_description="PIN number").first()
                for supplier_response in supplier_responses:
                    r = verify_pincert(document=f"{supplier_response.document_url}")
                    self.responses.append(r)
                    # if r['pin']:

                    tax_pin_outcome = "Fail"
                    if q is not None:
                        s_r = SupplierResponse.objects.filter(
                            question_id=q.id, supplier_id=supplier_response.supplier_id).first()
                        if s_r is not None:
                            if s_r.options:
                                if s_r.options == r['pin']:
                                    tax_pin_outcome = "Pass"

                    instance, created = QualityAssuranceResponse.objects.update_or_create(
                        supplier_id=supplier_response.supplier_id,
                        quality_assurance_question_id=pin_cert.id,
                        defaults={
                            "number": r['pin'],
                            "date": make_aware(
                                datetime.datetime.now(), timezone=pytz.timezone("Africa/Nairobi")),
                            "ocr_response": r, "outcome": tax_pin_outcome
                        }
                    )

                    QaPinCertificateResponse.objects.update_or_create(
                        qa_response_id=instance.id,
                        defaults={"pin_number": r['pin'], "tax_pin_outcome": tax_pin_outcome}
                    )
                    # QualityAssuranceResponse.objects.update_or_create()
                self.messages.append(f'Processing done for PIN Cert {supplier_responses.count()}')
            except Exception as e:
                # capture_exception(e)
                print(f"Pin cert {e}")
        else:
            self.messages.append("NO PIN CERT QUESTION IN QA")

    def ocr_incorporation_certificate(self, qa_instance, category_id):
        incorporation_cert = QualityAssuranceQuestion.objects.filter(
            question__short_description__icontains='Registration certificate',
            quality_assurance_id=qa_instance.id
        ).first()
        if incorporation_cert:
            try:
                supplier_responses = SupplierResponse.objects.filter(
                    question_id=incorporation_cert.question_id
                )
                q = Question.objects.filter(
                    section__category_id=category_id,
                    short_description="Company registration certificate number").first()

                for supplier_response in supplier_responses:
                    r = verify_incorporationcert(document=f"{supplier_response.document_url}")
                    self.responses.append(r)
                    # if r['number'] and r['date']:
                    date = date = make_aware(r['date'], timezone=pytz.timezone("Africa/Nairobi")) if r['date'] else None
                    
                    company_name_outcome = "Fail"
                    supplier = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                        id=self.context['participant_id']).first()
                    if supplier is not None:
                        if supplier.company_name.lower() == r['name'].lower():
                            company_name_outcome = "Pass"

                    company_number_outcome = "Fail"
                    if q is not None:
                        s_r = SupplierResponse.objects.filter(
                            question_id=q.id, supplier_id=self.context['participant_id']).first()
                        if s_r is not None:
                            if s_r.options:
                                if s_r.options == r['number']:
                                    company_number_outcome = "Pass"
                    
                    overall_outcome = "Pass"
                    if company_name_outcome == "Fail" or company_number_outcome == "Fail":
                        overall_outcome = "Fail"

                    instance, created = QualityAssuranceResponse.objects.update_or_create(
                        supplier_id=supplier_response.supplier_id,
                        quality_assurance_question_id=incorporation_cert.id,
                        defaults={
                            "number": r['number'],
                            "date": date, "ocr_response": r, "outcome": overall_outcome
                        }
                    )

                    QaIncorporationCertificateResponse.objects.update_or_create(
                        qa_response_id=instance.id,
                        defaults={
                            "company_name": r["name"],
                            "company_name_outcome": company_name_outcome,
                            "company_number": r["number"],
                            "company_number_outcome": company_number_outcome
                        }
                    )
                self.messages.append(f'Processing done for Incorporation Cert {supplier_responses.count()}')
            except Exception as e:
                print(f"Incorporation cert {e}")
        else:
            self.messages.append('NO INCORPORATION QUESTION IN QA')

    def ocr_nca_certificate(self, qa_instance, category_id):
        nca = QualityAssuranceQuestion.objects.filter(
            question__short_description__icontains='NCA certificate',
            quality_assurance_id=qa_instance.id
        ).first()
        if nca:
            try:
                supplier_responses = SupplierResponse.objects.filter(
                    question_id=nca.question_id
                )

                for supplier_response in supplier_responses:
                    r = verify_nca(document=f"{supplier_response.document_url}")
                    date = make_aware(r['expiry_date'], timezone=pytz.timezone("Africa/Nairobi")) if r[
                        'expiry_date'] else None

                    expiry_date_outcome = "Fail"
                    if date is not None:
                        if datetime.datetime.now().date() <= date.date():
                            expiry_date_outcome = "Pass"
                    
                    instance, created = QualityAssuranceResponse.objects.update_or_create(
                        supplier_id=supplier_response.supplier_id,
                        quality_assurance_question_id=nca.id,
                        defaults={
                            "number": r['kra_pin'], "date": date, "ocr_response": r,
                            "outcome": expiry_date_outcome
                        }
                    )

                    QaNcaaResponse.objects.update_or_create(
                        qa_response_id=instance.id,
                        defaults={
                            "serial_number": r["serial_no"],
                            "expiry_date": date, 
                            "expiry_date_outcome": expiry_date_outcome 
                        }
                    )
                self.messages.append(f'Processing done for NCA Cert {supplier_responses.count()}')
            except Exception as e:
                # capture_exception(e)
                print(f"Pin cert {e}")
        else:
            self.messages.append("NO PIN CERT QUESTION IN QA")

        # if nca:
        #     try:
        #         supplier_responses = SupplierResponse.objects.filter(
        #             question_id=nca.question_id
        #         )
        #         q = Question.objects.filter(
        #             section__category_id=category_id,
        #             short_description="Company registration certificate number").first()

        return


@shared_task(bind=True)
def existing_supplier_email_invite(self, supplier_id, category_id):
    # try:
    supplier = apps.apps.get_model('suppliers', 'Supplier').objects.filter(id=supplier_id).first()
    category = apps.apps.get_model('prequal', 'Category').objects.filter(id=category_id).first()
    job = category.prequalification
    company = job.company
    email_config = None
    to_email = supplier.email

    email_subject = job.title.upper() + " INVITATION"
    if email_config is not None:
        message = render_to_string(
            "prequal/emails/category_invitation_email.html",
            {
                "supplier": supplier, "category": category,
                "buyer_name": f"{company.company_name} Team",
                "buyer_logo": "company_logo", "company": company,
            },
        )
        email_config.send_email(email_subject, message, bcc=[to_email])
    else:
        if category.invite_only is True:
            body = render_to_string(
                "prequal/emails/category_invitation_email.html",
                {
                    "supplier": supplier, "category": category,
                    "buyer_name": "Tendersure Team", "buyer_logo": "tendersure_logo",
                },
            )
            # Attach bidder instructions
            if job.bidding_instructions:
                A = PrivateMediaStorage()
                headers = {"ResponseContentDisposition": f"attachment;"}
                time = datetime.datetime.now()
                file_url = A.url(
                    f"{job.bidding_instructions}",
                    expire=300,
                    parameters=headers,
                    http_method="GET",
                )
                dir_name = Path(
                    "media/temp/{}/{}/{}".format(time.year, time.month, time.day)
                )
                # folder structure
                dir_name.mkdir(parents=True, exist_ok=True)
                file_name = os.path.basename(f"{job.bidding_instructions}")
                filepath = "{}/{}".format(dir_name, file_name)
                r = requests.get(file_url)
                with open("{}".format(filepath), "wb") as f:
                    f.write(r.content)
                email = EmailMultiAlternatives(email_subject, body, to=[to_email])
                email.attach_alternative(body, "text/html")
                email.attach_file(filepath)
                email.send()
            else:
                email = EmailMultiAlternatives(email_subject, body, to=[to_email])
                email.attach_alternative(body, "text/html")
                email.send()

        else:
            message = render_to_string(
                "prequal/emails/category_invitation_email.html",
                {
                    "supplier": supplier,
                    "category": category,
                    "buyer_name": "Tendersure Team",
                    "buyer_logo": "tendersure_logo",
                },
            )
            email = EmailMultiAlternatives(email_subject, message, to=[to_email])
            email.attach_alternative(message, "text/html")
            email.send()
    # except Exception as e:
    #     capture_exception(e)
    #     print(e)


@shared_task(bind=True)
def non_existent_invites_email(self, user_email, category):
    try:
        supplier = "Esteemed Vendor"
        job = category.job
        company = job.company

        email_config = None
        to_email = user_email
        email_subject = "Prequalification Invitation"

        if category.invite_only is True:
            body = render_to_string(
                "prequal/emails/category_invitation_email_non_existent_supplier.html",
                {
                    "supplier": supplier, "category": category, "company": company,
                    "buyer_name": "Tendersure Team", "buyer_logo": "tendersure_logo",
                },
            )
            # Attach bidder instructions
            A = PrivateMediaStorage()
            headers = {"ResponseContentDisposition": f"attachment;"}
            time = datetime.datetime.now()
            file_url = A.url(
                f"{job.bidding_instructions}",
                expire=300,
                parameters=headers,
                http_method="GET",
            )
            dir_name = Path(
                "media/temp/{}/{}/{}".format(time.year, time.month, time.day)
            )
            # folder structure
            dir_name.mkdir(parents=True, exist_ok=True)
            file_name = os.path.basename(f"{job.bidding_instructions}")
            filepath = "{}/{}".format(dir_name, file_name)
            r = requests.get(file_url)
            with open("{}".format(filepath), "wb") as f:
                f.write(r.content)
            email = EmailMultiAlternatives(email_subject, body, to=[to_email])
            email.attach_alternative(body, "text/html")
            email.attach_file(filepath)
            email.send(fail_silently=True)

        else:
            body = render_to_string(
                "prequal/emails/category_invitation_email_non_existent_supplier.html",
                {
                    "supplier": supplier, "category": category, "company": company,
                    "buyer_name": "Tendersure Team", "buyer_logo": "tendersure_logo",
                },
            )
            email = EmailMultiAlternatives(email_subject, body, to=[to_email])
            email.attach_alternative(body, "text/html")
            email.send(fail_silently=True)
    except Exception as e:
        capture_exception(e)
        print(e)


class ZipQuestionFiles(Task):
    name = "ZipPrequalQuestionFiles"
    progress_recorder = None
    result = 0

    def run(self, *args, **kwargs):
        self.progress_recorder = ProgressRecorder(self)
        context = self.zip_files(kwargs["question_id"])
        return context

    def zip_files(self, question_id):
        progress_recorder = ProgressRecorder(self)
        result = 0
        question = Question.objects.filter(id=question_id).first()
        supplier_responses = (
            SupplierResponse.objects.filter(question_id=question_id).distinct()
        )
        print(supplier_responses)

        total_for_progress = supplier_responses.count()
        print(total_for_progress)
        # dir_name = Path(
        #     f"media/temp/zip/question_documents/{question_id}"
        # )  # folder structure
        dir_name = Path(
            f"media/question_documents/{question_id}"
        )  # folder structure
        dir_name.mkdir(parents=True, exist_ok=True)
        zip_name = f"{dir_name}/{question.short_description.replace('/', '_').replace(' ', '_')}.zip"
        # zip_name = (
        #     f"{question.short_description.replace('/', '_').replace(' ', '_')}.zip"
        # )
        with ZipFile(zip_name, "w") as zipObj2:
            for supplier_response in supplier_responses:
                progress_recorder.set_progress(result, total_for_progress)
                if supplier_response is not None:
                    try:
                        dir_name.mkdir(parents=True, exist_ok=True)
                        file_name = os.path.basename(
                            f"{supplier_response.document_url}"
                        )
                        filepath = f"{dir_name}/{file_name}"
                        if supplier_response.document_url.name != "":
                            r = requests.get(supplier_response.document_url.url)

                            with open("{}".format(filepath), "wb") as f:
                                f.write(r.content)

                            zipObj2.write(f"{filepath}")
                        else:
                            sup_prof = apps.apps.get_model('suppliers', 'SupplierCompanyProfile').objects.filter(
                                supplier__id=supplier_response.supplier.id).order_by("id").last()
                            if sup_prof.registration_cert_url.name != "":
                                n_file_name = os.path.basename(f"{sup_prof.registration_cert_url}")
                                n_filepath = f"{dir_name}/{n_file_name}"
                                r = requests.get(sup_prof.registration_cert_url.url)

                                with open("{}".format(n_filepath), "wb") as f:
                                    f.write(r.content)

                                zipObj2.write(f"{n_filepath}")

                    except Exception as e:
                        capture_exception(e)
                        print(e)
                result += 1
        # shutil.rmtree(
        #     f"media/temp/zip/question_documents/{question_id}", ignore_errors=True
        # )
        context = {
            "filepath": zip_name,
            "response_message": "File generated successfully",
        }

        return context

# @shared_task(bind=True)
def upload_job_current_suppliers(job_id):
    """
    Uploads current suppliers and saves to the Job Model
    """
    # progress_recorder = ProgressRecorder(self)
    result = 0

    messages = []
    job = Prequalification.objects.get(id=job_id)
    company = job.company
    
    try:
        workbook = load_workbook(job.current_suppliers, data_only=True)
    except Exception as e:
        capture_exception(e)
        context = {
            "messages": messages.append("Excel not found"),
            "response_message": "Error uploading current suppliers"
        }
        return context

    try:
        worksheet = workbook["current_suppliers"]
    except Exception as e:
        capture_exception(e)
        context = {
            "messages": messages.append("Invalid current_suppliers sheet not found"),
            "response_message": "Error uploading current suppliers"
        }
        return context

    try:
        total_for_progress = worksheet.max_row
        count = 1
        for i in range(4, worksheet.max_row + 1):
            result += 1
            # progress_recorder.set_progress(result, total_for_progress)

            qed_code = worksheet["C{}".format(i)].value
            company_name = worksheet["D{}".format(i)].value
            company_email = worksheet["E{}".format(i)].value
            phone = worksheet["F{}".format(i)].value

            category_type = apps.apps.get_model('core', 'CategoryType').objects.filter(
                innitials=qed_code
            ).first()
            
            # modify to pick based on category unique reference to avoid duplicate suppliers in categories
            categories = Category.objects.filter(prequalification=job, category_type=category_type)

            for category in categories:
                current_supplier = apps.apps.get_model('core', 'CurrentSupplier').objects.filter(
                    supplier_name=company_name, category_id=category.id).first()

                if current_supplier is not None:
                    current_supplier = apps.apps.get_model('core', 'CurrentSupplier').objects.update_or_create(
                        id=current_supplier.id,
                        supplier_name=company_name,
                        category_id=category.id,
                        job_id=job_id,
                        target=ContentType.objects.get_for_model(apps.apps.get_model('prequal', 'Category')),                   
                        defaults={"alternative_email": company_email},
                    )
                else:
                    current_supplier = apps.apps.get_model('core', 'CurrentSupplier').objects.create(
                        company=company,
                        category_id=category.id,
                        job_id=job_id,
                        supplier_name=company_name,
                        supplier_email=company_email,
                        supplier_phone=phone,
                        target=ContentType.objects.get_for_model(apps.apps.get_model('prequal', 'Category'))
                    )

            count += 1

        messages.append(
            "Imports succeeded. {} current suppliers added.".format(count)
        )
        context = {
            "response_message": "Current Suppliers Uploaded Successfully",
            "messages": messages
        }
        return context
    except Exception as e:
        capture_exception(e)
        context = {
            "messages": messages.append(f"{e}"),
            "response_message": "Error uploading current suppliers"
        }
        print(context)
        return context


@shared_task(bind=True)
def import_category_suppliers(self, job_id):
    messages = []

    job = Prequalification.objects.filter(id=job_id).first()
    filepath = get_file(job.category_suppliers)

    workbook = load_workbook(filepath, data_only=True)

    worksheets = workbook.sheetnames[1:]
    if len(worksheets) < 1:
        messages.append("There are no sheets to upload from the excel.")

        context = {"messages": messages}
        return context

    categories_count = 0
    suppliers = 0

    progress_recorder = ProgressRecorder(self)
    result = 0

    for sheet in worksheets:
        result += 1
        progress_recorder.set_progress(result, len(worksheets))

        sheet_name_list = sheet.split("_")
        category_sheet = workbook[sheet]
        CategoryType = apps.apps.get_model('core', 'CategoryType')
        category_type = CategoryType.objects.filter(innitials=sheet_name_list[-1])

        if category_type.count() == 1:
            categories_count += 1
            category_type = category_type.first()
            for i in range(7, category_sheet.max_row + 1):
                company_name = category_sheet["C{}".format(i)].value
                primary_email = category_sheet["E{}".format(i)].value
                alternative_email = category_sheet["F{}".format(i)].value
                primary_phone = category_sheet["G{}".format(i)].value
                alternative_phone = category_sheet["H{}".format(i)].value
                country = category_sheet["I{}".format(i)].value

                supplier = None
                supplier1 = None
                supplier2 = None

                if primary_email != None:
                    supplier1 = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                        username=primary_email).first()
                elif alternative_email != None:
                    supplier2 = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                        username=alternative_email
                    ).first()

                if supplier1:
                    supplier = supplier1
                elif supplier2:
                    supplier = supplier2

                if supplier is None:
                    suppliers += 1
                    try:
                        apps.apps.get_model('core', 'CategoryTypeSupplier').objects.update_or_create(
                            primary_email=primary_email, alternative_email=alternative_email,
                            category_type=category_type,
                            defaults={
                                "primary_email": primary_email, "alternative_email": alternative_email,
                                "primary_phone": primary_phone, "alternative_phone": alternative_phone,
                                "category_type": category_type, "company_name": company_name,
                                "country": country,
                            },
                        )
                    except Exception as e:
                        capture_exception(e)
                        print(e)
                else:
                    suppliers += 1
                    try:

                        apps.apps.get_model('core', 'CategoryTypeSupplier').objects.update_or_create(
                            supplier=supplier, category_type=category_type,
                            defaults={
                                "supplier": supplier, "primary_email": primary_email,
                                "alternative_email": alternative_email, "primary_phone": primary_phone,
                                "alternative_phone": alternative_phone, "category_type": category_type,
                                "company_name": company_name, "country": country,
                            },
                        )
                    except Exception as e:
                        capture_exception(e)
                        print(e)
    os.remove(filepath)
    messages.append(
        "Imports succeeded. {} suppliers added to {} categories.".format(
            suppliers, categories_count
        )
    )
    context = {"messages": messages}
    return context


@shared_task()
def calculate_financial_ratios_after_qa(instance_id, *args, **kwargs):
    """
    Calculate Ratios After QA
    debt/equity = long term loans/equity
    current ratio = current assets/current liabilities
    cash ratio = cash/current liabilities
    GP margin = GP/turnover
    NP margin = NP/Turnover
    """
    ratio_instance = apps.apps.get_model('prequal', 'FinancialRatio').objects.filter(id=instance_id).first()
    section = ratio_instance.section
    user = kwargs.get("user", None)
    if user is None:
        user = User.objects.filter(email="interns@qedsolutions.co.ke").first()

    all_questions = Question.objects.filter(
        section_id=ratio_instance.section_id).order_by("id")
    first_question = all_questions.first()
    other_questions = all_questions.exclude(id=first_question.id)
    supplier = ratio_instance.supplier

    section_score = 0
    try:
        debt_equity_ratio = float(ratio_instance.debtors_after_qa) / float(ratio_instance.equity_after_qa)
    except:
        debt_equity_ratio = 0
    try:
        # current_ratio = (float(ratio_instance.current_assets_after_qa) + float(ratio_instance.cash_after_qa)
        # ) / float(ratio_instance.curr_liabilities_after_qa)
        current_ratio = float(ratio_instance.current_assets_after_qa) / float(ratio_instance.curr_liabilities_after_qa)
    except:
        current_ratio = 0

    try:
        cash_ratio = float(ratio_instance.cash_after_qa) / float(ratio_instance.curr_liabilities_after_qa)
    except:
        cash_ratio = 0

    try:
        gp_margin = (
            float(ratio_instance.gross_profit_after_qa)
            / float(ratio_instance.turnover_after_qa)
        ) * 100
    except:
        gp_margin = 0

    try:
        np_margin = (
            float(ratio_instance.net_profit_after_qa)
            / float(ratio_instance.turnover_after_qa)
        ) * 100
    except:
        np_margin = 0

    ratio_list = [debt_equity_ratio, current_ratio, cash_ratio, gp_margin, np_margin]

    ac_question = Question.objects.filter(
        description="1. For limited liability companies, attach audited accounts for the last two years, "
                    "for sole proprietors and partnerships, attach your most recent management accounts",
        section__category_id=section.category_id
    ).first()

    if ac_question is not None:
        supplier_response = SupplierResponse.objects.filter(question_id=ac_question.id, supplier_id=supplier.id).first()
    else:
        supplier_response = None

    # create QAResponse
    c = 0
    for question in other_questions:
        ratio = ratio_list[c]
        question_options = question.options
        c += 1
        data = None
        if ratio is not None:
            for option in question_options:
                start = float(option.split("-", 1)[0])
                end = float(option.split("-", 1)[1])
                q_ratio = math.floor(float(ratio) * 10)/10

                if start <= q_ratio <= end:
                    my_index = question_options.index(option)
                    score_after_qa = float(question.scores[my_index])
                    data = {
                        "supplier": supplier,
                        "question": question,
                        "score_after_qa": score_after_qa,
                        "outcome": "Pass"
                    }
                    break
                else:
                    continue

        if data is None:
            data = {
                "supplier": supplier,
                "question": question,
                "score_after_qa": 0,
                "outcome": "Fail"
            }

        """ Check if supplier submitted their financials file if not give zero marks """
        if supplier_response is None:
            data = {
                "supplier": supplier,
                "question": question,
                "score_after_qa": 0,
                "outcome": "Fail"
            }
        elif supplier_response is not None:
            if supplier_response.options is None and supplier_response.document_url is None:
                data = {
                    "supplier": supplier,
                    "question": question,
                    "score_after_qa": 0,
                    "outcome": "Fail"
                }

        qa_question = QualityAssuranceQuestion.objects.filter(question=question).first()

        if qa_question is not None:
            qa_res = QualityAssuranceResponse.objects.filter(
                supplier=data["supplier"], quality_assurance_question=qa_question).last()

            if qa_res is not None:
                qa_res.score_after_qa = data["score_after_qa"]
                qa_res.created_by = user
                qa_res.outcome = data["outcome"]
                qa_res.save()
            else:
                qa_res = QualityAssuranceResponse.objects.create(
                    supplier=data["supplier"], quality_assurance_question=qa_question,
                    score_after_qa=data["score_after_qa"], created_by=user, outcome=data["outcome"]
                )
            context = {"qa_res": qa_res}
        else:
            context = {}

        print(f"Context: {context}")
    return context


app.register_task(SendSMSNotifications())
app.register_task(SendEmailNotifications())
app.register_task(FinancialRatiosReport())
app.register_task(ConductOcr())
app.register_task(ZipQuestionFiles())
app.register_task(EvaluatePrequal())
app.register_task(PrequalInterimReport())
app.register_task(QARankingReport())
app.register_task(DueDiligenceRankingReport())
app.register_task(FinancialRatiosReport())
