import datetime
from io import BytesIO
import os
import string

from pathlib import Path
from django.db.models import Q
from django import apps
from django.core.files import File

from openpyxl import Workbook
from openpyxl.drawing.image import Image

# celery
from celery import shared_task, Task
from celery_progress.backend import ProgressRecorder
from sentry_sdk import capture_exception


from apps.core.utils import (
    Render,
    delete_matching_files_in_directory,
    format_excel,
    get_file_path,
    insert_image,
    show,
)

from apps.rfq.models import (
    Category,
    RFQCategoryReport,
    RFQItemResponse,
    RFQJobReport,
    Rfq,
    SupplierRfqTotal,
    SupplierResponse,
)

@shared_task(bind=True)
def supplier_rfq_pdf_report(self,supplier_id, category_id):
    """
    Supplier RFQ PDF responses report
    """
    supplier = (
        apps.apps.get_model("suppliers", "Supplier")
        .objects.filter(id=supplier_id)
        .first()
    )
    category = Category.objects.filter(id=category_id).first()

    try:
        if supplier is not None and category is not None:
            time = datetime.datetime.now()
            dir_name = Path("reports/{}/{}".format(time.year, time.month))
            dir_name.mkdir(parents=True, exist_ok=True)

            time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
            filepath = "{}/{}_{}_{}.pdf".format(
                dir_name,
                category.id,
                supplier.company_name.replace(" ", "_").replace("&", "_"),
                time_only,
            )
            rfq_responses = RFQItemResponse.objects.filter(
                supplier=supplier, rfq_item__category_id=category.id
            )

            context = {
                "category": category,
                "supplier": supplier,
                "rfq_responses": rfq_responses,
                "time": time,
            }

            pdf = Render.render(
                "supplier_rfq_summary_report.html",
                context,
            )

            pdf_file = BytesIO(pdf.content)
            supplier_response, created = SupplierResponse.objects.update_or_create(
                supplier_id=supplier_id,
                category_id=category_id,
                defaults={"document_url": File(pdf_file, filepath)},
            )

            context = {"file": supplier_response.document_url}

            print(context)

    except Exception as e:
        print(e)



def rfq_job_summary_report(category_id):
    """
    RFQ Summary Job Level
    """
    category = Category.objects.filter(id=category_id).first()
    buyer = (
        apps.apps.get_model("buyer", "Company")
        .objects.filter(id=category.rfq.company.id)
        .first()
    )
    if category is not None and buyer is not None:
        time = datetime.datetime.now()
        dir_name = Path("rfq_reports/{}/{}".format(time.year, time.month))
        dir_name.mkdir(parents=True, exist_ok=True)

        time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
        filepath = "{}/{}_{}_{}.pdf".format(
            dir_name,
            category.id,
            time_only,
            buyer.company_name.replace(" ", "_").replace("&", "_"),
        )
        participants = category.participants

        context = {
            "category": category,
            "participants": participants["participants"],
            "time": time,
            "buyer": buyer,
        }

        pdf = Render.render(
            "rfq_job_summary_report.html",
            context,
        )

        pdf_file = BytesIO(pdf.content)
        rfq_category_report, created = RFQCategoryReport.objects.update_or_create(
            category_id=category_id,
            defaults={"category_rfq_pdf": File(pdf_file, filepath)},
        )

        context = {"file": rfq_category_report.category_rfq_pdf}

        return rfq_category_report


class RFQFinancialReport(Task):
    """
    Category RFQ Financial Report(Lowest Item, Lowest Supplier,Savings and Individual sheets)
    """

    name = "RFQFinancialReport"
    progress_recorder = None
    result = 0

    def run(self, *args, **kwargs):
        self.progress_recorder = ProgressRecorder(self)
        context = self.report(kwargs["category_id"])
        return context

    def report(self, category_id):
        category = Category.objects.filter(id=category_id).first()
        job = category.rfq
        items = category.items

        rfq_participants = category.participants

        # has_blank=False, has_outlier=False
        lowest_cumm_suppliers_ranked = SupplierRfqTotal.objects.filter(
            category=category
        ).order_by("score")
        lowest_cumm_suppliers_opp = SupplierRfqTotal.objects.filter(
            Q(category=category), ~Q(has_blank=False) | ~Q(has_outlier=False)
        ).order_by("score")
        lowest_cumm_suppliers = list(lowest_cumm_suppliers_ranked) + list(
            lowest_cumm_suppliers_opp
        )

        # edit
        total_for_progress = len(rfq_participants) + len(items)
        category_total_bidders = len(category.invited_suppliers)
        category_responsive_bidders = rfq_participants["count"]
        try:
            percentage_responses = "{}%".format(
                round(((category_responsive_bidders / category_total_bidders) * 100))
            )
        except Exception as e:
            percentage_responses = 0

        time = datetime.datetime.now()
        dir_name = Path("media/rfq_reports/{}/{}".format(time.year, time.month))
        dir_name.mkdir(parents=True, exist_ok=True)

        time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
        filepath = "{}/{}_{}_{}.xlsx".format(
            dir_name, job.id, job.company.company_name.replace(" ", "_"), time_only
        )
        # delete similar reports run this month
        match_string = "{}_{}".format(
            job.id, job.company.company_name.replace(" ", "_")
        )
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")
        workbook = Workbook()
        qed_logo = Image("static/core/img/Tendersure_Logo.png")

        if (
            job.company.company_logo_url is not None
            and job.company.company_logo_url != ""
        ):
            buyer_logo = get_file_path(job.company.company_logo_url)
        else:
            buyer_logo = "static/core/img/no-company-image128-128.png"

        qed_logo_anchor = "F1"
        buyer_logo_anchor = "B1"

        workbook.create_sheet("LOWEST ITEM COST REPORT")
        pricing_worksheet = workbook["LOWEST ITEM COST REPORT"]

        letters = list(string.ascii_uppercase)
        num_cols = 200
        excel_cols = []
        for i in range(0, num_cols - 1):
            n = i // 26
            m = n // 26
            i -= n * 26
            n -= m * 26
            col = (
                letters[m - 1] + letters[n - 1] + letters[i]
                if m > 0
                else letters[n - 1] + letters[i]
                if n > 0
                else letters[i]
            )
            excel_cols.append(col)

        data = {
            "category_responsive_bidders": category_responsive_bidders,
            "category_total_bidders": int(category_total_bidders),
            "category": category,
            "items": items,
            "lowest_cumm_suppliers": lowest_cumm_suppliers,
            "total_for_progress": total_for_progress,
            "excel_cols": excel_cols,
            "percentage_responses": percentage_responses,
            "participants": rfq_participants["participants"],
            "lowest_cumm_suppliers_ranked": lowest_cumm_suppliers_ranked,
            "lowest_cumm_suppliers_opp": lowest_cumm_suppliers_opp,
        }
        
        self.create_lowest_cost_report_worksheet(
            pricing_worksheet=pricing_worksheet, data=data
        )
        self.create_lowest_supplier_worksheet(workbook=workbook, data=data)
        self.create_savings_worksheet(workbook=workbook, data=data)
        # self.create_individual_worksheets(workbook=workbook, data=data)

        workbook.remove(workbook["Sheet"])
        workbook.save(filepath)

        f_data = {
            "SAVINGS REPORT_A": 5,
            "SAVINGS REPORT_B": 5,
            "SAVINGS REPORT_C": 31,
            "SAVINGS REPORT_D": 8,
            "SAVINGS REPORT_E": 13,
            "SAVINGS REPORT_F": 17,
            "SAVINGS REPORT_G": 17,
            "SAVINGS REPORT_H": 17,
            "SAVINGS REPORT_I": 19,
            "SAVINGS REPORT_J": 13,
            "SAVINGS REPORT_K": 13,
            "SAVINGS REPORT_L": 13,
            "SAVINGS REPORT_M": 13,
            "LOWEST ITEM COST REPORT_A": 5,
            "LOWEST ITEM COST REPORT_B": 5,
            "LOWEST ITEM COST REPORT_C": 31,
            "LOWEST ITEM COST REPORT_D": 8,
            "LOWEST ITEM COST REPORT_E": 13,
            "LOWEST ITEM COST REPORT_F": 13,
            "LOWEST ITEM COST REPORT_G": 14,
            "LOWEST ITEM COST REPORT_H": 24,
            "LOWEST ITEM COST REPORT_I": 14,
            "LOWEST ITEM COST REPORT_J": 24,
            "LOWEST ITEM COST REPORT_K": 14,
            "LOWEST ITEM COST REPORT_L": 24,
            "LOWEST ITEM COST REPORT_OTHER": 15,
            "LOWEST SUPPLIER REPORT_A": 5,
            "LOWEST SUPPLIER REPORT_B": 5,
            "LOWEST SUPPLIER REPORT_C": 31,
            "LOWEST SUPPLIER REPORT_D": 8,
            "LOWEST SUPPLIER REPORT_E": 13,
            "LOWEST SUPPLIER REPORT_F": 13,
            "LOWEST SUPPLIER REPORT_G": 24,
            "LOWEST SUPPLIER REPORT_H": 24,
            "LOWEST SUPPLIER REPORT_I": 24,
            "LOWEST SUPPLIER REPORT_J": 24,
            "LOWEST SUPPLIER REPORT_K": 24,
            "LOWEST SUPPLIER REPORT_OTHER": 24,
        }

        format_excel(filepath, data=f_data)
        insert_image(
            excel_url=filepath,
            worksheet_name="LOWEST ITEM COST REPORT",
            anchor=qed_logo_anchor,
        )
        insert_image(
            filepath,
            worksheet_name="LOWEST ITEM COST REPORT",
            anchor=buyer_logo_anchor,
            image_url=buyer_logo,
        )
        try:
            time = datetime.datetime.now()
            with open(filepath, "rb") as f:
                #save copy to S3
                # storage = PrivateMediaStorage()
                # url = f"rfq_reports/{time.year}/{time.month}/{time.day}/{os.path.basename(f.name)}"
                # storage.save(url, l)
                report = File(f, name=f.name)

                report, created = RFQCategoryReport.objects.update_or_create(
                    category_id=category_id, defaults={"financial": report}
                )
                context = {
                    "filepath": report.financial.url,
                    "response_message": "Report generated successfully",
                }
                print(context)
                return context

        except Exception as e:
            capture_exception(e)
            print(e)
            context = {
                "response_message": "Error generating report"
            }
            return context

        # context = {
        #     "response_message": "Report generated successfully",
        #     "filepath": filepath.split("/", 1)[1],
        # }
        # return context

    def create_lowest_cost_report_worksheet(self, pricing_worksheet, data):
        category_responsive_bidders = data["category_responsive_bidders"]
        category_total_bidders = data["category_total_bidders"]
        category = data["category"]
        items = data["items"]
        lowest_cumm_suppliers = data["lowest_cumm_suppliers"]
        total_for_progress = data["total_for_progress"]
        excel_cols = data["excel_cols"]
        percentage_responses = data["percentage_responses"]

        try:
            rfq_current_cost = 0
            rfq_eprocure_cost = 0
            rfq_total_potential_savngs = 0

            details_summary = [
                ("", "", "", ""),
                ("", "", "", ""),
                ("", "", "", ""),
                ("", "", "", ""),
                ("", "Bnb_Lowest Cost Report"),
                ("", f"Bnb_Category: {category.name.title()}", "", ""),
                ("", "", "", ""),
                ("", "Number of bidders", "", f"Ac_{category_total_bidders}"),
                (
                    "",
                    "Responsive bidders",
                    "",
                    "Ac_{}".format(category_responsive_bidders),
                ),
                ("", "Percentage responses", "", f"Ac_{percentage_responses}"),
                ("", "Currency", "", "Ac_{}".format(category.currency.initials)),
                ("", "", "", ""),
            ]
            for row in details_summary:
                pricing_worksheet.append(row)

            pricing_worksheet.merge_cells("B5:C5")
            pricing_worksheet.merge_cells("B6:H6")
            pricing_worksheet.merge_cells("B8:C8")
            pricing_worksheet.merge_cells("B9:C9")
            pricing_worksheet.merge_cells("B10:C10")
            pricing_worksheet.merge_cells("B11:C11")

            supplier_headers = ["", "", "", "", "", "", "", "", "", "", ""]

            items_header = [
                "",
                "hc_#",
                "h_Item Description",
                "hc_Qty",
                "h_UOM",
                "hc_Lowest Cost 1",
                "h_Lowest Supplier 1",
                "hc_Lowest Cost 2",
                "h_Lowest Supplier 2",
                "hc_Lowest Cost 3",
                "h_Lowest Supplier 3",
            ]

            no_of_items = len(items)
            no_of_rows_at_the_top = 14
            supplier_totals_row = no_of_rows_at_the_top + no_of_items
            formulas_row = []

            formulas_row.extend(
                ["qT_{}".format(f"= SUM(F15:F{supplier_totals_row})"), ""]
            )
            formulas_row.extend(
                ["qT_{}".format(f"= SUM(H15:H{supplier_totals_row})"), ""]
            )
            formulas_row.extend(
                ["qT_{}".format(f"= SUM(J15:J{supplier_totals_row})"), ""]
            )
            pricing_worksheet.merge_cells("C6:H6")
            c = 0

            for count, supplier in enumerate(lowest_cumm_suppliers, start=1):
                # self.result += 1
                # self.progress_recorder.set_progress(self.result, total_for_progress)

                supplier_total_row_name = excel_cols[11 + c]
                formulas_row.extend(
                    [
                        "qT_{}".format(
                            f"= SUM({supplier_total_row_name}15:{supplier_total_row_name}{supplier_totals_row})"
                        ),
                    ]
                )

                supplier_headers.append(
                    "hc_%s" % supplier.supplier.company_name.title()
                )
                items_header.append("hc_Cost")
                c += 1

            supplier_headers = tuple(supplier_headers)
            pricing_worksheet.append(supplier_headers)
            items_header = tuple(items_header)
            pricing_worksheet.append(items_header)

            j = 1
            for item in items:
                # self.result += 1
                # self.progress_recorder.set_progress(self.result, total_for_progress)

                outlier_score = item.outlier_score
                responses = item.responses

                try:
                    rfq_current_cost += float(item.current_price)
                    rfq_eprocure_cost += item.eprocure_supplier_cost(responses)
                    rfq_total_potential_savngs += item.savings(responses)
                except:
                    rfq_current_cost += 0
                    rfq_eprocure_cost += 0
                    rfq_total_potential_savngs += 0

                if item.item_savings_value is None:
                    item_savings = item.savings(responses)
                else:
                    item_savings = item.item_savings_value

                if item.eprocure_total_value is None:
                    item_eprocure_total = item.eprocure_total(responses)
                else:
                    item_eprocure_total = item.eprocure_total_value

                item_current_total = item.current_total
                total_items_current_total = item_current_total * item.quantity
                item_row = []
                
                tendersure_suppliers = item.multi_item_tender_eprocure_suppliers(
                    responses
                )
                price_1 = tendersure_suppliers[0]["price"]
                price_2 = tendersure_suppliers[1]["price"]
                price_3 = tendersure_suppliers[2]["price"]

                try:
                    if item.second_description is None:
                        second_description = "Ac_N/A"
                    else:
                        second_description = item.second_description
                except:
                    second_description = "Ac_N/A"

                try:
                    if (
                        item_eprocure_total > total_items_current_total
                        and total_items_current_total != 0
                    ):
                        savings_lowest_supplier = "Current Supplier"
                        price_3 = price_2
                        price_2 = price_1
                        price_1 = total_items_current_total
                        item_row.extend(
                            [
                                "",
                                "Ac_{}".format(item.item_number),
                                item.item_description,
                                "Ac_{}".format(item.quantity),
                                second_description,
                                "Tr_{}".format(f"{price_1:}"),
                                savings_lowest_supplier,
                                "Tr_{}".format(f"{price_2:}"),
                                tendersure_suppliers[0]["supplier"],
                                "Tr_{}".format(f"{price_3:}"),
                                tendersure_suppliers[1]["supplier"],
                            ]
                        )
                    else:
                        savings_lowest_supplier = tendersure_suppliers[0]["supplier"]
                        item_row.extend(
                            [
                                "",
                                "Ac_{}".format(item.item_number),
                                item.item_description,
                                "Ac_{}".format(item.quantity),
                                second_description,
                                "Tr_{}".format(f"{price_1:}"),
                                savings_lowest_supplier,
                                "Tr_{}".format(f"{price_2:}"),
                                tendersure_suppliers[1]["supplier"],
                                "Tr_{}".format(f"{price_3:}"),
                                tendersure_suppliers[2]["supplier"],
                            ]
                        )
                except:
                    item_row.extend(
                        [
                            "",
                            "Ac_{}".format(item.item_number),
                            item.item_description,
                            "Ac_{}".format(item.quantity),
                            second_description,
                            "Tr_{}".format(f"{price_1:}"),
                            tendersure_suppliers[0]["supplier"],
                            "Tr_{}".format(f"{price_2:}"),
                            tendersure_suppliers[1]["supplier"],
                            "Tr_{}".format(f"{price_3:}"),
                            tendersure_suppliers[2]["supplier"],
                        ]
                    )
                
                for count, supplier in enumerate(lowest_cumm_suppliers, start=1):
                    # self.result += 1
                    # self.progress_recorder.set_progress(self.result, total_for_progress)
                    if not supplier.supplier:
                        supplier_total = 0
                        item_row.append(supplier_total)
                    else:
                        # item_response = supplier.supplier.rfq_item_response(item=item)
                        item_response = RFQItemResponse.objects.filter(
                            supplier=supplier.supplier, rfq_item=item
                        ).first()
                        if item_response is not None:
                            # supplier_total = show(item_response.total)
                            supplier_total = item_response.total_price
                        else:
                            supplier_total = 0
                        s_total = round(supplier_total, 2)

                        if (
                            RFQItemResponse.objects.filter(rfq_item_id=item.id).count()
                            <= 2
                        ):
                            supplier_total = "Tr_{}".format(f"{s_total:}")
                        else:
                            if float(s_total) < float(outlier_score):
                                supplier_total = "O_{}".format(f"{s_total:}")
                            else:
                                supplier_total = "Tr_{}".format(f"{s_total:}")
                        item_row.append(supplier_total)

                item_row = tuple(item_row)
                pricing_worksheet.append(item_row)

            # rfq.total_savings = float(rfq_total_potential_savngs)
            # rfq.save()

            colums_totals = []
            colums_totals.extend(["", "", "qT_Total", "", ""])

            colums_totals.extend(formulas_row)

            pricing_worksheet.append(tuple(colums_totals))
            pricing_worksheet.append(("", "", ""))
            pricing_worksheet.append(("", ""))
            pricing_worksheet.append(("", "hc_#", "Nmrl_3_Notes"))
            pricing_worksheet.append(
                (
                    "",
                    "Ac_1",
                    "Nmrl_3_Outlier: The prices highlighted in this colour are outliers which means they are 50% lower than the median and are not considered in the analysis",
                )
            )
            pricing_worksheet.freeze_panes = pricing_worksheet["F15"]
            return
        except Exception as e:
            capture_exception(e)
            return

    def create_lowest_supplier_worksheet(self, workbook, data):
        try:
            category = data["category"]
            excel_cols = data["excel_cols"]
            rfq_items = data["items"]
            participants = data["participants"]
            category_total_bidders = data["category_total_bidders"]
            category_responsive_bidders = data["category_responsive_bidders"]
            percentage_responses = data["percentage_responses"]
            lowest_cumm_suppliers_ranked = data["lowest_cumm_suppliers_ranked"]
            lowest_cumm_suppliers_opp = data["lowest_cumm_suppliers_opp"]
            lowest_cumm_suppliers = data["lowest_cumm_suppliers"]
            total_for_progress = data["total_for_progress"]

            rfq_current_cost = 0
            rfq_eprocure_cost = 0
            rfq_total_potential_savngs = 0

            summary_rows = [
                ("", "Invited bidders", "", "Ac_{}".format(category_total_bidders)),
                (
                    "",
                    "Bidder responses",
                    "",
                    "Ac_{}".format(category_responsive_bidders),
                ),
                ("", "Percentage responses", "", "Ac_{}".format(percentage_responses)),
                ("", "Currency", "", "Ac_{}".format(category.currency.initials)),
            ]

            workbook.create_sheet("LOWEST SUPPLIER REPORT")
            lowest_supplier_worksheet = workbook["LOWEST SUPPLIER REPORT"]

            lowest_supplier_worksheet["B3"] = "Bnb_Category: {}".format(category.name)
            lowest_supplier_worksheet["B2"] = "Bnb_Lowest Supplier Report"
            lowest_supplier_worksheet["B4"] = ""

            for row in summary_rows:
                lowest_supplier_worksheet.append(row)

            empty_row = [""]
            empty_row = tuple(empty_row)
            lowest_supplier_worksheet.append(empty_row)
            lowest_supplier_worksheet.merge_cells("B2:C2")
            lowest_supplier_worksheet.merge_cells("B3:H3")
            lowest_supplier_worksheet.merge_cells("B5:C5")
            lowest_supplier_worksheet.merge_cells("B6:C6")
            lowest_supplier_worksheet.merge_cells("B7:C7")
            lowest_supplier_worksheet.merge_cells("B8:C8")

            s = 4
            supplier_totals_row = 12 + len(rfq_items)
            lowest_cumm_suppliers_totals = []
            # lowest_cumm_suppliers = list(lowest_cumm_suppliers_ranked) + list(
            #     lowest_cumm_suppliers_opp
            # )

            lowest_supplier_sub_headers = ["", "", "", "", ""]
            lowest_supplier_rank_headers = ["", "", "", "", ""]
            lowest_supplier_headers = [
                "",
                "hc_#",
                "h_Item Description",
                "hc_Qty",
                "h_UOM",
            ]

            scores = []
            for count, supplier in enumerate(lowest_cumm_suppliers, start=1):
                # self.result += 1
                # self.progress_recorder.set_progress(self.result, total_for_progress)

                row_name = excel_cols[s + count]
                lowest_cumm_suppliers_totals.append(
                    "qT_{}".format(
                        f"= SUM({row_name}13:{row_name}{supplier_totals_row})"
                    )
                )
                lowest_supplier_sub_headers.append(
                    "hc_%s" % supplier.supplier.company_name.title()
                )
                lowest_supplier_headers.append("hc_Total")
                if not supplier.has_outlier:
                    lowest_supplier_rank_headers.append(f"hc_Rank {count}")
                else:
                    lowest_supplier_rank_headers.append(f"hc_N/A")
                # scores.append(f"Tcb_{supplier.rfq_score}")
                scores.append(f"Tcb_{0}")

            lowest_supplier_headers = tuple(lowest_supplier_headers)
            lowest_supplier_sub_headers = tuple(lowest_supplier_sub_headers)
            lowest_supplier_rank_headers = tuple(lowest_supplier_rank_headers)
            lowest_supplier_worksheet.append(lowest_supplier_rank_headers)
            lowest_supplier_worksheet.append(lowest_supplier_sub_headers)
            lowest_supplier_worksheet.append(lowest_supplier_headers)

            j = 1
            for item in rfq_items:
                self.result += 1
                self.progress_recorder.set_progress(self.result, total_for_progress)

                outlier_score = item.outlier_score
                responses = item.responses
                try:
                    rfq_current_cost += float(item.current_price)
                    rfq_eprocure_cost += item.eprocure_supplier_cost(responses)
                    rfq_total_potential_savngs += item.savings(responses)
                except:
                    rfq_current_cost += 0
                    rfq_eprocure_cost += 0
                    rfq_total_potential_savngs += 0

                try:
                    if item.second_description is None:
                        second_description = "Ac_N/A"
                    else:
                        second_description = item.second_description
                except:
                    second_description = "Ac_N/A"
                   
                lowest_supplier_row = []

                lowest_supplier_row.extend(
                    [
                        "",
                        "Ac_{}".format(item.item_number),
                        item.item_description,
                        "Ac_{}".format(item.quantity),
                        second_description,
                    ]
                )

                for count, supplier in enumerate(lowest_cumm_suppliers, start=1):
                    if not supplier.supplier:
                        supplier_total = 0
                    else:
                        item_response = RFQItemResponse.objects.filter(supplier=supplier.supplier, rfq_item=item).first()
                        if item_response is not None:
                            # if item_response.total is not None:
                            #     supplier_total = show(item_response.total)
                            if item_response.total_price is not None:
                                supplier_total = item_response.total_price
                        else:
                            supplier_total = 0
                        s_total = round(supplier_total, 2)
                        if float(s_total) < float(outlier_score):
                            supplier_total = "O_{}".format(f"{s_total:}")
                        else:
                            supplier_total = "Tr_{}".format(f"{s_total:}")
                        lowest_supplier_row.append(supplier_total)

                lowest_supplier_row = tuple(lowest_supplier_row)
                lowest_supplier_worksheet.append(lowest_supplier_row)

                j += 1

            lowest_supplier_totals = []
            lowest_supplier_totals.extend(["", "", "qT_Total", "", ""])
            lowest_supplier_totals.extend(lowest_cumm_suppliers_totals)
            lowest_supplier_worksheet.append(tuple(lowest_supplier_totals))
            lowest_supplier_worksheet.append(empty_row)
            lowest_supplier_worksheet.append(empty_row)
            lowest_supplier_worksheet.append(("", "Ac_#", "Nmrl_3_Notes"))
            lowest_supplier_worksheet.append(
                (
                    "",
                    "Ac_1",
                    "Nmrl_3_Rank: This is the ranking of suppliers with the lowest cost and is not an outlier",
                )
            )
            lowest_supplier_worksheet.append(
                (
                    "",
                    "Ac_2",
                    "Nmrl_3_Outlier: The prices highlighted in this colour are "
                    "outliers which means that they are 50% lower than the median "
                    "and are not considered in the analysis.",
                )
            )
            lowest_supplier_worksheet.freeze_panes = lowest_supplier_worksheet["F13"]
            return
        except Exception as e:
            capture_exception(e)
            return

    def create_savings_worksheet(self, workbook, data):
        workbook.create_sheet("SAVINGS REPORT")
        savings_worksheet = workbook["SAVINGS REPORT"]

        try:
            category = data["category"]
            rfq_items = data["items"]
            category_total_bidders = data["category_total_bidders"]
            category_responsive_bidders = data["category_responsive_bidders"]
            percentage_responses = data["percentage_responses"]
            empty_row = [""]
            empty_row = tuple(empty_row)
            total_for_progress = data["total_for_progress"]

            details_summary = [
                ("", "Bnb_Savings Report"),
                ("", f"Bnb_Category: {category.name}", "", "", ""),
                ("", "", "", ""),
                ("", "Number of bidders", "", "Ac_{}".format(category_total_bidders)),
                (
                    "",
                    "Bidder responses",
                    "",
                    "Ac_{}".format(category_responsive_bidders),
                ),
                ("", "Percentage responses", "", "Ac_{}".format(percentage_responses)),
                ("", "Currency", "", "Ac_{}".format(category.currency.initials)),
            ]

            savings_worksheet.append(empty_row)

            for row in details_summary:
                savings_worksheet.append(row)

            savings_worksheet.append(empty_row)

            savings_items_header = [
                "",
                "hc_#",
                "h_Item Description",
                "hc_Qty",
                "h_UOM",
                "h_Current Supplier",
                "hc_Current Price",
                "hc_Current Cost",
                "h_Lowest Supplier",
                "hc_Lowest Cost",
                "hc_Unit Savings",
                "hc_Total Savings",
                "hc_% Savings",
            ]

            savings_worksheet.append(savings_items_header)
            savings_worksheet.merge_cells("B2:C2")
            savings_worksheet.merge_cells("B3:H3")
            savings_worksheet.merge_cells("B4:C4")
            savings_worksheet.merge_cells("B5:C5")
            savings_worksheet.merge_cells("B6:C6")
            savings_worksheet.merge_cells("B7:C7")
            savings_worksheet.merge_cells("B8:C8")

            j = 1
            for item in rfq_items:
                # self.result += 1
                # self.progress_recorder.set_progress(self.result, total_for_progress)
                responses = item.responses

                savings_item_row = []
                item_savings = item.savings(responses)

                if item.eprocure_total_value is None:
                    item_eprocure_total = item.eprocure_total(responses)
                else:
                    item_eprocure_total = item.eprocure_total_value

                item_current_total = item.current_total
                total_items_current_total = item_current_total * item.quantity
                unit_savings = item_savings / item.quantity
                rfq_item_current_total = "T_{}".format(f"{total_items_current_total:}")
                tendersure_suppliers = item.multi_item_tender_eprocure_suppliers(
                    responses
                )
                price_1 = tendersure_suppliers[0]["price"]
                savings_lowest_supplier = ""

                if (
                    item_eprocure_total > total_items_current_total
                    and total_items_current_total != 0
                ):
                    item_savings_cell = "Ar_0"
                    savings_lowest_supplier = "Current Supplier"
                else:
                    item_savings_cell = "Tr_{}".format(unit_savings)
                    savings_lowest_supplier = tendersure_suppliers[0]["supplier"]

                try:
                    if item.second_description is None:
                        second_description = "Ac_N/A"
                    else:
                        second_description = item.second_description
                except:
                    second_description = "Ac_N/A"

                try:
                    if item.current_supplier is None:
                        current_supplier = "Ac_N/A"
                    else:
                        current_supplier = item.current_supplier.title()
                except:
                    current_supplier = "Ac_N/A"

                savings_item_row.extend(
                    [
                        "",
                        "Ac_{}".format(item.item_number),
                        item.item_description,
                        "Ac_{}".format(item.quantity),
                        second_description,
                        current_supplier,
                        "T_{}".format(f"{item_current_total:}")
                        if item_current_total > 0
                        else "Ac_N/A",
                        rfq_item_current_total,
                        savings_lowest_supplier,
                        "Tr_{}".format(f"{price_1:}"),
                        item_savings_cell,
                        "T_{}".format(item_savings),
                        "Ac_{}".format(item.percentage_savings(responses)),
                    ]
                )

                savings_item_row = tuple(savings_item_row)
                savings_worksheet.append(savings_item_row)
                j += 1
            total_percentage_savings = (
                f"=L{10 + len(rfq_items)+1}/H{10 + len(rfq_items)+1}"
            )
            totals_row = [
                "",
                "",
                "B_Total",
                "",
                "",
                "",
                f"Tc_=SUM(G11:G{10+len(rfq_items)})",
                f"Tc_=SUM(H11:H{10 + len(rfq_items)})",
                "",
                f"Tr_=SUM(J11:J{10 + len(rfq_items)})",
                f"Tr_=SUM(K11:K{10 + len(rfq_items)})",
                f"Tr_=SUM(L11:L{10 + len(rfq_items)})",
                "Pc_{}".format(total_percentage_savings),
            ]
            savings_worksheet.append(totals_row)

            savings_worksheet.append(empty_row)
            savings_worksheet.append(["", "hc_#", "Nmrl_3_Notes"])
            savings_worksheet.append(
                [
                    "",
                    "Ac_1",
                    "Nmrl_3_Where the current prices are lower than the lowest price obtained, current price is retained",
                ]
            )
            savings_worksheet.append(
                [
                    "",
                    "Ac_2",
                    "Nmrl_3_N/A means that the current supplier has not been provided",
                ]
            )
            savings_worksheet.freeze_panes = savings_worksheet["D11"]
        except Exception as e:
            print(e)
            capture_exception(e)
        return


class RFQSummaryReport(Task):
    """
    RFQ Summary Report, summary savings and individual category savings worksheet
    :job-Rfq
    :rfq-category
    """

    name = "RFQSummaryReport"
    progress_recorder = None
    result = 0

    def run(self, *args, **kwargs):
        self.progress_recorder = ProgressRecorder(self)
        context = self.report(kwargs["rfq_id"])
        return context

    def report(self, rfq_id):
        job = Rfq.objects.filter(id=rfq_id).first()
        company = job.company
        rfqs = job.closed_rfqs
        rfqs_data_all = []

        total_for_progress = len(rfqs)

        time = datetime.datetime.now()
        dir_name = Path("media/rfq_reports/{}/{}".format(time.year, time.month))
        dir_name.mkdir(parents=True, exist_ok=True)

        time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
        filepath = "{}/{}_{}_{}.xlsx".format(
            dir_name, job.id, job.company.company_name.replace(" ", "_"), time_only
        )
        # delete similar reports run this month
        match_string = "{}_{}".format(
            job.id, job.company.company_name.replace(" ", "_")
        )
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")
        workbook = Workbook()
        qed_logo = Image("static/core/img/no-company-image128-128.png")

        workbook.create_sheet("SUMMARY")
        summary_worksheet = workbook["SUMMARY"]

        data = {
            "SUMMARY_B": 5,
            "SUMMARY_C": 49,
            "SUMMARY_D": 20,
            "SUMMARY_E": 16,
            "SUMMARY_F": 16,
            "SUMMARY_G": 16,
            "SUMMARY_H": 16,
            "SUMMARY_I": 16,
            "SUMMARY_J": 16,
            "SUMMARY_K": 16,
        }
        f_data = {}

        for category in rfqs:
            self.result += 1
            self.progress_recorder.set_progress(self.result, total_for_progress)
            rfq_participants = category.participants
            participants = rfq_participants["participants"]
            items = category.items
            category_total_bidders = len(category.invited_suppliers)
            category_responsive_bidders = rfq_participants["count"]
         
            rfq_data = {
                "company": company,
                "category": category,
                "category_total_bidders": category_total_bidders,
                "participants": participants,
                "items": items,
                "category_responsive_bidders": category_responsive_bidders,
            }
            rfqs_data_all.append(rfq_data)

            total_for_progress += len(rfqs_data_all)

            f_data[f"{category.unique_reference.upper()}_A"] = 5
            f_data[f"{category.unique_reference.upper()}_B"] = 5
            f_data[f"{category.unique_reference.upper()}_C"] = 31
            f_data[f"{category.unique_reference.upper()}_D"] = 8
            f_data[f"{category.unique_reference.upper()}_E"] = 13
            f_data[f"{category.unique_reference.upper()}_F"] = 17
            f_data[f"{category.unique_reference.upper()}_G"] = 17
            f_data[f"{category.unique_reference.upper()}_H"] = 17
            f_data[f"{category.unique_reference.upper()}_I"] = 19
            f_data[f"{category.unique_reference.upper()}_J"] = 13
            f_data[f"{category.unique_reference.upper()}_K"] = 13
            f_data[f"{category.unique_reference.upper()}_L"] = 13
            f_data[f"{category.unique_reference.upper()}_M"] = 13

        data.update(f_data)

        report_data = {
            "rfqs_data": rfqs_data_all,
            "total_for_progress": total_for_progress,
        }

        summary_report = self.create_summary_worksheet(
            company, summary_worksheet, report_data
        )
        savings_report = self.create_savings_worksheet(workbook, report_data)

        if summary_report is True and savings_report is True:
            if company.company_logo_url is not None and company.company_logo_url != "":
                buyer_logo_url = get_file_path(company.company_logo_url)
            else:
                buyer_logo_url = "static/core/img/no-company-image128-128.png"
            qed_logo_anchor = "L1"
            qed_logo_worksheet = "SUMMARY"
            buyer_logo_anchor = "B1"
            buyer_logo_worksheet = "SUMMARY"

            workbook.remove(workbook["Sheet"])
            workbook.save(filepath)
            format_excel(filepath, data=data)

            insert_image(
                filepath,
                worksheet_name="SUMMARY",
                anchor=qed_logo_anchor,
            )
            insert_image(
                filepath,
                worksheet_name="SUMMARY",
                anchor=buyer_logo_anchor,
                image_url=buyer_logo_url,
            )

            # context = {
            #     "response_message": "Report generated successfully",
            #     "filepath": filepath.split("/", 1)[1],
            # }
            try:
                time = datetime.datetime.now()
                with open(filepath, "rb") as f:
                    #save copy to S3
                    # storage = PrivateMediaStorage()
                    # url = f"rfq_reports/{time.year}/{time.month}/{time.day}/{os.path.basename(f.name)}"
                    # storage.save(url, l)
                    report = File(f, name=f.name)

                    report, created = RFQJobReport.objects.update_or_create(
                        job_id=rfq_id, defaults={"job_savings": report}
                    )
                    context = {
                        "filepath": report.job_savings.url,
                        "response_message": "Report generated successfully",
                    }
                    print(context)
                    return context

            except Exception as e:
                capture_exception(e)
                print(e)
                context = {
                    "response_message": "Error generating report"
                }
        else:
            if summary_report is False:
                context = {"response_message": "Failed to generate summary report"}
            elif savings_report is False:
                context = {"response_message": "Failed to generate savings report"}
            else:
                context = {"response_message": "Failed to generate report"}

        return context

    def create_summary_worksheet(self, company, summary_worksheet, report_data):
        try:
            rfqs_data = report_data["rfqs_data"]
            total_for_progress = report_data["total_for_progress"]

            summary_data = [
                ("", "", "", ""),
                ("", "", "", ""),
                ("", "", "", ""),
                ("", "", "", ""),
                ("", "Bnb_Summary of Savings Report", "", ""),
                ("", "", "", ""),
                ("", "Number of categories", "", f"Ac_{len(rfqs_data)}"),
                ("", "", "", ""),
                ("", "", "", ""),
                ("", "", "", ""),
            ]

            for row in summary_data:
                summary_worksheet.append(row)

            summary_worksheet.append(
                (
                    "",
                    "h_#",
                    "h_Category Description",
                    "h_Category Code",
                    "hc_Current Cost(KES)",
                    "hc_Lowest Cost(KES)",
                    "hc_Potential Savings",
                    "hc_% Savings",
                    "hc_Total Bidders",
                    "hc_Responsive Bidders",
                )
            )
            summary_worksheet.merge_cells("B7:C7")

            # column totals
            total_rfq_bidders = 0
            total_prequal_bidders = 0
            total_resp_bidders = 0
            total_curr_cost = 0
            total_low_cost = 0
            total_savings = 0
            count = 0

            for rfq in rfqs_data:
                count += 1
                rfq_current_cost = 0
                rfq_eprocure_cost = 0
                rfq_total_potential_savngs = 0
                items = rfq["items"]
                category = rfq["category"]
                category_total_bidders = rfq["category_total_bidders"]
                category_responsive_bidders = rfq["category_responsive_bidders"]

                for item in items:
                    responses = item.responses
                    try:
                        rfq_current_cost += item.actual_current_total
                        rfq_eprocure_cost += item.eprocure_supplier_cost(responses)
                        rfq_total_potential_savngs += item.savings(responses)
                    except:
                        rfq_current_cost += 0
                        rfq_eprocure_cost += 0
                        rfq_total_potential_savngs += 0
                try:
                    percentage_savings = "{}%".format(
                        round(((rfq_total_potential_savngs / rfq_current_cost) * 100))
                    )
                except Exception as e:
                    percentage_savings = 0

                summary_worksheet.append(
                    (
                        "",
                        "Ac_%d" % (count),
                        category.name,
                        category.unique_reference.upper(),
                        "T_{}".format(rfq_current_cost),
                        "T_{}".format(rfq_eprocure_cost),
                        "T_{}".format(rfq_total_potential_savngs),
                        "Ac_{}".format(percentage_savings),
                        "Ac_{}".format(category_total_bidders),
                        "Ac_{}".format(category_responsive_bidders),
                    )
                )
                summary_worksheet.merge_cells("B5:D5")

                total_rfq_bidders += category_total_bidders
                total_resp_bidders += category_responsive_bidders
                total_curr_cost += rfq_current_cost
                total_low_cost += rfq_eprocure_cost
                total_savings += rfq_total_potential_savngs

            try:
                buyer_percentage_savings = "{}%".format(
                    round(((total_savings / total_curr_cost) * 100))
                )

            except:
                buyer_percentage_savings = 0

            summary_worksheet.append(
                (
                    "",
                    "",
                    "",
                    "qT_Total",
                    "T_{}".format(total_curr_cost),
                    "T_{}".format(total_low_cost),
                    "T_{}".format(total_savings),
                    "Ac_{}".format(buyer_percentage_savings),
                    "Ac_{}".format(total_rfq_bidders),
                    "Ac_{}".format(total_resp_bidders),
                )
            )
            summary_worksheet.freeze_panes = summary_worksheet["E12"]
            return True
        except Exception as e:
            print(e)
            capture_exception(e)
            return False

    def create_savings_worksheet(self, workbook, report_data):
        rfqs_data = report_data["rfqs_data"]
        total_for_progress = report_data["total_for_progress"]

        try:
            for rfq in rfqs_data:
                category = rfq["category"]
                rfq_items = rfq["items"]
                category_total_bidders = rfq["category_total_bidders"]
                category_responsive_bidders = rfq["category_responsive_bidders"]
                # percentage_responses = rfq["percentage_responses"]
                empty_row = [""]
                empty_row = tuple(empty_row)
                # total_for_progress = rfq["total_for_progress"]
                try:
                    percentage_responses = "{}%".format(
                        round(
                            (
                                (category_responsive_bidders / category_total_bidders)
                                * 100
                            )
                        )
                    )
                except Exception as e:
                    percentage_responses = 0

                # self.result += 1
                # self.progress_recorder.set_progress(self.result, total_for_progress)

                workbook.create_sheet(f"{category.unique_reference.upper()}")
                savings_worksheet = workbook[f"{category.unique_reference.upper()}"]

                details_summary = [
                    ("", f"Bnb_{category.name} Savings Report", ""),
                    ("", "", "", ""),
                    (
                        "",
                        "Number of bidders",
                        "",
                        "Ac_{}".format(category_total_bidders),
                    ),
                    (
                        "",
                        "Bidder responses",
                        "",
                        "Ac_{}".format(category_responsive_bidders),
                    ),
                    (
                        "",
                        "Percentage responses",
                        "",
                        "Ac_{}".format(percentage_responses),
                    ),
                    ("", "Currency", "", "Ac_{}".format(category.currency.initials)),
                ]

                savings_worksheet.append(empty_row)

                for row in details_summary:
                    savings_worksheet.append(row)

                savings_worksheet.append(empty_row)

                savings_items_header = [
                    "",
                    "hc_#",
                    "h_Item Description",
                    "hc_Qty",
                    "h_UOM",
                    "h_Current Supplier",
                    "hc_Current Price",
                    "hc_Current Cost",
                    "h_Lowest Supplier",
                    "hc_Lowest Cost",
                    "hc_Unit Savings",
                    "hc_Total Savings",
                    "hc_% Savings",
                ]

                savings_worksheet.append(savings_items_header)
                savings_worksheet.merge_cells("B2:H2")
                savings_worksheet.merge_cells("B4:C4")
                savings_worksheet.merge_cells("B5:C5")
                savings_worksheet.merge_cells("B6:C6")
                savings_worksheet.merge_cells("B7:C7")

                j = 1
                for item in rfq_items:
                    # self.result += 1
                    # self.progress_recorder.set_progress(self.result, total_for_progress)
                    outlier_score = item.outlier_score
                    responses = item.responses

                    savings_item_row = []
                    if item.item_savings_value is None:
                        item_savings = item.savings(responses)
                    else:
                        item_savings = item.item_savings_value

                    if item.eprocure_total_value is None:
                        item_eprocure_total = item.eprocure_total(responses)
                    else:
                        item_eprocure_total = item.eprocure_total_value

                    try:
                        item_current_total = item.current_total
                        total_items_current_total = item_current_total * item.quantity
                        unit_savings = item_savings / item.quantity
                    except:
                        item_current_total = 0
                        total_items_current_total = 0

                    # rfq_item_current_total = item.quantity * item_current_total

                    rfq_item_current_total = "T_{}".format(
                        f"{total_items_current_total:}"
                    )
                    # if float(rfq_item_current_total) < float(outlier_score):
                    #     # rfq_item_current_total = "O_{}".format(f"{rfq_item_current_total:}")
                    #     rfq_item_current_total = "Ac_N/A".format(
                    #         f"{rfq_item_current_total:}"
                    #     )
                    # else:
                    #     rfq_item_current_total = "T_{}".format(
                    #         f"{rfq_item_current_total:}"
                    #     )

                    tendersure_suppliers = item.multi_item_tender_eprocure_suppliers(responses)
                    price_1 = tendersure_suppliers[0]["price"]
                    savings_lowest_supplier = ""

                    total_savings = item.quantity * item_savings

                    # if (
                    #     item_eprocure_total > item_current_total
                    #     and item_current_total != 0
                    # ):
                    if (
                        item_eprocure_total > total_items_current_total
                        and total_items_current_total != 0
                    ):
                        item_savings_cell = "Ar_0"
                        savings_lowest_supplier = "Current Supplier"
                    else:
                        item_savings_cell = "Tr_{}".format(unit_savings)
                        savings_lowest_supplier = tendersure_suppliers[0]["supplier"]

                    try:
                        if item.second_description is None:
                            second_description = "Ac_N/A"
                        else:
                            second_description = item.second_description
                    except:
                        second_description = "Ac_N/A"

                    try:
                        if item.current_supplier is None:
                            current_supplier = "Ac_N/A"
                        else:
                            current_supplier = item.current_supplier.title()
                    except:
                        current_supplier = "Ac_N/A"

                    # if item_current_total != 0:
                    savings_item_row.extend(
                        [
                            "",
                            "Ac_{}".format(item.item_number),
                            item.item_description,
                            "Ac_{}".format(item.quantity),
                            second_description,
                            current_supplier,
                            "T_{}".format(f"{item_current_total:}")
                            if item_current_total > 0
                            else "Ac_N/A",
                            rfq_item_current_total,
                            savings_lowest_supplier,
                            "Tr_{}".format(f"{price_1:}"),
                            item_savings_cell,
                            "T_{}".format(item_savings),
                            "Ac_{}".format(item.percentage_savings(responses)),
                        ]
                    )

                    savings_item_row = tuple(savings_item_row)
                    savings_worksheet.append(savings_item_row)
                    j += 1
                total_percentage_savings = (
                    f"=L{9 + len(rfq_items)+1}/H{9 + len(rfq_items)+1}"
                )
                totals_row = [
                    "",
                    "",
                    "B_Total",
                    "",
                    "",
                    "",
                    f"Tc_=SUM(G10:G{9 + len(rfq_items)})",
                    f"Tc_=SUM(H10:H{9 + len(rfq_items)})",
                    "",
                    f"Tr_=SUM(J10:J{9 + len(rfq_items)})",
                    f"Tr_=SUM(K10:K{9 + len(rfq_items)})",
                    f"Tc_=SUM(L10:L{9 + len(rfq_items)})",
                    "Pc_{}".format(total_percentage_savings),
                ]
                savings_worksheet.append(totals_row)

                savings_worksheet.append(empty_row)
                savings_worksheet.append(["", "hc_#", "Nmrl_3_Notes"])
                savings_worksheet.append(
                    [
                        "",
                        "Ac_1",
                        "Nmrl_3_Where the current prices are lower than the lowest price obtained, current price is retained",
                    ]
                )
                savings_worksheet.append(
                    [
                        "",
                        "Ac_2",
                        "Nmrl_3_N/A means that the current supplier has not been provided",
                    ]
                )
                savings_worksheet.freeze_panes = savings_worksheet["D10"]
            return True

        except Exception as e:
            print(e)
            capture_exception(e)
            return False


@shared_task(bind=True)
def rfq_participation_status(self, rfq_id=None, category_id=None):
    """
    Get Participation Status for RFQ Invited suppliers
    """
    try:
        if rfq_id is not None:
            job = Rfq.objects.filter(id=rfq_id).first()
            company = job.company
            data = {}
            time = datetime.datetime.now()
            if job is None:
                return {"message": "Job not an RFQ"}
            else:
                # categories = job.categories.filter(invite_only=True)
                categories = job.categories
                if len(categories) < 1:
                    return {"message": "RFQ job has no closed categories"}
                else:
                    rfq_suppliers = []
                    for cat in categories:
                        rfq = cat
                        if rfq is not None:
                            rfq_participants = rfq.participants
                            rfq_invited_suppliers = rfq.invited_suppliers

                        rfq_data = {
                            "category": cat,
                            "participants": rfq_participants,
                            "invitees": rfq_invited_suppliers,
                        }

                        rfq_suppliers.append(rfq_data)

        # elif category_id is not None:
        #     category = Category.objects.select_related("job").get(id=category_id)
        #     company = category.job.company
        #     job = category.job
        #     data = {}
        #     time = datetime.datetime.now()

        #     if category is None:
        #         return {"message": "RFQ Category does not exist"}
        #     else:
        #         rfq_suppliers = []
        #         rfq = category.rfq
        #         if rfq is not None:
        #             rfq_participants = rfq.participants
        #             rfq_invited_suppliers = rfq.invited_suppliers

        #         rfq_data = {
        #             "category": category,
        #             "participants": list(rfq_participants),
        #             "invitees": rfq_invited_suppliers,
        #         }
        #         rfq_suppliers.append(rfq_data)

        else:
            return {"message": "RFQ job/category does not exist "}

        # Generate Report
        dir_name = Path(
            "media/rfq_participants/%s/%s/%s"
            % (company.company_name, time.year, job.title)
        )  # folder structure
        dir_name.mkdir(parents=True, exist_ok=True)
        time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
        filepath = "{}/{}_{}_{}.xlsx".format(
            dir_name, job.unique_reference, job.title, time_only
        )
        # delete similar reports run this month
        match_string = "{}_{}".format(job.id, company.company_name.replace(" ", "_"))
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")
        # create and save a workbook to later populate
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f"Summary"

        if company.company_logo_url is not None and company.company_logo_url != "":
            buyer_logo_url = get_file_path(company.company_logo_url)
        else:
            buyer_logo_url = "static/core/img/no-company-image128-128.png"
        qed_logo_anchor = "D1"
        qed_logo_worksheet = worksheet.title
        buyer_logo_anchor = "C1"
        buyer_logo_worksheet = worksheet.title

        worksheet["C5"] = "Nb_{}".format(job.title)
        worksheet["C6"] = "Nb_{}".format(job.unique_reference)
        worksheet["C7"] = "Nb_RFQ Participation Report"
        worksheet["C8"] = "Nb_"

        worksheet.append(
            (
                "",
                "",
                "",
                "hc_Participants",
            )
        )
        worksheet.append(("", "hc_No", "h_Category Name", "h_Invited", "h_Yes", "h_No"))
        worksheet.merge_cells("D9:F9")
        data["Summary_C"] = 62
        data["Summary_D"] = 10
        data["Summary_E"] = 10
        data["Summary_F"] = 10

        total_invited = 0
        total_participants = 0
        total_non_participants = 0
        result = 0
        progress_recorder = ProgressRecorder(self)
        progress_total = len(rfq_suppliers)

        for count, rfq_data in enumerate(rfq_suppliers, start=1):
            result += 1
            progress_recorder.set_progress(result, progress_total)

            category = rfq_data["category"]
            invitees = rfq_data["invitees"]
            parts = rfq_data["participants"]
            participants = parts["participants"]
            participants_count = parts["count"]
            # percentage_participation = "{:.2f}".format(
            #     len(participants) / len(invitees)
            # )
            percentage_participation = "{}%".format(
                round(((participants_count / len(invitees)) * 100))
            )
            non_participants = len(rfq_data["invitees"]) - participants_count
            worksheet.append(
                (
                    "",
                    "Ac_%d" % (count),
                    rfq_data["category"].name,
                    "Ac_{}".format(len(invitees)),
                    "Ac_{}".format(participants_count),
                    non_participants,
                )
            )

            total_invited += len(invitees)
            total_participants += participants_count
            total_non_participants += non_participants

            # create individual sheets
            workbook.create_sheet(category.unique_reference.replace("/", "_"))
            details_worksheet = workbook[category.unique_reference.replace("/", "_")]
            # formating_data = {
            #     "Summary_C": 60,
            # }
            empty_row = [""]
            empty_row = tuple(empty_row)
            details_worksheet.append(empty_row)
            details_worksheet["C2"] = "Nb_{}".format(category.name)
            details_worksheet["C3"] = "Nb_"
            details_worksheet["C4"] = "Number of bidders invited"
            details_worksheet["C5"] = "Number of bidders participated"
            details_worksheet["C6"] = "% Participation"
            details_worksheet["D4"] = "Ac_{}".format(len(invitees))
            details_worksheet["D5"] = "Ac_{}".format(participants_count)
            details_worksheet["D6"] = "Ac_{}".format(percentage_participation)
            details_worksheet["C7"] = "Nb_"

            details_worksheet.append(
                (
                    "",
                    "",
                    "",
                    "",
                    "hc_Participation Status",
                )
            )

            details_worksheet.append(
                (
                    "",
                    "hc_No",
                    "h_Supplier",
                    "h_Phone Number",
                    "h_Yes",
                    "h_No",
                )
            )
            details_worksheet.merge_cells("E8:F8")
            details_worksheet.merge_cells("C2:E2")
            # formating
            data[f'{category.unique_reference.replace("/", "_")}_C'] = 35
            data[f'{category.unique_reference.replace("/", "_")}_D'] = 15
            data[f'{category.unique_reference.replace("/", "_")}_E'] = 10
            data[f'{category.unique_reference.replace("/", "_")}_F'] = 10

            status_yes = ""
            status_no = ""
            for count, supplier in enumerate(invitees, start=1):
                if supplier is not None:
                    if supplier in participants:
                        status_yes = "Yes"
                        status_no = " "
                    else:
                        status_no = "No"
                        status_yes = " "

                    if isinstance(supplier, str):
                        company = supplier
                        phone = "N/A"
                    else:
                        company = supplier.company_name
                        phone = supplier.phone_number

                    details_worksheet.append(
                        (
                            "",
                            "Ac_%d" % (count),
                            company,
                            phone,
                            "Ac_{}".format(status_yes),
                            "Ac_{}".format(status_no),
                        )
                    )

        worksheet.append(
            (
                "",
                "",
                "Total",
                "Ac_{}".format(total_invited),
                "Ac_{}".format(total_participants),
                "Ac_{}".format(total_non_participants),
            )
        )
        workbook.save(filepath)
        format_excel(filepath, data=data)
        insert_image(
            excel_url=filepath,
            worksheet_name=qed_logo_worksheet,
            anchor=qed_logo_anchor,
        )
        insert_image(
            filepath,
            buyer_logo_worksheet,
            buyer_logo_anchor,
            image_url=buyer_logo_url,
        )

        context = {
            "response_message": "Report generated successfully",
            "filepath": filepath.split("/", 1)[1],
        }

        return context

    except Exception as e:
        print(e)
        capture_exception(e)
        context = {
            "messages": "Failed to generate report",
        }
        return context
