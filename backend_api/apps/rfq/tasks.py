import datetime
from email import message
import os
from pathlib import Path
from unicodedata import category
from django import apps
from django.conf import settings
from django.template.loader import render_to_string
from django.core.mail import EmailMultiAlternatives

from celery import Task, shared_task
from celery_progress.backend import ProgressRecorder
from backend.celery import app

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from sentry_sdk import capture_exception
from apps.common.utils import get_local_filepath
from apps.core.utils import format_excel, insert_image, get_file_path
from backend.storage_backends import PrivateMediaStorage
import requests
from apps.rfq.models import (
    Category,
    RFQItem,
    RFQItemResponse,
    Rfq,
    RfqInvitee,
    SupplierResponse,
    SupplierRfqTotal,
)
from apps.rfq.reports import RFQFinancialReport, RFQSummaryReport


@shared_task(bind=True)
def create_rfq_items(self, category_id):
    category = Category.objects.filter(id=category_id).first()
    if category is not None:
        if category.self_evaluate:
            """
            Self Evaluate, evaluated by system
            """
            A = PrivateMediaStorage()
            headers = {"ResponseContentDisposition": f"attachment;"}
            file_url = A.url(
                f"{category.items_template}",
                expire=300,
                parameters=headers,
                http_method="GET",
            )
            time = datetime.datetime.now()
            dir_name = Path(
                "media/temp/{}/{}/{}".format(time.year, time.month, time.day)
            )  # folder structure
            dir_name.mkdir(parents=True, exist_ok=True)
            file_name = os.path.basename(f"{category.items_template}")
            filepath = "{}/{}".format(dir_name, file_name)
            r = requests.get(file_url)
            filepath = os.path.basename(f"{category.items_template}")

            with open("{}".format(filepath), "wb") as f:
                f.write(r.content)

            # filepath = get_local_filepath(category.items_template.url)

            workbook = load_workbook(filepath, data_only=True)
            try:
                worksheet = workbook.get_sheet_by_name("RFQ ITEMS")

                for i in range(15, worksheet.max_row + 1):
                    try:
                        if len(worksheet.cell(row=i, column=2).value) > 1:
                            if worksheet.cell(row=14, column=3).value != "UoM":
                                unit_of_measure = "N/A"
                            else:
                                unit_of_measure = worksheet.cell(row=i, column=3).value

                            if worksheet.cell(row=14, column=5).value != "Specification 1":
                                specification_1 = "N/A"
                            else:
                                specification_1 = worksheet.cell(row=i, column=5).value

                            if worksheet.cell(row=14, column=6).value != "Specification 2":
                                specification_2 = "N/A"
                            else:
                                specification_2 = worksheet.cell(row=i, column=6).value
                                
                            if worksheet.cell(row=14, column=4).value != "Qty":
                                qty = 1
                            else:
                                qty = worksheet.cell(row=i, column=4).value
                            data = {
                                "category_id": category.id,
                                "item_description": worksheet.cell(row=i, column=2).value,
                                "unit_of_measure": unit_of_measure,
                                "specification_1": specification_1,
                                "specification_2": specification_2,
                                "item_number": worksheet.cell(row=i, column=1).value,
                                "quantity": qty,
                                "current_price": 0,
                                "price_validity_months": 12,
                            }
                            new_rfq_category = RFQItem(**data)
                            new_rfq_category.save()
                            print("created got here")

                        else:
                            message = "Column 2 must not be empty"
                            return {"result": "error", "error_message": message}

                    except Exception as e:
                        print(e)
                        message = str(e)
                        return {"result": "error", "error_message": message}

                message = "RFQ Created successfully"
                return {
                    "result": "success",
                    "success_message": message,
                    "category": new_rfq_category.id,
                }
            except Exception as e:
                message = "Column 2 must not be empty"
                return {"result": "error", "error_message": message}

        else:
            RFQItem.objects.create(
                category=category,
                item_description="Test Item",
                item_code="N/A",
                unit_of_measure="N/A",
                item_number=1,
                quantity=1,
                current_price=0,
                price_validity_months=12,
            )
    message = "RFQ or Category is None"
    return {"result": "error", "error_message": message}


@shared_task(bind=True)
def generate_supplier_rfq_report():
    pass


class EvaluateRFQ(Task):
    name = "EvaluateRFQ"

    def run(self, *args, **kwargs):
        context = self.calculate_category_scores(kwargs["category_id"])
        return context

    def calculate_rfq_score(self, category_id):
        """
        Calculate supplier RFQ Scores and rank participants
        """
        category = Category.objects.filter(id=category_id).first()
        if category is not None:
            suppliers = category.participants["participants"]
            for supplier in suppliers:
                responses = RFQItemResponse.objects.filter(
                    supplier_id=supplier.id, rfq_item__category_id=category.id
                )
                rfq_score = SupplierRfqTotal.objects.filter(
                    supplier_id=supplier.id, category_id=category.id
                ).first()
                score = 0
                is_outlier = False
                is_blank = False
                has_outlier = False
                has_blank = False

                for response in responses:
                    outlier_score = response.rfq_item.outlier_score
                    # upper_outlier_score = response.rfq_item.upper_outlier_score
                    # value = show(response.total)
                    total = response.total_price
                    if total == 0:
                        is_blank = True
                        if is_blank is True:
                            has_blank = True
                    if float(total) < float(outlier_score):
                        is_outlier = True
                        if is_outlier is True:
                            has_outlier = True

                    score += total

                if rfq_score is None:
                    obj, created = SupplierRfqTotal.objects.update_or_create(
                        supplier_id=supplier.id,
                        category_id=category.id,
                        defaults={
                            "score": score,
                            "has_outlier": has_outlier,
                            "has_blank": has_blank,
                        },
                    )
                else:
                    rfq_score.has_outlier = has_outlier
                    rfq_score.has_blank = has_blank
                    rfq_score.score = score
                    rfq_score.save()

            self.rank_participants(category_id=category.id)
            return True
        pass

    def rank_participants(self, category_id):
        category = Category.objects.filter(id=category_id).first()
        scores = SupplierRfqTotal.objects.filter(
            category_id=category.id, has_outlier=False
        ).order_by("score")
        if scores.count() == 0:
            return True
        first_score = scores.first()
        first_score.rank = 1
        first_score.save()
        prev_score = first_score
        for rfq_score in scores:
            if rfq_score.score == prev_score.score:
                rfq_score.rank = prev_score.rank
            else:
                rfq_score.rank = rank + 1

            if rfq_score.has_outlier:
                rfq_score.rank = 0
            rank += 1
            rfq_score.save()
            prev_score = rfq_score

        return True

    # def refresh_scores(self, category_id):
    #     try:
    #         category = Category.objects.filter(id=category_id).first()
    #         items = RFQItem.objects.filter(category_id=category_id)
    #         for item in items:
    #             item.outlier_score_final = None
    #             item.upper_outlier_score_final = None
    #             item.median_price_final = None
    #             item.save()

    #         suppliers = category.participants["participants"]

    #         lowest_price = 100000000000000000
    #         for supplier in suppliers:
    #             # resolve_rfq_score(supplier_id=supplier.id, rfq=self)
    #             rfq_score = SupplierRfqTotal.objects.filter(supplier_id=supplier.id,category_id=category.id).first()
    #             if rfq_score < lowest_price and rfq_score.has_outlier is False:
    #                 lowest_price = rfq_score.score

    #         self.rank_participants()
    #     except Exception as e:
    #         print(e)


# @shared_task(bind=True)
def invite_suppliers_rfq(category_id):
    category = Category.objects.filter(id=category_id).first()
    for invitee in RfqInvitee.objects.filter(category=category):
        if invitee.supplier is None:
            supplier = "Esteemed Vendor"
        else:
            supplier = invitee.supplier

        email_subject = "Tendersure RFQ Invitation"
        to_email = invitee.email

        # send default from Tendersure
        message = render_to_string(
            "emails/rfq_invitation_email.html",
            {
                "supplier": supplier,
                "category": category,
                "buyer_name": "Tendersure Team",
                "buyer_logo": "tendersure_logo",
            },
        )
        email = EmailMultiAlternatives(email_subject, message, to=[to_email])
        email.attach_alternative(message, "text/html")
        email.send(fail_silently=True)

    context = {"message": "Invitations to participate sent to suppliers."}

    return context


@shared_task(bind=True)
def import_current_supplier_prices(self, file_url):
    messages = []

    filepath = f"{settings.BASE_DIR}{file_url}"
    workbook = load_workbook(filepath, data_only=True)

    worksheets = workbook.sheetnames[0]
    if len(worksheets) < 1:
        messages.append("There are no sheets to upload from the excel.")
        context = {"result": "error", "response_message": messages}
        return context

    items_count = 0

    # progress_recorder = ProgressRecorder(self)
    # result = 0

    # result += 1
    # progress_recorder.set_progress(result, len(worksheets))
    try:

        category_sheet = workbook["ITEM PRICES"]

        for i in range(7, category_sheet.max_row + 1):
            items_count += 1
            item_id = category_sheet["B{}".format(i)].value
            item_number = category_sheet["C{}".format(i)].value
            item_description = category_sheet["D{}".format(i)].value
            current_supplier = category_sheet["E{}".format(i)].value
            quantity = category_sheet["F{}".format(i)].value
            current_price = category_sheet["G{}".format(i)].value

            if RFQItem.objects.filter(
                id=item_id, item_number=item_number, item_description=item_description
            ).exists():
                try:
                    rfq_item = RFQItem.objects.update_or_create(
                        id=item_id,
                        item_number=item_number,
                        item_description=item_description,
                        defaults={
                            "current_price": current_price,
                            "current_supplier": current_supplier,
                            "quantity": quantity,
                        },
                    )
                except Exception as e:
                    # capture_exception(e)
                    print(e)
            else:
                messages.append({"error": "The item does not exists"})

        os.remove(filepath)
        messages.append(
            "Imports succeeded. current suppliers added to {} rfq items.".format(
                items_count
            )
        )
        context = {"result": "success", "response_message": messages}
        return context

    except Exception as e:
        # capture_exception(e)
        print(e)
        messages.append("Failed to import current suppliers template")
        context = {"result": "error", "response_message": messages}
        return context


@shared_task(bind=True)
def download_current_prices_import(self, category_id):
    """
    Download template to import current prices & current suppliers
    """
    time = datetime.datetime.now()
    dir_name = Path(
        "media/current_supplier_prices/%s/%s/%s" % (time.year, time.month, time.day)
    )  # folder structure
    dir_name.mkdir(parents=True, exist_ok=True)
    time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
    current_prices = "RFQ_Current_Prices"
    filepath = "{}/{}_{}.xlsx".format(dir_name, current_prices, time_only)

    workbook = Workbook()
    items_sheet = workbook.active
    items_sheet.title = "ITEM PRICES"
    buyer_logo_url = "static/core/img/Tendersure_Logo.png"
    buyer_logo = Image("static/core/img/Tendersure_Logo.png")
    buyer_logo_height, buyer_logo_width = buyer_logo.height, buyer_logo.width
    max_height = 75
    ratio = buyer_logo_height / max_height
    buyer_logo.height, buyer_logo.width = buyer_logo_height, buyer_logo_width

    buyer_logo_anchor = "B1"

    items_sheet.title = "ITEM PRICES"
    logo_worksheet = items_sheet.title
    items_sheet.append(("", "", ""))
    items_sheet.append(("", "", ""))
    items_sheet.append(("", "", ""))
    items_sheet.append(("", "", ""))
    items_sheet.append(
        (
            "",
            "hc_Item ID",
            "hc_No",
            "h_Item Description",
            "h_Current Supplier",
            "h_Annual Qty",
            "h_Current Price",
            "h_Item Code",
            "h_Price Validity (Months)",
        )
    )
    data = {}
    data["ITEM PRICES_B"] = 8
    data["ITEM PRICES_C"] = 5
    data["ITEM PRICES_D"] = 39
    data["ITEM PRICES_E"] = 20
    data["ITEM PRICES_F"] = 15
    data["ITEM PRICES_G"] = 15
    data["ITEM PRICES_H"] = 15
    data["ITEM PRICES_I"] = 15

    rfq_items = (
        RFQItem.objects.filter(category_id=category_id)
        .order_by("item_number")
        .only("id")
    )

    try:
        result = 0
        progress_recorder = ProgressRecorder(self)
        total_for_progress = rfq_items.count()

        for count, item in enumerate(rfq_items):
            result += 1
            progress_recorder.set_progress(result, total_for_progress)

            items_sheet.append(
                (
                    "",
                    f"Ac_{item.id}",
                    f"Ac_{item.item_number}",
                    item.item_description,
                    "",
                    "",
                ),
            )
        workbook.save(filepath)
        format_excel(filepath, data=data)
        filepath = insert_image(
            filepath, logo_worksheet, buyer_logo_anchor, image_url=buyer_logo_url
        )
        context = {
            "result": "success",
            "response_message": "Template generated successfully",
            # "filepath": filepath.split("/", 1)[1],
            "filepath": filepath,
        }

        return context

    except Exception as e:
        # capture_exception(e)
        print(e)
        context = {
            "result": "error",
            "response_message": "Error generating template",
        }
        return context


@shared_task(bind=True)
def submit_rfq(self, category_id, supplier_id, price_template):
    messages = []
    category = Category.objects.filter(id=category_id).first()
    supplier = apps.apps.get_model("suppliers", "Supplier").objects.get(id=supplier_id)

    supplier_response = SupplierResponse.objects.filter(category_id=category_id, supplier_id=supplier_id).first()

    A = PrivateMediaStorage()
    headers = {"ResponseContentDisposition": f"attachment;"}
    time = datetime.datetime.now()
    file_url = A.url(
        f"{supplier_response.document_url}", expire=300, parameters=headers, http_method="GET"
    )
    dir_name = Path(
        "media/temp/{}/{}/{}".format(time.year, time.month, time.day)
    )  # folder structure
    dir_name.mkdir(parents=True, exist_ok=True)
    file_name = os.path.basename(f"{supplier_response.document_url}")
    filepath = "{}/{}".format(dir_name, file_name)
    r = requests.get(file_url)

    with open("{}".format(filepath), "wb") as f:
        f.write(r.content)

    # filepath = get_local_filepath(category.items_template.url)
    # filepath = get_local_filepath(price_template)
    print(filepath)

    if category.self_evaluate == False:
        try:
            workbook = load_workbook(filepath, data_only=True)
        except Exception as e:
            workbook = None

        if not workbook:
            context = {
                "rfq": category,
                "job": category.rfq,
                "company": category.rfq.company,
            }
            messages.append(
                "Please make sure the document you are uploading is an excel sheet."
            )

        old_response = None
        rfq_item = RFQItem.objects.filter(
            category=category, item_description="Test Item", item_number=1
        ).first()
        # check for old response on the item and update or create a new one.
        old_responses = RFQItemResponse.objects.filter(
            rfq_item=rfq_item, supplier=supplier
        )
        if old_responses.count() > 0:
            old_response = old_responses.first()
            total = 1
            if total is not None:
                old_response.total = hash(total)
            else:
                old_response.total = hash(0)
            old_response.cell_data = "cell data"
            old_response.column_data = "column data"
            # old_response.excel_url = excel_url
            old_response.save()
            updated_items = True
        else:
            new_total = ""
            total = 1
            if total is not None:
                new_total = hash(total)
            else:
                new_total = hash(0)
            RFQItemResponse.objects.create(
                supplier=supplier,
                rfq_item=rfq_item,
                total=new_total,
                cell_data="cell data",
                column_data="column data",
                item_number=1,
                # excel_url=excel_url,
            )
    else:
        workbook = load_workbook(filepath, data_only=True)
        if len(workbook.worksheets) > 2:
            messages.append(
                "Please make sure the document you are uploading has the same format as the downloaded template"
            )
            context = {
                "result": "error",
                "response_message": messages,
            }
            return context

        if not "RFQ ITEMS" in workbook.sheetnames:
            messages.append(
                'Please make sure the document you are uploading has a worksheet with the name "RFQ ITEMS"'
            )
            context = {
                "result": "error",
                "response_message": messages,
            }
            return context

        try:
            worksheet = workbook["RFQ ITEMS"]
            old_response = None

            # capture submissions for the rest of the items
            for i in range(15, worksheet.max_row + 1):
                if (
                    worksheet.cell(row=i, column=2).value != None
                    and worksheet.cell(row=i, column=1).value != "No"
                ):
                    print(worksheet.cell(row=i, column=worksheet.max_column).value)
                    column_data = ""
                    cell_data = ""
                    for k in range(5, worksheet.max_column):
                        column_data += "%s," % str(
                            worksheet.cell(row=14, column=k).value
                        )
                        cell_data += "%s," % str(worksheet.cell(row=i, column=k).value)

                    rfq_item = RFQItem.objects.filter(
                        category=category,
                        item_description=worksheet.cell(row=i, column=2).value,
                        item_number=worksheet.cell(row=i, column=1).value,
                    ).first()
                    # if rfq_item is None:
                    #     raise ValueError("Item mismatch")
                    if rfq_item is not None:
                        total = worksheet.cell(row=i, column=worksheet.max_column).value
                        item_number = worksheet.cell(row=i, column=1).value

                        # check for old response on the item and update or create a new one.
                        old_responses = RFQItemResponse.objects.filter(
                            rfq_item=rfq_item, supplier=supplier
                        )
                        if old_responses.count() > 0:
                            old_response = old_responses.first()
                            if total is not None:
                                old_response.total = hash(total)
                            else:
                                old_response.total = hash(0)
                            old_response.cell_data = cell_data
                            old_response.column_data = column_data
                            old_response.save()
                            updated_items = True
                        else:
                            new_total = ""
                            if total is not None:
                                new_total = hash(total)
                            else:
                                new_total = hash(0)
                            item_response = RFQItemResponse.objects.filter(
                                rfq_item=rfq_item, supplier=supplier
                            ).first()
                            if item_response is not None:
                                item_response.total = new_total
                                item_response.cell_data = cell_data
                                item_response.column_data = column_data
                                item_response.save()
                            else:
                                RFQItemResponse.objects.create(
                                    supplier=supplier,
                                    rfq_item=rfq_item,
                                    total=new_total,
                                    cell_data=cell_data,
                                    column_data=column_data,
                                    item_number=item_number,
                                )
            # create supplier response
            messages.append(
                "RFQ Item Responses created successfuly"
            )
            context = {
                "result": "success",
                "response_message": messages,
            }
            return context
            # supplier.send_financial_responses(rfq.id, type)
        except Exception as e:
            print(e)
            # todo delete that excel from the server
            # "Rfq / 2019 / 08 / 24 / RFQ_Template_for_Ouma_Company_Buyers.xlsx"

            # dir_name = "media/%s" % excel_url.rsplit('/', 1)[0]
            # match_string = excel_url.rsplit('/', 1)[-1]
            # delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

            messages.append(
                "There is a problem with your uploaded Excel. Verify and retry."
            )
            context = {
                "result": "error",
                "response_message": messages,
            }
            return context
        

@shared_task(bind=True)
def import_category_suppliers(self, job_id):
    messages = []

    job = Rfq.objects.filter(id=job_id).first()
    filepath = get_file_path(job.category_suppliers)
    workbook = load_workbook(filepath, data_only=True)
    worksheets = workbook.sheetnames[1:]
    if len(worksheets) < 1:
        messages.append("There are no sheets to upload from the excel.")

        context = {"messages": messages}
        return context

    categories_count = 0
    suppliers = 0

    progress_recorder = ProgressRecorder(self)
    result = 0

    for sheet in worksheets:
        result += 1
        progress_recorder.set_progress(result, len(worksheets))

        sheet_name_list = sheet.split("_")
        category_sheet = workbook[sheet]
        CategoryType = apps.apps.get_model('core', 'CategoryType')
        category_type = CategoryType.objects.filter(innitials=sheet_name_list[-1])

        if category_type.count() == 1:
            categories_count += 1
            category_type = category_type.first()
            for i in range(7, category_sheet.max_row + 1):
                company_name = category_sheet["C{}".format(i)].value
                primary_email = category_sheet["E{}".format(i)].value
                alternative_email = category_sheet["F{}".format(i)].value
                primary_phone = category_sheet["G{}".format(i)].value
                alternative_phone = category_sheet["H{}".format(i)].value
                country = category_sheet["I{}".format(i)].value

                supplier = None
                supplier1 = None
                supplier2 = None

                if primary_email != None:
                    supplier1 = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                        username=primary_email).first()
                elif alternative_email != None:
                    supplier2 = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                        username=alternative_email
                    ).first()

                if supplier1:
                    supplier = supplier1
                elif supplier2:
                    supplier = supplier2

                if supplier is None:
                    suppliers += 1
                    try:
                        apps.apps.get_model('core', 'CategoryTypeSupplier').objects.update_or_create(
                            primary_email=primary_email, alternative_email=alternative_email,
                            category_type=category_type,
                            defaults={
                                "primary_email": primary_email, "alternative_email": alternative_email,
                                "primary_phone": primary_phone, "alternative_phone": alternative_phone,
                                "category_type": category_type, "company_name": company_name,
                                "country": country,
                            },
                        )
                    except Exception as e:
                        capture_exception(e)
                        print(e)
                else:
                    suppliers += 1
                    try:

                        apps.apps.get_model('core', 'CategoryTypeSupplier').objects.update_or_create(
                            supplier=supplier, category_type=category_type,
                            defaults={
                                "supplier": supplier, "primary_email": primary_email,
                                "alternative_email": alternative_email, "primary_phone": primary_phone,
                                "alternative_phone": alternative_phone, "category_type": category_type,
                                "company_name": company_name, "country": country,
                            },
                        )
                    except Exception as e:
                        capture_exception(e)
                        print(e)
    os.remove(filepath)
    messages.append(
        "Imports succeeded. {} suppliers added to {} categories.".format(
            suppliers, categories_count
        )
    )
    context = {"messages": messages}
    return context


# register RFQ TasksS
app.register_task(RFQFinancialReport())
app.register_task(EvaluateRFQ())
app.register_task(RFQSummaryReport())
