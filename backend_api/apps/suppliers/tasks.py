import datetime
import math
from io import BytesIO
from pathlib import Path
from django.utils.translation import gettext_lazy as _
from celery import shared_task
from django import apps
from django.core.exceptions import ValidationError
from django.core.files import File
from django.core.mail import EmailMultiAlternatives
from django.core.validators import validate_email
from django.template.loader import render_to_string
from sentry_sdk import capture_exception
from openpyxl import Workbook, load_workbook
from apps.core.models import CategoryTypeSupplier, CategoryTypeSupplierLocation
from apps.core.utils import Render, delete_matching_files_in_directory, get_file_path, get_local_filepath, insert_image
import copy
from apps.suppliers.models import Supplier
from openpyxl.workbook.protection import WorkbookProtection
from backend.storage_backends import PrivateMediaStorage
import os
import requests
from apps.core.utils import hash


@shared_task(bind=True)
def process_supplier_category_type_location_information(
    self, supplier_id, category_type_ids, locations
):
    supplier = Supplier.objects.filter(id=supplier_id).first()

    # old_suppliers = CategoryTypeSupplier.objects.filter(
    #     Q(Q(primary_email=supplier.email) | Q(alternative_email=supplier.email))
    # )
    # for old_supplier in old_suppliers:
    #     old_supplier.supplier = supplier
    #     old_supplier.save()

    if supplier is not None:
        for category_type_id in category_type_ids:
            category_type_supplier,created = CategoryTypeSupplier.objects.update_or_create(
                supplier_id=supplier_id,
                category_type_id=category_type_id,
                defaults={
                    "company_name": supplier.company_name,
                    "primary_email": supplier.email,
                    "primary_phone": supplier.phone_number,
                    "country": supplier.country,
                },
            )

            for location in locations:
                CategoryTypeSupplierLocation.objects.update_or_create(
                    category_type_supplier_id=category_type_supplier.id,
                    location=location,
                )
    context = {"response_message": "Task Completed"}
    return context


@shared_task()
def send_tender_responses(supplier_id, category_id):
    """
    Sends the reponses submitted for the tender technical part
    :param category_id:
    :return:
    """
    supplier = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
        id=supplier_id
    ).first()
    countries = ["Uganda", "Malawi", "Zambia", "Tanzania", "Mozambique"]
    try:
        validate_email(supplier.email)
        category = apps.apps.get_model('tender', 'Category').objects.filter(
            id=category_id).prefetch_related('tender').first()

        job = category.tender
        company = job.company
        country = company.country
        time = datetime.datetime.now()
        if country not in countries:

            file_path = "media/%s/%s/%s/participations/%s" % (
                company.company_name, job.unique_reference,
                category.unique_reference, supplier.company_name,
            )

            # folder structure
            Path(file_path).mkdir(parents=True, exist_ok=True)
            template_path = "tender/emails/supplier/tender_responses.html"
            date_submitted = time.strftime("%Y-%m-%d %H:%M:%S")
            french_countries = ["Cameroon"]
            portuguese_countries = ["Mozambique"]

            context = {
                "category": category,
                "supplier": supplier,
                time: datetime.datetime.now(),
                "date_submitted": date_submitted,
            }

            pdf_file_path = Render.render(path=template_path, params=context)

            email_subject = _(
                "Tendersure: Your participation responses %s "
                % category.name
            )

            if country in french_countries:
                body = render_to_string(
                    "tender/emails/supplier/french_tender_responses.html",
                    {"category": category, "supplier": supplier},
                )
                message = render_to_string(
                    "tender/emails/supplier/french_tender_responses_body.html",
                    {"supplier": supplier, "category": category},
                )
            elif country in portuguese_countries:
                body = render_to_string(
                    "tender/emails/supplier/portuguese_tender_responses.html",
                    {"category": category, "supplier": supplier},
                )
                message = render_to_string(
                    "supplier_portal/emails/countries/portuguese_prequal_responses_body.html",
                    {"supplier": supplier, "category": category},
                )
            else:
                body = render_to_string(
                    "tender/emails/supplier/tender_responses_body.html",
                    {"category": category, "supplier": supplier},
                )

                message = render_to_string(
                    "tender/emails/supplier/tender_responses_body.html",
                    {"supplier": supplier, "category": category},
                )

            pdf_file = BytesIO(pdf_file_path.content)
            category_report, created = apps.apps.get_model('tender', 'SupplierPDFResponse').objects.update_or_create(
                category_id=category_id,
                supplier_id=supplier_id,
                defaults={
                    "document_url": File(pdf_file, "SupplierResponses.pdf")
                }
            )

            email = EmailMultiAlternatives(
                subject=email_subject, body=body, to=[supplier.email]
            )
            email.attach_alternative(message, "text/html")
            try:
                email.attach_file(category_report.document_url.url)
            except:
                pass
            email.send(fail_silently=True)

            context = {
                "response_message": "Email sent successfully",
                "filepath": category_report.document_url.url
            }
            return context
        else:
            # send with  personalized phone and email
            help_contacts = apps.apps.get_model('core', 'HelpContact').objects.get(country=country)
            file_path = "media/%s/%s/%s/participations/%s" % (
                company.company_name, job.unique_reference, category.unique_reference,
                supplier.company_name,
            )
            filename = "%s_%d_%d.pdf" % (
                category.unique_reference.replace(" ", "_"),
                time.year, time.month,
            )

            # folder structure
            Path(file_path).mkdir(parents=True, exist_ok=True)
            date_submitted = time.strftime("%Y-%m-%d %H:%M:%S")
            template_path = "tender/emails/supplier/tender_responses.html"
            context = {
                "category": category, "supplier": supplier,
                time: datetime.datetime.now(), "date_submitted": date_submitted,
            }

            pdf_file_path = Render.render(path=template_path, params=context)

            email_subject = _(
                "Tendersure: Your participation responses %s "
                % category.name
            )

            body = render_to_string(
                "tender/emails/supplier/custom/tender_responses_body.html",
                {
                    "category": category,
                    "supplier": supplier,
                    "phone": help_contacts.contact_phone,
                    "help_email": help_contacts.helpemail,
                },
            )

            message = render_to_string(
                "tender/emails/supplier/custom/tender_responses_body.html",
                {
                    "supplier": supplier,
                    "category": category,
                    "phone": help_contacts.contact_phone,
                    "help_email": help_contacts.helpemail,
                },
            )
            email = EmailMultiAlternatives(
                subject=email_subject, body=body, to=[supplier.email]
            )
            email.attach_alternative(message, "text/html")
            # email.attach_file(pdf_file_path)

            pdf_file = BytesIO(pdf_file_path.content)
            category_report, created = apps.apps.get_model('tender', 'SupplierPDFResponse').objects.update_or_create(
                category_id=category_id,
                supplier_id=supplier_id,
                defaults={
                    "document_url": File(pdf_file, "SupplierResponses.pdf")
                }
            )

            try:
                email.attach_file(category_report.document_url.url)
            except:
                pass

            email.send(fail_silently=True)

            context = {
                "response_message": "Email sent successfully",
                "filepath": category_report.document_url.url
            }
            return context
    except ValidationError as e:
        capture_exception(e)
        print(e)
        return {"error": "Invalid email address"}


def save_prequal_financial_ratio_responses(instance_id):
    """
    debt/equity = long term loans/equity
    current ratio = current assets/current liabilities
    cash ratio = cash/current liabilities
    GP margin = GP/turnover
    NP margin = NP/Turnover
    """
    ratio_instance = apps.apps.get_model(
        'prequal', 'FinancialRatio').objects.filter(id=instance_id).first()

    # section = ratio_instance.section

    all_questions = apps.apps.get_model(
        'prequal', 'Question').objects.filter(section_id=ratio_instance.section_id).order_by("id")
    first_question = all_questions.first()
    other_questions = all_questions.exclude(id=first_question.id)
    supplier = ratio_instance.supplier

    # section_score = 0

    try:
        debt_equity_ratio = float(ratio_instance.debtors) / float(ratio_instance.equity)
    except:
        debt_equity_ratio = 0

    try:
        current_ratio = (float(ratio_instance.current_assets)) / float(ratio_instance.curr_liabilities)
    except:
        current_ratio = 0

    try:
        cash_ratio = float(ratio_instance.cash) / float(ratio_instance.curr_liabilities)
    except:
        cash_ratio = 0

    try:
        gp_margin = (
            float(ratio_instance.gross_profit) / float(ratio_instance.turnover)) * 100
    except:
        gp_margin = 0

    try:
        np_margin = (
            float(ratio_instance.net_profit) / float(ratio_instance.turnover)) * 100
    except:
        np_margin = 0

    c = 0
    ratio_list = [debt_equity_ratio, current_ratio, cash_ratio, gp_margin, np_margin]

    for question in other_questions:
        ratio = ratio_list[c]
        r = apps.apps.get_model('prequal', 'SupplierResponse').objects.filter(
            question_id=question.id, supplier_id=supplier.id
        ).first()
        if ratio is not None:
            if r.first() is not None:
                r.update(options=ratio)
            else:
                apps.apps.get_model('prequal', 'SupplierResponse').objects.create(
                    supplier_id=supplier.id, question_id=question.id,
                    options=ratio
                )
        else:
            apps.apps.get_model('prequal', 'SupplierResponse').objects.update_or_create(
                supplier_id=supplier.id, question_id=question.id,
            )
        c += 1
    context = {
        'response_message': 'response stored'
    }
    return context


@shared_task(bind=True)
def save_tender_financial_ratio_responses(self, instance_id):
    """
    debt/equity = long term loans/equity
    current ratio = current assets/current liabilities (current assets+ cash)/current liabilities
    cash ratio = cash/current liabilities
    GP margin = GP/turnover
    NP margin = NP/Turnover
    """
    ratio_instance = apps.apps.get_model(
        'tender', 'FinancialRatio'
    ).objects.filter(id=instance_id).first()

    # section = ratio_instance.section

    all_questions = apps.apps.get_model(
        'tender', 'Question'
    ).objects.filter(section_id=ratio_instance.section_id).order_by("id")

    first_question = all_questions.first()
    other_questions = all_questions.exclude(id=first_question.id)
    supplier = ratio_instance.supplier

    # section_score = 0

    try:
        debt_equity_ratio = float(ratio_instance.debtors) / float(ratio_instance.equity)
    except:
        debt_equity_ratio = 0

    try:
        current_ratio = float(ratio_instance.current_assets) / float(ratio_instance.curr_liabilities)
    except:
        current_ratio = 0

    try:
        cash_ratio = float(ratio_instance.cash) / float(ratio_instance.curr_liabilities)
    except:
        cash_ratio = 0

    try:
        gp_margin = (
            float(ratio_instance.gross_profit) / float(ratio_instance.turnover)) * 100
    except:
        gp_margin = 0

    try:
        np_margin = (
            float(ratio_instance.net_profit) / float(ratio_instance.turnover)) * 100
    except:
        np_margin = 0

    c = 0
    ratio_list = [debt_equity_ratio, current_ratio, cash_ratio, gp_margin, np_margin]

    for question in other_questions:
        ratio = ratio_list[c]
        # question_options = question.options
        r = apps.apps.get_model('tender', 'SupplierResponse').objects.filter(
            question_id=question.id, supplier_id=supplier.id
        )
        if ratio is not None:
            if r.first() is not None:
                r.update(options=ratio)
            else:
                apps.apps.get_model('tender', 'SupplierResponse').objects.create(
                    supplier_id=supplier.id, question_id=question.id,
                    options=ratio
                )
        else:
            apps.apps.get_model('tender', 'SupplierResponse').objects.update_or_create(
                supplier_id=supplier.id, question_id=question.id,
            )
        c += 1
    context = {
        'response_message': 'response stored'
    }
    return context


@shared_task(bind=True)
def create_rfq_items_template(supplier_id, category_id):
    supplier = apps.apps.get_model("suppliers", "Supplier").objects.filter(id=supplier_id).first()
    category = apps.apps.get_model("rfq", "Category").objects.filter(id=category_id).first()

    time = datetime.datetime.now()
    dir_name = Path("media/temp/{}/{}".format(time.year, time.month))  
    dir_name.mkdir(parents=True, exist_ok=True)

    time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
    filepath = "{}/{}_{}_{}_ITEMS_TEMPLATE.xlsx".format(
        dir_name, supplier.company_name.replace(" ", "_").replace("/", "_"), category.unique_reference, time_only)

    match_string = "{}".format(
        supplier.company_name.replace(" ", "_").replace("/", "_"))
    delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

    # system_workbook = load_workbook(get_file_path(rfq.excel_url))
    system_workbook = load_workbook(get_file_path(category.items_template))
    system_worksheet = system_workbook["RFQ ITEMS"]

    supplier_workbook = Workbook()
    supplier_worksheet = supplier_workbook.active
    supplier_worksheet.title = "RFQ ITEMS"

    for row in system_worksheet:
        for cell in row:
            quote_cell = supplier_worksheet[cell.coordinate]
            quote_cell.value = cell.value
            if cell.has_style:
                quote_cell.font = copy.copy(cell.font)
                quote_cell.border = copy.copy(cell.border)
                quote_cell.fill = copy.copy(cell.fill)
                quote_cell.number_format = copy.copy(cell.number_format)
                quote_cell.protection = copy.copy(cell.protection)
                quote_cell.alignment = copy.copy(cell.alignment)

    for colLetter, colDimension in system_worksheet.column_dimensions.items():
        if colDimension.hidden == True:
            supplier_worksheet.column_dimensions[colLetter].hidden = True

    for idx, rd in system_worksheet.column_dimensions.items():
        supplier_worksheet.column_dimensions[idx] = copy.copy(rd)

    supplier_worksheet['C6'] = f"{supplier.company_name.title()}"
    supplier_worksheet['C7'] = f"{supplier.email_address}"
    supplier_worksheet['C8'] = f"{supplier.address}"
    supplier_worksheet['C9'] = f"{supplier.phone_number}"
    supplier_worksheet['C10'] = f"{supplier.contact_name.title()}"
    supplier_worksheet['C11'] = f"{datetime.datetime.now().date()}"

    supplier_worksheet.merge_cells("A5:B5")
    supplier_worksheet.merge_cells("A6:B6")
    supplier_worksheet.merge_cells("A7:B7")
    supplier_worksheet.merge_cells("A8:B8")
    supplier_worksheet.merge_cells("A9:B9")
    supplier_worksheet.merge_cells("A10:B10")
    supplier_worksheet.merge_cells("A11:B11")

    supplier_worksheet.merge_cells("C5:D5")
    supplier_worksheet.merge_cells("C6:D6")
    supplier_worksheet.merge_cells("C7:D7")
    supplier_worksheet.merge_cells("C8:D8")
    supplier_worksheet.merge_cells("C9:D9")
    supplier_worksheet.merge_cells("C10:D10")
    supplier_worksheet.merge_cells("C11:D11")

    supplier_worksheet.protection.password = 'Temp7Pass'
    supplier_worksheet.protection.selectUnlockedCells = False
    supplier_worksheet.protection.selectLockedCells = True
    supplier_worksheet.protection.sheet = True

    supplier_workbook.security = WorkbookProtection(
        workbookPassword='Temp7Pass', revisionsPassword='Temp7Pass', lockWindows=True,
        lockStructure=True, lockRevision=True)
    supplier_workbook.save(filepath)

    buyer_logo_anchor = "B1"

    if supplier.logo_url is not None and supplier.logo_url != "":
        logo_url = get_file_path(supplier.logo_url)
    else:
        logo_url = "static/core/img/no-company-image128-128.png"

    insert_image(
        excel_url=filepath, worksheet_name="RFQ ITEMS", anchor=buyer_logo_anchor,
        image_url=logo_url)

    context = {
        "response_message": "Template generated successfully",
        "filepath": filepath
    }
    return context
    # except Exception as e:
    #     capture_exception(e)
    #     context = {
    #         "response_message": "An error occurred. Please try again!",
    #         "errors": f"{e}"
    #     }
    #     return context


@shared_task()
def create_tender_items_template(supplier_id, category_id):
    # try:
    supplier = apps.apps.get_model("suppliers", "Supplier").objects.filter(id=supplier_id).first()
    category = apps.apps.get_model("tender", "Category").objects.filter(id=category_id).first()

    time = datetime.datetime.now()
    dir_name = Path("media/temp/{}/{}".format(time.year, time.month))  
    dir_name.mkdir(parents=True, exist_ok=True)

    time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
    filepath = "{}/{}_{}_{}_ITEMS_TEMPLATE.xlsx".format(
        dir_name, supplier.company_name.replace(" ", "_").replace("/", "_"), category.unique_reference, time_only)

    match_string = "{}".format(
        supplier.company_name.replace(" ", "_").replace("/", "_"))
    delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

    # system_workbook = load_workbook(get_file_path(rfq.excel_url))
    system_workbook = load_workbook(get_file_path(category.items_template))
    system_worksheet = system_workbook["RFQ ITEMS"]

    supplier_workbook = Workbook()
    supplier_worksheet = supplier_workbook.active
    supplier_worksheet.title = "RFQ ITEMS"

    for row in system_worksheet:
        for cell in row:
            quote_cell = supplier_worksheet[cell.coordinate]
            quote_cell.value = cell.value
            if cell.has_style:
                quote_cell.font = copy.copy(cell.font)
                quote_cell.border = copy.copy(cell.border)
                quote_cell.fill = copy.copy(cell.fill)
                quote_cell.number_format = copy.copy(cell.number_format)
                quote_cell.protection = copy.copy(cell.protection)
                quote_cell.alignment = copy.copy(cell.alignment)

    for colLetter, colDimension in system_worksheet.column_dimensions.items():
        if colDimension.hidden == True:
            supplier_worksheet.column_dimensions[colLetter].hidden = True

    for idx, rd in system_worksheet.column_dimensions.items():
        supplier_worksheet.column_dimensions[idx] = copy.copy(rd)

    supplier_worksheet['C6'] = f"{supplier.company_name.title()}"
    supplier_worksheet['C7'] = f"{supplier.email_address}"
    supplier_worksheet['C8'] = f"{supplier.address}"
    supplier_worksheet['C9'] = f"{supplier.phone_number}"
    supplier_worksheet['C10'] = f"{supplier.contact_name.title()}"
    supplier_worksheet['C11'] = f"{datetime.datetime.now().date()}"

    supplier_worksheet.merge_cells("A5:B5")
    supplier_worksheet.merge_cells("A6:B6")
    supplier_worksheet.merge_cells("A7:B7")
    supplier_worksheet.merge_cells("A8:B8")
    supplier_worksheet.merge_cells("A9:B9")
    supplier_worksheet.merge_cells("A10:B10")
    supplier_worksheet.merge_cells("A11:B11")

    supplier_worksheet.merge_cells("C5:D5")
    supplier_worksheet.merge_cells("C6:D6")
    supplier_worksheet.merge_cells("C7:D7")
    supplier_worksheet.merge_cells("C8:D8")
    supplier_worksheet.merge_cells("C9:D9")
    supplier_worksheet.merge_cells("C10:D10")
    supplier_worksheet.merge_cells("C11:D11")

    supplier_worksheet.protection.password = 'Temp7Pass'
    supplier_worksheet.protection.selectUnlockedCells = False
    supplier_worksheet.protection.selectLockedCells = True
    supplier_worksheet.protection.sheet = True

    supplier_workbook.security = WorkbookProtection(
        workbookPassword='Temp7Pass', revisionsPassword='Temp7Pass', lockWindows=True,
        lockStructure=True, lockRevision=True)
    supplier_workbook.save(filepath)

    buyer_logo_anchor = "B1"

    if supplier.logo_url is not None and supplier.logo_url != "":
        logo_url = get_file_path(supplier.logo_url)
    else:
        logo_url = "static/core/img/no-company-image128-128.png"

    insert_image(
        excel_url=filepath, worksheet_name="RFQ ITEMS", anchor=buyer_logo_anchor,
        image_url=logo_url)

    context = {
        "response_message": "Template generated successfully",
        "filepath": filepath
    }
    return context
    # except Exception as e:
    #     capture_exception(e)
    #     context = {
    #         "response_message": "An error occurred. Please try again!",
    #         "errors": f"{e}"
    #     }
    #     return context


def submit_tender_financial_responses(financial_response):
    category = financial_response.category
    A = PrivateMediaStorage()
    headers = {"ResponseContentDisposition": f"attachment;"}
    time = datetime.datetime.now()
    file_url = A.url(
        f"{financial_response.excel_url}", expire=300, parameters=headers, http_method="GET"
    )
    dir_name = Path(
        "media/temp/{}/{}/{}".format(time.year, time.month, time.day)
    )  # folder structure
    dir_name.mkdir(parents=True, exist_ok=True)
    file_name = os.path.basename(f"{financial_response.excel_url}")
    filepath = "{}/{}".format(dir_name, file_name)
    r = requests.get(file_url)
    with open("{}".format(filepath), "wb") as f:
        f.write(r.content)

    try:
        workbook = load_workbook(filepath, data_only=True)
    except Exception as e:
        workbook = None

    if not workbook:
        context = {
            "response_message": "Please make sure the document you are uploading is an excel sheet."
        }
        return context

    if category.self_evaluate == False:
        updated_items = False
        old_response = None
        item = apps.apps.get_model("tender", "Item").objects.filter(
            category_id=category.id, item_description="Test Item", item_number=1).first()

        old_responses = apps.apps.get_model("tender", "ItemResponse").objects.filter(
            item_id=item.id, supplier_id=financial_response.supplier_id)

        if old_responses.count() > 0:
            old_response = old_responses.first()
            total = 1
            if total is not None:
                old_response.total = hash(total)
            else:
                old_response.total = hash(0)
            old_response.cell_data = "cell data"
            old_response.column_data = "column data"
            old_response.save()
        else:
            new_total = ""
            total = 1
            if total is not None:
                new_total = hash(total)
            else:
                new_total = hash(0)
            apps.apps.get_model("tender", "ItemResponse").objects.create(
                supplier_id=financial_response.supplier_id, item_id=item.id,
                total=new_total, cell_data="cell data", column_data="column data",
                item_number=1,
            )
        context = {
            "response_message": "Your financial evaluation for the tender updated successfully"
        }
        # send responses through mail
        send_financial_responses.delay(category.id, financial_response.supplier_id)
        return context
    else:
        if not "RFQ ITEMS" in workbook.sheetnames:
            context = {
                "response_message": "Please make sure the document you are uploading has a worksheet with the name RFQ ITEMS"
            }
            return context
        try:
            worksheet = workbook["RFQ ITEMS"]
            old_response = None

            for i in range(15, worksheet.max_row + 1):
                if worksheet.cell(row=i, column=2).value != None and worksheet.cell(row=i, column=1).value != "No":
                    print(worksheet.cell(row=i, column=2).value)
                    column_data = ""
                    cell_data = ""
                    for k in range(5, worksheet.max_column):
                        column_data += "%s," % str(
                            worksheet.cell(row=14, column=k).value
                        )
                        cell_data += "%s," % str(
                            worksheet.cell(row=i, column=k).value
                        )

                    item = apps.apps.get_model("tender", "Item").objects.filter(
                        category_id=category.id, description=worksheet.cell(row=i, column=2).value,
                        number=worksheet.cell(row=i, column=1).value).first()
                    
                    if item is not None:
                        total = worksheet.cell(
                            row=i, column=worksheet.max_column
                        ).value
                        item_number = worksheet.cell(row=i, column=1).value

                        old_responses = apps.apps.get_model("tender", "ItemResponse").objects.filter(
                            item_id=item.id, supplier_id=financial_response.supplier_id
                        )
                        if old_responses.count() > 0:
                            old_response = old_responses.first()
                            if total is not None:
                                old_response.total = hash(total)
                            else:
                                old_response.total = hash(0)
                            old_response.cell_data = cell_data
                            old_response.column_data = column_data
                            old_response.save()
                        else:
                            new_total = ""
                            if total is not None:
                                new_total = hash(total)
                            else:
                                new_total = hash(0)
                            apps.apps.get_model("tender", "ItemResponse").objects.create(
                                supplier_id=financial_response.supplier_id, item_id=item.id,
                                total=new_total, cell_data=cell_data, column_data=column_data,
                                item_number=item_number
                            )
            send_financial_responses.delay(category.id, financial_response.supplier_id)
            context = {
                "response_message": "Your financial evaluation for the tender was submitted successfully."
            }
            return context
        except Exception as e:
            print(f"{e}")
            responses_to_delete = apps.apps.get_model("tender", "ItemResponse").objects.filter(
                supplier_id=financial_response.supplier_id, item__category_id=category.id
            )
            for response in responses_to_delete:
                response.delete()

            context = {
                "response_message": "There is a problem with your uploaded Excel. Verify and retry."
            }
            return context


@shared_task()
def send_financial_responses(category_id, supplier_id):
    """
        Sends the reponses submitted for the financial evaluation
    """
    try:
        job_type = "Tender"
        category = apps.apps.get_model("tender", "Category").objects.filter(id=category_id).first()
        supplier = apps.apps.get_model("suppliers", "Supplier").objects.filter(id=supplier_id).first()
        validate_email(supplier.email_address)

        company = category.tender.company
        email_subject = (
            "Tendersure: Your financial bid for %s " % category.name)

        email_config = apps.apps.get_model("core", "EmailConfiguration").objects.filter(company=company).first()
        if email_config is not None:
            message = render_to_string(
                "suppliers/emails/rfq_responses_body.html",
                {
                    "supplier": supplier, "category": category, "buyer_name": f"{company.company_name} Team",
                    "buyer_logo": "company_logo", "company": company, "type": job_type
                })
            email_config.send_email(email_subject, message, bcc=[supplier.email_address])
        else:
            body = render_to_string(
                "suppliers/emails/rfq_responses_body.html",
                {
                    "category": category, "supplier": supplier,
                    "buyer_name": "Tendersure Team", "buyer_logo": "tendersure_logo", "type": job_type
                })

            message = render_to_string(
                "suppliers/emails/rfq_responses_body.html",
                {"supplier": supplier, "category": category, "type": job_type},
            )

            email = EmailMultiAlternatives(
                subject=email_subject, body=body, to=[supplier.email_address]
            )
            email.attach_alternative(message, "text/html")
            email.send()

    except ValidationError as e:
        return {"error": "Invalid email address"}
    return {"success": "Participation email sent"}