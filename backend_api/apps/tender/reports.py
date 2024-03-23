import datetime
import os
import string
from copy import copy
from pathlib import Path
from celery import Task, shared_task
from celery_progress.backend import ProgressRecorder
from django.db.models import Q
from django.core.files import File
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Side
from sentry_sdk import capture_exception
from .models import *
from apps.core.utils import (
    format_excel, insert_image, delete_matching_files_in_directory, get_file_path, show, Render
    )
from .utils import (
    c_suppliers, get_locations, get_supplier_data, get_cat_supplier_locations, ratio_scores_before_after_qa,
    variance_ratios, ratios_before_after_qa
)
from ..core.models import CategoryTypeSupplier
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


class TenderSummaryReport(Task):
    # ignore_result = True
    name = "TenderSummaryReport"
    progress_recorder = None
    result = 0
    format_data = {
        "TENDER SUMMARY REPORT_A": 5, "TENDER SUMMARY REPORT_B": 5, "TENDER SUMMARY REPORT_C": 31,
        "TENDER SUMMARY REPORT_D": 8, "TENDER SUMMARY REPORT_E": 13, "TENDER SUMMARY REPORT_F": 21,
        "TENDER SUMMARY REPORT_G": 11, "TENDER SUMMARY REPORT_H": 12, "TENDER SUMMARY REPORT_I": 21,
        "TENDER SUMMARY REPORT_J": 11, "TENDER SUMMARY REPORT_K": 12, "TENDER SUMMARY REPORT_L": 21,
        "TENDER SUMMARY REPORT_M": 11, "TENDER SUMMARY REPORT_N": 12, "TENDER SUMMARY REPORT_O": 21,
        "TENDER SUMMARY REPORT_P": 11, "TENDER SUMMARY REPORT_Q": 12, "TENDER SUMMARY REPORT_R": 21,
        "TENDER SUMMARY REPORT_S": 11, "TENDER SUMMARY REPORT_T": 12, "TECHNICAL_REPORT_A": 5,
        "TECHNICAL_REPORT_B": 5, "TECHNICAL_REPORT_C": 31, "TECHNICAL_REPORT_D": 8,
        "TECHNICAL_REPORT_E": 11, "TECHNICAL_REPORT_F": 11, "TECHNICAL_REPORT_G": 11,
        "TECHNICAL_REPORT_H": 11, "TECHNICAL_REPORT_I": 11, "TECHNICAL_REPORT_J": 11,
        "TECHNICAL_REPORT_K": 11, "TECHNICAL_REPORT_L": 11, "TECHNICAL_REPORT_M": 11,
        "TECHNICAL_REPORT_OTHER": 21, "SAVINGS REPORT_A": 5, "SAVINGS REPORT_B": 5, "SAVINGS REPORT_C": 31,
        "SAVINGS REPORT_D": 8, "SAVINGS REPORT_E": 13, "SAVINGS REPORT_F": 17, "SAVINGS REPORT_G": 17,
        "SAVINGS REPORT_H": 17, "SAVINGS REPORT_I": 19, "SAVINGS REPORT_J": 17, "SAVINGS REPORT_K": 17,
        "LOWEST ITEM COST REPORT_A": 5, "LOWEST ITEM COST REPORT_B": 5, "LOWEST ITEM COST REPORT_C": 31,
        "LOWEST ITEM COST REPORT_D": 8, "LOWEST ITEM COST REPORT_E": 13, "LOWEST ITEM COST REPORT_F": 13,
        "LOWEST ITEM COST REPORT_G": 14, "LOWEST ITEM COST REPORT_H": 24, "LOWEST ITEM COST REPORT_I": 14,
        "LOWEST ITEM COST REPORT_J": 24, "LOWEST ITEM COST REPORT_K": 14, "LOWEST ITEM COST REPORT_L": 24,
        "LOWEST ITEM COST REPORT_OTHER": 15, "LOWEST SUPPLIER REPORT_A": 5, "LOWEST SUPPLIER REPORT_B": 5,
        "LOWEST SUPPLIER REPORT_C": 31, "LOWEST SUPPLIER REPORT_D": 8, "LOWEST SUPPLIER REPORT_E": 13,
        "LOWEST SUPPLIER REPORT_F": 13, "LOWEST SUPPLIER REPORT_G": 24, "LOWEST SUPPLIER REPORT_H": 24,
        "LOWEST SUPPLIER REPORT_I": 24, "LOWEST SUPPLIER REPORT_J": 24, "LOWEST SUPPLIER REPORT_K": 24,
        "LOWEST SUPPLIER REPORT_OTHER": 24,
    }

    def run(self, *args, **kwargs):
        self.progress_recorder = ProgressRecorder(self)
        context = self.report(kwargs["category_id"])
        print(context)
        return context

    def report(self, category_id):

        category = Category.objects.filter(id=category_id).first()
        tender = category.tender
        company = tender.company

        time = datetime.datetime.now()
        dir_name = Path("media/tender_reports/{}/{}".format(time.year, time.month))
        dir_name.mkdir(parents=True, exist_ok=True)

        # time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
        filepath = "{}/{}_MultiItem_Tender_Summary_Report.xlsx".format(
            dir_name, tender.id
        )

        # delete similar reports run this month
        match_string = f"{tender.id}_{company.company_name.replace(' ', '_')}"
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

        qed_logo = Image("static/core/img/Tendersure_Logo.png")

        if company.company_logo_url is not None and company.company_logo_url != "":
            buyer_logo = get_file_path(company.company_logo_url)
        else:
            buyer_logo = "static/core/img/no-company-image128-128.png"

        qed_logo_anchor = "L1"
        buyer_logo_anchor = "B1"

        workbook = Workbook()

        items = Item.objects.filter(category_id=category_id)

        # technical_participants = category.technical_participants
        # sorted_participants = prequal.sorted_participants(participants)
        #
        # rfq_participants = rfq.participants

        # try to get the rfq report
        workbook.create_sheet("TENDER SUMMARY REPORT")
        rfq_summary = workbook["TENDER SUMMARY REPORT"]

        letters = list(string.ascii_uppercase)
        num_cols = 200
        excel_cols = []
        for i in range(0, num_cols - 1):
            n = i // 26
            m = n // 26
            i -= n * 26
            n -= m * 26
            col = (
                letters[m - 1] + letters[n - 1] + letters[i]
                if m > 0
                else letters[n - 1] + letters[i]
                if n > 0
                else letters[i]
            )
            excel_cols.append(col)

        category_order = apps.apps.get_model('core', 'CategoryOrder')
        category_total_bidders = category_order.objects.filter(
            category_id=category_id, payment_status=category_order.PAID, target__model='tender'
        ).count()

        category_prequal_bidders = category.qualified_bidders
        category_responsive_bidders = category.responsive_bidders

        lowest_cumm_suppliers_ranked = SupplierFinancialTotal.objects.filter(
            category_id=category.id, has_blank=False, has_outlier=False
        ).order_by("score")
        lowest_cumm_suppliers_opp = SupplierFinancialTotal.objects.filter(
            Q(category_id=category.id), ~Q(has_blank=False) | ~Q(has_outlier=False)
        ).order_by("score")

        lowest_cumm_suppliers = list(lowest_cumm_suppliers_ranked) + list(
            lowest_cumm_suppliers_opp
        )

        sections = Section.objects.filter(
            category_id=category.id, parent_section_id__isnull=True
        )
        questions = Question.objects.filter(section__category=category, is_qa=True)

        total_for_progress = 6

        try:
            percentage_responses = "{}%".format(
                round(((category_responsive_bidders / category_total_bidders) * 100))
            )
        except Exception as e:
            percentage_responses = 0

        data = {
            "category_responsive_bidders": category_responsive_bidders,
            "category_total_bidders": int(category_total_bidders),
            "category": category,
            "category_prequal_bidders": category_prequal_bidders,
            "tender": tender,
            "items": items,
            "lowest_cumm_suppliers": lowest_cumm_suppliers,
            "total_for_progress": total_for_progress,
            "excel_cols": excel_cols,
            "sorted_participants": category.sorted_technical_participants,
            "percentage_responses": percentage_responses,
            "participants": lowest_cumm_suppliers,
            "sections": sections,
            "questions": questions,
            "lowest_cumm_suppliers_ranked": lowest_cumm_suppliers_ranked,
            "lowest_cumm_suppliers_opp": lowest_cumm_suppliers_opp,
        }

        self.create_summary_worksheet(rfq_summary=rfq_summary, data=data)
        self.create_technical_worksheet(workbook=workbook, data=data)
        self.create_savings_worksheet(workbook=workbook, data=data)
        self.create_lowest_cost_report_worksheet(workbook=workbook, data=data)
        self.create_lowest_supplier_worksheet(workbook=workbook, data=data)
        # individual_sheets = self.create_individual_worksheets(
        #     workbook=workbook, data=data
        # )
        # print(individual_sheets)

        workbook.remove(workbook["Sheet"])
        workbook.save(filepath)

        # f_data.update(individual_sheets)

        format_excel(filepath, data=self.format_data)
        insert_image(
            excel_url=filepath, worksheet_name="TENDER SUMMARY REPORT", anchor=qed_logo_anchor)
        insert_image(
            excel_url=filepath, worksheet_name="TENDER SUMMARY REPORT", anchor=buyer_logo_anchor,
            image_url=buyer_logo)

        try:
            with open(filepath, "rb") as l:
                storage = PrivateMediaStorage()
                url = f"tender/category/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
                storage.save(url, l)

                category_report, created = CategoryReport.objects.update_or_create(
                    category_id=category_id, defaults={"multi_item_tender_summary_report": url}
                )
                context = {
                    "response_message": "Report generated successfully",
                    "filepath": category_report.multi_item_tender_summary_report.url,
                }
                return context
        except Exception as e:
            capture_exception(e)
            print(e)
            context = {
                "error": f"{e}"
            }
            return context

    def create_summary_worksheet(self, rfq_summary, data):
        try:
            category_total_bidders = data["category_total_bidders"]
            category = data["category"]
            tender = data["tender"]
            items = data["items"]
            lowest_cumm_suppliers = data["lowest_cumm_suppliers"]
            total_for_progress = data["total_for_progress"]
            excel_cols = data["excel_cols"]
            percentage_responses = data["percentage_responses"]

            rfq_current_cost = 0
            rfq_eprocure_cost = 0
            rfq_total_potential_savngs = 0

            spaces = [
                ("", "", "", ""), ("", "", "", ""),
                ("", "", "", ""), ("", "", "", ""),
                ("", "", "", ""),
            ]

            for row in spaces:
                rfq_summary.append(row)

            details_summary = [
                ("", "", "", ""),
                ("", f"Bnb_Category: {category.name}", "", ""),
                ("", "", ""),
                ("", "Number of bidders", "", f"Ac_{category_total_bidders}"),
                ("", "Responsive bidders", "", f"Ac_{category.responsive_bidders.count()}"),
                ("", "Percentage responses", "", f"Ac_{percentage_responses}"),
                ("", "Currency", "", "Ac_{}".format(category.currency)),
                (
                    "", "Weighting (Technical:Financial)",
                    "", f"Ac_{int(float(category.technical_weight))}:{int(float(category.financial_weight))}",
                ),
            ]

            for row in details_summary:
                rfq_summary.append(row)

            empty_row = [""]
            empty_row = tuple(empty_row)

            # List of items and submissions
            rfq_summary.append(empty_row)
            rfq_summary.merge_cells("B7:H7")
            rfq_summary.merge_cells("B9:C9")
            rfq_summary.merge_cells("B10:C10")
            rfq_summary.merge_cells("B11:C11")
            rfq_summary.merge_cells("B12:C12")
            rfq_summary.merge_cells("B13:C13")

            supplier_headers = [
                "", "", "", "", "", "Hmr_2_Best Supplier 1", "", "",
                "Hmr_2_Best Supplier 2", "", "", "Hmr_2_Best Supplier 3",
                "", "", "Hmr_2_Best Supplier 4", "", "", "Hmr_2_Best Supplier 5",
                "", "",
            ]
            items_header = [
                "", "hc_#", "hl_Item Description", "hc_Qty", "hc_Specification",
            ]

            for i in range(0, 5):
                items_header.extend(["hl_Name", "hc_TW Score", "hc_Cost KES"])

            no_of_items = len(items)
            no_of_rows_at_the_top = 16
            supplier_totals_row = no_of_rows_at_the_top + no_of_items
            suppliers_totals_formula = []

            c = 4
            for count, supplier in enumerate(lowest_cumm_suppliers[:5], start=1):
                supplier_total_row_name = excel_cols[3 + c]
                suppliers_totals_formula.extend(
                    [
                        "", "",
                        "qT_{}".format(
                            f"= SUM({supplier_total_row_name}17:{supplier_total_row_name}{supplier_totals_row})"
                        ),
                    ]
                )
                c += 3

            supplier_headers = tuple(supplier_headers)
            rfq_summary.append(supplier_headers)
            items_header = tuple(items_header)
            rfq_summary.append(items_header)
            rfq_summary.merge_cells("C7:H7")

            j = 1
            for item in items:

                outlier_score = item.outlier_score
                responses = item.responses

                rfq_current_cost += float(item.current_price)
                rfq_eprocure_cost += item.eprocure_supplier_cost(responses)
                rfq_total_potential_savngs += item.savings(responses)

                item_row = []
                if item.savings_value is None:
                    item_savings = item.savings(responses)
                else:
                    item_savings = item.savings_value

                if item.eprocure_total_value is None:
                    item_eprocure_total = item.eprocure_total(responses)
                else:
                    item_eprocure_total = item.eprocure_total_value

                item_current_total = item.current_total
                rfq_item_current_total = item.quantity * item_current_total

                if float(rfq_item_current_total) < float(outlier_score):
                    rfq_item_current_total = "O_{}".format(f"{rfq_item_current_total:}")
                else:
                    rfq_item_current_total = "T_{}".format(f"{rfq_item_current_total:}")

                tendersure_suppliers = (
                    item.multi_item_tender_eprocure_suppliers_with_tt_tech_weight(
                        responses=responses))

                try:
                    if item.second_description is None:
                        second_description = "Ac_N/A"
                    else:
                        second_description = item.second_description
                except:
                    second_description = "Ac_N/A"

                price_1 = tendersure_suppliers[0]["price"]
                price_2 = tendersure_suppliers[1]["price"]
                price_3 = tendersure_suppliers[2]["price"]
                price_4 = tendersure_suppliers[3]["price"]
                price_5 = tendersure_suppliers[4]["price"]
                savings_lowest_supplier = ""
                try:
                    if item_eprocure_total > item_current_total and item_current_total != 0:
                        item_savings_cell = "Ar_0"
                        savings_lowest_supplier = "Current Supplier"
                        price_5 = price_4
                        price_4 = price_3
                        price_3 = price_2
                        price_2 = price_1
                        price_1 = item_current_total
                        item_row.extend(
                            [
                                "", "Ac_{}".format(item.number), item.description,
                                "Tc_{}".format(item.quantity), second_description,
                                savings_lowest_supplier, "", "Tr_{}".format(f"{price_1:}"),
                                tendersure_suppliers[0]["supplier"], tendersure_suppliers[0]["weighted_score"],
                                "Tr_{}".format(f"{price_2:}"), tendersure_suppliers[1]["supplier"],
                                tendersure_suppliers[1]["weighted_score"], "Tr_{}".format(f"{price_3:}"),
                                tendersure_suppliers[2]["supplier"], tendersure_suppliers[2]["weighted_score"],
                                "Tr_{}".format(f"{price_4:}"), tendersure_suppliers[3]["supplier"],
                                tendersure_suppliers[3]["weighted_score"], "Tr_{}".format(f"{price_5:}"),
                            ]
                        )
                    else:
                        item_savings_cell = "T_{}".format(item_savings)
                        savings_lowest_supplier = tendersure_suppliers[0]["supplier"]
                        item_row.extend(
                            [
                                "", "Ac_{}".format(item.number), item.description,
                                "Tc_{}".format(item.quantity), second_description,
                                savings_lowest_supplier, tendersure_suppliers[0]["weighted_score"],
                                "Tr_{}".format(f"{price_1:}"), tendersure_suppliers[1]["supplier"],
                                tendersure_suppliers[1]["weighted_score"], "Tr_{}".format(f"{price_2:}"),
                                tendersure_suppliers[2]["supplier"], tendersure_suppliers[2]["weighted_score"],
                                "Tr_{}".format(f"{price_3:}"), tendersure_suppliers[3]["supplier"],
                                tendersure_suppliers[3]["weighted_score"], "Tr_{}".format(f"{price_4:}"),
                                tendersure_suppliers[4]["supplier"], tendersure_suppliers[4]["weighted_score"],
                                "Tr_{}".format(f"{price_5:}"),
                            ]
                        )
                except:
                    item_row.extend(
                        [
                            "", "Ac_{}".format(item.number), item.description,
                            "Tc_{}".format(item.quantity), second_description,
                            tendersure_suppliers[0]["supplier"], tendersure_suppliers[0]["weighted_score"],
                            "Tr_{}".format(f"{price_1:}"), tendersure_suppliers[1]["supplier"],
                            tendersure_suppliers[1]["weighted_score"], "Tr_{}".format(f"{price_2:}"),
                            tendersure_suppliers[2]["supplier"], tendersure_suppliers[2]["weighted_score"],
                            "Tr_{}".format(f"{price_3:}"), tendersure_suppliers[3]["supplier"],
                            tendersure_suppliers[3]["weighted_score"], "Tr_{}".format(f"{price_4:}"),
                            tendersure_suppliers[4]["supplier"], tendersure_suppliers[4]["weighted_score"],
                            "Tr_{}".format(f"{price_5:}"),
                        ]
                    )

                item_row = tuple(item_row)
                rfq_summary.append(item_row)

            tender.total_savings = float(rfq_total_potential_savngs)
            tender.save()

            colums_totals = []
            colums_totals.extend(["", "", "qT_Total", ""])

            colums_totals.extend(suppliers_totals_formula)
            rfq_summary.append(tuple(colums_totals))
            rfq_summary.append(("", "", ""))
            rfq_summary.append(("", "", ""))
            rfq_summary.append(("", "Ac_#", "Notes"))
            rfq_summary.append(
                ("", "Ac_1",
                 "Nmrl_3_TW Score : This means the total (technical + financial) weighted score")
            )
            rfq_summary.append(
                ("", "Ac_2",
                 "Nmrl_3_Best Supplier: This means the supplier with the best total weighted score (TW Score)"))
            rfq_summary.freeze_panes = rfq_summary["F17"]
            self.result += 1
            self.progress_recorder.set_progress(self.result, total_for_progress)
            return
        except Exception as e:
            capture_exception(e)
            print(e)
            return

    def create_technical_worksheet(self, workbook, data):
        try:
            category = data["category"]
            total_for_progress = data["total_for_progress"]
            sorted_participants = data["sorted_participants"]
            sections = data["sections"]
            questions = data["questions"]
            tender = data["tender"]

            percentage_responses = data["percentage_responses"]

            workbook.create_sheet("TECHNICAL_REPORT")
            technical_summary = workbook["TECHNICAL_REPORT"]
            category_order = apps.apps.get_model('core', 'CategoryOrder')
            all = category_order.objects.filter(
                category_id=category.id, payment_status=category_order.PAID, target__model='tender'
            ).count()

            details_summary = [
                ("", "", "", ""),
                ("", f"Bnb_Category: {category.name}", "", ""),
                ("", "", "", ""),
                ("", "Number of bidders", "", f"Ac_{all}"),
                ("", "Number of responsive bidders", "", f"Ac_{category.responsive_bidders.count()}"),
                ("", "Percentage responses", "", f"Ac_{percentage_responses}"),
                (
                    "", "Weighting (Technical:Financial)", "",
                    f"Ac_{int(float(category.technical_weight))}:{int(float(category.financial_weight))}"),
                ("", "", "", ""),
            ]
            print(details_summary)
            for row in details_summary:
                technical_summary.append(row)

            technical_summary.merge_cells("B2:H2")
            technical_summary.merge_cells("B4:C4")
            technical_summary.merge_cells("B5:C5")
            technical_summary.merge_cells("B6:C6")
            technical_summary.merge_cells("B7:C7")

            category_headers = [
                "", "hc_#", "hl_Bidder", "hc_Rank", "hc_W Score", "hc_Score"]

            for section in sections:
                self.result += 1
                self.progress_recorder.set_progress(self.result, total_for_progress)
                if section.has_child_sections:
                    if section.is_scored:
                        s_total = section.section_score
                        for s in section.child_sections:
                            s_total += s.section_score
                        category_headers.extend([f"hc_{section.short_name}"])
                    else:
                        evaluate_this = False
                        for s in section.child_sections:
                            if s.is_scored:
                                evaluate_this = True
                                break
                        if evaluate_this == True:
                            s_total = 0
                            for s in section.child_sections:
                                s_total += s.section_score
                            category_headers.extend([f"hc_{section.short_name}"])
                elif section.is_scored:
                    category_headers.extend([f"hc_{section.short_name}"])

            for question in questions:
                category_headers.extend([f"hc_{question.short_description}"])
            technical_summary.append(category_headers)

            c = 1
            for supplier in sorted_participants:
                technical_score = SupplierTechnicalScore.objects.filter(
                    category_id=category.id, supplier_id=supplier.id).first()
                # prequal_score = supplier.tt_prequal_score(prequal.id)
                # prequalified_bidders += 1

                supplier_info = [
                    "", "Ac_{}".format(c), supplier.company_name.title(),
                    f"Ac_{technical_score.rank}",
                    f"Tc_{technical_score.weighted_score}",
                    f"Tc_{technical_score.score}",
                ]

                for section in sections:
                    if section.has_child_sections:
                        if section.is_scored:
                            section_total = supplier.tender_section_score(section.id)
                            for s in section.child_sections:
                                if s.is_scored:
                                    section_total += supplier.tender_section_score(s.id)
                            supplier_info.extend([f"Tc_{section_total}"])
                        else:
                            evaluate_this = False
                            for s in section.child_sections:
                                if s.is_scored:
                                    evaluate_this = True
                                    break
                            if evaluate_this == True:
                                section_total = supplier.tender_section_score(section.id)
                                for s in section.child_sections:
                                    if s.is_scored:
                                        section_total += supplier.tender_section_score(s.id)
                                supplier_info.extend([f"Tc_{section_total}"])
                    else:
                        if section.is_scored:
                            section_total = supplier.tender_section_score(section.id)
                            supplier_info.extend([f"Tc_{section_total}"])

                for question in questions:
                    qa_response = QualityAssuranceResponse.objects.filter(
                        supplier=supplier, quality_assurance_question__question=question
                    ).first()
                    if qa_response is not None:
                        supplier_info.extend(
                            [
                                f"Qac_{qa_response.outcome + '(' + qa_response.comment + ')'}"
                            ]
                        )
                    else:
                        supplier_info.extend([" "])
                technical_summary.append(supplier_info)
                c += 1

            technical_summary.append(("", "", ""))
            technical_summary.append(("", "Ac_#", "Nmrl_3_Notes"))
            technical_summary.append(
                (
                    "", "Ac_1",
                    "Nmrl_3_TW Score : This means the weighted technical score",
                )
            )
            technical_summary.append(
                ("", "Ac_2", "Nmrl_3_BCC - Business compliance and continuity")
            )
            technical_summary.append(
                ("", "Ac_3", "Nmrl_3_Technical - Technical Specifications")
            )
            technical_summary.append(("", "Ac_4", "Nmrl_3_Staff - Staffing"))
            technical_summary.append(
                ("", "Ac_5", "Nmrl_3_Recommendation - Recommendation")
            )
            technical_summary.append(
                ("", "Ac_6", "Nmrl_3_Years in Business - Years in Business")
            )
            technical_summary.append(
                (
                    "", "Ac_7",
                    "Nmrl_3_Financial - Financials and business sustainability",
                )
            )
            technical_summary.append(("", "Ac_8", "Nmrl_3_Declaration - Declaration"))
            technical_summary.freeze_panes = technical_summary["E9"]
            self.result += 1
            self.progress_recorder.set_progress(self.result, total_for_progress)
            return
        except Exception as e:
            capture_exception(e)
            print(e)
            return

    def create_lowest_cost_report_worksheet(self, workbook, data):
        category_responsive_bidders = data["category_responsive_bidders"]
        category_total_bidders = data["category_total_bidders"]
        category = data["category"]
        items = data["items"]
        lowest_cumm_suppliers = data["lowest_cumm_suppliers"]
        total_for_progress = data["total_for_progress"]
        excel_cols = data["excel_cols"]
        percentage_responses = data["percentage_responses"]

        try:
            workbook.create_sheet("LOWEST ITEM COST REPORT")
            pricing_worksheet = workbook["LOWEST ITEM COST REPORT"]

            rfq_current_cost = 0
            rfq_eprocure_cost = 0
            rfq_total_potential_savngs = 0

            pricing_worksheet["B2"] = f"Bnb_Category: {category.name}"
            pricing_worksheet["B3"] = "Bnb_Lowest Cost Report"
            pricing_worksheet["B4"] = ""
            pricing_worksheet["B5"] = "Number of bidders"
            pricing_worksheet["D5"] = f"Ac_{category_total_bidders}"
            pricing_worksheet["B6"] = "Responsive bidders"
            pricing_worksheet["D6"] = "Ac_{}".format(category_responsive_bidders)
            pricing_worksheet["B7"] = "Percentage responses"
            pricing_worksheet["D7"] = f"Ac_{percentage_responses}"
            pricing_worksheet["B8"] = "Currency"
            pricing_worksheet["D8"] = "Ac_{}".format(category.currency)

            pricing_worksheet.merge_cells("B2:H2")
            pricing_worksheet.merge_cells("B4:C4")
            pricing_worksheet.merge_cells("B5:C5")
            pricing_worksheet.merge_cells("B6:C6")
            pricing_worksheet.merge_cells("B7:C7")
            pricing_worksheet.merge_cells("B8:C8")

            empty_row = [""]
            empty_row = tuple(empty_row)
            pricing_worksheet.append(empty_row)

            supplier_headers = ["", "", "", "", "", "", "", "", "", "", "", ""]

            items_header = [
                "", "hc_#", "h_Item Description", "hc_Qty",
                "hc_Price Validity (Months)", "h_Specification","h_Lowest Cost 1",
                "h_Lowest Supplier 1", "h_Lowest Cost 2", "h_Lowest Supplier 2",
                "h_Lowest Cost 3", "h_Lowest Supplier 3",
            ]

            no_of_items = len(items)
            no_of_rows_at_the_top = 11
            supplier_totals_row = no_of_rows_at_the_top + no_of_items
            formulas_row = []

            formulas_row.extend(
                ["qT_{}".format(f"= SUM(G12:G{supplier_totals_row})"), ""]
            )
            formulas_row.extend(
                ["qT_{}".format(f"= SUM(I12:I{supplier_totals_row})"), ""]
            )
            formulas_row.extend(
                ["qT_{}".format(f"= SUM(K12:K{supplier_totals_row})"), ""]
            )

            c = 1
            for count, supplier in enumerate(lowest_cumm_suppliers, start=1):
                supplier_total_row_name = excel_cols[10 + c]
                formulas_row.extend(
                    [
                        "qT_{}".format(
                            f"= SUM({supplier_total_row_name}12:{supplier_total_row_name}{supplier_totals_row})"
                        ),
                        "",
                    ]
                )

                supplier_headers.append(f"Hmr_1_{supplier.supplier.company_name.title()}")
                supplier_headers.extend(
                    [
                        "",
                    ]
                )
                items_header.append("hc_COST")
                items_header.append("hc_WF SCORE")
                c += 2

            supplier_headers = tuple(supplier_headers)
            pricing_worksheet.append(supplier_headers)
            items_header = tuple(items_header)
            pricing_worksheet.append(items_header)

            j = 1
            for item in items:
                outlier_score = item.outlier_score
                responses = item.responses

                rfq_current_cost += float(item.current_price)
                rfq_eprocure_cost += item.eprocure_supplier_cost(responses)
                rfq_total_potential_savngs += item.savings(responses)

                if item.savings_value is None:
                    item_savings = item.savings(responses)
                else:
                    item_savings = item.savings_value

                if item.eprocure_total_value is None:
                    item_eprocure_total = item.eprocure_total(responses)
                else:
                    item_eprocure_total = item.eprocure_total_value

                item_current_total = item.current_total
                item_row = []

                tendersure_suppliers = item.multi_item_tender_eprocure_suppliers(
                    responses
                )
                price_1 = tendersure_suppliers[0]["price"]
                price_2 = tendersure_suppliers[1]["price"]
                price_3 = tendersure_suppliers[2]["price"]

                try:
                    if item.second_description is None:
                        second_description = "Ac_N/A"
                    else:
                        second_description = item.second_description

                    if item.price_validity_months is None:
                        price_validity = "Ac_{}".format(12)
                    else:
                        price_validity = "Ac_{}".format(item.price_validity_months)

                except:
                    second_description = "Ac_N/A"
                    price_validity = "Ac_{}".format(12)

                try:
                    if item_eprocure_total > item_current_total and item_current_total != 0:
                        savings_lowest_supplier = "Current Supplier"
                        price_3 = price_2
                        price_2 = price_1
                        price_1 = item_current_total
                        item_row.extend(
                            [
                                "", "Ac_{}".format(item.number),
                                item.description, "Tc_{}".format(item.quantity),
                                price_validity, second_description, "Tr_{}".format(f"{price_1:}"),
                                savings_lowest_supplier, "Tr_{}".format(f"{price_2:}"),
                                tendersure_suppliers[0]["supplier"], "Tr_{}".format(f"{price_3:}"),
                                tendersure_suppliers[1]["supplier"],
                            ]
                        )
                    else:
                        savings_lowest_supplier = tendersure_suppliers[0]["supplier"]
                        item_row.extend(
                            [
                                "", "Ac_{}".format(item.number), item.description,
                                "Tc_{}".format(item.quantity), price_validity, second_description,
                                "Tr_{}".format(f"{price_1:}"), savings_lowest_supplier,
                                "Tr_{}".format(f"{price_2:}"), tendersure_suppliers[1]["supplier"],
                                "Tr_{}".format(f"{price_3:}"), tendersure_suppliers[2]["supplier"],
                            ]
                        )
                except:
                    item_row.extend(
                        [
                            "", "Ac_{}".format(item.number), item.description,
                            "Tc_{}".format(item.quantity), price_validity, second_description,
                            "Tr_{}".format(f"{price_1:}"), tendersure_suppliers[0]["supplier"],
                            "Tr_{}".format(f"{price_2:}"), tendersure_suppliers[1]["supplier"],
                            "Tr_{}".format(f"{price_3:}"), tendersure_suppliers[2]["supplier"],
                        ]
                    )

                for count, supplier in enumerate(lowest_cumm_suppliers, start=1):
                    if not supplier:
                        supplier_total = 0
                        item_row.append(supplier_total)
                    else:
                        item_response = ItemResponse.objects.filter(
                            supplier_id=supplier.supplier_id, item_id=item.id).first()
                        if item_response is not None:
                            supplier_total = show(item_response.total)
                        else:
                            supplier_total = 0
                        s_total = round(supplier_total, 2)

                        if ItemResponse.objects.filter(item_id=item.id).count() <= 2:
                            supplier_total = "Tr_{}".format(f"{s_total:}")
                        else:
                            if float(s_total) < float(outlier_score):
                                supplier_total = "O_{}".format(f"{s_total:}")
                            else:
                                supplier_total = "Tr_{}".format(f"{s_total:}")
                        item_row.append(supplier_total)
                        item_row.append(
                            f"{supplier.supplier.tender_item_weighted_score(item=item)}"
                        )

                item_row = tuple(item_row)
                pricing_worksheet.append(item_row)

            category.total_savings = float(rfq_total_potential_savngs)
            category.save()

            colums_totals = []
            colums_totals.extend(["", "", "qT_Total", "", "", ""])

            colums_totals.extend(formulas_row)

            pricing_worksheet.append(tuple(colums_totals))
            pricing_worksheet.append(("", "", ""))
            pricing_worksheet.append(("", ""))
            pricing_worksheet.append(("", "hc_#", "Nmrl_3_Notes"))
            pricing_worksheet.append(
                ("", "Ac_1", "Nmrl_3_Outliers are highlighted in red.")
            )
            pricing_worksheet.append(
                (
                    "",
                    "Ac_2",
                    "Nmrl_3_Note: Orange highlight indicates that the current cost is lower than the cheapest bid.",
                )
            )
            pricing_worksheet.append(
                (
                    "",
                    "Ac_3",
                    "Nmrl_3_Outlier: The prices highlighted in this colour are outliers which means that they are 25% lower than the median and are not considered in the analysis.",
                )
            )
            pricing_worksheet.freeze_panes = pricing_worksheet["E12"]
            self.result += 1
            self.progress_recorder.set_progress(self.result, total_for_progress)
            return
        except Exception as e:
            capture_exception(e)
            print(e)
            return

    def create_lowest_supplier_worksheet(self, workbook, data):
        try:
            category = data["category"]
            excel_cols = data["excel_cols"]
            items = data["items"]
            category_total_bidders = data["category_total_bidders"]
            category_responsive_bidders = data["category_responsive_bidders"]
            percentage_responses = data["percentage_responses"]
            lowest_cumm_suppliers_ranked = data["lowest_cumm_suppliers_ranked"]
            lowest_cumm_suppliers_opp = data["lowest_cumm_suppliers_opp"]
            total_for_progress = data["total_for_progress"]

            rfq_current_cost = 0
            rfq_eprocure_cost = 0
            rfq_total_potential_savngs = 0

            workbook.create_sheet("LOWEST SUPPLIER REPORT")
            lowest_supplier_worksheet = workbook["LOWEST SUPPLIER REPORT"]

            lowest_supplier_worksheet["B2"] = "Bnb_Category: {}".format(category.name)
            lowest_supplier_worksheet["B3"] = "Bnb_Lowest Supplier Report"
            lowest_supplier_worksheet["B4"] = ""
            lowest_supplier_worksheet["B5"] = "Invited bidders"
            lowest_supplier_worksheet["D5"] = "Ac_{}".format(category_total_bidders)
            lowest_supplier_worksheet["B6"] = "Bidder responses"
            lowest_supplier_worksheet["D6"] = "Ac_{}".format(
                category_responsive_bidders
            )
            lowest_supplier_worksheet["B7"] = "Percentage responses"
            lowest_supplier_worksheet["D7"] = "Ac_{}".format(percentage_responses)
            lowest_supplier_worksheet["B8"] = "Currency"
            lowest_supplier_worksheet["D8"] = "Ac_{}".format(category.currency)

            empty_row = [""]
            empty_row = tuple(empty_row)
            lowest_supplier_worksheet.append(empty_row)
            lowest_supplier_worksheet.merge_cells("B2:H2")
            lowest_supplier_worksheet.merge_cells("B4:C4")
            lowest_supplier_worksheet.merge_cells("B5:C5")
            lowest_supplier_worksheet.merge_cells("B6:C6")
            lowest_supplier_worksheet.merge_cells("B7:C7")
            lowest_supplier_worksheet.merge_cells("B8:C8")

            s = 5
            supplier_totals_row = 12 + len(items)
            lowest_cumm_suppliers_totals = []
            lowest_cumm_suppliers = list(lowest_cumm_suppliers_ranked) + list(
                lowest_cumm_suppliers_opp
            )

            lowest_supplier_sub_headers = ["", "", "", "", "", ""]
            lowest_supplier_rank_headers = ["", "", "", "", "", ""]
            lowest_supplier_headers = [
                "", "hc_#", "h_Item Description", "hc_Qty", "h_Specification",
            ]

            scores = []
            for count, supplier in enumerate(lowest_cumm_suppliers, start=1):
                row_name = excel_cols[s + count]
                lowest_cumm_suppliers_totals.append(
                    "qT_{}".format(
                        f"= SUM({row_name}13:{row_name}{supplier_totals_row})"
                    )
                )
                lowest_supplier_sub_headers.append(
                    "hc_%s" % supplier.supplier.company_name.title()
                )
                lowest_supplier_headers.append("hc_TOTAL")
                if not supplier.has_outlier:
                    lowest_supplier_rank_headers.append(f"hc_Rank {count}")
                else:
                    lowest_supplier_rank_headers.append(f"hc_N/A")
                scores.append(f"Tcb_{supplier.score}")

            lowest_supplier_headers = tuple(lowest_supplier_headers)
            lowest_supplier_sub_headers = tuple(lowest_supplier_sub_headers)
            lowest_supplier_rank_headers = tuple(lowest_supplier_rank_headers)
            lowest_supplier_worksheet.append(lowest_supplier_rank_headers)
            lowest_supplier_worksheet.append(lowest_supplier_sub_headers)
            lowest_supplier_worksheet.append(lowest_supplier_headers)

            j = 1
            for item in items:
                outlier_score = item.outlier_score
                responses = item.responses

                rfq_current_cost += float(item.current_price)
                rfq_eprocure_cost += item.eprocure_supplier_cost(responses)
                rfq_total_potential_savngs += item.savings(responses)

                lowest_supplier_row = []

                try:
                    if item.second_description is None:
                        second_description = "Ac_N/A"
                    else:
                        second_description = item.second_description

                    if item.price_validity_months is None:
                        price_validity = "Ac_{}".format(12)
                    else:
                        price_validity = "Ac_{}".format(item.price_validity_months)
                except:
                    second_description = "Ac_N/A"

                lowest_supplier_row.extend(
                    [
                        "", "Ac_{}".format(item.number),
                        item.description, "Tc_{}".format(item.quantity),
                        price_validity, second_description,
                    ]
                )

                for count, lowest_cumm_supplier in enumerate(
                    lowest_cumm_suppliers, start=1
                ):
                    if not lowest_cumm_supplier:
                        supplier_total = 0
                    else:
                        item_response = ItemResponse.objects.filter(
                            item_id=item.id, supplier_id=lowest_cumm_supplier.supplier_id).first()
                        if item_response is not None:
                            if item_response.total is not None:
                                supplier_total = show(item_response.total)
                        else:
                            supplier_total = 0
                        s_total = round(supplier_total, 2)
                        if float(s_total) < float(outlier_score):
                            supplier_total = "O_{}".format(f"{s_total:}")
                        else:
                            supplier_total = "Tr_{}".format(f"{s_total:}")
                        lowest_supplier_row.append(supplier_total)

                lowest_supplier_row = tuple(lowest_supplier_row)
                lowest_supplier_worksheet.append(lowest_supplier_row)

                j += 1

            lowest_supplier_totals = []
            lowest_supplier_totals.extend(["", "", "qT_Total", "", "", ""])
            weighted_financial_scores = [
                "", "", "qT_Weighted Financial Score", "",
                "", "",
            ]
            weighted_financial_scores.extend(scores)
            lowest_supplier_totals.extend(lowest_cumm_suppliers_totals)
            lowest_supplier_worksheet.append(tuple(lowest_supplier_totals))
            lowest_supplier_worksheet.append(empty_row)
            lowest_supplier_worksheet.append(weighted_financial_scores)
            lowest_supplier_worksheet.append(empty_row)
            lowest_supplier_worksheet.append(("", "Ac_#", "Nmrl_3_Notes"))
            lowest_supplier_worksheet.append(
                (
                    "", "Ac_1",
                    "Nmrl_3_Rank: This is the ranking of suppliers with the lowest cost and is not an outlier",
                )
            )
            lowest_supplier_worksheet.append(
                (
                    "", "Ac_2",
                    "Nmrl_3_Outlier: The prices highlighted in this colour are "
                    "outliers which means that they are 50% lower than the median "
                    "and are not considered in the analysis.",
                )
            )
            lowest_supplier_worksheet.freeze_panes = lowest_supplier_worksheet["G13"]

            self.result += 1
            self.progress_recorder.set_progress(self.result, total_for_progress)
            return
        except Exception as e:
            capture_exception(e)
            print(e)
            return

    def create_savings_worksheet(self, workbook, data):
        workbook.create_sheet("SAVINGS REPORT")
        savings_worksheet = workbook["SAVINGS REPORT"]

        try:
            category = data["category"]
            items = data["items"]
            category_total_bidders = data["category_total_bidders"]
            category_responsive_bidders = data["category_responsive_bidders"]
            percentage_responses = data["percentage_responses"]
            empty_row = [""]
            empty_row = tuple(empty_row)
            total_for_progress = data["total_for_progress"]

            savings_worksheet.append(empty_row)
            savings_worksheet["B2"] = f"Bnb_Category: {category.name}"
            savings_worksheet["B3"] = "Bnb_Savings Report"
            savings_worksheet["B4"] = ""
            savings_worksheet["D4"] = ""
            savings_worksheet["B5"] = "Number of bidders"
            savings_worksheet["D5"] = "Ac_{}".format(category_total_bidders)
            savings_worksheet["B6"] = "Bidder responses"
            savings_worksheet["D6"] = "Ac_{}".format(category_responsive_bidders)
            savings_worksheet["B7"] = "Percentage responses"
            savings_worksheet["D7"] = "Ac_{}".format(percentage_responses)
            savings_worksheet["B8"] = "Currency"
            savings_worksheet["D8"] = "Ac_{}".format(category.currency)
            savings_worksheet.append(empty_row)

            savings_worksheet.merge_cells("B2:H2")
            savings_worksheet.merge_cells("B4:C4")
            savings_worksheet.merge_cells("B5:C5")
            savings_worksheet.merge_cells("B6:C6")
            savings_worksheet.merge_cells("B7:C7")
            savings_worksheet.merge_cells("B8:C8")

            savings_items_header = [
                "", "hc_#", "h_Item Description", "hc_Qty",
                "h_Specification", "h_Current Supplier", "hc_Current Price",
                "h_Current Cost", "h_Lowest Supplier", "hc_Lowest Cost",
                "hc_Unit Savings", "hc_Total Savings", "hc_% Savings",
            ]

            savings_worksheet.append(savings_items_header)

            j = 1
            for item in items:
                outlier_score = item.outlier_score
                responses = item.responses

                savings_item_row = []
                if item.savings_value is None:
                    item_savings = item.savings(responses)
                else:
                    item_savings = item.savings_value

                if item.eprocure_total_value is None:
                    item_eprocure_total = item.eprocure_total(responses)
                else:
                    item_eprocure_total = item.eprocure_total_value

                item_current_total = item.current_total
                item_current_total = item.quantity * item_current_total

                if float(item_current_total) < float(outlier_score):
                    rfq_item_current_total = "Ac_N/A".format(f"{item_current_total:}")
                else:
                    rfq_item_current_total = "T_{}".format(f"{item_current_total:}")

                tendersure_suppliers = item.eprocure_suppliers(responses)
                price_1 = tendersure_suppliers[0]["price"]
                savings_lowest_supplier = ""

                if item_eprocure_total > item_current_total and item_current_total != 0:
                    item_savings_cell = "Ar_0"
                    savings_lowest_supplier = "Current Supplier"
                else:
                    item_savings_cell = "Tr_{}".format(item_savings)
                    savings_lowest_supplier = tendersure_suppliers[0]["supplier"]

                total_savings = item_savings * item.quantity

                try:
                    if item.second_description is None:
                        second_description = "Ac_N/A"
                    else:
                        second_description = item.second_description
                except:
                    second_description = "Ac_N/A"

                # if item_current_total != 0:
                savings_item_row.extend(
                    [
                        "",
                        "Ac_{}".format(item.number),
                        item.description,
                        "Tc_{}".format(item.quantity),
                        second_description,
                        "Ac_N/A",
                        "T_{}".format(f"{item_current_total:}")
                        if item_current_total > 0
                        else "Ac_N/A",
                        rfq_item_current_total,
                        savings_lowest_supplier,
                        "Tr_{}".format(f"{price_1:}"),
                        item_savings_cell,
                        f"T_{total_savings}",
                        "Ac_{}".format(item.percentage_savings(responses)),
                    ]
                )

                savings_item_row = tuple(savings_item_row)
                savings_worksheet.append(savings_item_row)
                j += 1
            totals_row = [
                "",
                "",
                "B_Total",
                "",
                "",
                "",
                f"Tc_=SUM(G10:G{9 + len(items)})",
                f"Tc_=SUM(H10:H{9 + len(items)})",
                "",
                f"Tr_=SUM(J10:J{9 + len(items)})",
                f"Tr_=SUM(K10:K{9 + len(items)})",
                f"Tc_=SUM(L10:L{9 + len(items)})",
                f"Ac_=SUM(L10:L{9 + len(items)})/SUM(H10:H{9 + len(items)})",
            ]
            savings_worksheet.append(totals_row)

            savings_worksheet.append(empty_row)
            savings_worksheet.append(["", "hc_#", "Nmrl_3_Notes"])
            savings_worksheet.append(
                [
                    "", "Ac_1",
                    "Nmrl_3_Where the current prices are lower than the lowest price obtained, current price is retained",
                    "", "",
                ]
            )
            savings_worksheet.append(
                [
                    "", "Ac_2",
                    "Nmrl_3_N/A means that the current supplier has not been provided",
                    "", "",
                ]
            )
            savings_worksheet.freeze_panes = savings_worksheet["E11"]
            self.result += 1
            self.progress_recorder.set_progress(self.result, total_for_progress)
        except Exception as e:
            capture_exception(e)
            print(e)
        return

    # def create_individual_worksheets(self, workbook, data):
    #     try:
    #         total_for_progress = data["total_for_progress"]
    #         participants = data["participants"]
    #         rfq = data["rfq"]
    #
    #         c_data = {}
    #         for count, supplier in enumerate(participants, start=1):
    #             try:
    #                 workbook.create_sheet(f"{supplier.company_name.upper()}")
    #                 supplier_quotation = workbook[f"{supplier.company_name.upper()}"]
    #                 supplier_workbook = load_workbook(
    #                     supplier.relative_quotation_url(rfq), data_only=True
    #                 )
    #                 bid_worksheet = supplier_workbook["RFQ ITEMS"]
    #                 c_data[f"{supplier.company_name.upper()}_A"] = 5
    #                 c_data[f"{supplier.company_name.upper()}_B"] = 27
    #                 c_data[f"{supplier.company_name.upper()}_C"] = 27
    #                 c_data[f"{supplier.company_name.upper()}_D"] = 27
    #
    #                 for row in bid_worksheet:
    #                     for cell in row:
    #                         quote_cell = supplier_quotation[cell.coordinate]
    #                         quote_cell.value = cell.value
    #                         if cell.has_style:
    #                             quote_cell.font = copy(cell.font)
    #                             quote_cell.border = copy(cell.border)
    #                             quote_cell.fill = copy(cell.fill)
    #                             quote_cell.number_format = copy(cell.number_format)
    #                             quote_cell.protection = copy(cell.protection)
    #                             quote_cell.alignment = copy(cell.alignment)
    #             except:
    #                 pass
    #
    #         self.result += 1
    #         self.progress_recorder.set_progress(self.result, total_for_progress)
    #         return c_data
    #     except Exception as e:
    #         capture_exception(e)


class ConsolidatedTenderSummaryReport(Task):
    name = "ConsolidatedTenderSummaryReport"
    progress_recorder = None
    result = 0

    def run(self, *args, **kwargs):
        self.progress_recorder = ProgressRecorder(self)
        context = self.report(kwargs["category_id"])
        print(context)
        return context

    def report(self, category_id):
        category = Category.objects.filter(id=category_id).first()
        job = category.tender
        company = job.company

        time = datetime.datetime.now()
        dir_name = Path(
            "media/tender_reports/{}/{}".format(time.year, time.month))
        dir_name.mkdir(parents=True, exist_ok=True)
        # time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
        filepath = "{}/{}_Consolidated_Tender_Summary_Report.xlsx".format(
            dir_name, job.id
        )

        # delete similar reports run this month
        match_string = "{}_{}".format(
            job.id, job.company.company_name.replace(" ", "_")
        )
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

        # create and save a workbook to later populate
        workbook = Workbook()
        financial_weight = category.financial_weight
        technical_weight = category.technical_weight

        items = Item.objects.filter(category_id=category_id).order_by('number')

        successful_participants = category.sorted_successful_participants

        category_total_bidders = category.total_bidders
        category_responsive_bidders = category.responsive_bidders

        try:
            percentage_responses = "{}%".format(
                round(((category_responsive_bidders.count() / category_total_bidders) * 100))
            )
        except Exception as e:
            percentage_responses = 0

        letters = list(string.ascii_uppercase)
        # if rfq:
        num_cols = items.count() + 100
        # else:
        #     num_cols = 100

        excel_cols = []
        for i in range(0, num_cols - 1):
            n = i // 26
            m = n // 26
            i -= n * 26
            n -= m * 26
            col = (
                letters[m - 1] + letters[n - 1] + letters[i]
                if m > 0
                else letters[n - 1] + letters[i]
                if n > 0
                else letters[i]
            )
            excel_cols.append(col)

        scored_sections = category.scored_sections
        scored_sections_count = scored_sections.count()

        lowest_cumm_suppliers_ranked = SupplierFinancialTotal.objects.filter(
            category_id=category_id, has_blank=False, has_outlier=False,
            meets_all_mandatory_requirements=True, technically_qualified=True).order_by("rank")

        lowest_cumm_suppliers_opp = SupplierFinancialTotal.objects.filter(
            Q(has_blank=True) | Q(has_outlier=True) | Q(meets_all_mandatory_requirements=False) |
            Q(technically_qualified=False),
            category_id=category_id).order_by("-score")

        questions = Question.objects.filter(section__category=category, is_qa=True)

        financial_participants = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
            id__in=SupplierFinancialTotal.objects.filter(
                category_id=category_id).order_by('rank').only('supplier_id').values('supplier_id')
        )

        total_for_progress = 4
        data = {
            "job": job,
            "category": category,
            "technical_weight": technical_weight,
            "financial_weight": financial_weight,
            "excel_cols": excel_cols,
            "successful_participants": successful_participants,
            "total_for_progress": total_for_progress,
            "items": items,
            "scored_sections_count": scored_sections_count,
            "scored_sections": scored_sections,
            "participants": financial_participants,
            "category_prequal_bidders": category.technical_participants,
            "category_total_bidders": category_total_bidders,
            "category_responsive_bidders": category_responsive_bidders,
            "percentage_responses": percentage_responses,
            "lowest_cumm_suppliers_opp": lowest_cumm_suppliers_opp,
            "lowest_cumm_suppliers_ranked": lowest_cumm_suppliers_ranked,
            "questions": questions,
        }
        self.create_tender_summary_worksheet(workbook, data)
        self.create_technical_worksheet(workbook, data)
        self.create_savings_worksheet(workbook, data)
        self.create_lowest_supplier_worksheet(workbook, data)
        # self.create_lowest_item_worksheet(workbook, data)
        # self.create_individual_worksheets(workbook, data)

        qed_logo_anchor = "H1"
        buyer_logo_anchor = "B1"

        if company.company_logo_url is not None and company.company_logo_url != "":
            buyer_logo_url = get_file_path(company.company_logo_url)
        else:
            buyer_logo_url = "static/core/img/no-company-image128-128.png"

        workbook.save(filepath)

        f_data = {
            "TENDER SUMMARY REPORT_A": 5, "TENDER SUMMARY REPORT_B": 5, "TENDER SUMMARY REPORT_C": 31,
            "TENDER SUMMARY REPORT_D": 8, "TENDER SUMMARY REPORT_E": 4, "TENDER SUMMARY REPORT_F": 14,
            "TENDER SUMMARY REPORT_G": 4, "TENDER SUMMARY REPORT_H": 14, "TENDER SUMMARY REPORT_I": 14,
            "TENDER SUMMARY REPORT_J": 4, "TENDER SUMMARY REPORT_K": 14, "TENDER SUMMARY REPORT_L": 14,
            "TENDER SUMMARY REPORT_N": 50, "TENDER SUMMARY REPORT_M": 4, "TECHNICAL REPORT_A": 5,
            "TECHNICAL REPORT_B": 5, "TECHNICAL REPORT_C": 31, "TECHNICAL REPORT_D": 6,
            "TECHNICAL REPORT_E": 7.5, "TECHNICAL REPORT_F": 11, "TECHNICAL REPORT_G": 9,
            "TECHNICAL REPORT_H": 9, "TECHNICAL REPORT_I": 9, "TECHNICAL REPORT_J": 9,
            "TECHNICAL REPORT_K": 9, "TECHNICAL REPORT_L": 9, "TECHNICAL REPORT_M": 9,
            "TECHNICAL REPORT_OTHER": 21, "SAVINGS REPORT_A": 5, "SAVINGS REPORT_B": 5,
            "SAVINGS REPORT_C": 31, "SAVINGS REPORT_D": 8, "SAVINGS REPORT_E": 13,
            "SAVINGS REPORT_F": 17, "SAVINGS REPORT_G": 17, "SAVINGS REPORT_H": 17,
            "SAVINGS REPORT_I": 17, "SAVINGS REPORT_J": 17, "SAVINGS REPORT_K": 17,
            "LOWEST COST REPORT_A": 5, "LOWEST COST REPORT_B": 5, "LOWEST COST REPORT_C": 31,
            "LOWEST COST REPORT_D": 8, "LOWEST COST REPORT_E": 13, "LOWEST COST REPORT_F": 13,
            "LOWEST COST REPORT_G": 24, "LOWEST COST REPORT_H": 24, "LOWEST COST REPORT_I": 24,
            "LOWEST COST REPORT_J": 24, "LOWEST COST REPORT_K": 24, "LOWEST COST REPORT_OTHER": 24,
        }

        format_excel(excel_url=filepath, data=f_data)
        insert_image(
            excel_url=filepath,
            worksheet_name="TENDER SUMMARY REPORT", anchor=qed_logo_anchor,
        )
        insert_image(
            excel_url=filepath, worksheet_name="TENDER SUMMARY REPORT",
            anchor=buyer_logo_anchor, image_url=buyer_logo_url,
        )

        try:
            with open(filepath, "rb") as l:
                storage = PrivateMediaStorage()
                url = f"tender/category/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
                storage.save(url, l)

                category_report, created = CategoryReport.objects.update_or_create(
                    id=category_id, defaults={"consolidated_tender_summary_report": url}
                )
                context = {
                    "response_message": "Report generated successfully",
                    "filepath": category_report.consolidated_tender_summary_report.url,
                }
                return context
        except Exception as e:
            context = {
                "response_message": "Report generation error",
                "errors": f"{e}",
            }
            return context

    def create_tender_summary_worksheet(self, workbook, data):
        try:
            category = data["category"]
            technical_weight = data["technical_weight"]
            financial_weight = data["financial_weight"]
            excel_cols = data["excel_cols"]
            successful_participants = data["successful_participants"]
            total_for_progress = data["total_for_progress"]
            items = data["items"]
            percentage_responses = data["percentage_responses"]

            category_worksheet = workbook.active
            category_worksheet.title = "TENDER SUMMARY REPORT"

            category_worksheet["B5"] = f"Bnb_Category: {category.name}"

            category_worksheet["B7"] = "Number of bidders"
            category_worksheet["D7"] = "Ac_%d" % category.total_bidders

            category_worksheet["B8"] = "Number of responsive bidders"
            category_worksheet["D8"] = f"Ac_{category.responsive_bidders.count()}"

            category_worksheet["B9"] = "Percentage response"
            category_worksheet["D9"] = f"Ac_{percentage_responses}"

            category_worksheet["B10"] = "Weighting(Technical:Financial)"
            category_worksheet["D10"] = "Ac_%d:%d" % (
                technical_weight,
                financial_weight,
            )

            category_worksheet["B11"] = "Currency"
            category_worksheet["D11"] = f"Ac_{category.currency.initials}"

            # merge_note = "C6" + ":" + "D6"
            # category_worksheet.merge_cells(merge_note)
            category_worksheet.merge_cells("B5:H5")
            category_worksheet.merge_cells("B7:C7")
            category_worksheet.merge_cells("B8:C8")
            category_worksheet.merge_cells("B9:C9")
            category_worksheet.merge_cells("B10:C10")
            category_worksheet.merge_cells("B11:C11")
            category_worksheet.merge_cells("H13:I13")
            category_worksheet.merge_cells("K13:L13")

            header_1 = ["", "", "", "", "", "", ""]
            header_2 = [
                "", "", "", "", "", "hc_Total", "", "hc_Technical Score"]

            header_3 = [
                "", "hc_#", "h_Name of Supplier", "hc_Rank",
                "", "hc_W Score", ""]

            header_2.append("hc_")
            header_3.append("hc_Total Score")
            header_3.append("hc_W Score")

            technical_start = "H12"
            technical_score_end = 8 + 2
            technical_score_end_position = "I12"
            merge_technical = technical_start + ":" + technical_score_end_position
            category_worksheet.merge_cells(merge_technical)

            header_1.append("")
            header_2.append("")
            header_3.append("")
            header_2.append("hc_Financial Score")
            header_1.append("")
            header_2.append("hc_")
            header_3.append("hc_Total")
            header_3.append("hc_W Score")

            header_3.append("")
            header_3.append("hc_Comment")

            financial_start = str(excel_cols[(technical_score_end + 2)]) + str(12)
            financial_end = str(excel_cols[(technical_score_end + 3)]) + str(12)
            merge_financial = financial_start + ":" + financial_end
            category_worksheet.merge_cells(merge_financial)

            category_worksheet.append(tuple(header_1))
            category_worksheet.append(tuple(header_2))
            category_worksheet.append(tuple(header_3))

            # if category is not None:
            ids_to_exclude = []
            count = 1
            tender_totals = SupplierCategoryScore.objects.filter(category_id=category.id)
            for supplier in successful_participants:
                supplier_tender_total = tender_totals.filter(supplier_id=supplier.id).first()
                supplier_row = []
                category_rank = supplier_tender_total.rank
                if category_rank != 0:
                    rank = category_rank
                else:
                    rank = "N/A"
                supplier_row.extend(
                    [
                        "", "Ac_{}".format(count),
                        supplier.company_name.title(), "Ac_{}".format(rank), "",
                        "Tc_{}".format(
                            supplier_tender_total.financial_score + supplier_tender_total.technical_score
                        ), "",
                    ]
                )

                supplier_row.append(
                    "Tc_{}".format(supplier_tender_total.technical_score)
                )
                supplier_row.append(
                    "Tc_{}".format(supplier_tender_total.technical_score)
                )
                supplier_row.append("")

                total_rfq_amount = 0

                for item in items:
                    supplier_total = 0
                    item_response = ItemResponse.objects.filter(item_id=item.id, supplier_id=supplier.id).first()
                    if item_response is not None:
                        supplier_total = show(item_response.total)
                    else:
                        supplier_total = 0
                    s_total = round(supplier_total, 2)
                    total_rfq_amount += s_total

                supplier_row.append("Tr_{}".format(total_rfq_amount))
                supplier_row.append("Tc_{}".format(supplier_tender_total.financial_score))
                category_worksheet.append(tuple(supplier_row))
                ids_to_exclude.append(supplier.id)
                count += 1

            # not_in_financial = SupplierCategoryScore.objects.filter(
            #     prequal_id=prequal.id
            # ).exclude(supplier_id__in=ids_to_exclude)
            # for supplier in not_in_financial:
            #     ids_to_exclude.append(supplier.supplier_id)
            #     supplier_row = [
            #         "",
            #         "Ac_{}".format(count),
            #         supplier.supplier.short_name.title(),
            #         "Ac_N/A",
            #         "",
            #         f"Tc_{supplier.supplier.total_prequal_score(prequal)}",
            #         "",
            #         "Tc_{}".format(supplier.supplier.tt_prequal_score(prequal.id)),
            #         "Tc_{}".format(supplier.supplier.total_prequal_score(prequal)),
            #         "",
            #         "Tr_0",
            #         "Tc_0",
            #     ]
            #     category_worksheet.append(tuple(supplier_row))
            #     count += 1

            # not_in_technical = SupplierFinancialTotal.objects.filter(rfq_id=rfq.id).exclude(
            #     supplier_id__in=ids_to_exclude
            # )
            # for supplier in not_in_technical:
            #     ids_to_exclude.append(supplier.supplier_id)
            #     total_rfq_amount = 0
            #
            #     for item in rfq_items:
            #         supplier_total = 0
            #         item_response = supplier.supplier.rfq_item_response(item=item)
            #         if item_response is not None:
            #             supplier_total = show(item_response.total)
            #         else:
            #             supplier_total = 0
            #         s_total = round(supplier_total, 2)
            #         total_rfq_amount += s_total
            #
            #     supplier_row = [
            #         "",
            #         "Ac_{}".format(count),
            #         supplier.supplier.short_name.title(),
            #         "Ac_N/A",
            #         "",
            #         f"Tc_{supplier.supplier.total_rfq_score(rfq)}",
            #         "",
            #         "Tc_0",
            #         "Tc_0",
            #         "",
            #         "Tr_{}".format(total_rfq_amount),
            #         "Tc_{}".format(supplier.supplier.total_rfq_score(rfq)),
            #     ]
            #     category_worksheet.append(tuple(supplier_row))
            #     count += 1

            category_order = apps.apps.get_model('core', 'CategoryOrder')
            for order in category_order.objects.filter(
                category_id=category.id, payment_status=category_order.PAID, target__model='tender'
            ).exclude(supplier_id__in=ids_to_exclude):

                supplier_row = [
                    "",
                    "Ac_{}".format(count),
                    order.supplier.company_name.title(),
                    "Ac_N/R",
                ]
                category_worksheet.append(tuple(supplier_row))
                count += 1

            empty_row = [""]
            empty_row = tuple(empty_row)

            # List of items and submissions
            category_worksheet.append(empty_row)
            category_worksheet.append(empty_row)

            category_worksheet.append(("", "Ac_#", "Nmrl_3_Notes"))
            category_worksheet.append(("", "Ac_1", "Nmrl_3_W Score - Weighted Score"))

            category_worksheet.append(
                (
                    "",
                    "Ac_2",
                    "Nmrl_3_Responsive bidders represents the suppliers who provided both technical "
                    "and financial proposals.",
                )
            )
            category_worksheet.append(
                (
                    "",
                    "Ac_3",
                    "Nmrl_3_The total score is a combination of technical and financial scores.",
                )
            )
            category_worksheet.append(
                (
                    "",
                    "Ac_4",
                    "Nmrl_3_N/A means not ranked.",
                )
            )
            category_worksheet.append(
                (
                    "",
                    "Ac_5",
                    "Nmrl_3_N/R means not responsive.",
                )
            )
            category_worksheet.freeze_panes = category_worksheet["E17"]
            self.result += 1
            self.progress_recorder.set_progress(self.result, total_for_progress)
        except Exception as e:
            capture_exception(e)
            print(e)

    def create_technical_worksheet(self, workbook, data):
        try:
            category = data["category"]
            technical_weight = data["technical_weight"]
            financial_weight = data["financial_weight"]
            questions = data["questions"]
            total_for_progress = data["total_for_progress"]
            category_total_bidders = data["category_total_bidders"]
            category_responsive_bidders = data["category_responsive_bidders"]
            percentage_responses = data["percentage_responses"]

            workbook.create_sheet("TECHNICAL REPORT")
            technical_summary = workbook["TECHNICAL REPORT"]

            resp = SupplierResponse.objects.filter(
                question__section__category_id=category.id
            ).values_list("supplier", flat=True).distinct().count()

            category_order = apps.apps.get_model('core', 'CategoryOrder')
            all = category_order.objects.filter(
                category_id=category.id, payment_status=category_order.PAID,
                target__model='tender').count()

            details_summary = [
                ("", "", "", ""),
                ("", f"Bnb_Category: {category.name}", "", "", ""),
                ("", "", "", ""),
                ("", "Number of bidders", "", f"Ac_{category_total_bidders}"),
                ("", "Number of responsive bidders", "",
                    "Ac_{}".format(len(category_responsive_bidders))),
                ("", "Percentage responses", "","Ac_{}".format(percentage_responses)),
                ("", "Weighting(Technical:Financial)", "",
                 f"Ac_{int(technical_weight)}:{int(financial_weight)}",),
                ("", "", "", ""),
            ]

            empty_row = [""]
            empty_row = tuple(empty_row)

            for row in details_summary:
                technical_summary.append(row)

            technical_summary.append(empty_row)

            technical_summary.merge_cells("B2:H2")
            technical_summary.merge_cells("B4:C4")
            technical_summary.merge_cells("B5:C5")
            technical_summary.merge_cells("B6:C6")
            technical_summary.merge_cells("B7:C7")
            category_headers = [
                "", "hc_#", "hl_Bidder", "hc_Rank", "hc_W Score", "hc_Score",
            ]
            # prequalified_bidders = 0

            sections = category.scored_sections

            for section in sections:
                if section.has_child_sections:
                    if section.is_scored:
                        s_total = section.section_score
                        for s in section.child_sections:
                            s_total += s.section_score
                        category_headers.extend([f"hc_{section.short_name}"])
                    else:
                        evaluate_this = False
                        for s in section.child_sections:
                            if s.is_scored:
                                evaluate_this = True
                                break
                        if evaluate_this == True:
                            s_total = 0
                            for s in section.child_sections:
                                s_total += s.section_score
                            category_headers.extend([f"hc_{section.short_name}"])
                elif section.is_scored:
                    category_headers.extend([f"hc_{section.short_name}"])

            for question in questions:
                category_headers.extend([f"hl_{question.short_description}"])
            technical_summary.append(category_headers)

            # participants = prequal.participants
            # sorted_participants = prequal.sorted_participants(participants)
            sorted_participant_scores = SupplierTechnicalScore.objects.filter(
                category_id=category.id).prefetch_related('supplier').order_by('rank')

            ids_to_exclude = []
            number_count = 1
            for sorted_participant_score in sorted_participant_scores:
                ids_to_exclude.append(sorted_participant_score.supplier_id)

                supplier = sorted_participant_score.supplier
                score = sorted_participant_score.score

                supplier_info = [
                    "", f"Ac_{number_count}", supplier.company_name.title(),
                    f"Ac_{sorted_participant_score.rank}",
                    f"Tc_{sorted_participant_score.weighted_score}",
                    f"Tc_{score}",
                ]

                for section in sections:
                    if section.has_child_sections:
                        if section.is_scored:
                            section_total = supplier.tender_section_score(section)
                            for s in section.child_sections:
                                if s.is_scored:
                                    section_total += supplier.tender_section_score(s)
                            supplier_info.extend([f"Tc_{section_total}"])
                        else:
                            evaluate_this = False
                            for s in section.child_sections:
                                if s.is_scored:
                                    evaluate_this = True
                                    break
                            if evaluate_this == True:
                                section_total = supplier.tender_section_score(section)
                                for s in section.child_sections:
                                    if s.is_scored:
                                        section_total += supplier.tender_section_score(
                                            s
                                        )
                                supplier_info.extend([f"Tc_{section_total}"])
                    else:
                        if section.is_scored:
                            section_total = supplier.tender_section_score(section)
                            supplier_info.extend([f"Tc_{section_total}"])

                for question in questions:
                    qa_response = QualityAssuranceResponse.objects.filter(
                        supplier_id=supplier.id, quality_assurance_question__question_id=question.id
                    ).first()
                    if qa_response is not None:
                        supplier_info.extend(
                            [
                                f"Qac_{qa_response.outcome + '(' + qa_response.comment + ')'}"
                            ]
                        )
                    else:
                        supplier_info.extend([" "])
                technical_summary.append(supplier_info)
                number_count += 1

            category_order = apps.apps.get_model('core', 'CategoryOrder')

            for order in category_order.objects.filter(
                category_id=category.id, payment_status=category_order.PAID, target__model='prequalification'
            ).exclude(supplier_id__in=ids_to_exclude):
                supplier_info = [
                    "",
                    f"Ac_{number_count}",
                    order.supplier.company_name.title(),
                    f"Ac_N/R",
                    f"Ac_N/R",
                    f"Ac_N/R",
                ]
                technical_summary.append(supplier_info)
                number_count += 1

            technical_summary.append(("", "", ""))
            technical_summary.append(("", "Ac_#", "Nmrl_3_Notes"))
            technical_summary.append(
                (
                    "",
                    "Ac_1",
                    "Nmrl_3_TW Score : This means the weighted technical score",
                )
            )
            technical_summary.append(
                ("", "Ac_2", "Nmrl_3_BCC - Business compliance and continuity")
            )
            technical_summary.append(
                ("", "Ac_3", "Nmrl_3_Technical - Technical Specifications")
            )
            technical_summary.append(("", "Ac_4", "Nmrl_3_Staff - Staffing"))
            technical_summary.append(
                ("", "Ac_5", "Nmrl_3_Recommendation - Recommendation")
            )
            technical_summary.append(
                ("", "Ac_6", "Nmrl_3_Years in Business - Years in Business")
            )
            technical_summary.append(
                (
                    "",
                    "Ac_7",
                    "Nmrl_3_Financial - Financials and business sustainability",
                )
            )
            technical_summary.append(("", "Ac_8", "Nmrl_3_Declaration - Declaration"))
            technical_summary.append(
                (
                    "",
                    "Ac_9",
                    "Nmrl_3_N/A means not ranked.",
                )
            )
            technical_summary.append(
                (
                    "",
                    "Ac_10",
                    "Nmrl_3_N/R means not responsive.",
                )
            )
            technical_summary.freeze_panes = technical_summary["G11"]
            self.result += 1
            self.progress_recorder.set_progress(self.result, total_for_progress)
            return
        except Exception as e:
            capture_exception(e)
            print(e)
            return

    def create_savings_worksheet(self, workbook, data):
        workbook.create_sheet("SAVINGS REPORT")
        savings_worksheet = workbook["SAVINGS REPORT"]

        try:
            category = data["category"]
            items = data["items"]
            category_total_bidders = data["category_total_bidders"]
            category_responsive_bidders = data["category_responsive_bidders"]
            percentage_responses = data["percentage_responses"]
            empty_row = [""]
            empty_row = tuple(empty_row)
            total_for_progress = data["total_for_progress"]
            lowest_cumm_suppliers_ranked = data["lowest_cumm_suppliers_ranked"]

            details_summary = [
                ("", "", "", ""),
                ("", "Bnb_Savings Report", "", ""),
                ("", f"Bnb_Category: {category.name}", "", "", ""),
                ("", "", "", ""),
                ("", "Invited bidders", "", "Ac_{}".format(category_total_bidders)),
                ("", "Bidder responses", "", f"Ac_{len(category_responsive_bidders)}"),
                ("", "Percentage responses", "", "Ac_{}".format(percentage_responses)),
                ("", "Currency", "", "Ac_{}".format(category.currency.initials)),
                ("", "", "", ""),
            ]

            for row in details_summary:
                savings_worksheet.append(row)
            savings_worksheet.append(empty_row)

            savings_worksheet.merge_cells("B2:C2")
            savings_worksheet.merge_cells("B3:H3")
            savings_worksheet.merge_cells("B5:C5")
            savings_worksheet.merge_cells("B6:C6")
            savings_worksheet.merge_cells("B7:C7")
            savings_worksheet.merge_cells("B8:C8")
            savings_items_header = [
                "", "hc_#", "h_Item Description", "hc_Qty", "h_Specification",
                "h_Current Supplier", "hc_Current Price", "hc_Current Cost",
                "hc_Lowest Cost", "hc_Unit Savings", "hc_Total Savings", "hc_% Savings",
            ]

            savings_worksheet.append(savings_items_header)
            lowest_supplier = lowest_cumm_suppliers_ranked.first()

            j = 1
            for item in items:
                outlier_score = item.outlier_score
                responses = item.responses

                savings_item_row = []
                if item.savings_value is None:
                    item_savings = item.savings(responses)
                else:
                    item_savings = item.savings_value

                if item.eprocure_total_value is None:
                    item_eprocure_total = item.eprocure_total(responses)
                else:
                    item_eprocure_total = (
                        item.eprocure_total_value
                    )  # check on this, likely to change for consolidated

                item_current_price = float(item.current_price)
                rfq_item_current_cost = item.quantity * item_current_price

                if float(rfq_item_current_cost) < float(outlier_score):
                    rfq_item_current_cost = "O_{}".format(f"{rfq_item_current_cost:}")
                else:
                    rfq_item_current_cost = "Tr_{}".format(f"{rfq_item_current_cost:}")

                savings_lowest_supplier = ""
                if lowest_supplier is not None:
                    lowest_supplier_item_response = responses.filter(
                        supplier_id=lowest_supplier.supplier_id
                    ).first()
                else:
                    lowest_supplier_item_response = None

                if item_eprocure_total > item_current_price and item_current_price != 0:
                    item_savings_cell = "Ar_0"
                    savings_lowest_supplier = "Current Supplier"
                else:
                    item_savings_cell = "T_{}".format(item_savings)
                    savings_lowest_supplier = (
                        lowest_supplier.supplier.company_name if lowest_supplier else ""
                    )
                total_savings = item_savings * item.quantity
                try:
                    if item.second_description is None:
                        second_description = "Ac_N/A"
                    else:
                        second_description = item.second_description
                except:
                    second_description = "Ac_N/A"

                savings_item_row.extend(
                    [
                        "", "Ac_{}".format(item.number), item.description,
                        "Ac_{}".format(item.quantity), second_description,
                        "Ac_N/A", "Tr_{}".format(f"{item_current_price:}"),
                        rfq_item_current_cost, f"T_{show(lowest_supplier_item_response.total)}"
                        if lowest_supplier_item_response else "",
                        item_savings_cell, f"T_{total_savings}",
                        "Ar_{}".format(item.percentage_savings(responses)),
                    ]
                )

                savings_item_row = tuple(savings_item_row)
                savings_worksheet.append(savings_item_row)
                j += 1
            totals_row = [
                "", "", "", "", "", "", f"T_=SUM(G12:G{10 + len(items)})",
                f"T_=SUM(H12:H{10 + len(items)})", f"Tr_=SUM(J12:J{10 + len(items)})",
                f"T_=SUM(K12:K{10 + len(items)})", f"T_=SUM(L12:L{10 + len(items)})",
                f"Ac_=SUM(K12:K{10 + len(items)})/SUM(L12:L{10 + len(items)})",
            ]
            savings_worksheet.append(totals_row)
            savings_worksheet.append(empty_row)
            savings_worksheet.append(("", "Ac_#", "Nmrl_3_Notes"))
            savings_worksheet.append(
                ("", "Ac_1",
                 "Nmrl_3_Where the current prices are lower than the lowest price obtained, current price is retained",
                 "", "")
            )
            savings_worksheet.append(
                (
                    "", "Ac_2" if lowest_supplier else "",
                    f"Nmrl_3_The lowest supplier is {lowest_supplier.supplier.company_name.lower()}"
                    if lowest_supplier else "", "", "",
                )
            )
            savings_worksheet.freeze_panes = savings_worksheet["E12"]

            self.result += 1
            self.progress_recorder.set_progress(self.result, total_for_progress)
        except Exception as e:
            capture_exception(e)
            print(e)

    def create_lowest_supplier_worksheet(self, workbook, data):
        try:
            category = data["category"]
            excel_cols = data["excel_cols"]
            items = data["items"]
            category_total_bidders = data["category_total_bidders"]
            category_responsive_bidders = data["category_responsive_bidders"]
            percentage_responses = data["percentage_responses"]
            lowest_cumm_suppliers_ranked = data["lowest_cumm_suppliers_ranked"]
            lowest_cumm_suppliers_opp = data["lowest_cumm_suppliers_opp"]
            total_for_progress = data["total_for_progress"]
            technical_weight = data["technical_weight"]
            financial_weight = data["financial_weight"]

            rfq_current_cost = 0
            rfq_eprocure_cost = 0
            rfq_total_potential_savngs = 0

            summary_rows = [
                ("", "Invited bidders", "", "Ac_{}".format(category_total_bidders)),
                ("","Bidder responses", "", f"Ac_{len(category_responsive_bidders)}",
                ),
                ("", "Percentage responses", "", "Ac_{}".format(percentage_responses)),
                (
                    "",
                    "Weighting(Technical:Financial)",
                    "",
                    f"Ac_{int(technical_weight)}:{int(financial_weight)}",
                ),
                ("", "Currency", "", "Ac_{}".format(category.currency.initials)),
            ]

            workbook.create_sheet("LOWEST COST REPORT")
            lowest_supplier_worksheet = workbook["LOWEST COST REPORT"]

            lowest_supplier_worksheet["B3"] = "Bnb_Category: {}".format(category.name)
            lowest_supplier_worksheet["B2"] = "Bnb_Lowest Cost Report"
            lowest_supplier_worksheet["B4"] = ""

            for row in summary_rows:
                lowest_supplier_worksheet.append(row)

            empty_row = [""]
            empty_row = tuple(empty_row)
            lowest_supplier_worksheet.append(empty_row)

            lowest_supplier_worksheet.merge_cells("B2:C2")
            lowest_supplier_worksheet.merge_cells("B3:H3")
            lowest_supplier_worksheet.merge_cells("B4:C4")
            lowest_supplier_worksheet.merge_cells("B5:C5")
            lowest_supplier_worksheet.merge_cells("B6:C6")
            lowest_supplier_worksheet.merge_cells("B7:C7")
            lowest_supplier_worksheet.merge_cells("B8:C8")
            lowest_supplier_worksheet.merge_cells("B9:C9")

            s = 5
            supplier_totals_row = 13 + len(items)
            lowest_cumm_suppliers_totals = []
            lowest_cumm_suppliers = list(lowest_cumm_suppliers_ranked) + list(
                lowest_cumm_suppliers_opp
            )

            lowest_supplier_sub_headers = ["", "", "", "", "", ""]
            lowest_supplier_rank_headers = ["", "", "", "", "", ""]
            lowest_supplier_headers = [
                "", "hc_#", "h_Item Description", "hc_Qty",
                "hc_Price Validity (Months)", "h_Specification",
            ]

            scores = []
            for count, supplier in enumerate(lowest_cumm_suppliers, start=1):
                row_name = excel_cols[s + count]
                lowest_cumm_suppliers_totals.append(
                    "qT_{}".format(
                        f"= SUM({row_name}14:{row_name}{supplier_totals_row})"
                    )
                )
                lowest_supplier_sub_headers.append(
                    "hc_%s" % supplier.supplier.company_name.title()
                )
                lowest_supplier_headers.append("hc_TOTAL")
                if (
                    not supplier.has_outlier
                    and supplier.technically_qualified
                    and supplier.meets_all_mandatory_requirements
                ):
                    lowest_supplier_rank_headers.append(f"hc_Rank {count}")
                else:
                    lowest_supplier_rank_headers.append(f"hc_N/A")
                scores.append(f"Tcb_{supplier.score}")

            lowest_supplier_headers = tuple(lowest_supplier_headers)
            lowest_supplier_sub_headers = tuple(lowest_supplier_sub_headers)
            lowest_supplier_rank_headers = tuple(lowest_supplier_rank_headers)
            lowest_supplier_worksheet.append(lowest_supplier_rank_headers)
            lowest_supplier_worksheet.append(lowest_supplier_sub_headers)
            lowest_supplier_worksheet.append(lowest_supplier_headers)

            j = 1
            for item in items:
                outlier_score = item.outlier_score
                responses = item.responses

                rfq_current_cost += float(item.current_price)
                rfq_eprocure_cost += item.eprocure_supplier_cost(responses)
                rfq_total_potential_savngs += item.savings(responses)

                try:
                    if item.second_description is None:
                        second_description = "Ac_N/A"
                    else:
                        second_description = item.second_description

                    if item.price_validity_months is None:
                        price_validity = "Ac_{}".format(12)
                    else:
                        price_validity = "Ac_{}".format(item.price_validity_months)

                except:
                    second_description = "Ac_N/A"
                    price_validity = "Ac_{}".format(12)

                lowest_supplier_row = []

                lowest_supplier_row.extend(
                    [
                        "", "Ac_{}".format(item.number), item.description,
                        "Ac_{}".format(item.quantity), price_validity, second_description,
                    ]
                )

                for count, lowest_cumm_supplier in enumerate(
                    lowest_cumm_suppliers, start=1
                ):
                    if not lowest_cumm_supplier:
                        supplier_total = 0
                    else:
                        item_response = ItemResponse.objects.filter(
                            supplier_id=lowest_cumm_supplier.supplier_id, item_id=item.id).first()
                        if item_response is not None:
                            if item_response.total is not None:
                                supplier_total = show(item_response.total)
                        else:
                            supplier_total = 0
                        s_total = round(supplier_total, 2)
                        if float(s_total) < float(outlier_score):
                            supplier_total = "O_{}".format(f"{s_total:}")
                        else:
                            supplier_total = "Tr_{}".format(f"{s_total:}")
                        lowest_supplier_row.append(supplier_total)

                lowest_supplier_row = tuple(lowest_supplier_row)
                lowest_supplier_worksheet.append(lowest_supplier_row)

                j += 1

            lowest_supplier_totals = []
            lowest_supplier_totals.extend(["", "", "qT_Total", "", "", ""])
            weighted_financial_scores = [
                "", "", "qT_Weighted Financial Score", "", "", "",
            ]
            weighted_financial_scores.extend(scores)
            lowest_supplier_totals.extend(lowest_cumm_suppliers_totals)
            lowest_supplier_worksheet.append(tuple(lowest_supplier_totals))
            lowest_supplier_worksheet.append(empty_row)
            lowest_supplier_worksheet.append(weighted_financial_scores)
            lowest_supplier_worksheet.append(empty_row)
            lowest_supplier_worksheet.append(("", "Ac_#", "Nmrl_3_Notes"))
            lowest_supplier_worksheet.append(
                (
                    "",
                    "Ac_1",
                    "Nmrl_3_Rank: This is the ranking of suppliers with the lowest cost and is not an outlier",
                )
            )
            lowest_supplier_worksheet.append(
                (
                    "",
                    "Ac_2",
                    "Nmrl_3_Outlier: The prices highlighted in this colour are "
                    "outliers which means that they are 50% lower than the median "
                    "and are not considered in the analysis.",
                )
            )
            lowest_supplier_worksheet.append(
                ("", "Ac_4", "Nmrl_3_N/A means not ranked.")
            )
            lowest_supplier_worksheet.append(
                ("", "Ac_5", "Nmrl_3_N/R means not responsive.")
            )
            lowest_supplier_worksheet.freeze_panes = lowest_supplier_worksheet["G13"]

            self.result += 1
            self.progress_recorder.set_progress(self.result, total_for_progress)
            return
        except Exception as e:
            capture_exception(e)
            print(e)

    # def create_individual_worksheets(self, workbook, data):
    #     try:
    #         total_for_progress = data["total_for_progress"]
    #         participants = data["participants"]
    #         rfq = data["rfq"]
    #
    #         for count, supplier in enumerate(participants, start=1):
    #             try:
    #                 workbook.create_sheet(f"{supplier.company_name.upper()}")
    #                 supplier_quotation = workbook[f"{supplier.company_name.upper()}"]
    #                 supplier_workbook = load_workbook(
    #                     supplier.relative_quotation_url(rfq)
    #                 )
    #                 bid_worksheet = supplier_workbook["RFQ ITEMS"]
    #
    #                 for row in bid_worksheet:
    #                     for cell in row:
    #                         quote_cell = supplier_quotation[cell.coordinate]
    #                         quote_cell.value = cell.value
    #                         if cell.has_style:
    #                             quote_cell.font = copy(cell.font)
    #                             quote_cell.border = copy(cell.border)
    #                             quote_cell.fill = copy(cell.fill)
    #                             quote_cell.number_format = copy(cell.number_format)
    #                             quote_cell.protection = copy(cell.protection)
    #                             quote_cell.alignment = copy(cell.alignment)
    #             except:
    #                 pass
    #
    #         self.result += 1
    #         self.progress_recorder.set_progress(self.result, total_for_progress)
    #         return
    #     except Exception as e:
    #         capture_exception(e)


class TenderQARankingReport(Task):
    name = "TenderQARankingReport"

    def run(self, *args, **kwargs):
        context = self.qa_ranking_report(kwargs["job_id"])
        return context

    def supplier_total_section_score(self, parent_section, section_scores):
        t = 0
        if parent_section.is_scored:
            score = section_scores.filter(section_id=parent_section.id).first()
            if score is not None:
                t += score.score_after_qa
            if parent_section.has_child_sections:
                for child_section in parent_section.child_sections:
                    t += self.supplier_total_section_score(child_section, section_scores)
        return t

    def questions_short_name(self):
        data = {
            "Tax compliance": "TCC", "PIN certificate": "PIN", "Certificate of Incorporation": "COI",
            "CR12 Certificate or equivalent": "CR12", "NCA certificate": "NCA-C", "NCA license": "NCA-L",
            "Institute of Certified Public Accountants of Kenya Certificate": "ICPAK",
            "Institute of  Certified Public Secretaries of Kenya": "ICPSK", "Registration certificate": "COI",
            "Energy Petroleum Regulatory Authority License": "EPRA", "County government business permit": "BP",
            "Pharmacy and poisons board License": "PPB", "Insurance and Regulatory Authority License": "IRA",
            "National identification number": "ID", "Police Clearance Certificate": "PCC",
            "Water Services Regulatory Board License": "WASREB",
            "National Industrial Training Authority Certificate": "NITA",
            "Certified Public Accountants of Kenya Certificate": "CPAK",
            "Association of Chartered Certified Accountants Ceritificate": "ACCA",
            "Law Society of Kenya Registration Certificate": "LSK",
            "National Transport and Safety Authority Registration Ceritificate": "NTSA",
            "Marketing Society of Kenya Certificate": "MSK",
            "Association of Practitioners in Advertising Certificate": "APA",
            "International Air Transport Association License": "IATA-L",
            "International Air Transport Association Certificate": "IATA-C",
            "Kenya Association of Travel Agents Certificate": "KATA", "Engineers Board of Kenya License": "EBK-L",
            "Engineers Board of Kenya certificate": "EBK-C", "Institution of Engineers of Kenya Certificate": "IEK",
            "Kenya Engineering Technology Registration Board License": "KETRB",
            "Board of Registration of Architects and Quantity Surveyors Certificate": "BORAQS-C",
            "Board of Registration of Architects and Quantity Surveyors License": "BORAQS-L",
            "Architectural Association of Kenya Certificate": "AAK",
            "Institute of Quantity Surveyors of Kenya Certificate": "IQSK",
            "Institute of Human Resources Management Certificate": "IHRM",
            "Directorate Of Occupational Safety And Health Services Certificate": "DOSH",
            "Marketing and Social Research Association Certificate": "MSRA",
            "Work Injury Benefits Act": "WIBA", "National Environmental Management Authority Certificate": "NEMA-C",
            "National Environmental Management Authority License": "NEMA-L", "Environmental Impact Assessment": "EIA",
            "Kenya Bureau of Standards": "KEBS", "Energy Regulatory Commission License": "ERC",
            "Institution of Surveyors of Kenya": "ISK", "Kenya Film Classification Board": "KFCB",
            "Estate Agents Registration Board": "EARB", "Valuers Registration Board": "VRB",
            "Authorised Economic Operator": "AEO",
            "Institute of Certified Safety and Security Proffesionals of Kenya": "ICSSPK",
            "Kenya Security Industry Association": "KSIA", "Protective Security Industry Association": "PSIA",
            "Private Security Regulatory Authority": "PSRA", "Kenya Animal Genetic Resources Centre": "KAGRC",
            "Kenya Veterinary Board": "KVB"

        }
        return data

    def qa_ranking_report(self, job_id):
        try:
            progress_recorder = ProgressRecorder(self)
            result = 0

            job = Tender.objects.get(id=job_id)
            company = job.company

            categories = Category.objects.filter(
                tender_id=job.id).only("id", "unique_reference", "name").order_by("unique_reference")

            time = datetime.datetime.now()
            dir_name = Path("media/prequal/job/reports/{}/{}/{}".format(time.year, time.month, time.day))

            dir_name.mkdir(parents=True, exist_ok=True)
            filepath = f"{dir_name}/{job_id}_QA_Ranking_Report_{time.hour}{time.minute}{time.second}.xlsx"

            workbook = Workbook()
            summary_sheet = workbook["Sheet"]
            summary_sheet.title = "Summary"

            buyer_logo_anchor = "B1"

            if company.company_logo_url is not None and company.company_logo_url != "":
                buyer_logo_url = get_file_path(company.company_logo_url)
            else:
                buyer_logo_url = "static/core/img/no-company-image128-128.png"

            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(
                (
                    " ", "hc_#", "hc_Category Code", "hl_Category Title", "hhc_Responsive",
                    "hhc_Non-Responsive", "hhc_Total Bidders", "hhc_Prequalified Bidders", "hhc_Letters",
                )
            )

            data = {
                "Summary_B": 5, "Summary_C": 21, "Summary_D": 41, "Summary_E": 6, "Summary_F": 6,
                "Summary_G": 6, "Summary_H": 6, "Summary_I": 6, "QA_OTHER_B": 41, "QA_OTHER_C": 8,
                "QA_OTHER_D": 8, "QA_OTHER_E": 8, "QA_OTHER_F": 8,
            }

            # totals
            total_responsive = 0
            total_non_responsive = 0
            total_bidders = 0
            total_prequalified = 0
            total_letters = 0
            total_category = categories.count()
            category_count = 1
            for category in categories:
                data[f"{category.unique_reference.replace('/', '_')}_C"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_D"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_E"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_F"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_G"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_H"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_I"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_J"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_K"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_L"] = 5
                # data[f"{category.unique_reference.replace('/', '_')}_M"] = 8
                data[f'{category.unique_reference.replace("/", "_")}_B'] = 31

                result += 1
                progress_recorder.set_progress(result, total_category)
                resp = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                    id__in=SupplierResponse.objects.filter(
                    question__section__category_id=category.id).only('supplier_id').values('supplier_id').distinct()
                ).count()

                all = apps.apps.get_model('core', 'CategoryOrder').objects.filter(
                    category_id=category.id, payment_status=apps.apps.get_model('core', 'CategoryOrder').PAID, 
                    target__model='category', 
                    target__app_label='tender').count()
                non_resp = all - resp

                if non_resp < 0:
                    non_resp = 0

                audited_accounts_question = Question.objects.filter(
                    description="For limited liability companies, attach audited accounts for the last two years, for "
                                "sole proprietors and partnerships, attach your most recent management accounts",
                    section__category_id=category.id).first()

                summary_sheet.append(
                    (
                        " ", f"Ac_{category_count}", f"Al_{category.unique_reference}",
                        f"Al_{category.name}",
                        f"Nc_{resp}",
                        f"Nc_{non_resp}",
                        f"Nc_{all}",
                        f"Nc_{category.suppliers_in_qa}",
                        f"Nc_{AwardLetter.objects.filter(category_id=category.id).count()}",
                    )
                )
                try:
                    total_responsive += resp
                    total_non_responsive += non_resp
                    total_bidders += all
                    total_prequalified += category.suppliers_in_qa
                    total_letters += AwardLetter.objects.filter(
                        category_id=category.id
                    ).count()
                except:
                    total_responsive = 0
                    total_non_responsive = 0
                    total_bidders = 0
                    total_prequalified = 0
                    total_letters = 0

                category_type = category.category_type
                category_count += 1

                category_worksheet = workbook.create_sheet(category.unique_reference.replace("/", "_"))
                category_worksheet.append(("", ""))
                category_worksheet["B2"] = f"Bnb_Category: {category.name}"
                category_worksheet["B3"] = ""
                category_worksheet["B4"] = "Number of bidders"
                category_worksheet["C4"] = f"Nc_{all}"
                category_worksheet["B5"] = "Number of pre-qualified bidders"
                category_worksheet["B6"] = "Number of responsive bidders"
                category_worksheet["C6"] = f"Nc_{resp}"
                category_worksheet["B7"] = "Pass mark"
                category_worksheet["C7"] = f"Tcnd_{category.pass_score}"
                category_worksheet.append(("", ""))

                category_headers = [
                    " ", "hl_Bidder", "hhc_Rank After QA", "hhc_Score After QA",
                    "hhc_Rank Before QA", "hhc_Score Before QA",
                ]

                category_worksheet.merge_cells("B2:H2")
                prequalified_bidders = 0

                short_descriptions_to_exclude = [
                    "Debt/equity ratio", "NP margin", "GP margin", "Cash ratio", "Current ratio"]

                questions = Question.objects.filter(
                    section__category_id=category.id, is_qa=True).exclude(
                    short_description__in=short_descriptions_to_exclude).only(
                    "id", "short_description", "description")

                sections = Section.objects.filter(
                    category_id=category.id, parent_section__isnull=True).order_by("id")

                self.append_category_headers(sections, category_headers, questions, category_worksheet)
                responsive_bidders = Supplier.objects.filter(
                    id__in=SupplierTechnicalScore.objects.filter(
                    category_id=category.id).order_by("rank_after_qa").only("supplier_id").values('supplier_id')
                )

                for supplier in responsive_bidders:
                    supplier_technical_response = SupplierTechnicalScore.objects.filter(
                            supplier_id=supplier.id, category_id=category.id).first()
                    section_scores = SupplierSectionScore.objects.filter(
                        supplier_id=supplier.id, section__category_id=category.id)

                    technical_score = supplier_technical_response.score

                    if technical_score < category.pass_score:
                        score_after_qa = "Ac_N/A"
                    else:
                        prequalified_bidders += 1
                        score_after_qa = (
                            f"Tcnd_{supplier_technical_response.score_after_qa}"
                            if supplier_technical_response.score_after_qa is not None
                            else "Ac_N/A"
                        )

                    c_data = {
                        "supplier": supplier, "supplier_technical_response": supplier_technical_response,
                        "sections": sections, "technical_score": technical_score,
                        "score_after_qa": score_after_qa, "section_scores": section_scores,
                        "category": category
                    }


                    if technical_score >= category.pass_score:
                        # get section score for supplier over pass score
                        supplier_info = self.section_score_for_supplier_over_pass_score(data=c_data)
                    else:
                        # get section score for supplier under pass score
                        supplier_info = self.section_score_for_supplier_under_pass_score(data=c_data)

                    for question in questions:
                        qa_response = QualityAssuranceResponse.objects.filter(
                                supplier_id=supplier.id, quality_assurance_question__question_id=question.id,
                            ).only("id", "outcome", "comment").first()

                        if qa_response is not None:
                            supplier_info.extend([f"Qac_{qa_response.outcome} ({qa_response.comment})"])
                        else:
                            supplier_info.extend([" "])

                    try:
                        location = "N/A"
                        pass
                    except:
                        location = "N/A"

                    supplier_info.extend(
                        [
                            f"Al_{supplier.contact_name.title()}", f"Ar_{supplier.phone_number}", f"Al_{supplier.email}",
                            f"{ILLEGAL_CHARACTERS_RE.sub(r'', str(location))}",
                        ]
                    )


                    rank_comment = ""
                    if not technical_score < float(category.pass_score):
                        section_score_instances = SupplierSectionScore.objects.filter(
                            supplier_id=supplier.id, section__category_id=category.id).prefetch_related('section')
                        for instance in section_score_instances:
                            if instance.score_after_qa != instance.score and instance.score_after_qa is not None:
                                difference = instance.score_after_qa - instance.score
                                if difference > 0:
                                    difference = f"+{difference}"
                                rank_comment = rank_comment + f"{instance.section.short_name}: {difference}, "

                    if audited_accounts_question:
                        ac_supplier_response = SupplierResponse.objects.filter(
                            question_id=audited_accounts_question.id, supplier_id=supplier.id).first()
                        if ac_supplier_response is None:
                            rank_comment = rank_comment + "Audited Accounts Not Attached, "
                        elif ac_supplier_response is not None:
                            if ac_supplier_response.options is None and ac_supplier_response.document_url is None:
                                rank_comment = rank_comment + "Audited Accounts Not Attached, "

                    if rank_comment != "":
                        rank_data = supplier_info[3]
                        supplier_info[3] = f"{rank_data}_{rank_comment}"

                    category_worksheet.append(supplier_info)
                # Append number of prequalified bidders
                category_worksheet["C5"] = f"Nc_{prequalified_bidders}"

                # non responsive bidders
                non_responsive_bidders = Supplier.objects.filter(
                    id__in=apps.apps.get_model('core', 'CategoryOrder').objects.filter(
                        category_id=category.id, payment_status=apps.apps.get_model('core', 'CategoryOrder').PAID,
                        target__model='category', target__app_label='tender').exclude(
                        supplier_id__in=[r.id for r in responsive_bidders]).only("supplier_id")
                )

                for supplier in non_responsive_bidders:
                    supplier_info = [
                        "",
                        supplier.company_name.title(),
                        "Ac_N/R",
                        "Ac_N/R",
                        "Ac_N/R",
                        "Ac_N/R",
                    ]

                    for section in sections:
                        if section.has_child_sections:
                            if section.is_scored:
                                supplier_info.extend(["Ac_N/R"])
                            else:
                                evaluate_this = False
                                for s in section.child_sections:
                                    if s.is_scored:
                                        evaluate_this = True
                                        break
                                if evaluate_this == True:
                                    supplier_info.extend(["Ac_N/R"])
                        else:
                            if section.is_scored:
                                supplier_info.extend(["Ac_N/R"])

                    for question in questions:
                        qa_response = (
                            QualityAssuranceResponse.objects.filter(
                                supplier_id=supplier.id,quality_assurance_question__question_id=question.id,
                            ).only("id", "outcome", "comment").first()
                        )
                        if qa_response is not None:
                            supplier_info.extend(["Ac_N/R"])
                        else:
                            supplier_info.extend(["Ac_N/R"])
                    supplier_info.extend(["Ac_N/R", "Ac_N/R", "Ac_N/R", "Ac_N/R"])
                    category_worksheet.append(supplier_info)

                # append notes
                category_worksheet.append(("", ""))
                category_worksheet.append(("", "Notes"))
                category_worksheet.append(("", "N/A - Not Applicable"))
                category_worksheet.append(("", "N/R - None responsive"))
                category_worksheet.freeze_panes = category_worksheet["D10"]

            summary_sheet.append(
                (
                    " ", "", "", "qT_Total", "Ac_{}".format(total_responsive),
                    "Ac_{}".format(total_non_responsive), "Ac_{}".format(total_bidders),
                    "Ac_{}".format(total_prequalified), "Ac_{}".format(total_letters),
                )
            )

            summary_sheet.freeze_panes = summary_sheet["D10"]
            workbook.save(filepath)
            format_excel(filepath, data=data)

            # insert_image(excel_url=filepath, worksheet_name="Summary", anchor=qed_logo_anchor)
            try:
                insert_image(
                    excel_url=filepath, worksheet_name="Summary", anchor=buyer_logo_anchor, image_url=buyer_logo_url)
            except:
                insert_image(
                    excel_url=filepath, worksheet_name="Summary", anchor=buyer_logo_anchor, image_url="",
                )

            try:
                time = datetime.datetime.now()
                with open(filepath, "rb") as l:
                    storage = PrivateMediaStorage()
                    url = f"prequal/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
                    storage.save(url, l)

                    job_report, created = JobReport.objects.update_or_create(
                        tender_id=job_id, defaults={"qa_ranking_report": url}
                    )
                    path = job_report.qa_ranking_report.url
                    context = {
                        "filepath": path,
                        "response_message": "Report generated successfully",
                    }
                    return context
            except Exception as e:
                print(e)
                capture_exception(e)
                context = {"response_message": "Error generating qa ranking report"}
                return context
        except Exception as e:
            print(str(e))
            # capture_exception(e)
            context = {"response_message": "Error generating qa ranking report", "errors": f"{str(e)}"}
            return context

    def append_category_headers(self, prequal_sections, category_headers, questions, category_worksheet):
        question_short_names = self.questions_short_name()
        for section in prequal_sections:
            if section.has_child_sections:
                if section.is_scored:
                    s_total = section.section_score
                    for s in section.child_sections:
                        if s.has_child_sections:
                            for x in s.child_sections:
                                s_total += x.section_score
                        else:
                            s_total += s.section_score
                    category_headers.extend(
                        [f"hhc_{section.short_name}({int(s_total)})"]
                    )
                else:
                    evaluate_this = False
                    for s in section.child_sections:
                        if s.has_child_sections:
                            for x in s.child_sections:
                                if x.is_scored:
                                    evaluate_this = True
                                    break
                        else:
                            if s.is_scored:
                                evaluate_this = True
                                break
                    if evaluate_this == True:
                        s_total = 0
                        for s in section.child_sections:
                            if s.has_child_sections:
                                for x in s.child_sections:
                                    s_total += x.section_score
                            else:
                                s_total += s.section_score
                        category_headers.extend(
                            [f"hhc_{section.short_name}({int(s_total)})"]
                        )
            elif section.is_scored:
                category_headers.extend(
                    [
                        f"hhc_{section.short_name}({int(section.section_score)})"
                    ]
                )

        for question in questions:
            if question.short_description is not None:
                if question.short_description in question_short_names:
                    category_headers.extend(
                        [f"hl_{question_short_names[question.short_description]}"]
                    )
                else:
                    category_headers.extend(
                        [f"hl_{question.short_description}"]
                    )
            else:
                category_headers.extend([f"hl_{question.description}"])

        category_headers.extend(
            ["hl_Contact Person", "hl_Phone Number", "hl_Email Address", "hl_Main Location", ]
        )
        category_worksheet.append(category_headers)
        category_worksheet.merge_cells("B2:H2")

        return

    def section_score_for_supplier_over_pass_score(self, data):
        print("got here")
        sections = data["sections"]
        supplier = data["supplier"]
        supplier_technical_response = data["supplier_technical_response"]
        technical_score = data["technical_score"]
        score_after_qa = data["score_after_qa"]
        category = data["category"]
        section_scores = data["section_scores"]

        supplier_info = [
            "",
            supplier.short_name if supplier.short_name is not None else supplier.company_name, 
            f"Ac_{supplier_technical_response.rank_after_qa}"
            if supplier_technical_response.rank_after_qa and not technical_score < category.pass_score
            else "Ac_N/A",
            score_after_qa, f"Nc_{supplier_technical_response.rank}", f"Tcnd_{technical_score}",
        ]
        for section in sections:
            if section.is_parent_section_scored:
                section_total = self.supplier_total_section_score(section, section_scores)
                supplier_info.extend([f"Tcnd_{section_total}"])
        print(supplier_info)
        return supplier_info

    def section_score_for_supplier_under_pass_score(self, data):
        sections = data["sections"]
        supplier = data["supplier"]
        supplier_technical_response = data["supplier_technical_response"]
        technical_score = data["technical_score"]
        score_after_qa = data["score_after_qa"]
        category = data["category"]
        section_scores = data["section_scores"]

        supplier_info = [
            "",
            supplier.short_name if supplier.short_name is not None else supplier.company_name,
            f"Ac_{supplier_technical_response.rank_after_qa}"
            if supplier_technical_response.rank_after_qa and not technical_score < category.pass_score
            else "Ac_N/A",
            score_after_qa, 
            f"Nc_{supplier_technical_response.rank}", f"Tcnd_{technical_score}",
        ]
        print(supplier_info)
        for section in sections:
            if section.is_parent_section_scored:
                section_total = self.supplier_total_section_score(section, section_scores)
                supplier_info.extend([f"Tcnd_{section_total}"])
        print(supplier_info)
        return supplier_info




class DueDiligenceRankingReport(Task):
    name = "TenderDueDiligenceRankingReport"

    def run(self, *args, **kwargs):
        context = self.report(kwargs["job_id"])
        return context

    def report(self, job_id):
        try:
            progress_recorder = ProgressRecorder(self)
            result = 0

            data = {}
            job = Tender.objects.get(id=job_id)
            categories = Category.objects.filter(tender_id=job_id)
            total_for_progress = categories.count()
            company = job.company

            time = datetime.datetime.now()
            dir_name = Path("media/job/reports/{}/{}".format(time.year, time.month))
            dir_name.mkdir(parents=True, exist_ok=True)
            # time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
            filepath = "{}/{}_DueDiligence_Report.xlsx".format(
                dir_name, job_id
            )

            workbook = Workbook()
            summary_sheet = workbook.get_sheet_by_name("Sheet")
            summary_sheet.title = "Summary"

            qed_logo_anchor = "G2"
            buyer_logo_anchor = "B2"

            if company.company_logo_url is not None and company.company_logo_url != "":
                buyer_logo_url = get_file_path(company.company_logo_url)
            else:
                buyer_logo_url = "static/core/img/no-company-image128-128.png"

            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))

            summary_sheet.append(
                (
                    " ", "hc_#", "hc_Category No", "hc_Category Title",
                    "hhc_Responsive", "hhc_Non-Responsive", "hhc_Total Bidders",
                    "hhc_Prequalified Bidders QA", "hhc_Due Diligence", "hhc_Success/Regret Letters",
                )
            )

            additional_data = {
                "Summary_C": 21, "Summary_D": 62, "Summary_E": 6.2, "Summary_F": 6.2,
                "Summary_G": 6.2, "Summary_H": 6.2, "Summary_I": 6.2, "Summary_J": 6.2
            }
            data.update(additional_data)

            responsive_bidders = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                id__in=SupplierResponse.objects.filter(
                    question__section__category__tender_id=job_id).only('supplier_id').values('supplier_id')
            )
            category_order = apps.apps.get_model('core', 'CategoryOrder')
            paid_bidders = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                id__in=category_order.objects.filter(
                    payment_status=category_order.PAID, job_id=job_id, target__model='tender'
                ).only('supplier_id').values('supplier_id')
            )
            non_responsive_bidders = paid_bidders.difference(responsive_bidders)

            responsive_bidders_count = responsive_bidders.count()
            non_responsive_bidders_count = paid_bidders.count() - responsive_bidders.count()
            total_bidders = paid_bidders.count()

            suppliers_in_qa = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                id__in=QualityAssuranceResponse.objects.filter(
                    quality_assurance_question__quality_assurance__category__tender_id=job_id
                ).only('supplier_id').values('supplier_id')
            ).count()

            dd_suppliers = DueDiligenceSupplier.objects.filter(
                due_diligence__category__tender_id=job_id).count()

            qa_and_dd_suppliers = (AwardLetter.objects.filter(category__tender_id=job_id).count()
                                   + RegretLetter.objects.filter(category__tender_id=job_id).count())

            section_scores = SupplierSectionScore.objects.filter(section__category__tender_id=job_id)
            category_count = 1
            for category in categories:
                result += 1
                progress_recorder.set_progress(result, total_for_progress)

                summary_sheet.append(
                    (
                        "", category_count, category.unique_reference, category.name,
                        responsive_bidders_count, non_responsive_bidders_count,
                        total_bidders, suppliers_in_qa, dd_suppliers, qa_and_dd_suppliers
                    )
                )
                category_count += 1

                category_worksheet = workbook.create_sheet(
                    category.unique_reference.replace("/", "_")
                )
                category_worksheet.append((" ", " "))
                category_worksheet["B2"] = f"Category: {category.name}"
                category_worksheet["B3"] = "Number of bidders"
                category_worksheet["C3"] = f"Ac_{category.total_bidders}"
                category_worksheet["B4"] = "Number of pre-qualified bidders"
                category_worksheet["C4"] = f"Ac_{category.count_qualified_bidders}"
                category_worksheet["B5"] = "Number of responsive bidders"
                category_worksheet["C5"] = f"Ac_{category.responsive_bidder_count}"
                category_worksheet["B6"] = "Pass mark"
                category_worksheet["C6"] = f"Tcnd_{category.pass_score}"
                category_worksheet.append(("", ""))

                main_headers = ["", "", "hc_Prequalification Results", "", ""]

                category_headers = [
                    " ", "hc_Bidder", "hhc_Rank", "hhc_Total", "hhc_Score After QA",
                ]

                data[f'{category.unique_reference.replace("/", "_")}_B'] = 36

                questions = Question.objects.filter(
                    section__category_id=category.id, is_qa=True).order_by("id")

                sections = category.sections
                no_of_sections = 0
                for section in sections:
                    if section.is_scored:
                        no_of_sections += 1
                        category_headers.append(f"hhc_{section.short_name}")
                        main_headers.append("")
                main_headers.append("hc_Quality Assurance")

                no_of_questions = 0
                for question in questions:
                    no_of_questions += 1
                    category_headers.append(f"hhc_{question.short_description}")
                    main_headers.append("")
                main_headers.append("hhc_Tender Committee")

                category_worksheet.append(main_headers)
                category_worksheet.append(category_headers)
                category_worksheet.merge_cells(
                    start_row=8, start_column=3,
                    end_row=8, end_column=5 + no_of_sections,
                )

                category_worksheet.merge_cells(
                    start_row=8, start_column=6 + no_of_sections,
                    end_row=8, end_column=6 + no_of_sections + no_of_questions,
                )

                sorted_participants = apps.apps.get_model('prequal', 'SupplierCategoryScore').objects.filter(
                    category_id=category.id).prefetch_related('supplier').order_by("rank")

                for sorted_participant in sorted_participants:
                    p_score = sorted_participant.score
                    s_qa = sorted_participant.score_after_qa
                    supplier = sorted_participant.supplier
                    supplier_info = [
                        " ",
                        supplier.company_name.upper(),
                        f"Ac_{sorted_participant.rank}",
                        f"Tc_{p_score}" if p_score is not None else " ",
                        f"Tc_{s_qa}" if s_qa is not None else " ",
                    ]

                    for section in sections:
                        section_score = section_scores.filter(id=section.id, supplier_id=supplier.id).first()
                        s_score = section_score.score if section_score is not None else 0
                        if section.is_scored:
                            supplier_info.extend(
                                [f"Tcnd_{s_score}" if s_score is not None else " "]
                            )

                    for question in questions:
                        qa_response = QualityAssuranceResponse.objects.filter(
                            supplier=supplier,
                            quality_assurance_question__question=question,
                        ).first()
                        if qa_response is not None:
                            supplier_info.extend(
                                [qa_response.outcome + "(" + qa_response.comment + ")"]
                            )
                        else:
                            supplier_info.extend([" "])
                    category_worksheet.append(supplier_info)
                category_worksheet.freeze_panes = category_worksheet["F10"]
            summary_sheet.freeze_panes = summary_sheet["E7"]
            workbook.save(filepath)
            format_excel(filepath, data=data)
            insert_image(
                excel_url=filepath, worksheet_name="Summary", anchor=qed_logo_anchor
            )

            try:
                insert_image(
                    excel_url=filepath, worksheet_name="Summary",
                    anchor=buyer_logo_anchor, image_url=buyer_logo_url,
                )
            except:
                insert_image(
                    excel_url=filepath, worksheet_name="Summary",
                    anchor=buyer_logo_anchor, image_url="",
                )

            with open(filepath, "rb") as l:
                storage = PrivateMediaStorage()
                url = f"tender/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
                storage.save(url, l)

                job_report, created = JobReport.objects.update_or_create(
                    tender_id=job_id, defaults={"dd_ranking_report": url}
                )
                path = job_report.dd_ranking_report.url
                context = {
                    "filepath": path,
                    "response_message": "Report generated successfully",
                }
                return context
        except Exception as e:
            capture_exception(e)
            print(e)
            context = {
                "response_message": "Report generation error",
                "messages": [f"{e}",]
            }
            return context


class TenderInterimReport(Task):
    """
    Interim prequal report before QA
    """

    name = "TenderInterimReport"
    progress_recorder = None
    result = 0

    def run(self, *args, **kwargs):
        self.progress_recorder = ProgressRecorder(self)
        context = self.report(kwargs["job_id"])
        return context

    def report(self, job_id):
        try:
            job = Tender.objects.filter(id=job_id).first()
            company = job.company

            categories = Category.objects.filter(
                tender_id=job.id).only("id", "unique_reference", "name").order_by("unique_reference")

            time = datetime.datetime.now()
            dir_name = Path(
                "media/job/reports/{}/{}/{}".format(time.year, time.month, time.day)
            )
            dir_name.mkdir(parents=True, exist_ok=True)
            time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
            filepath = "{}/{}_Interim_Technical_Report.xlsx".format(
                dir_name, job_id
            )

            workbook = Workbook()
            summary_sheet = workbook["Sheet"]
            summary_sheet.title = "Summary"

            qed_logo_anchor = "H1"
            buyer_logo_anchor = "B1"

            if company.company_logo_url is not None and company.company_logo_url != "":
                buyer_logo_url = get_file_path(company.company_logo_url)
            else:
                buyer_logo_url = f"static/core/img/no-company-image128-128.png"

            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(
                (
                    " ", "hc_#", "hc_Category Code", "hl_Category Title", "hhc_Responsive",
                    "hhc_Non-Responsive", "hhc_Total Bidders",
                )
            )

            data = {
                "Summary_B": 5, "Summary_C": 21, "Summary_D": 41,
                "Summary_E": 6, "Summary_F": 6, "Summary_G": 6,
            }

            progress_recorder = ProgressRecorder(self)
            result = 0

            # totals
            total_responsive = 0
            total_non_responsive = 0
            total_bidders = 0
            total_prequalified = 0
            total_letters = 0

            category_count = 1
            for category in categories:
                data[f"{category.unique_reference.replace('/', '_')}_C"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_D"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_E"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_F"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_G"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_H"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_I"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_J"] = 5
                data[f"{category.unique_reference.replace('/', '_')}_K"] = 18
                data[f"{category.unique_reference.replace('/', '_')}_L"] = 18
                data[f'{category.unique_reference.replace("/", "_")}_B'] = 31

                result += 1
                progress_recorder.set_progress(result, categories.count())
                resp = (
                    SupplierResponse.objects.filter(
                        question__section__category_id=category.id
                    )
                    .only('supplier_id').values('supplier_id').distinct()
                    .count()
                )
                category_order = apps.apps.get_model('core', 'CategoryOrder')
                all = category_order.objects.filter(
                    category_id=category.id, payment_status=category_order.PAID, target__model='tender'
                ).count()

                non_resp = all - resp
                if non_resp < 0:
                    non_resp = 0

                summary_sheet.append(
                    (
                        " ", f"Ac_{category_count}", f"Al_{category.unique_reference}",
                        f"Al_{category.name}", f"Nc_{resp}", f"Nc_{non_resp}", f"Nc_{all}",
                    )
                )
                try:
                    total_responsive += resp
                    total_non_responsive += non_resp
                    total_bidders += all
                except:
                    total_responsive = 0
                    total_non_responsive = 0
                    total_bidders = 0

                category_type = category.category_type

                category_count += 1
                category_worksheet = workbook.create_sheet(
                    category.unique_reference.replace("/", "_")
                )
                category_worksheet.append(("", ""))
                category_worksheet["B2"] = f"Bnb_Category: {category.name}"
                category_worksheet["B3"] = ""
                category_worksheet["B4"] = "Number of bidders"
                category_worksheet["C4"] = f"Nc_{all}"
                category_worksheet["B5"] = "Number of responsive bidders"
                category_worksheet["C5"] = f"Nc_{resp}"
                category_worksheet["B6"] = "Proposed pass mark"
                category_worksheet["C6"] = f"Tcnd_{category.pass_score}"
                category_worksheet.append(("", ""))

                category_headers = [" ", "hl_Bidder", "hhc_Rank", "hhc_Score Before QA"]

                category_worksheet.merge_cells("B2:E2")
                prequalified_bidders = 0

                questions = Question.objects.filter(
                    section__category_id=category.id, is_qa=True).only("id", "short_description", "description")
                sections = Section.objects.filter(
                    category_id=category.id, parent_section__isnull=True).order_by("id")

                for section in sections:
                    if section.has_child_sections:
                        if section.is_scored:
                            s_total = section.section_score
                            for s in section.child_sections:
                                if s.has_child_sections:
                                    for x in s.child_sections:
                                        s_total += x.section_score
                                else:
                                    s_total += s.section_score
                            category_headers.extend(
                                [f"hhc_{section.short_name}({int(s_total)})"]
                            )
                        else:
                            evaluate_this = False
                            for s in section.child_sections:
                                if s.has_child_sections:
                                    for x in s.child_sections:
                                        if x.is_scored:
                                            evaluate_this = True
                                            break
                                else:
                                    if s.is_scored:
                                        evaluate_this = True
                                        break
                            if evaluate_this == True:
                                s_total = 0
                                for s in section.child_sections:
                                    if s.has_child_sections:
                                        for x in s.child_sections:
                                            s_total += x.section_score
                                    else:
                                        s_total += s.section_score
                                category_headers.extend(
                                    [f"hhc_{section.short_name}({int(s_total)})"]
                                )
                    elif section.is_scored:
                        category_headers.extend(
                            [
                                f"hhc_{section.short_name}({int(section.section_score)})"
                            ]
                        )

                category_headers.extend(
                    ["hl_Contact Person", "hl_Phone Number", "hl_Email Address", "hl_Main Location"]
                )
                category_worksheet.append(category_headers)

                responsive_bidders = c_suppliers(category, interim=True)

                for supplier in responsive_bidders:
                    supplier_prequal_response = (
                        SupplierTechnicalScore.objects.filter(
                            supplier_id=supplier.id, category_id=category.id
                        ).only("score_after_qa", "rank", "rank_after_qa").first()
                    )
                    prequal_score = supplier_prequal_response.score

                    supplier_info = [
                        "",
                        supplier.company_name,
                        f"Nc_{supplier_prequal_response.rank}",
                        f"Tcnd_{prequal_score}",
                    ]

                    for section in sections:
                        if section.has_child_sections:
                            if section.is_scored:
                                section_total = supplier.tender_section_score(section)
                                for s in section.child_sections:
                                    if s.has_child_sections:
                                        if s.is_scored:
                                            section_total += (
                                                supplier.tender_section_score(s)
                                            )
                                            for x in s.child_sections:
                                                if x.is_scored:
                                                    section_total += supplier.tender_section_score(x)
                                        else:
                                            evaluate_this = False
                                            for x in s.child_sections:
                                                if x.is_scored:
                                                    evaluate_this = True
                                                    break
                                            if evaluate_this == True:
                                                section_total += (
                                                    supplier.tender_section_score(s)
                                                )
                                                for x in s.child_sections:
                                                    if x.is_scored:
                                                        section_total += supplier.tender_section_score(x)

                                    else:
                                        if s.is_scored:
                                            section_total += (
                                                supplier.tender_section_score(s)
                                            )
                                supplier_info.extend([f"Tcnd_{section_total}"])
                            else:
                                evaluate_this = False
                                for s in section.child_sections:
                                    if s.is_scored:
                                        evaluate_this = True
                                        break
                                if evaluate_this == True:
                                    section_total = supplier.tender_section_score(
                                        section
                                    )
                                    for s in section.child_sections:
                                        if s.has_child_sections:
                                            if s.is_scored:
                                                section_total += (
                                                    supplier.tender_section_score(
                                                        s
                                                    )
                                                )
                                                for x in s.child_sections:
                                                    if x.is_scored:
                                                        section_total += supplier.tender_section_score(
                                                            x
                                                        )
                                            else:
                                                evaluate_this = False
                                                for x in s.child_sections:
                                                    if x.is_scored:
                                                        evaluate_this = True
                                                        break
                                                if evaluate_this == True:
                                                    section_total += supplier.tender_section_score(
                                                        s
                                                    )
                                                    for x in s.child_sections:
                                                        if x.is_scored:
                                                            section_total += supplier.tender_section_score(
                                                                x
                                                            )
                                        else:
                                            if s.is_scored:
                                                section_total += (
                                                    supplier.tender_section_score(
                                                        s
                                                    )
                                                )
                                    supplier_info.extend([f"Tcnd_{section_total}"])
                        else:
                            if section.is_scored:
                                section_total = supplier.tender_section_score(
                                    section
                                )
                                supplier_info.extend([f"Tcnd_{section_total}"])

                    try:
                        # supplier_locations = get_cat_supplier_locations(cat_type_supplier)
                        supplier_locations = get_locations(
                            category=category, supplier_id=supplier.id
                        )
                        location = supplier_locations.get("main_location", "Ar_N/A")
                    except:
                        location = "N/A"

                    try:
                        supplier_info.extend(
                            [
                                f"Al_{supplier.contact_name}",
                                f"Al_{supplier.phone_number}",
                                f"Al_{supplier.email}",
                                f"{location}",
                            ]
                        )
                    except Exception as e:
                        capture_exception(e)

                    category_worksheet.append(supplier_info)

                # Append number of prequalified bidders
                # category_worksheet["C4"] = f"Nc_{prequalified_bidders}"

                # non responsive bidders
                category_order = apps.apps.get_model('core', 'CategoryOrder')
                non_responsive_bidders = Supplier.objects.filter(
                    id__in=category_order.objects.filter(
                        category_id=category.id, payment_status=category_order.PAID,
                        target__model="tender").exclude(
                        supplier_id__in=[r.id for r in responsive_bidders]).only("supplier_id")
                )

                for supplier in non_responsive_bidders:
                    supplier_info = [
                        "",
                        supplier.company_name,
                        "Ac_N/R",
                        "Ac_N/R",
                    ]
                    for section in sections:
                        if section.has_child_sections:
                            if section.is_scored:
                                supplier_info.extend(["Ac_N/R"])
                            else:
                                evaluate_this = False
                                for s in section.child_sections:
                                    if s.is_scored:
                                        evaluate_this = True
                                        break
                                if evaluate_this == True:
                                    supplier_info.extend(["Ac_N/R"])
                        else:
                            if section.is_scored:
                                supplier_info.extend(["Ac_N/R"])

                    supplier_info.extend(["Ac_N/R", "Ac_N/R", "Ac_N/R", "Ac_N/R"])
                    category_worksheet.append(supplier_info)

                # append notes
                category_worksheet.append(("", ""))
                category_worksheet.append(("", "Notes"))
                category_worksheet.append(("", "N/A - Not Applicable"))
                category_worksheet.append(("", "N/R - None responsive"))
                category_worksheet.freeze_panes = category_worksheet["C9"]

            summary_sheet.append(
                (
                    " ", "", "", "qT_Total", "Ac_{}".format(total_responsive),
                    "Ac_{}".format(total_non_responsive), "Ac_{}".format(total_bidders),
                )
            )
            summary_sheet.freeze_panes = summary_sheet["E7"]
            workbook.save(filepath)
            format_excel(filepath, data=data)
            insert_image(
                excel_url=filepath, worksheet_name="Summary", anchor=qed_logo_anchor
            )
            insert_image(
                excel_url=filepath, worksheet_name="Summary",
                anchor=buyer_logo_anchor, image_url=buyer_logo_url,
            )

            with open(filepath, "rb") as l:
                storage = PrivateMediaStorage()
                url = f"tender/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
                storage.save(url, l)

                job_report, created = JobReport.objects.update_or_create(
                    tender_id=job_id, defaults={"interim_report": url}
                )
                path = job_report.interim_report.url
                context = {
                    "filepath": path,
                    "response_message": "Report generated successfully",
                }
                return context
        except Exception as e:
            print(e)
            capture_exception(e)
            context = {
                "response_message": "Report generation error",
                "messages": [f"{e}", ]
            }
            return context


@shared_task(bind=True)
def tender_participation_status(self, job_id):
    """
    Checks status of Invited suppliers against participation status
    """
    try:
        job = Tender.objects.select_related("company").get(id=job_id)
        company = job.company
        data = {}
        time = datetime.datetime.now()

        categories = job.categories
        if len(categories) < 1:
            return {"message": "Job has no invite only categories"}
        else:
            total = categories.count()
            result = 0
            progress_recorder = ProgressRecorder(self)

            tender_participants = []
            for cat in categories:
                result += 1
                progress_recorder.set_progress(result, total)

                technical_participants = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                    id__in=SupplierResponse.objects.filter(
                        question__section__category_id=cat.id
                    ).only('supplier_id').values('supplier_id').distinct()
                )
                financial_participants = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                    id__in=ItemResponse.objects.filter(
                        item__category_id=cat.id
                    ).only('supplier_id').values('supplier_id').distinct()
                )

                if cat.invite_only is True:
                    invitees = cat.invited_suppliers
                else:
                    invitees = list(cat.paid_bidders)

                tender_data = {
                    "category_code": cat.unique_reference,
                    "category": cat,
                    "prequal_participants": technical_participants,
                    "rfq_participants": financial_participants,
                    "participants": technical_participants.union(financial_participants),
                    "invitees": invitees,
                }

                tender_participants.append(tender_data)

            # Report
            dir_name = Path(
                "media/tender_participants/%s/%s/%s" % (company.company_name, time.year, job.title)
            )  # folder structure
            dir_name.mkdir(parents=True, exist_ok=True)
            # time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
            filepath = "{}/{}_Tender_Participation_Report.xlsx".format(
                dir_name, job.unique_reference
            )
            # delete similar reports run this month
            match_string = "{}_{}".format(
                job.id, company.company_name.replace(" ", "_")
            )
            delete_matching_files_in_directory(dir_name, match_string, ".xlsx")
            # create and save a workbook to later populate
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = f"Summary"

            if (
                company.company_logo_url is not None
                and company.company_logo_url != ""
            ):
                buyer_logo_url = get_file_path(company.company_logo_url)
            else:
                buyer_logo_url = "static/core/img/no-company-image128-128.png"
            qed_logo_anchor = "D1"
            qed_logo_worksheet = worksheet.title
            buyer_logo_anchor = "C1"
            buyer_logo_worksheet = worksheet.title

            worksheet["C5"] = "Nb_{}".format(job.title)
            worksheet["C6"] = "Nb_{}".format(job.unique_reference)
            worksheet["C7"] = "Nb_Tender Participation Report"
            worksheet["C8"] = "Nb_"

            worksheet.append(
                ("", "", "", "", "hc_Participants")
            )
            worksheet.append(
                (
                    "", "h_#", "h_Category Code", "h_Category Name",
                    "h_Invited/Paid", "h_Yes", "h_No",
                )
            )
            worksheet.merge_cells("C5:D5")
            worksheet.merge_cells("E9:G9")

            data["Summary_C"] = 25
            data["Summary_D"] = 62
            data["Summary_E"] = 10
            data["Summary_F"] = 10
            data["Summary_G"] = 10
            data["Summary_H"] = 10

            total_invited = 0
            total_participants = 0
            total_non_participants = 0

            total += len(tender_participants)

            for count, tender_data in enumerate(tender_participants, start=1):
                result += 1
                progress_recorder.set_progress(result, total)
                category = tender_data["category"]
                invitees = tender_data["invitees"]
                participants = tender_data["participants"]
                prequal_participants = tender_data["prequal_participants"]
                rfq_participants = tender_data["rfq_participants"]
                # percentage_participation = "{:.2f}".format(
                #     len(participants) / len(invitees)
                # )
                try:
                    percentage_participation = "{}%".format(
                        round(((len(participants) / len(invitees)) * 100))
                    )
                except Exception as e:
                    percentage_participation = 0

                if len(tender_data["invitees"]) > 0:
                    non_participants = len(tender_data["invitees"]) - len(
                        tender_data["participants"]
                    )
                else:
                    non_participants = 0

                worksheet.append(
                    (
                        "", "Ac_%d" % (count), tender_data["category_code"],
                        tender_data["category"].name, "Ac_{}".format(len(invitees)),
                        "Ac_{}".format(len(participants)), "Ac_{}".format(non_participants),
                    )
                )

                total_invited += len(invitees)
                total_participants += len(participants)
                total_non_participants += non_participants

                # create individual sheets
                workbook.create_sheet(category.unique_reference.replace("/", "_"))
                details_worksheet = workbook[
                    category.unique_reference.replace("/", "_")
                ]
                empty_row = [""]
                empty_row = tuple(empty_row)
                details_worksheet.append(empty_row)
                details_worksheet["C2"] = "Nb_{}".format(category.name)
                details_worksheet["C3"] = "Nb_"
                details_worksheet["C4"] = "Number of bidders invited"
                details_worksheet["C5"] = "Number of bidders participated"
                details_worksheet["C6"] = "% Participation"
                details_worksheet["D4"] = "Ac_{}".format(len(invitees))
                details_worksheet["D5"] = "Ac_{}".format(len(participants))
                details_worksheet["D6"] = "Ac_{}".format(percentage_participation)
                details_worksheet["C7"] = "Nb_"

                details_worksheet.append(
                    ("", "", "", "", "hc_Participation Status")
                )

                details_worksheet.append(
                    ("", "", "", "", "hc_Technical", "", "hc_Financial", "")
                )

                details_worksheet.append(
                    (
                        "", "hc_No", "h_Supplier", "h_Phone Number", "h_Yes",
                        "h_No", "h_Yes", "h_No",
                    )
                )
                details_worksheet.merge_cells("E8:H8")
                details_worksheet.merge_cells("E9:F9")
                details_worksheet.merge_cells("G9:H9")
                details_worksheet.merge_cells("C2:E2")
                # formating
                data[f'{category.unique_reference.replace("/", "_")}_C'] = 35
                data[f'{category.unique_reference.replace("/", "_")}_D'] = 15
                data[f'{category.unique_reference.replace("/", "_")}_E'] = 10
                data[f'{category.unique_reference.replace("/", "_")}_F'] = 10
                data[f'{category.unique_reference.replace("/", "_")}_G'] = 10
                data[f'{category.unique_reference.replace("/", "_")}_H'] = 10

                prequal_yes = ""
                prequal_no = ""
                rfq_yes = ""
                rfq_no = ""
                for count, supplier in enumerate(invitees, start=1):
                    if supplier is not None:
                        if supplier in prequal_participants:
                            prequal_yes = "Yes"
                            prequal_no = " "
                        else:
                            prequal_no = "No"
                            prequal_yes = " "

                        if supplier in rfq_participants:
                            rfq_yes = "Yes"
                            rfq_no = " "
                        else:
                            rfq_no = "No"
                            rfq_yes = " "

                        if isinstance(supplier, str):
                            company = supplier
                            phone = "N/A"
                        else:
                            company = supplier.company_name
                            phone = supplier.phone_number
                        details_worksheet.append(
                            (
                                "", "Ac_%d" % (count), company, phone,
                                "Ac_{}".format(prequal_yes), "Ac_{}".format(prequal_no),
                                "Ac_{}".format(rfq_yes), "Ac_{}".format(rfq_no),
                            )
                        )

            worksheet.append(
                (
                    "", "", "", "Total", "Ac_{}".format(total_invited),
                    "Ac_{}".format(total_participants), "Ac_{}".format(total_non_participants),
                )
            )
            workbook.save(filepath)
            format_excel(filepath, data=data)
            insert_image(
                excel_url=filepath, worksheet_name=qed_logo_worksheet,
                anchor=qed_logo_anchor,
            )
            insert_image(
                filepath, buyer_logo_worksheet, buyer_logo_anchor,
                image_url=buyer_logo_url,
            )

        with open(filepath, "rb") as l:
            storage = PrivateMediaStorage()
            url = f"tender/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
            storage.save(url, l)

            job_report, created = JobReport.objects.update_or_create(
                tender_id=job_id, defaults={"participation_status_report": url}
            )
            path = job_report.participation_status_report.url

            context = {
                "filepath": path,
                "response_message": "Report generated successfully",
            }
            return context

    except Exception as e:
        capture_exception(e)
        context = {
            "response_message": "An error occurred",
            "messages": [f"{e}", ]
        }
        print(e)
        return context


@shared_task(bind=True)
def suppliers_list_report(self, job_id):
    try:
        job = Tender.objects.filter(id=job_id).prefetch_related('company').first()
        company = job.company
        time = datetime.datetime.now()
        data = {}
        dir_name = Path(
            "media/suppliers_list_report/%s/%s/%s"
            % (company.company_name, time.year, job.title)
        )  # folder structure
        dir_name.mkdir(parents=True, exist_ok=True)
        # time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
        filepath = "{}/{}_Supplier_List_Report.xlsx".format(
            dir_name, job.unique_reference
        )

        # delete similar reports run this month
        match_string = "{}_{}".format(job.id, company.company_name.replace(" ", "_"))
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

        # create and save a workbook to later populate
        workbook = Workbook()

        # create and save a workbook to later populate
        suppliers_list_worsheet = workbook.active
        suppliers_list_worsheet.title = "SUPPLIERS"
        if company.company_logo_url is not None and company.company_logo_url != "":
            buyer_logo_url = get_file_path(company.company_logo_url)
        else:
            buyer_logo_url = "static/core/img/no-company-image128-128.png"

        qed_logo_anchor = "D1"
        qed_logo_worksheet = suppliers_list_worsheet.title
        buyer_logo_anchor = "B1"
        buyer_logo_worksheet = suppliers_list_worsheet.title
        suppliers_list_worsheet["B5"] = job.title
        suppliers_list_worsheet["B6"] = job.unique_reference
        suppliers_list_worsheet["B7"] = "Supplier Details"
        suppliers_list_worsheet["B8"] = "Nb_"

        data["SUPPLIERS_C"] = 35
        data["SUPPLIERS_D"] = 25
        data["SUPPLIERS_E"] = 30
        data["SUPPLIERS_F"] = 25
        data["SUPPLIERS_G"] = 20
        data["SUPPLIERS_H"] = 62

        suppliers_list_worsheet.merge_cells("B5:D5")

        categories = job.categories.only("id", "unique_reference", "name").order_by(
            "unique_reference"
        )
        total = categories.count()
        result = 0
        progress_recorder = ProgressRecorder(self)

        for count, category in enumerate(categories, start=1):
            suppliers_list_worsheet.append(
                ("", f"Category: {category.name}", "", "", "", "", "", "")
            )
            suppliers_list_worsheet.append(
                (
                    "", "h_Bidder", "h_Contact Person", "h_Phone Number",
                    "h_Email Address", "h_Postal Address", "h_Tax PIN",
                    "h_Physical Address", "h_Main Location", "h_Branches",
                )
            )
            result += 1
            progress_recorder.set_progress(result, total)

            suppliers = Supplier.objects.filter(
                id__in=SupplierResponse.objects.filter(question__section__category_id=category.id
                ).only('supplier_id').values('supplier_id').distinct()
            )

            for count, supplier in enumerate(suppliers, start=1):
                supplier_data = get_supplier_data(
                    category_id=category.id, supplier_id=supplier.id
                )
                try:
                    suppliers_list_worsheet.append(
                        (
                            "", supplier.company_name, supplier.contact_name,
                            supplier.phone_number, supplier.email,
                            supplier_data["Section"].get("P.O Box address", "N/A"),
                            supplier_data.get("Tax_ID", "N/A"),
                            supplier_data["Section"].get("Office_Address", "N/A"),
                            supplier_data["Section"].get("Main_Office", "N/A"),
                            supplier_data["Section"].get("Branch_Office", "N/A"),
                        )
                    )
                except Exception as e:
                    capture_exception(e)
                    print(e)

            suppliers_list_worsheet.append(
                ("", "", "", "", "", "", "", "")
            )

        suppliers_list_worsheet.freeze_panes = suppliers_list_worsheet["C8"]

        workbook.save(filepath)
        format_excel(filepath, data=data)
        insert_image(
            excel_url=filepath,
            worksheet_name=qed_logo_worksheet,
            anchor=qed_logo_anchor,
        )
        insert_image(
            filepath, buyer_logo_worksheet, buyer_logo_anchor, image_url=buyer_logo_url
        )

        try:
            with open(filepath, "rb") as l:
                storage = PrivateMediaStorage()
                url = f"tender/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
                storage.save(url, l)

                job_report, created = JobReport.objects.update_or_create(
                    tender_id=job_id, defaults={"supplier_details_report": url}
                )
                path = job_report.supplier_details_report.url
                context = {
                    "filepath": path,
                    "response_message": "Report generated successfully",
                }
                return context
        except Exception as e:
            context = {
                "response_message": "Report generation error",
                "messages": [f"{e}",]
            }
            return context

    except Exception as e:
        capture_exception(e)
        context = {
            "response_message": "Report generation error",
            "messages": [f"{e}",]
        }
        return context


@shared_task(bind=True)
def job_qed_category_suppliers(self, job_id):
    data = {}
    job = Tender.objects.filter(id=job_id).first()
    company = job.company
    categories = Category.objects.filter(tender_id=job_id)

    time = datetime.datetime.now()
    dir_name = Path(
        "media/category_types/jobs/%d/suppliers/%s" % (job_id, time.year)
    )  # folder structure
    dir_name.mkdir(parents=True, exist_ok=True)
    time_only = "%d_%d_%d" % (time.month, time.day, time.hour)
    all_suppliers = "QED_Category_Suppliers"
    filepath = "{}/{}_{}.xlsx".format(dir_name, all_suppliers, time_only)

    # delete similar reports run this month
    match_string = "{}_{}".format(all_suppliers, time_only)
    delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

    # create and save a workbook to later populate
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    buyer_logo_url = "static/core/img/QED2.png"
    buyer_logo = Image("static/core/img/QED2.png")
    buyer_logo_height, buyer_logo_width = buyer_logo.height, buyer_logo.width
    max_height = 75
    ratio = buyer_logo_height / max_height
    buyer_logo.height, buyer_logo.width = buyer_logo_height, buyer_logo_width

    buyer_logo.anchor = summary_sheet["C1"]
    summary_sheet.add_image(buyer_logo, "C1")
    summary_sheet.title = "Summary"
    logo_worksheet = summary_sheet.title
    summary_sheet.append(("", "", ""))
    summary_sheet.append(("", "", ""))
    summary_sheet.append(("", "", ""))
    summary_sheet.append(("", "", ""))
    summary_sheet.append(
        (
            "",
            "hc_No",
            "h_QED Code",
            "h_Category Name",
            "h_Category Code",
            "h_Suppliers Count",
            "h_Paid Bidders",
        )
    )
    data["Summary_C"] = 18
    data["Summary_D"] = 62
    data["Summary_E"] = 7.7
    data["Summary_F"] = 7.7
    data["Summary_G"] = 7.7

    # try:
    result = 0
    progress_recorder = ProgressRecorder(self)
    total_for_progress = categories.count()

    total_suppliers_count = 0
    total_paid_bidders = 0
    for count, category in enumerate(categories, start=1):
        result += 1
        progress_recorder.set_progress(result, total_for_progress)

        suppliers = category.category_type.suppliers_list(job.company.country)
        category_total_bidders = category.total_bidders
        category_type = category.category_type
        summary_sheet.append(
            (
                "",
                "Ac_%d" % (count),
                category.category_type.innitials,
                category.name,
                category.unique_reference,
                suppliers["supplier_count"],
                category_total_bidders,
            )
        )
        total_paid_bidders += category_total_bidders
        total_suppliers_count += int(suppliers["supplier_count"])

        # create and save a workbook to later populate
        category_name = (
            f"{category.unique_reference}_{category_type.innitials}"
        )

        category_type_worksheet = workbook.create_sheet(category_name)
        category_type_worksheet.title = category_name
        category_type_worksheet["C2"] = "Nb_Category Name"
        category_type_worksheet["D2"] = "Nb_{}".format(category.name)
        category_type_worksheet["C3"] = "Nb_QED Category".format(
            category_type.name)

        category_type_worksheet["D3"] = "Nb_{}".format(category_type.name)
        category_type_worksheet["C4"] = "Nb_QED Code".format(
            category_type.innitials)
        category_type_worksheet["D4"] = "Nb_{}".format(
            category_type.innitials)
        category_type_worksheet["C5"] = "Nb_"

        category_type_worksheet.merge_cells("D2:F2")
        category_type_worksheet.merge_cells("D3:F3")

        data[f"category.unique_reference_category.category_type.innitials_C"] = 30
        data[f"category.unique_reference_category.category_type.innitials_D"] = 15
        data[f"category.unique_reference_category.category_type.innitials_E"] = 28
        data[f"category.unique_reference_category.category_type.innitials_F"] = 28
        data[f"category.unique_reference_category.category_type.innitials_G"] = 28
        data[f"category.unique_reference_category.category_type.innitials_H"] = 28
        data[f"category.unique_reference_category.category_type.innitials_I"] = 10

        category_type_worksheet.append(
            (
                "", "hc_No", "h_Company_Name", "h_Status (Paid/Not_Paid)",
                "h_Primary Email", "h_Alternative Email", "h_Primary Phone",
                "h_Alternative Phone", "h_Country", "h_Location 1",
                "h_Location 2", "h_Location 3", "h_Location 4", "h_Location 5",
                "h_Location 6", "h_Location 7", "h_Location 8", "h_Location 9",
                "h_Location 10", "h_Other Locations",
            )
        )

        total_for_progress += len(suppliers["registered_suppliers"])
        for count, supplier in enumerate(
            suppliers["registered_suppliers"], start=1
        ):
            result += 1
            progress_recorder.set_progress(result, total_for_progress)
            category_type = category.category_type

            supplier_paid_categories = Category.objects.filter(
                id__in=apps.apps.get_model('core', 'CategoryOrder').objects.filter(
                    target__model='tender', supplier_id=supplier.id
                ).only('category_id').values('category_id').distinct()
            )

            cat_type_supplier = CategoryTypeSupplier.objects.filter(
                supplier=supplier, category_type=category_type
            ).first()

            supplier_locations = get_cat_supplier_locations(cat_type_supplier)
            try:
                location_1 = supplier_locations[0]
            except:
                location_1 = "N/A"
            try:
                location_2 = supplier_locations[1]
            except:
                location_2 = "N/A"
            try:
                location_3 = supplier_locations[2]
            except:
                location_3 = "N/A"
            try:
                location_4 = supplier_locations[3]
            except:
                location_4 = "N/A"
            try:
                location_5 = supplier_locations[4]
            except:
                location_5 = "N/A"
            try:
                location_6 = supplier_locations[5]
            except:
                location_6 = "N/A"
            try:
                location_7 = supplier_locations[6]
            except:
                location_7 = "N/A"
            try:
                location_8 = supplier_locations[7]
            except:
                location_8 = "N/A"
            try:
                location_9 = supplier_locations[8]
            except:
                location_9 = "N/A"
            try:
                location_10 = supplier_locations[9]
            except:
                location_10 = "N/A"

            if len(supplier_locations) > 10:
                try:
                    others = supplier_locations[10:]
                except:
                    others = ["NA"]
                others = ",".join(others)
            others = "NA"

            if category in supplier_paid_categories:
                status = "Paid"
            else:
                status = "Not Paid"

            category_type_worksheet.append(
                (
                    "", "Ac_%d" % (count), supplier.company_name, status,
                    supplier.email, " ", supplier.phone_number, " ", supplier.country,
                    location_1, location_2, location_3, location_4, location_5,
                    location_6, location_7, location_8, location_9, location_10, others,
                )
            )

        for count, supplier in enumerate(
            suppliers["old_suppliers"],
            start=len(suppliers["registered_suppliers"]) + 1,
        ):
            status = "Not Paid"

            if supplier.alternative_email == "":
                alternative_email = " "
            else:
                alternative_email = supplier.alternative_email

            if supplier.primary_phone == "":
                primary_phone = " "
            else:
                primary_phone = supplier.primary_phone

            if supplier.alternative_phone == "":
                alternative_phone = " "
            else:
                alternative_phone = supplier.alternative_phone

            category_type_worksheet.append(
                (
                    "", "Ac_%d" % (count), supplier.company_name, status,
                    supplier.primary_email, alternative_email, primary_phone,
                    alternative_phone, supplier.country,
                )
            )
        category_type_worksheet.freeze_panes = category_type_worksheet["D7"]
    summary_sheet.append(
        ("", "", "", "", "Total", total_suppliers_count, total_paid_bidders)
    )
    summary_sheet.freeze_panes = summary_sheet["E7"]
    workbook.save(filepath)
    format_excel(filepath)
    filepath = insert_image(
        filepath, logo_worksheet, buyer_logo.anchor, image_url=buyer_logo_url
    )

    try:
        with open(filepath, "rb") as l:
            storage = PrivateMediaStorage()
            url = f"tender/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
            storage.save(url, l)

            job_report, created = JobReport.objects.update_or_create(
                tender_id=job_id, defaults={"category_suppliers_report": url}
            )
            context = {
                "response_message": "Report generated successfully",
                "filepath": job_report.category_suppliers_report.url
            }
            return context
    except Exception as e:
        capture_exception(e)
        context = {
            "response_message": "Error generating report",
            "errors": f"{e}"
        }
        return context


@shared_task(bind=True)
def tender_bidder_locations_report(self, job_id):
    """
    Prequalified suppliers location and contact details
    """
    try:
        job = Tender.objects.filter(id=job_id).prefetch_related('company').first()
        company = job.company
        time = datetime.datetime.now()
        data = {}

        dir_name = Path(
            "media/bidder_locations/%s/%s/%s"
            % (company.company_name, time.year, job.title)
        )  # folder structure
        dir_name.mkdir(parents=True, exist_ok=True)
        # time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
        filepath = "{}/{}_Bidder_Locations_Report.xlsx".format(
            dir_name, job.unique_reference
        )

        # delete similar reports run this month
        match_string = "{}_{}".format(job.id, company.company_name.replace(" ", "_"))
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

        # create and save a workbook to later populate
        workbook = Workbook()

        # create and save a workbook to later populate
        bidder_location_details_worksheet = workbook.active
        bidder_location_details_worksheet.title = "Summary"
        if company.company_logo_url is not None and company.company_logo_url != "":
            buyer_logo_url = get_file_path(company.company_logo_url)
        else:
            buyer_logo_url = "static/core/img/no-company-image128-128.png"

        qed_logo_anchor = "K1"
        qed_logo_worksheet = bidder_location_details_worksheet.title
        buyer_logo_anchor = "C1"
        buyer_logo_worksheet = bidder_location_details_worksheet.title
        bidder_location_details_worksheet["C5"] = "Nb_{}".format(job.title)
        bidder_location_details_worksheet["C6"] = "Nb_{}".format(job.unique_reference)
        bidder_location_details_worksheet["C7"] = "Nb_Bidder Location Details"
        bidder_location_details_worksheet["C8"] = "Nb_"

        bidder_location_details_worksheet.append(
            ("", "hc_No", "h_Category Code", "h_Category Name")
        )
        bidder_location_details_worksheet.merge_cells("C5:D5")

        data["Summary_C"] = 25
        data["Summary_D"] = 62

        categories = Category.objects.filter(tender_id=job_id).only("id", "unique_reference", "name").order_by(
            "unique_reference"
        )
        total = categories.count()
        result = 0
        progress_recorder = ProgressRecorder(self)

        for count, category in enumerate(categories, start=1):
            result += 1
            progress_recorder.set_progress(result, total)
            bidder_location_details_worksheet.append(
                (
                    "",
                    "Ac_%d" % (count),
                    category.unique_reference,
                    category.name,
                )
            )

            workbook.create_sheet(category.unique_reference.replace("/", "_"))
            category_bidder_details_worksheet = workbook[
                category.unique_reference.replace("/", "_")
            ]
            empty_row = [""]
            empty_row = tuple(empty_row)
            category_bidder_details_worksheet.append(empty_row)
            category_bidder_details_worksheet["C2"] = "Nb_{}".format(category.name)
            category_bidder_details_worksheet["C3"] = "Nb_{}".format(
                category.unique_reference
            )
            category_bidder_details_worksheet["C4"] = "Nb_"

            data[f'category.unique_reference.replace("/", "_")_C'] = 25
            data[f'category.unique_reference.replace("/", "_")_D'] = 5
            data[f'category.unique_reference.replace("/", "_")_E'] = 14.6
            data[f'category.unique_reference.replace("/", "_")_F'] = 25
            data[f'category.unique_reference.replace("/", "_")_G'] = 14.6
            data[f'category.unique_reference.replace("/", "_")_H'] = 30
            data[f'category.unique_reference.replace("/", "_")_I'] = 35

            category_bidder_details_worksheet.merge_cells("C2:E2")
            category_bidder_details_worksheet.append(
                (
                    "", "hc_No", "h_Company Name", "h_Rank", "h_Marks After QA",
                    "h_Contact Person", "h_Phone Number", "h_Email Address",
                    "h_Physical Address", "h_Main Location", "h_Branches",
                    "h_Location 3", "h_Location 4", "h_Location 5", "h_Location 6",
                    "h_Location 7", "h_Location 8", "h_Location 9","h_Location 10",
                )
            )

            supplier_totals = SupplierTechnicalScore.objects.filter(category_id=category.id).order_by(
                "rank"
            ).prefetch_related('supplier')

            for count, supplier_total in enumerate(supplier_totals, start=1):
                supplier = supplier_total.supplier
                score_after_qa = supplier_total.score_after_qa
                rank_after_qa = supplier_total.rank_after_qa

                if score_after_qa is None:
                    score_after_qa = "Ar_N/A"
                    rank_after_qa = "Ar_N/A"

                if rank_after_qa is None:
                    rank_after_qa = "Ar_N/A"

                locations = get_locations(category=category, supplier_id=supplier_total.supplier_id)

                category_bidder_details_worksheet.append(
                    (
                        "", "Ac_%d" % (count), supplier.company_name,rank_after_qa, score_after_qa,
                        supplier.contact_name, supplier.phone_number, supplier.address,
                        " ".join(locations.get("main_location", "Ar_N/A").split()),
                        " ".join(locations.get("location_1", "Ar_N/A").split()),
                        " ".join(locations.get("location_2", "Ar_N/A").split()),
                        " ".join(locations.get("location_3", "Ar_N/A").split()),
                        " ".join(locations.get("location_4", "Ar_N/A").split()),
                        " ".join(locations.get("location_5", "Ar_N/A").split()),
                        " ".join(locations.get("location_6", "Ar_N/A").split()),
                        " ".join(locations.get("location_7", "Ar_N/A").split()),
                        " ".join(locations.get("location_8", "Ar_N/A").split()),
                        " ".join(locations.get("location_9", "Ar_N/A").split()),
                        " ".join(locations.get("location_10", "Ar_N/A").split()),
                    )
                )
                category_bidder_details_worksheet.freeze_panes = (
                    category_bidder_details_worksheet["D6"]
                )

        bidder_location_details_worksheet.freeze_panes = (
            bidder_location_details_worksheet["D10"]
        )
        workbook.save(filepath)
        format_excel(filepath, data=data)
        insert_image(
            excel_url=filepath,
            worksheet_name=qed_logo_worksheet,
            anchor=qed_logo_anchor,
        )
        insert_image(
            filepath, buyer_logo_worksheet, buyer_logo_anchor, image_url=buyer_logo_url
        )

        with open(filepath, "rb") as l:
            storage = PrivateMediaStorage()
            url = f"tender/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
            storage.save(url, l)

            job_report, created = JobReport.objects.update_or_create(
                tender_id=job_id, defaults={"bidder_locations_report": url}
            )
            context = {
                "response_message": "Report generated successfully",
                "filepath": job_report.bidder_locations_report.url,
            }
            return context

    except Exception as e:
        capture_exception(e)
        print(e)
        context = {
            "response_message": "Report generation error",
            "errors": f"{e}"
        }
        return context


@shared_task(bind=True)
def tender_responsive_bidders_report(self, job_id):
    try:
        result = 0
        progress_recorder = ProgressRecorder(self)

        job = Tender.objects.filter(id=job_id).prefetch_related('company').first()
        company = job.company
        time = datetime.datetime.now()

        dir_name = Path(
            "media/responsive_bidders/%s/%s/%s" % (company.company_name, time.year, job.title)
        )  # folder structure
        dir_name.mkdir(parents=True, exist_ok=True)
        time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
        filepath = "{}/{}_Responsive_Bidders_Report.xlsx".format(
            dir_name, job.unique_reference
        )

        # delete similar reports run this month
        match_string = "{}_{}".format(job.id, company.company_name.replace(" ", "_"))
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

        # create and save a workbook to later populate
        workbook = Workbook()

        responsive_bidders_worksheet = workbook.active
        responsive_bidders_worksheet.title = "Summary"

        if company.company_logo_url is not None and company.company_logo_url != "":
            buyer_logo_url = company.company_logo_url
        else:
            buyer_logo_url = "static/core/img/no-company-image128-128.png"

        qed_logo_anchor = "K1"
        qed_logo_worksheet = responsive_bidders_worksheet.title
        buyer_logo_anchor = "C1"
        buyer_logo_worksheet = responsive_bidders_worksheet.title
        responsive_bidders_worksheet["C5"] = "Nb_{}".format(job.title)
        responsive_bidders_worksheet["C6"] = "Nb_{}".format(job.unique_reference)
        responsive_bidders_worksheet["C7"] = "Nb_Responsive Bidders"
        responsive_bidders_worksheet["C8"] = "Nb_"

        responsive_bidders_worksheet.append(
            ("", "hc_No", "h_Category Name", "h_Category Code", "h_Responsive Bidders")
        )
        categories = Category.objects.filter(id=job_id)
        total = categories.count()
        total_responsive_bidders = 0

        for count, category in enumerate(categories, start=1):
            result += 1
            progress_recorder.set_progress(result, total)

            responsive_bidders = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                id__in=SupplierResponse.objects.filter(
                    question__section__category_id=category.id).only('supplier_id').values('supplier_id')
            )
            responsive_bidders_worksheet.append(
                ("", "Ac_%d" % (count), category.name,
                 category.unique_reference, responsive_bidders.count(),
                 )
            )
            workbook.create_sheet(category.unique_reference.replace("/", "_"))
            category_responsive_bidders_worksheet = workbook[
                category.unique_reference.replace("/", "_")
            ]
            empty_row = [""]
            empty_row = tuple(empty_row)
            category_responsive_bidders_worksheet.append(empty_row)

            category_responsive_bidders_worksheet["C2"] = "Nb_{}".format(category.name)
            category_responsive_bidders_worksheet["C3"] = "Nb_{}".format(
                category.unique_reference
            )
            category_responsive_bidders_worksheet["C4"] = "Nb_"

            category_responsive_bidders_worksheet.append(
                (
                    "", "hc_No", "h_Company Name", "h_Contact Person",
                    "h_Email Address", "h_Phone Number",
                )
            )
            total_responsive_bidders += responsive_bidders.count()
            for count, supplier in enumerate(responsive_bidders, start=1):
                category_responsive_bidders_worksheet.append(
                    (
                        "", "Ac_%d" % (count), supplier.company_name,
                        supplier.contact_name, supplier.address,
                        supplier.phone_number,
                    )
                )

        responsive_bidders_worksheet.append(
            ("", "", "", "h_Total", total_responsive_bidders)
        )
        workbook.save(filepath)
        format_excel(filepath)
        insert_image(
            excel_url=filepath,
            worksheet_name=qed_logo_worksheet,
            anchor=qed_logo_anchor,
        )
        insert_image(
            filepath, buyer_logo_worksheet, buyer_logo_anchor, image_url=buyer_logo_url
        )

        with open(filepath, "rb") as l:
            storage = PrivateMediaStorage()
            url = f"tender/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
            storage.save(url, l)

            job_report, created = JobReport.objects.update_or_create(
                tender_id=job_id, defaults={"responsive_bidders_report": url}
            )
            context = {
                "response_message": "Report generated successfully",
                "filepath": job_report.responsive_bidders_report.url,
            }
            return context
    except Exception as e:
        capture_exception(e)
        print(e)
        context = {
            "response_message": "Report generation error"
        }
        return context


@shared_task(bind=True)
def tender_non_responsive_bidders_report(self, job_id):
    try:
        result = 0
        progress_recorder = ProgressRecorder(self)
        job = Tender.objects.filter(id=job_id).prefetch_related('company').first()
        company = job.company
        time = datetime.datetime.now()
        dir_name = Path(
            "media/non_responsive_bidders/%s/%s/%s"
            % (company.company_name, time.year, job.title)
        )  # folder structure
        dir_name.mkdir(parents=True, exist_ok=True)
        # time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
        filepath = "{}/{}_Non_Responsive_Bidders_Report.xlsx".format(
            dir_name, job.unique_reference
        )

        # delete similar reports run this month
        match_string = "{}_{}".format(job.id, company.company_name.replace(" ", "_"))
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

        # create and save a workbook to later populate
        workbook = Workbook()

        responsive_bidders_worksheet = workbook.active
        responsive_bidders_worksheet.title = "Summary"
        total_non_responsive = 0

        if company.company_logo_url is not None and company.company_logo_url != "":
            buyer_logo_url = company.company_logo_url
        else:
            buyer_logo_url = "static/core/img/no-company-image128-128.png"

        qed_logo_anchor = "K1"
        qed_logo_worksheet = responsive_bidders_worksheet.title
        buyer_logo_anchor = "C1"
        buyer_logo_worksheet = responsive_bidders_worksheet.title
        responsive_bidders_worksheet["C5"] = "Nb_{}".format(job.title)
        responsive_bidders_worksheet["C6"] = "Nb_{}".format(job.unique_reference)
        responsive_bidders_worksheet["C7"] = "Nb_Non-Responsive Bidders"
        responsive_bidders_worksheet["C8"] = "Nb_"

        responsive_bidders_worksheet.append(
            (
                "", "hc_No", "h_Category Name", "h_Category Code", "h_Non-Responsive Bidders",
            )
        )
        categories = Category.objects.filter(tender_id=job.id)
        total_for_progress = categories.count()
        co = apps.apps.get_model('core', 'CategoryOrder')
        for count, category in enumerate(categories, start=1):
            result += 1
            progress_recorder.set_progress(result, total_for_progress)

            responsive_bidders = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                id__in=SupplierResponse.objects.filter(
                    question__section__category_id=category.id).only('supplier_id').values('supplier_id')
            )
            paid_bidders = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                id__in=co.objects.filter(
                    payment_status=co.PAID, category_id=category.id, target__model='tender'
                ).only('supplier_id').values('supplier_id')
            )
            non_responsive_bidders_count = paid_bidders.count() - responsive_bidders.count()
            non_responsive_bidders = paid_bidders.difference(responsive_bidders)

            responsive_bidders_worksheet.append(
                ("", "Ac_%d" % (count), category.name, category.unique_reference,
                 non_responsive_bidders_count,
                 )
            )
            total_non_responsive += non_responsive_bidders_count

            workbook.create_sheet(category.unique_reference.replace("/", "_"))
            category_responsive_bidders_worksheet = workbook[
                category.unique_reference.replace("/", "_")
            ]
            empty_row = [""]
            empty_row = tuple(empty_row)
            category_responsive_bidders_worksheet.append(empty_row)

            category_responsive_bidders_worksheet["C2"] = "Nb_{}".format(category.name)
            category_responsive_bidders_worksheet["C3"] = "Nb_{}".format(
                category.unique_reference
            )
            category_responsive_bidders_worksheet["C4"] = "Nb_"

            category_responsive_bidders_worksheet.append(
                (
                    "", "hc_No", "h_Company Name", "h_Contact Person", "h_Email Address", "h_Phone Number",
                )
            )

            for count, supplier in enumerate(non_responsive_bidders, start=1):
                category_responsive_bidders_worksheet.append(
                    (
                        "", "Ac_%d" % (count), supplier.company_name, supplier.contact_name,
                        supplier.address, supplier.phone_number,
                    )
                )
        responsive_bidders_worksheet.append(
            (
                "", "", "", "hc_Total", total_non_responsive,
            )
        )
        workbook.save(filepath)
        format_excel(filepath)
        insert_image(
            excel_url=filepath,
            worksheet_name=qed_logo_worksheet,
            anchor=qed_logo_anchor,
        )
        insert_image(
            filepath, buyer_logo_worksheet, buyer_logo_anchor, image_url=buyer_logo_url
        )

        with open(filepath, "rb") as l:
            storage = PrivateMediaStorage()
            url = f"tender/job/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
            storage.save(url, l)

            job_report, created = JobReport.objects.update_or_create(
                tender_id=job_id, defaults={"non_responsive_bidders_report": url}
            )
            context = {
                "response_message": "Report generated successfully",
                "filepath": job_report.non_responsive_bidders_report.url,
            }
            return context
    except Exception as e:
        capture_exception(e)
        print(e)
        context = {
            "response_message": "Report generation error"
        }
        return context


def tender_bidders_information_report(data):
    participants = data['participants']
    category = data['category']
    sections = data['sections']

    time = datetime.datetime.now()
    # bd = Side(style='thin', color="000000")
    dir_name = Path(
        "media/tender_bidders_info/{}/{}".format(time.year, time.month)
    )  # folder structure
    dir_name.mkdir(parents=True, exist_ok=True)

    time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
    filepath = "{}/{}_Bidders_Information_{}_{}.xlsx".format(
        dir_name,
        category.id,
        category.name.replace(" ", "_")
        .replace(":", "")
        .replace(",", "")
        .replace("-", "_")
        .replace("/", "_"),
        time_only,
    )

    workbook = Workbook()

    for section in sections:
        workbook.create_sheet(
            section.short_name_version.replace(" ", "_").replace(":", "_")
            .replace("(", "_").replace(")", "_")
        )
        section_name = section.short_name
        section_questions = section.questions
        current_worksheet = workbook[
            section.short_name_version.replace(" ", "_").replace(":", "_")
            .replace("(", "_").replace(")", "_")
        ]

        current_worksheet["B2"] = "Nb_{}".format(category.name)
        current_worksheet["B3"] = "Nb_%s" % section_name
        current_worksheet["B4"] = ""
        headers = [" ", "h_%s" % section_name]
        for supplier in participants:
            headers.append("h_%s" % supplier.company_name)
        current_worksheet.append(tuple(headers))
        i = 1
        for question in section_questions:
            question_row = [" ", question.short_description_value]
            for supplier in participants:
                response = supplier.prequal_question_response(question)
                question_row.append(response)
            current_worksheet.append(tuple(question_row))
            i += 1

    workbook.remove_sheet(workbook["Sheet"])
    workbook.save(filepath)
    format_excel(filepath)
    return filepath


def tender_evaluation_report_context(data):
    participants = data['participants']
    category = data['category']
    sections = data['sections']

    data = {
        "questions": category.questions,
        "category": category,
        "company": category.tender.company,
        "time": timezone.now(),
        "sections": sections,
        "suppliers": participants,
    }
    return data

@shared_task(bind=True)
def download_current_suppliers(self, job_id):
    data = {}
    job = Tender.objects.filter(id=job_id).first()
    company = job.company
    time = datetime.datetime.now()
    dir_name = Path(
        "media/current_suppliers/%s/%s/%s"
        % (company.company_name, time.year, job.job_title)
    )  # folder structure
    dir_name.mkdir(parents=True, exist_ok=True)
    time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
    filepath = "{}/{}_{}_{}.xlsx".format(
        dir_name, job.job_code, job.job_title, time_only
    )

    # delete similar reports run this month
    match_string = "{}_{}".format(job.id, company.company_name.replace(" ", "_"))
    delete_matching_files_in_directory(dir_name, match_string, ".xlsx")
    # create and save a workbook to later populate
    current_suppliers_workbook = Workbook()

    # create and save a workbook to later populate
    current_suppliers_worksheet = current_suppliers_workbook.active
    current_suppliers_worksheet.title = f"Summary"

    if company.company_logo_url is not None and company.company_logo_url != "":
        buyer_logo_url = get_file_path(company.company_logo_url)
    else:
        buyer_logo_url = "qed/static/img/no-company-image128-128.png"
    qed_logo_anchor = "D1"
    qed_logo_worksheet = current_suppliers_worksheet.title
    buyer_logo_anchor = "B1"
    buyer_logo_worksheet = current_suppliers_worksheet.title

    # current_suppliers_worksheet["C2"] = f"{job.company.company_name}"
    # current_suppliers_worksheet["C3"] = "Current Suppliers"
    # current_suppliers_worksheet["C4"] = ""
    current_suppliers_worksheet["C5"] = "Nb_{}".format(job.job_title)
    current_suppliers_worksheet["C6"] = "Nb_{}".format(job.job_code)
    current_suppliers_worksheet["C7"] = "Nb_Current Suppliers Report"
    current_suppliers_worksheet["C8"] = "Nb_"

    current_suppliers_worksheet.append(
        (
            "",
            "",
            "",
            "",
            "hc_Participation",
            "",
            "",
        )
    )
    current_suppliers_worksheet.merge_cells("C5:D5")
    current_suppliers_worksheet.merge_cells("E9:G9")
    current_suppliers_worksheet.append(
        ("", "hc_No", "h_Category Code", "h_Category Name", "h_Yes", "h_No", "h_Total")
    )
    data["Summary_C"] = 18
    data["Summary_D"] = 62
    data["Summary_E"] = 7.7
    data["Summary_F"] = 7.7
    data["Summary_G"] = 7.7
    current_suppliers = (
        apps.apps.get_model('core', 'CurrentSupplier').objects.select_related("category")
        .filter(company=company,  job_id=job.id)
        .order_by("category__unique_reference")
    )
    try:
        current_supplier_categories = []
        for cs in current_suppliers:
            if cs.category is not None:
                current_supplier_categories.append(cs.category)

        cat_list = []
        for x in current_supplier_categories:
            if x not in cat_list:
                cat_list.append(x)

        total_partcipants = 0
        total_non_participants = 0
        total_suppliers = 0

        result = 0
        progress_recorder = ProgressRecorder(self)
        total_for_progress = len(cat_list)

        for count, category in enumerate(cat_list, start=1):
            result += 1
            progress_recorder.set_progress(result, total_for_progress)

            sups = category_participation_count(category)
            current_suppliers_worksheet.append(
                (
                    "",
                    "Ac_%d" % (count),
                    category.unique_reference,
                    category.name,
                    "Ac_{}".format(sups["participants"]),
                    "Ac_{}".format(sups["non_participants"]),
                    "Ac_{}".format(sups["total"]),
                )
            )
            total_partcipants += sups["participants"]
            total_non_participants += sups["non_participants"]
            total_suppliers += sups["total"]

            current_suppliers_workbook.create_sheet(
                category.unique_reference.replace("/", "_")
            )
            current_suppliers_details_worksheet = current_suppliers_workbook[
                category.unique_reference.replace("/", "_")
            ]
            percentage_participation = "{}%".format(
                round(((sups["participants"] / sups["total"]) * 100))
            )
            empty_row = [""]
            empty_row = tuple(empty_row)
            current_suppliers_details_worksheet.append(empty_row)
            current_suppliers_details_worksheet["C2"] = "Nb_{}".format(category.name)
            current_suppliers_details_worksheet["C3"] = "Nb_{}".format(
                category.unique_reference
            )
            current_suppliers_details_worksheet["C4"] = "Nb_"
            current_suppliers_details_worksheet["C5"] = "Number of bidders invited"
            current_suppliers_details_worksheet["C6"] = "Number of bidders participated"
            current_suppliers_details_worksheet["C7"] = "% Participation"
            current_suppliers_details_worksheet["D5"] = "Ac_{}".format(sups["total"])
            current_suppliers_details_worksheet["D6"] = "Ac_{}".format(
                sups["participants"]
            )
            current_suppliers_details_worksheet["D7"] = "Ac_{}".format(
                percentage_participation
            )
            current_suppliers_details_worksheet["C8"] = "Nb_"

            current_suppliers_details_worksheet.merge_cells("C2:E2")
            data[f'category.unique_reference.replace("/", "_")_C'] = 25
            data[f'category.unique_reference.replace("/", "_")_D'] = 15
            data[f'category.unique_reference.replace("/", "_")_E'] = 16
            data[f'category.unique_reference.replace("/", "_")_F'] = 28
            data[f'category.unique_reference.replace("/", "_")_G'] = 28
            data[f'category.unique_reference.replace("/", "_")_H'] = 25

            current_suppliers_details_worksheet.append(
                (
                    "",
                    "hc_No",
                    "h_Company Name",
                    "h_Status",
                    "h_Tax PIN Number",
                    "h_Email",
                    "h_Alternative Email",
                    "h_Phone",
                )
            )

            current_suppliers_by_category = current_suppliers.filter(category=category)

            total_for_progress += len(current_suppliers_by_category)

            for count, category_supplier in enumerate(
                current_suppliers_by_category, start=1
            ):
                result += 1
                progress_recorder.set_progress(result, total_for_progress)
                current_suppliers_details_worksheet.append(
                    (
                        "",
                        "Ac_%d" % (count),
                        category_supplier.supplier_name,
                        category_supplier.participation_status,
                        category_supplier.get_tax_pin(),
                        category_supplier.supplier_email,
                        category_supplier.alternative_email,
                        category_supplier.supplier_phone,
                    )
                )
        current_suppliers_worksheet.append(
            (
                "",
                "",
                "",
                "h_Total",
                total_partcipants,
                total_non_participants,
                total_suppliers,
            )
        )
        current_suppliers_worksheet.freeze_panes = current_suppliers_worksheet["D11"]
        current_suppliers_workbook.save(filepath)
        format_excel(filepath, data=data)
        insert_image(
            excel_url=filepath,
            worksheet_name=qed_logo_worksheet,
            anchor=qed_logo_anchor,
        )
        insert_image(
            filepath, buyer_logo_worksheet, buyer_logo_anchor, image_url=buyer_logo_url
        )
        context = {
            "response_message": "Report generated successfully",
            "filepath": filepath.split("/", 1)[1],
        }

        return context

    except Exception as e:
        capture_exception(e)
        return True

    
class FinancialRatiosReport(Task):
    name = "FinancialRatiosReport"
    progress_recorder = None
    result = 0

    def run(self, *args, **kwargs):
        self.progress_recorder = ProgressRecorder(self)
        context = self.report(kwargs["category_id"])
        return context

    def report(self, category_id):
        try:
            category = Category.objects.get(id=category_id)
            company = category.tender.company

            section = category.sections.filter(name="Financial Ratios").first()

            time = datetime.datetime.now()
            dir_name = Path(
                "media/category/ratios/reports/{}/{}/{}".format(
                    time.year, time.month, time.day
                )
            )
            dir_name.mkdir(parents=True, exist_ok=True)
            time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
            filepath = "{}/{}_{}_{}.xlsx".format(
                dir_name, category_id, company.company_name.replace(" ", "_"), time_only
            )

            workbook = Workbook()
            summary_sheet = workbook["Sheet"]
            summary_sheet.title = "Summary"

            qed_logo_anchor = "H1"
            buyer_logo_anchor = "B1"
            qed_logo = Image("static/core/img/Tendersure_Logo.png")

            if company.company_logo_url is not None and company.company_logo_url != "":
                buyer_logo = get_file_path(company.company_logo_url)
            else:
                buyer_logo = "static/core/img/no-company-image128-128.png"

            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet.append(("", "", ""))
            summary_sheet["B6"] = f"Bnb_{category.name.title()}"
            summary_sheet["B7"] = ""
            summary_sheet["B8"] = f"Bnb_{'Financial Ratio Summary Report'}"
            summary_sheet.append(("", "", ""))
            summary_sheet.merge_cells("B6:E6")
            summary_sheet.merge_cells("B8:E8")
            summary_sheet.append(
                (
                    " ", "hc_#", "hl_Supplier Name", "hc_Score Before QA",
                    "hc_Score After QA", "hc_Variance", "hl_Comments",
                )
            )

            data = {
                "Summary_B": 5, "Summary_C": 41, "Summary_D": 15,
                "Summary_E": 15, "Summary_F": 15, "Summary_G": 15,
            }

            progress_recorder = ProgressRecorder(self)
            result = 0

            responsive_bidders = category.technical_participants

            supplier_count = 1
            for supplier in responsive_bidders:
                if section is not None:
                    score_before_qa = SupplierSectionScore.objects.filter(
                        supplier=supplier, section=section
                    ).first()

                    question_in_qa = QualityAssuranceResponse.objects.filter(
                        quality_assurance_question__question__section=section,
                        supplier__id=supplier.id,
                    )
                    if question_in_qa:
                        total_score = 0
                        for question in question_in_qa:
                            if question.score_after_qa is not None:
                                total_score += float(question.score_after_qa)
                        score_after_qa = total_score
                    else:
                        score_after_qa = 0

                supplier_name = supplier.short_name.replace("/", "_") if supplier.short_name is not None else supplier.company_name.replace("/", "_")
                supplier_worksheet = workbook.create_sheet(
                    supplier_name
                )

                data[f"{supplier_name}_B"] = 20
                data[f"{supplier_name}_C"] = 15
                data[f"{supplier_name}_D"] = 15
                data[f"{supplier_name}_E"] = 15
                data[f"{supplier_name}_F"] = 15
                data[f"{supplier_name}_G"] = 15
                data[f"{supplier_name}_H"] = 15

                result += 1
                progress_recorder.set_progress(result, len(responsive_bidders))

                if score_before_qa is not None:
                    variance = float(score_after_qa) - float(score_before_qa.score)
                else:
                    variance = 0

                summary_sheet.append(
                    (
                        " ",
                        f"Ac_{supplier_count}",
                        f"Al_{supplier_name}",
                        f"Nc_{float(score_before_qa.score)}"
                        if score_before_qa is not None
                        else f"Nc_{0}",
                        f"Nc_{score_after_qa}",
                        f"Nc_{variance}",
                    )
                )

                supplier_count += 1

                supplier_worksheet.append(("", ""))
                supplier_worksheet[
                    "B2"
                ] = f"Bnb_Supplier: {supplier.company_name.title()}"
                supplier_worksheet["B3"] = ""
                supplier_worksheet["B4"] = f"Bnb_{'Latest Financial Statements'}"
                supplier_worksheet.append(("", ""))
                supplier_worksheet.append(("", ""))

                worksheet_headers = [
                    " ",
                    "hl_Description",
                    "hc_Supplier (KES)",
                    "hc_QA (KES)",
                    "hc_Variance (KES)",
                    "hc_Comment",
                ]

                supplier_worksheet.merge_cells("B2:E2")
                supplier_worksheet.merge_cells("B4:E4")

                supplier_worksheet.append(worksheet_headers)
                financial_ratio = FinancialRatio.objects.filter(
                    supplier=supplier, section=section
                ).first()
                if financial_ratio is not None:
                    equity = [
                        "",
                        "Equity",
                        f"Tr_{float(financial_ratio.equity)}"
                        if financial_ratio.equity is not None
                        else 0,
                        f"Tr_{float(financial_ratio.equity_after_qa)}"
                        if financial_ratio.equity_after_qa is not None
                        else 0,
                    
                    ]
                    long_term_loans = [
                        "",
                        "Long Term Loans (Debt)",
                        f"Tr_{float(financial_ratio.debtors)}"
                        if financial_ratio.debtors is not None
                        else 0,
                        f"Tr_{float(financial_ratio.debtors_after_qa)}"
                        if financial_ratio.debtors_after_qa is not None
                        else 0,
                    ]
                    curr_liabilities = [
                        "",
                        "Current Liabilities",
                        f"Tr_{float(financial_ratio.curr_liabilities)}"
                        if financial_ratio.curr_liabilities is not None
                        else 0,
                        f"Tr_{float(financial_ratio.curr_liabilities_after_qa)}"
                        if financial_ratio.curr_liabilities_after_qa is not None
                        else 0,
                    
                    ]
                    fixed_assets = [
                        "",
                        "Fixed Assets",
                        f"Tr_{float(financial_ratio.fixed_assets)}"
                        if financial_ratio.fixed_assets is not None
                        else 0,
                        f"Tr_{float(financial_ratio.fixed_assets_after_qa)}"
                        if financial_ratio.fixed_assets_after_qa is not None
                        else 0,
                        
                    ]
                    curr_assets = [
                        "",
                        "Curent Assets",
                        f"Tr_{float(financial_ratio.current_assets)}"
                        if financial_ratio.current_assets is not None
                        else 0,
                        f"Tr_{float(financial_ratio.current_assets_after_qa)}"
                        if financial_ratio.current_assets_after_qa is not None
                        else 0,
                    ]
                    cash = [
                        "",
                        "Cash",
                        f"Tr_{float(financial_ratio.cash)}"
                        if financial_ratio.cash is not None
                        else 0,
                        f"Tr_{float(financial_ratio.cash_after_qa)}"
                        if financial_ratio.cash_after_qa is not None
                        else 0,
                    ]
                    turnover = [
                        "",
                        "Turnover",
                        f"Tr_{float(financial_ratio.turnover)}"
                        if financial_ratio.turnover is not None
                        else 0,
                        f"Tr_{float(financial_ratio.turnover_after_qa)}"
                        if financial_ratio.turnover_after_qa is not None
                        else 0,
                    ]
                    gross_profit = [
                        "",
                        "Gross Profit",
                        f"Tr_{float(financial_ratio.gross_profit)}"
                        if financial_ratio.gross_profit is not None
                        else 0,
                        f"Tr_{float(financial_ratio.gross_profit_after_qa)}"
                        if financial_ratio.gross_profit_after_qa is not None
                        else 0,
                    ]
                    net_profit = [
                        "",
                        "Net Profit",
                        f"Tr_{float(financial_ratio.net_profit)}"
                        if financial_ratio.net_profit is not None
                        else 0,
                        f"Tr_{float(financial_ratio.net_profit_after_qa)}"
                        if financial_ratio.net_profit_after_qa is not None
                        else 0,
                    ]

                    variances = variance_ratios(financial_ratio.id)

                    equity.extend([f'Tr_{variances["equity"]["variance"]}', variances["equity"]["comment"]])
                    long_term_loans.extend([f'Tr_{variances["debtors"]["variance"]}', variances["debtors"]["comment"]])
                    curr_liabilities.extend([f'Tr_{variances["curr_liabilities"]["variance"]}', variances["curr_liabilities"]["comment"]])
                    fixed_assets.extend([f'Tr_{variances["f_assets"]["variance"]}', variances["f_assets"]["comment"]])
                    curr_assets.extend([f'Tr_{variances["c_assets"]["variance"]}', variances["c_assets"]["comment"]])
                    cash.extend([f'Tr_{variances["cash"]["variance"]}', variances["cash"]["comment"]])
                    turnover.extend([f'Tr_{variances["turnover"]["variance"]}', variances["turnover"]["comment"]])
                    gross_profit.extend([f'Tr_{variances["g_profit"]["variance"]}', variances["g_profit"]["comment"]])
                    net_profit.extend([f'Tr_{variances["n_profit"]["variance"]}', variances["n_profit"]["comment"]])

                    supplier_worksheet.append(equity)
                    supplier_worksheet.append(long_term_loans)
                    supplier_worksheet.append(curr_liabilities)
                    supplier_worksheet.append(fixed_assets)
                    supplier_worksheet.append(curr_assets)
                    supplier_worksheet.append(cash)
                    supplier_worksheet.append(turnover)
                    supplier_worksheet.append(gross_profit)
                    supplier_worksheet.append(net_profit)
                    supplier_worksheet.append(("", "", ""))
                    supplier_worksheet.append(("", "", ""))
                    supplier_worksheet.append(("", "", ""))

                    ratio_headers = [
                        " ",
                        "hl_Ratios",
                        "hc_Before QA",
                        "hc_After QA",
                        "hc_Variance",
                        "hc_Score Before QA",
                        "hc_Score After QA",
                        "hc_Variance",
                    ]
                    supplier_worksheet.append(ratio_headers)
                    ratios = ratios_before_after_qa(instance_id=financial_ratio.id)
                    debt_equity = [
                        "",
                        "Debt/equity ratio",
                        "Tc_{}".format(ratios["debt_equity"]["before_qa"]),
                        "Tc_{}".format(ratios["debt_equity"]["after_qa"]),
                        "Tc_{}".format(ratios["debt_equity"]["variance"]),
                    ]
                    current_ratio = [
                        "",
                        "Current ratio",
                        "Tc_{}".format(ratios["current_ratio"]["before_qa"]),
                        "Tc_{}".format(ratios["current_ratio"]["after_qa"]),
                        "Tc_{}".format(ratios["current_ratio"]["variance"]),
                    ]
                    cash_ratio = [
                        "",
                        "Cash ratio",
                        "Tc_{}".format(ratios["cash_ratio"]["before_qa"]),
                        "Tc_{}".format(ratios["cash_ratio"]["after_qa"]),
                        "Tc_{}".format(ratios["cash_ratio"]["variance"]),
                    ]
                    gp_margin = [
                        "",
                        "GP Margin",
                        "Tc_{}".format(ratios["gp_margin"]["before_qa"]),
                        "Tc_{}".format(ratios["gp_margin"]["after_qa"]),
                        "Tc_{}".format(ratios["gp_margin"]["variance"]),
                    ]
                    np_margin = [
                        "",
                        "NP Margin",
                        "Tc_{}".format(ratios["np_margin"]["before_qa"]),
                        "Tc_{}".format(ratios["np_margin"]["after_qa"]),
                        "Tc_{}".format(ratios["np_margin"]["variance"]),
                    ]

                    scores = ratio_scores_before_after_qa(
                        supplier_id=supplier.id, section_id=section.id
                    )

                    for score in scores:
                        if score["index"] == 0:
                            ratio_debt_equity = [
                                "Tc_{}".format(score["before_qa"]),
                                "Tc_{}".format(score["after_qa"]),
                                "Tc_{}".format(score["variance"]),
                            ]
                            debt_equity.extend(ratio_debt_equity)
                        if score["index"] == 1:
                            ratio_current = [
                                "Tc_{}".format(score["before_qa"]),
                                "Tc_{}".format(score["after_qa"]),
                                "Tc_{}".format(score["variance"]),
                            ]
                            current_ratio.extend(ratio_current)
                        if score["index"] == 2:
                            ratio_cash = [
                                "Tc_{}".format(score["before_qa"]),
                                "Tc_{}".format(score["after_qa"]),
                                "Tc_{}".format(score["variance"]),
                            ]
                            cash_ratio.extend(ratio_cash)
                        if score["index"] == 3:
                            ratio_gp = [
                                "Tc_{}".format(score["before_qa"]),
                                "Tc_{}".format(score["after_qa"]),
                                "Tc_{}".format(score["variance"]),
                            ]
                            gp_margin.extend(ratio_gp)
                        if score["index"] == 4:
                            ratio_np = [
                                "Tc_{}".format(score["before_qa"]),
                                "Tc_{}".format(score["after_qa"]),
                                "Tc_{}".format(score["variance"]),
                            ]
                            np_margin.extend(ratio_np)

                    supplier_worksheet.append(debt_equity)
                    supplier_worksheet.append(current_ratio)
                    supplier_worksheet.append(cash_ratio)
                    supplier_worksheet.append(gp_margin)
                    supplier_worksheet.append(np_margin)

                formulas_row = []
                formulas_row.extend(["", "", "", "", "qT_Total"])

                formulas_row.extend(["qT_{}".format(f"= SUM(F21:F25)")])
                formulas_row.extend(["qT_{}".format(f"= SUM(G21:G25)")])
                formulas_row.extend(["qT_{}".format(f"= SUM(H21:H25)")])
                supplier_worksheet.append(tuple(formulas_row))

            workbook.save(filepath)
            format_excel(filepath, data=data)
            # insert_image(
            #     excel_url=filepath,
            #     worksheet_name="Summary",
            #     anchor=qed_logo_anchor,
            # )

            insert_image(
                excel_url=filepath,
                worksheet_name="Summary",
                anchor=buyer_logo_anchor,
                image_url=buyer_logo,
            )

            time = datetime.datetime.now()
            with open(filepath, "rb") as l:
                storage = PrivateMediaStorage()
                url = f"tender/category/reports/{time.year}/{time.month}/{time.day}/{os.path.basename(l.name)}"
                storage.save(url, l)

                category_report, created = CategoryReport.objects.update_or_create(
                    category_id=category_id, 
                    defaults={"financial_ratios": url}
                )
                path = category_report.financial_ratios .url
                context = {
                    "filepath": path,
                    "response_message": "Report generated successfully",
                }
                return context

        except Exception as e:
            capture_exception(e)
            context = {"response_message": "Error generating report", "errors": f"{e}"}
            return context


# financial reports
@shared_task(bind=True)
def financial_pdf_report(self,supplier_id, category_id):
    """
    Supplier RFQ PDF responses report
    """
    supplier = (
        apps.apps.get_model("suppliers", "Supplier").objects.filter(id=supplier_id).first()
    )
    category = Category.objects.filter(id=category_id).first()

    try:
        if supplier is not None and category is not None:
            time = datetime.datetime.now()
            dir_name = Path("reports/{}/{}".format(time.year, time.month))
            dir_name.mkdir(parents=True, exist_ok=True)

            time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
            filepath = "{}/{}_{}_{}.pdf".format(
                dir_name, category.id, supplier.company_name.replace(" ", "_").replace("&", "_"),
                time_only)

            rfq_responses = ItemResponse.objects.filter(
                supplier=supplier, item__category_id=category.id)

            context = {
                "category": category, "supplier": supplier, "rfq_responses": rfq_responses,
                "time": time,
            }

            pdf = Render.render(
                "financial_summary_report.html", context,
            )

            pdf_file = BytesIO(pdf.content)

            category_report, created = CategoryReport.objects.update_or_create(
                category_id=category_id,
                defaults={"financial_summary_report": File(pdf_file, filepath)},
            )

            # supplier_response, created = SupplierResponse.objects.update_or_create(
            #     supplier_id=supplier_id, category_id=category_id,
            #     defaults={"document_url": File(pdf_file, filepath)},
            # )

            context = {
                "file": category_report.financial_summary_report,
                "response_message": "Report Generated Successfully"
            }

            print(context)
            return context
    except Exception as e:
        print(e)
        capture_exception(e)
        context = {
            "response_message": "Report generation error",
            "errors": f"{e}"
        }
        return context


def rfq_job_summary_report(category_id):
    """
    RFQ Summary Job Level
    """
    category = Category.objects.filter(id=category_id).first()
    buyer = apps.apps.get_model("buyer", "Company").objects.filter(
        id=category.rfq.company.id).first()
    
    if category is not None and buyer is not None:
        time = datetime.datetime.now()
        dir_name = Path("rfq_reports/{}/{}".format(time.year, time.month))
        dir_name.mkdir(parents=True, exist_ok=True)
        time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
        filepath = "{}/{}_{}_{}.pdf".format(
            dir_name, category.id, time_only,
            buyer.company_name.replace(" ", "_").replace("&", "_"))
        participants = category.participants

        context = {
            "category": category, "participants": participants["participants"],
            "time": time, "buyer": buyer,
        }

        pdf = Render.render("financial_job_summary_report.html", context)

        pdf_file = BytesIO(pdf.content)
        category_report, created = CategoryReport.objects.update_or_create(
            category_id=category_id,
            defaults={"category_rfq_pdf": File(pdf_file, filepath)},
        )

        context = {"file": category_report.financial_job_summary_report}

        return context


class TenderFinancialReport(Task):
    """
    Category RFQ Financial Report(Lowest Item, Lowest Supplier,Savings and Individual sheets)
    """

    name = "TenderFinancialReport"
    progress_recorder = None
    result = 0

    def run(self, *args, **kwargs):
        self.progress_recorder = ProgressRecorder(self)
        context = self.report(kwargs["category_id"])
        return context

    def report(self, category_id):
        category = Category.objects.filter(id=category_id).first()
        job = category.rfq
        items = category.items

        financial_participants = category.financial_participants

        # has_blank=False, has_outlier=False
        lowest_cumm_suppliers_ranked = SupplierFinancialTotal.objects.filter(
            category=category).order_by("score")
        lowest_cumm_suppliers_opp = SupplierFinancialTotal.objects.filter(
            Q(category=category), ~Q(has_blank=False) | ~Q(has_outlier=False)).order_by("score")
        lowest_cumm_suppliers = list(lowest_cumm_suppliers_ranked) + list(
            lowest_cumm_suppliers_opp)

        # edit
        total_for_progress = len(financial_participants) + len(items)
        
        # get tender total bidders 
        category_total_bidders = len(category.invited_suppliers)
        
        category_responsive_bidders = financial_participants["count"]
        try:
            percentage_responses = "{}%".format(
                round(((category_responsive_bidders / category_total_bidders) * 100))
            )
        except Exception as e:
            percentage_responses = 0

        time = datetime.datetime.now()
        dir_name = Path("media/rfq_reports/{}/{}".format(time.year, time.month))
        dir_name.mkdir(parents=True, exist_ok=True)

        time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
        filepath = "{}/{}_{}_{}.xlsx".format(
            dir_name, job.id, job.company.company_name.replace(" ", "_"), time_only
        )
        # delete similar reports run this month
        match_string = "{}_{}".format(
            job.id, job.company.company_name.replace(" ", "_")
        )
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")
        workbook = Workbook()
        qed_logo = Image("static/core/img/Tendersure_Logo.png")

        if (
            job.company.company_logo_url is not None
            and job.company.company_logo_url != ""
        ):
            buyer_logo = get_file_path(job.company.company_logo_url)
        else:
            buyer_logo = "static/core/img/no-company-image128-128.png"

        qed_logo_anchor = "F1"
        buyer_logo_anchor = "B1"

        workbook.create_sheet("LOWEST ITEM COST REPORT")
        pricing_worksheet = workbook["LOWEST ITEM COST REPORT"]

        letters = list(string.ascii_uppercase)
        num_cols = 200
        excel_cols = []
        for i in range(0, num_cols - 1):
            n = i // 26
            m = n // 26
            i -= n * 26
            n -= m * 26
            col = (
                letters[m - 1] + letters[n - 1] + letters[i]
                if m > 0
                else letters[n - 1] + letters[i]
                if n > 0
                else letters[i]
            )
            excel_cols.append(col)

        data = {
            "category_responsive_bidders": category_responsive_bidders,
            "category_total_bidders": int(category_total_bidders),
            "category": category, "items": items,
            "lowest_cumm_suppliers": lowest_cumm_suppliers,
            "total_for_progress": total_for_progress,
            "excel_cols": excel_cols,
            "percentage_responses": percentage_responses,
            "participants": financial_participants["participants"],
            "lowest_cumm_suppliers_ranked": lowest_cumm_suppliers_ranked,
            "lowest_cumm_suppliers_opp": lowest_cumm_suppliers_opp,
        }
        
        self.create_lowest_cost_report_worksheet(
            pricing_worksheet=pricing_worksheet, data=data
        )
        self.create_lowest_supplier_worksheet(workbook=workbook, data=data)
        self.create_savings_worksheet(workbook=workbook, data=data)
        # self.create_individual_worksheets(workbook=workbook, data=data)

        workbook.remove(workbook["Sheet"])
        workbook.save(filepath)

        f_data = {
            "SAVINGS REPORT_A": 5, "SAVINGS REPORT_B": 5, "SAVINGS REPORT_C": 31,
            "SAVINGS REPORT_D": 8, "SAVINGS REPORT_E": 13, "SAVINGS REPORT_F": 17,
            "SAVINGS REPORT_G": 17, "SAVINGS REPORT_H": 17, "SAVINGS REPORT_I": 19,
            "SAVINGS REPORT_J": 13, "SAVINGS REPORT_K": 13, "SAVINGS REPORT_L": 13,
            "SAVINGS REPORT_M": 13, "LOWEST ITEM COST REPORT_A": 5, "LOWEST ITEM COST REPORT_B": 5,
            "LOWEST ITEM COST REPORT_C": 31, "LOWEST ITEM COST REPORT_D": 8, "LOWEST ITEM COST REPORT_E": 13,
            "LOWEST ITEM COST REPORT_F": 13, "LOWEST ITEM COST REPORT_G": 14, "LOWEST ITEM COST REPORT_H": 24,
            "LOWEST ITEM COST REPORT_I": 14, "LOWEST ITEM COST REPORT_J": 24, "LOWEST ITEM COST REPORT_K": 14,
            "LOWEST ITEM COST REPORT_L": 24, "LOWEST ITEM COST REPORT_OTHER": 15, "LOWEST SUPPLIER REPORT_A": 5,
            "LOWEST SUPPLIER REPORT_B": 5, "LOWEST SUPPLIER REPORT_C": 31, "LOWEST SUPPLIER REPORT_D": 8,
            "LOWEST SUPPLIER REPORT_E": 13, "LOWEST SUPPLIER REPORT_F": 13, "LOWEST SUPPLIER REPORT_G": 24,
            "LOWEST SUPPLIER REPORT_H": 24, "LOWEST SUPPLIER REPORT_I": 24, "LOWEST SUPPLIER REPORT_J": 24,
            "LOWEST SUPPLIER REPORT_K": 24, "LOWEST SUPPLIER REPORT_OTHER": 24,
        }

        format_excel(filepath, data=f_data)
        # insert_image(
        #     excel_url=filepath, worksheet_name="LOWEST ITEM COST REPORT",
        #     anchor=qed_logo_anchor)
        insert_image(
            filepath, worksheet_name="LOWEST ITEM COST REPORT",
            anchor=buyer_logo_anchor, image_url=buyer_logo)

        try:
            time = datetime.datetime.now()
            with open(filepath, "rb") as f:
                report = File(f, name=f.name)

                report, created = CategoryReport.objects.update_or_create(
                    category_id=category_id, defaults={"financial_report_excel": report}
                )
                context = {
                    "filepath": report.financial_report_excel.url,
                    "response_message": "Report generated successfully",
                }
                return context

        except Exception as e:
            capture_exception(e)
            print(e)
            context = {
                "response_message": "Error generating report"
            }
            return context

    def create_lowest_cost_report_worksheet(self, pricing_worksheet, data):
        category_responsive_bidders = data["category_responsive_bidders"]
        category_total_bidders = data["category_total_bidders"]
        category = data["category"]
        items = data["items"]
        lowest_cumm_suppliers = data["lowest_cumm_suppliers"]
        total_for_progress = data["total_for_progress"]
        excel_cols = data["excel_cols"]
        percentage_responses = data["percentage_responses"]

        try:
            rfq_current_cost = 0
            rfq_eprocure_cost = 0
            rfq_total_potential_savngs = 0

            details_summary = [
                ("", "", "", ""), ("", "", "", ""), ("", "", "", ""),
                ("", "", "", ""), ("", "Bnb_Lowest Cost Report"),
                ("", f"Bnb_Category: {category.name.title()}", "", ""),
                ("", "", "", ""),
                ("", "Number of bidders", "", f"Ac_{category_total_bidders}"),
                (
                    "", "Responsive bidders", "",
                    "Ac_{}".format(category_responsive_bidders),
                ),
                ("", "Percentage responses", "", f"Ac_{percentage_responses}"),
                ("", "Currency", "", "Ac_{}".format(category.currency.initials)),
                ("", "", "", ""),
            ]
            for row in details_summary:
                pricing_worksheet.append(row)

            pricing_worksheet.merge_cells("B5:C5")
            pricing_worksheet.merge_cells("B6:H6")
            pricing_worksheet.merge_cells("B8:C8")
            pricing_worksheet.merge_cells("B9:C9")
            pricing_worksheet.merge_cells("B10:C10")
            pricing_worksheet.merge_cells("B11:C11")

            supplier_headers = ["", "", "", "", "", "", "", "", "", "", ""]

            items_header = [
                "", "hc_#", "h_Item Description", "hc_Qty", "h_UOM",
                "hc_Lowest Cost 1", "h_Lowest Supplier 1", "hc_Lowest Cost 2",
                "h_Lowest Supplier 2", "hc_Lowest Cost 3", "h_Lowest Supplier 3",
            ]

            no_of_items = len(items)
            no_of_rows_at_the_top = 14
            supplier_totals_row = no_of_rows_at_the_top + no_of_items
            formulas_row = []

            formulas_row.extend(
                ["qT_{}".format(f"= SUM(F15:F{supplier_totals_row})"), ""]
            )
            formulas_row.extend(
                ["qT_{}".format(f"= SUM(H15:H{supplier_totals_row})"), ""]
            )
            formulas_row.extend(
                ["qT_{}".format(f"= SUM(J15:J{supplier_totals_row})"), ""]
            )
            pricing_worksheet.merge_cells("C6:H6")
            c = 0

            for count, supplier in enumerate(lowest_cumm_suppliers, start=1):
                supplier_total_row_name = excel_cols[11 + c]
                formulas_row.extend(
                    [
                        "qT_{}".format(
                            f"= SUM({supplier_total_row_name}15:{supplier_total_row_name}{supplier_totals_row})"
                        ),
                    ]
                )

                supplier_headers.append(
                    "hc_%s" % supplier.supplier.company_name.title()
                )
                items_header.append("hc_Cost")
                c += 1

            supplier_headers = tuple(supplier_headers)
            pricing_worksheet.append(supplier_headers)
            items_header = tuple(items_header)
            pricing_worksheet.append(items_header)

            j = 1
            for item in items:
                outlier_score = item.outlier_score
                responses = item.responses

                try:
                    rfq_current_cost += float(item.current_price)
                    rfq_eprocure_cost += item.eprocure_supplier_cost(responses)
                    rfq_total_potential_savngs += item.savings(responses)
                except:
                    rfq_current_cost += 0
                    rfq_eprocure_cost += 0
                    rfq_total_potential_savngs += 0

                if item.item_savings_value is None:
                    item_savings = item.savings(responses)
                else:
                    item_savings = item.item_savings_value

                if item.eprocure_total_value is None:
                    item_eprocure_total = item.eprocure_total(responses)
                else:
                    item_eprocure_total = item.eprocure_total_value

                item_current_total = item.current_total
                total_items_current_total = item_current_total * item.quantity
                item_row = []
                
                tendersure_suppliers = item.multi_item_tender_eprocure_suppliers(
                    responses
                )
                price_1 = tendersure_suppliers[0]["price"]
                price_2 = tendersure_suppliers[1]["price"]
                price_3 = tendersure_suppliers[2]["price"]

                try:
                    if item.second_description is None:
                        second_description = "Ac_N/A"
                    else:
                        second_description = item.second_description
                except:
                    second_description = "Ac_N/A"

                try:
                    if (
                        item_eprocure_total > total_items_current_total
                        and total_items_current_total != 0
                    ):
                        savings_lowest_supplier = "Current Supplier"
                        price_3 = price_2
                        price_2 = price_1
                        price_1 = total_items_current_total
                        item_row.extend(
                            [
                                "",
                                "Ac_{}".format(item.item_number),
                                item.item_description,
                                "Ac_{}".format(item.quantity),
                                second_description,
                                "Tr_{}".format(f"{price_1:}"),
                                savings_lowest_supplier,
                                "Tr_{}".format(f"{price_2:}"),
                                tendersure_suppliers[0]["supplier"],
                                "Tr_{}".format(f"{price_3:}"),
                                tendersure_suppliers[1]["supplier"],
                            ]
                        )
                    else:
                        savings_lowest_supplier = tendersure_suppliers[0]["supplier"]
                        item_row.extend(
                            [
                                "",
                                "Ac_{}".format(item.item_number),
                                item.item_description,
                                "Ac_{}".format(item.quantity),
                                second_description,
                                "Tr_{}".format(f"{price_1:}"),
                                savings_lowest_supplier,
                                "Tr_{}".format(f"{price_2:}"),
                                tendersure_suppliers[1]["supplier"],
                                "Tr_{}".format(f"{price_3:}"),
                                tendersure_suppliers[2]["supplier"],
                            ]
                        )
                except:
                    item_row.extend(
                        [
                            "",
                            "Ac_{}".format(item.item_number),
                            item.item_description,
                            "Ac_{}".format(item.quantity),
                            second_description,
                            "Tr_{}".format(f"{price_1:}"),
                            tendersure_suppliers[0]["supplier"],
                            "Tr_{}".format(f"{price_2:}"),
                            tendersure_suppliers[1]["supplier"],
                            "Tr_{}".format(f"{price_3:}"),
                            tendersure_suppliers[2]["supplier"],
                        ]
                    )
                
                for count, supplier in enumerate(lowest_cumm_suppliers, start=1):
                    if not supplier.supplier:
                        supplier_total = 0
                        item_row.append(supplier_total)
                    else:
                        item_response = ItemResponse.objects.filter(
                            supplier=supplier.supplier, rfq_item=item).first()
                        if item_response is not None:
                            supplier_total = item_response.total_price
                        else:
                            supplier_total = 0
                        s_total = round(supplier_total, 2)

                        if (
                            ItemResponse.objects.filter(rfq_item_id=item.id).count()
                            <= 2
                        ):
                            supplier_total = "Tr_{}".format(f"{s_total:}")
                        else:
                            if float(s_total) < float(outlier_score):
                                supplier_total = "O_{}".format(f"{s_total:}")
                            else:
                                supplier_total = "Tr_{}".format(f"{s_total:}")
                        item_row.append(supplier_total)

                item_row = tuple(item_row)
                pricing_worksheet.append(item_row)

            colums_totals = []
            colums_totals.extend(["", "", "qT_Total", "", ""])
            colums_totals.extend(formulas_row)

            pricing_worksheet.append(tuple(colums_totals))
            pricing_worksheet.append(("", "", ""))
            pricing_worksheet.append(("", ""))
            pricing_worksheet.append(("", "hc_#", "Nmrl_3_Notes"))
            pricing_worksheet.append(
                (
                    "", "Ac_1",
                    "Nmrl_3_Outlier: The prices highlighted in this colour are outliers which means they are 50% lower than the median and are not considered in the analysis",
                )
            )
            pricing_worksheet.freeze_panes = pricing_worksheet["F15"]
            return
        except Exception as e:
            capture_exception(e)
            return

    def create_lowest_supplier_worksheet(self, workbook, data):
        try:
            category = data["category"]
            excel_cols = data["excel_cols"]
            rfq_items = data["items"]
            participants = data["participants"]
            category_total_bidders = data["category_total_bidders"]
            category_responsive_bidders = data["category_responsive_bidders"]
            percentage_responses = data["percentage_responses"]
            lowest_cumm_suppliers_ranked = data["lowest_cumm_suppliers_ranked"]
            lowest_cumm_suppliers_opp = data["lowest_cumm_suppliers_opp"]
            lowest_cumm_suppliers = data["lowest_cumm_suppliers"]
            total_for_progress = data["total_for_progress"]

            rfq_current_cost = 0
            rfq_eprocure_cost = 0
            rfq_total_potential_savngs = 0

            summary_rows = [
                ("", "Invited bidders", "", "Ac_{}".format(category_total_bidders)),
                (
                    "",
                    "Bidder responses",
                    "",
                    "Ac_{}".format(category_responsive_bidders),
                ),
                ("", "Percentage responses", "", "Ac_{}".format(percentage_responses)),
                ("", "Currency", "", "Ac_{}".format(category.currency.initials)),
            ]

            workbook.create_sheet("LOWEST SUPPLIER REPORT")
            lowest_supplier_worksheet = workbook["LOWEST SUPPLIER REPORT"]

            lowest_supplier_worksheet["B3"] = "Bnb_Category: {}".format(category.name)
            lowest_supplier_worksheet["B2"] = "Bnb_Lowest Supplier Report"
            lowest_supplier_worksheet["B4"] = ""

            for row in summary_rows:
                lowest_supplier_worksheet.append(row)

            empty_row = [""]
            empty_row = tuple(empty_row)
            lowest_supplier_worksheet.append(empty_row)
            lowest_supplier_worksheet.merge_cells("B2:C2")
            lowest_supplier_worksheet.merge_cells("B3:H3")
            lowest_supplier_worksheet.merge_cells("B5:C5")
            lowest_supplier_worksheet.merge_cells("B6:C6")
            lowest_supplier_worksheet.merge_cells("B7:C7")
            lowest_supplier_worksheet.merge_cells("B8:C8")

            s = 4
            supplier_totals_row = 12 + len(rfq_items)
            lowest_cumm_suppliers_totals = []
            # lowest_cumm_suppliers = list(lowest_cumm_suppliers_ranked) + list(
            #     lowest_cumm_suppliers_opp
            # )

            lowest_supplier_sub_headers = ["", "", "", "", ""]
            lowest_supplier_rank_headers = ["", "", "", "", ""]
            lowest_supplier_headers = [
                "",
                "hc_#",
                "h_Item Description",
                "hc_Qty",
                "h_UOM",
            ]

            scores = []
            for count, supplier in enumerate(lowest_cumm_suppliers, start=1):
                # self.result += 1
                # self.progress_recorder.set_progress(self.result, total_for_progress)

                row_name = excel_cols[s + count]
                lowest_cumm_suppliers_totals.append(
                    "qT_{}".format(
                        f"= SUM({row_name}13:{row_name}{supplier_totals_row})"
                    )
                )
                lowest_supplier_sub_headers.append(
                    "hc_%s" % supplier.supplier.company_name.title()
                )
                lowest_supplier_headers.append("hc_Total")
                if not supplier.has_outlier:
                    lowest_supplier_rank_headers.append(f"hc_Rank {count}")
                else:
                    lowest_supplier_rank_headers.append(f"hc_N/A")
                # scores.append(f"Tcb_{supplier.rfq_score}")
                scores.append(f"Tcb_{0}")

            lowest_supplier_headers = tuple(lowest_supplier_headers)
            lowest_supplier_sub_headers = tuple(lowest_supplier_sub_headers)
            lowest_supplier_rank_headers = tuple(lowest_supplier_rank_headers)
            lowest_supplier_worksheet.append(lowest_supplier_rank_headers)
            lowest_supplier_worksheet.append(lowest_supplier_sub_headers)
            lowest_supplier_worksheet.append(lowest_supplier_headers)

            j = 1
            for item in rfq_items:
                self.result += 1
                self.progress_recorder.set_progress(self.result, total_for_progress)

                outlier_score = item.outlier_score
                responses = item.responses
                try:
                    rfq_current_cost += float(item.current_price)
                    rfq_eprocure_cost += item.eprocure_supplier_cost(responses)
                    rfq_total_potential_savngs += item.savings(responses)
                except:
                    rfq_current_cost += 0
                    rfq_eprocure_cost += 0
                    rfq_total_potential_savngs += 0

                try:
                    if item.second_description is None:
                        second_description = "Ac_N/A"
                    else:
                        second_description = item.second_description
                except:
                    second_description = "Ac_N/A"
                   
                lowest_supplier_row = []

                lowest_supplier_row.extend(
                    [
                        "",
                        "Ac_{}".format(item.item_number),
                        item.item_description,
                        "Ac_{}".format(item.quantity),
                        second_description,
                    ]
                )

                for count, supplier in enumerate(lowest_cumm_suppliers, start=1):
                    if not supplier.supplier:
                        supplier_total = 0
                    else:
                        item_response = ItemResponse.objects.filter(supplier=supplier.supplier, rfq_item=item).first()
                        if item_response is not None:
                            # if item_response.total is not None:
                            #     supplier_total = show(item_response.total)
                            if item_response.total_price is not None:
                                supplier_total = item_response.total_price
                        else:
                            supplier_total = 0
                        s_total = round(supplier_total, 2)
                        if float(s_total) < float(outlier_score):
                            supplier_total = "O_{}".format(f"{s_total:}")
                        else:
                            supplier_total = "Tr_{}".format(f"{s_total:}")
                        lowest_supplier_row.append(supplier_total)

                lowest_supplier_row = tuple(lowest_supplier_row)
                lowest_supplier_worksheet.append(lowest_supplier_row)

                j += 1

            lowest_supplier_totals = []
            lowest_supplier_totals.extend(["", "", "qT_Total", "", ""])
            lowest_supplier_totals.extend(lowest_cumm_suppliers_totals)
            lowest_supplier_worksheet.append(tuple(lowest_supplier_totals))
            lowest_supplier_worksheet.append(empty_row)
            lowest_supplier_worksheet.append(empty_row)
            lowest_supplier_worksheet.append(("", "Ac_#", "Nmrl_3_Notes"))
            lowest_supplier_worksheet.append(
                (
                    "",
                    "Ac_1",
                    "Nmrl_3_Rank: This is the ranking of suppliers with the lowest cost and is not an outlier",
                )
            )
            lowest_supplier_worksheet.append(
                (
                    "",
                    "Ac_2",
                    "Nmrl_3_Outlier: The prices highlighted in this colour are "
                    "outliers which means that they are 50% lower than the median "
                    "and are not considered in the analysis.",
                )
            )
            lowest_supplier_worksheet.freeze_panes = lowest_supplier_worksheet["F13"]
            return
        except Exception as e:
            capture_exception(e)
            return

    def create_savings_worksheet(self, workbook, data):
        workbook.create_sheet("SAVINGS REPORT")
        savings_worksheet = workbook["SAVINGS REPORT"]

        try:
            category = data["category"]
            rfq_items = data["items"]
            category_total_bidders = data["category_total_bidders"]
            category_responsive_bidders = data["category_responsive_bidders"]
            percentage_responses = data["percentage_responses"]
            empty_row = [""]
            empty_row = tuple(empty_row)
            total_for_progress = data["total_for_progress"]

            details_summary = [
                ("", "Bnb_Savings Report"),
                ("", f"Bnb_Category: {category.name}", "", "", ""),
                ("", "", "", ""),
                ("", "Number of bidders", "", "Ac_{}".format(category_total_bidders)),
                (
                    "",
                    "Bidder responses",
                    "",
                    "Ac_{}".format(category_responsive_bidders),
                ),
                ("", "Percentage responses", "", "Ac_{}".format(percentage_responses)),
                ("", "Currency", "", "Ac_{}".format(category.currency.initials)),
            ]

            savings_worksheet.append(empty_row)

            for row in details_summary:
                savings_worksheet.append(row)

            savings_worksheet.append(empty_row)

            savings_items_header = [
                "",
                "hc_#",
                "h_Item Description",
                "hc_Qty",
                "h_UOM",
                "h_Current Supplier",
                "hc_Current Price",
                "hc_Current Cost",
                "h_Lowest Supplier",
                "hc_Lowest Cost",
                "hc_Unit Savings",
                "hc_Total Savings",
                "hc_% Savings",
            ]

            savings_worksheet.append(savings_items_header)
            savings_worksheet.merge_cells("B2:C2")
            savings_worksheet.merge_cells("B3:H3")
            savings_worksheet.merge_cells("B4:C4")
            savings_worksheet.merge_cells("B5:C5")
            savings_worksheet.merge_cells("B6:C6")
            savings_worksheet.merge_cells("B7:C7")
            savings_worksheet.merge_cells("B8:C8")

            j = 1
            for item in rfq_items:
                responses = item.responses

                savings_item_row = []
                item_savings = item.savings(responses)

                if item.eprocure_total_value is None:
                    item_eprocure_total = item.eprocure_total(responses)
                else:
                    item_eprocure_total = item.eprocure_total_value

                item_current_total = item.current_total
                total_items_current_total = item_current_total * item.quantity
                unit_savings = item_savings / item.quantity
                rfq_item_current_total = "T_{}".format(f"{total_items_current_total:}")
                tendersure_suppliers = item.multi_item_tender_eprocure_suppliers(
                    responses
                )
                price_1 = tendersure_suppliers[0]["price"]
                savings_lowest_supplier = ""

                if (
                    item_eprocure_total > total_items_current_total
                    and total_items_current_total != 0
                ):
                    item_savings_cell = "Ar_0"
                    savings_lowest_supplier = "Current Supplier"
                else:
                    item_savings_cell = "Tr_{}".format(unit_savings)
                    savings_lowest_supplier = tendersure_suppliers[0]["supplier"]

                try:
                    if item.second_description is None:
                        second_description = "Ac_N/A"
                    else:
                        second_description = item.second_description
                except:
                    second_description = "Ac_N/A"

                try:
                    if item.current_supplier is None:
                        current_supplier = "Ac_N/A"
                    else:
                        current_supplier = item.current_supplier.title()
                except:
                    current_supplier = "Ac_N/A"

                savings_item_row.extend(
                    [
                        "",
                        "Ac_{}".format(item.item_number),
                        item.item_description,
                        "Ac_{}".format(item.quantity),
                        second_description,
                        current_supplier,
                        "T_{}".format(f"{item_current_total:}")
                        if item_current_total > 0
                        else "Ac_N/A",
                        rfq_item_current_total,
                        savings_lowest_supplier,
                        "Tr_{}".format(f"{price_1:}"),
                        item_savings_cell,
                        "T_{}".format(item_savings),
                        "Ac_{}".format(item.percentage_savings(responses)),
                    ]
                )

                savings_item_row = tuple(savings_item_row)
                savings_worksheet.append(savings_item_row)
                j += 1
            total_percentage_savings = (
                f"=L{10 + len(rfq_items)+1}/H{10 + len(rfq_items)+1}"
            )
            totals_row = [
                "", "", "B_Total", "", "", "", f"Tc_=SUM(G11:G{10+len(rfq_items)})",
                f"Tc_=SUM(H11:H{10 + len(rfq_items)})", "",
                f"Tr_=SUM(J11:J{10 + len(rfq_items)})", f"Tr_=SUM(K11:K{10 + len(rfq_items)})",
                f"Tr_=SUM(L11:L{10 + len(rfq_items)})", "Pc_{}".format(total_percentage_savings),
            ]
            savings_worksheet.append(totals_row)

            savings_worksheet.append(empty_row)
            savings_worksheet.append(["", "hc_#", "Nmrl_3_Notes"])
            savings_worksheet.append(
                [
                    "", "Ac_1",
                    "Nmrl_3_Where the current prices are lower than the lowest price obtained, current price is retained",
                ]
            )
            savings_worksheet.append(
                [
                    "", "Ac_2",
                    "Nmrl_3_N/A means that the current supplier has not been provided",
                ]
            )
            savings_worksheet.freeze_panes = savings_worksheet["D11"]
        except Exception as e:
            print(e)
            capture_exception(e)
        return


class TenderFinancialSummaryReport(Task):
    """
    RFQ Summary Report, summary savings and individual category savings worksheet
    :job-Rfq
    :rfq-category
    """

    name = "TenderFinancialSummaryReport"
    progress_recorder = None
    result = 0

    def run(self, *args, **kwargs):
        self.progress_recorder = ProgressRecorder(self)
        context = self.report(kwargs["rfq_id"])
        return context

    def report(self, job_id):
        job = Tender.objects.filter(id=job_id).first()
        company = job.company
        categories = job.categories
        job_data_all = []

        total_for_progress = len(categories)

        time = datetime.datetime.now()
        dir_name = Path("media/rfq_reports/{}/{}".format(time.year, time.month))
        dir_name.mkdir(parents=True, exist_ok=True)

        time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
        filepath = "{}/{}_{}_{}.xlsx".format(
            dir_name, job.id, job.company.company_name.replace(" ", "_"), time_only)
        # delete similar reports run this month
        match_string = "{}_{}".format(
            job.id, job.company.company_name.replace(" ", "_"))
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")
        workbook = Workbook()
        qed_logo = Image("static/core/img/no-company-image128-128.png")

        workbook.create_sheet("SUMMARY")
        summary_worksheet = workbook["SUMMARY"]

        data = {
            "SUMMARY_B": 5, "SUMMARY_C": 49, "SUMMARY_D": 20, "SUMMARY_E": 16,
            "SUMMARY_F": 16, "SUMMARY_G": 16, "SUMMARY_H": 16, "SUMMARY_I": 16,
            "SUMMARY_J": 16, "SUMMARY_K": 16,
        }
        f_data = {}

        for category in categories:
            self.result += 1
            self.progress_recorder.set_progress(self.result, total_for_progress)
            rfq_participants = category.financial_participants
            participants = rfq_participants["participants"]
            items = category.items

            # get tender participants
            category_total_bidders = len(category.invited_suppliers)
            category_responsive_bidders = rfq_participants["count"]
         
            category_data = {
                "company": company,
                "category": category,
                "category_total_bidders": category_total_bidders,
                "participants": participants,
                "items": items,
                "category_responsive_bidders": category_responsive_bidders,
            }
            job_data_all.append(category_data)

            total_for_progress += len(job_data_all)

            f_data[f"{category.unique_reference.upper()}_A"] = 5
            f_data[f"{category.unique_reference.upper()}_B"] = 5
            f_data[f"{category.unique_reference.upper()}_C"] = 31
            f_data[f"{category.unique_reference.upper()}_D"] = 8
            f_data[f"{category.unique_reference.upper()}_E"] = 13
            f_data[f"{category.unique_reference.upper()}_F"] = 17
            f_data[f"{category.unique_reference.upper()}_G"] = 17
            f_data[f"{category.unique_reference.upper()}_H"] = 17
            f_data[f"{category.unique_reference.upper()}_I"] = 19
            f_data[f"{category.unique_reference.upper()}_J"] = 13
            f_data[f"{category.unique_reference.upper()}_K"] = 13
            f_data[f"{category.unique_reference.upper()}_L"] = 13
            f_data[f"{category.unique_reference.upper()}_M"] = 13

        data.update(f_data)

        report_data = {
            "job_data": job_data_all,
            "total_for_progress": total_for_progress,
        }

        summary_report = self.create_summary_worksheet(
            company, summary_worksheet, report_data
        )
        savings_report = self.create_savings_worksheet(workbook, report_data)

        if summary_report is True and savings_report is True:
            if company.company_logo_url is not None and company.company_logo_url != "":
                buyer_logo_url = get_file_path(company.company_logo_url)
            else:
                buyer_logo_url = "static/core/img/no-company-image128-128.png"
            qed_logo_anchor = "L1"
            qed_logo_worksheet = "SUMMARY"
            buyer_logo_anchor = "B1"
            buyer_logo_worksheet = "SUMMARY"

            workbook.remove(workbook["Sheet"])
            workbook.save(filepath)
            format_excel(filepath, data=data)

            # insert_image(
            #     filepath,
            #     worksheet_name="SUMMARY",
            #     anchor=qed_logo_anchor,
            # )
            insert_image(
                filepath, worksheet_name="SUMMARY", anchor=buyer_logo_anchor,
                image_url=buyer_logo_url)

            try:
                time = datetime.datetime.now()
                with open(filepath, "rb") as f:
                    report = File(f, name=f.name)

                    report, created = JobReport.objects.update_or_create(
                        job_id=job_id, defaults={"job_savings": report})
                    context = {
                        "filepath": report.job_savings.url,
                        "response_message": "Report generated successfully",
                    }
                    return context

            except Exception as e:
                capture_exception(e)
                print(e)
                context = {
                    "response_message": "Error generating report"
                }
        else:
            if summary_report is False:
                context = {"response_message": "Failed to generate summary report"}
            elif savings_report is False:
                context = {"response_message": "Failed to generate savings report"}
            else:
                context = {"response_message": "Failed to generate report"}

        return context

    def create_summary_worksheet(self, company, summary_worksheet, report_data):
        try:
            job_data = report_data["job_data"]
            total_for_progress = report_data["total_for_progress"]

            summary_data = [
                ("", "", "", ""), ("", "", "", ""), ("", "", "", ""),
                ("", "", "", ""), ("", "Bnb_Summary of Savings Report", "", ""),
                ("", "", "", ""), ("", "Number of categories", "", f"Ac_{len(job_data)}"),
                ("", "", "", ""), ("", "", "", ""), ("", "", "", ""),
            ]

            for row in summary_data:
                summary_worksheet.append(row)

            summary_worksheet.append(
                (
                    "", "h_#", "h_Category Description", "h_Category Code",
                    "hc_Current Cost(KES)", "hc_Lowest Cost(KES)", "hc_Potential Savings",
                    "hc_% Savings", "hc_Total Bidders", "hc_Responsive Bidders",
                )
            )
            summary_worksheet.merge_cells("B7:C7")

            # column totals
            total_rfq_bidders = 0
            total_prequal_bidders = 0
            total_resp_bidders = 0
            total_curr_cost = 0
            total_low_cost = 0
            total_savings = 0
            count = 0

            for rfq in job_data:
                count += 1
                rfq_current_cost = 0
                rfq_eprocure_cost = 0
                rfq_total_potential_savngs = 0
                items = rfq["items"]
                category = rfq["category"]
                category_total_bidders = rfq["category_total_bidders"]
                category_responsive_bidders = rfq["category_responsive_bidders"]

                for item in items:
                    responses = item.responses
                    try:
                        rfq_current_cost += item.actual_current_total
                        rfq_eprocure_cost += item.eprocure_supplier_cost(responses)
                        rfq_total_potential_savngs += item.savings(responses)
                    except:
                        rfq_current_cost += 0
                        rfq_eprocure_cost += 0
                        rfq_total_potential_savngs += 0
                try:
                    percentage_savings = "{}%".format(
                        round(((rfq_total_potential_savngs / rfq_current_cost) * 100))
                    )
                except Exception as e:
                    percentage_savings = 0

                summary_worksheet.append(
                    (
                        "",
                        "Ac_%d" % (count),
                        category.name,
                        category.unique_reference.upper(),
                        "T_{}".format(rfq_current_cost),
                        "T_{}".format(rfq_eprocure_cost),
                        "T_{}".format(rfq_total_potential_savngs),
                        "Ac_{}".format(percentage_savings),
                        "Ac_{}".format(category_total_bidders),
                        "Ac_{}".format(category_responsive_bidders),
                    )
                )
                summary_worksheet.merge_cells("B5:D5")

                total_rfq_bidders += category_total_bidders
                total_resp_bidders += category_responsive_bidders
                total_curr_cost += rfq_current_cost
                total_low_cost += rfq_eprocure_cost
                total_savings += rfq_total_potential_savngs

            try:
                buyer_percentage_savings = "{}%".format(
                    round(((total_savings / total_curr_cost) * 100))
                )

            except:
                buyer_percentage_savings = 0

            summary_worksheet.append(
                (
                    "",
                    "",
                    "",
                    "qT_Total",
                    "T_{}".format(total_curr_cost),
                    "T_{}".format(total_low_cost),
                    "T_{}".format(total_savings),
                    "Ac_{}".format(buyer_percentage_savings),
                    "Ac_{}".format(total_rfq_bidders),
                    "Ac_{}".format(total_resp_bidders),
                )
            )
            summary_worksheet.freeze_panes = summary_worksheet["E12"]
            return True
        except Exception as e:
            print(e)
            capture_exception(e)
            return False

    def create_savings_worksheet(self, workbook, report_data):
        rfqs_data = report_data["rfqs_data"]
        total_for_progress = report_data["total_for_progress"]

        try:
            for rfq in rfqs_data:
                category = rfq["category"]
                rfq_items = rfq["items"]
                category_total_bidders = rfq["category_total_bidders"]
                category_responsive_bidders = rfq["category_responsive_bidders"]
                # percentage_responses = rfq["percentage_responses"]
                empty_row = [""]
                empty_row = tuple(empty_row)
                # total_for_progress = rfq["total_for_progress"]
                try:
                    percentage_responses = "{}%".format(
                        round(
                            (
                                (category_responsive_bidders / category_total_bidders)
                                * 100
                            )
                        )
                    )
                except Exception as e:
                    percentage_responses = 0

                # self.result += 1
                # self.progress_recorder.set_progress(self.result, total_for_progress)

                workbook.create_sheet(f"{category.unique_reference.upper()}")
                savings_worksheet = workbook[f"{category.unique_reference.upper()}"]

                details_summary = [
                    ("", f"Bnb_{category.name} Savings Report", ""),
                    ("", "", "", ""),
                    (
                        "",
                        "Number of bidders",
                        "",
                        "Ac_{}".format(category_total_bidders),
                    ),
                    (
                        "",
                        "Bidder responses",
                        "",
                        "Ac_{}".format(category_responsive_bidders),
                    ),
                    (
                        "",
                        "Percentage responses",
                        "",
                        "Ac_{}".format(percentage_responses),
                    ),
                    ("", "Currency", "", "Ac_{}".format(category.currency.initials)),
                ]

                savings_worksheet.append(empty_row)

                for row in details_summary:
                    savings_worksheet.append(row)

                savings_worksheet.append(empty_row)

                savings_items_header = [
                    "",
                    "hc_#",
                    "h_Item Description",
                    "hc_Qty",
                    "h_UOM",
                    "h_Current Supplier",
                    "hc_Current Price",
                    "hc_Current Cost",
                    "h_Lowest Supplier",
                    "hc_Lowest Cost",
                    "hc_Unit Savings",
                    "hc_Total Savings",
                    "hc_% Savings",
                ]

                savings_worksheet.append(savings_items_header)
                savings_worksheet.merge_cells("B2:H2")
                savings_worksheet.merge_cells("B4:C4")
                savings_worksheet.merge_cells("B5:C5")
                savings_worksheet.merge_cells("B6:C6")
                savings_worksheet.merge_cells("B7:C7")

                j = 1
                for item in rfq_items:
                    # self.result += 1
                    # self.progress_recorder.set_progress(self.result, total_for_progress)
                    outlier_score = item.outlier_score
                    responses = item.responses

                    savings_item_row = []
                    if item.item_savings_value is None:
                        item_savings = item.savings(responses)
                    else:
                        item_savings = item.item_savings_value

                    if item.eprocure_total_value is None:
                        item_eprocure_total = item.eprocure_total(responses)
                    else:
                        item_eprocure_total = item.eprocure_total_value

                    try:
                        item_current_total = item.current_total
                        total_items_current_total = item_current_total * item.quantity
                        unit_savings = item_savings / item.quantity
                    except:
                        item_current_total = 0
                        total_items_current_total = 0

                    # rfq_item_current_total = item.quantity * item_current_total

                    rfq_item_current_total = "T_{}".format(
                        f"{total_items_current_total:}"
                    )
                    # if float(rfq_item_current_total) < float(outlier_score):
                    #     # rfq_item_current_total = "O_{}".format(f"{rfq_item_current_total:}")
                    #     rfq_item_current_total = "Ac_N/A".format(
                    #         f"{rfq_item_current_total:}"
                    #     )
                    # else:
                    #     rfq_item_current_total = "T_{}".format(
                    #         f"{rfq_item_current_total:}"
                    #     )

                    tendersure_suppliers = item.multi_item_tender_eprocure_suppliers(responses)
                    price_1 = tendersure_suppliers[0]["price"]
                    savings_lowest_supplier = ""

                    total_savings = item.quantity * item_savings

                    # if (
                    #     item_eprocure_total > item_current_total
                    #     and item_current_total != 0
                    # ):
                    if (
                        item_eprocure_total > total_items_current_total
                        and total_items_current_total != 0
                    ):
                        item_savings_cell = "Ar_0"
                        savings_lowest_supplier = "Current Supplier"
                    else:
                        item_savings_cell = "Tr_{}".format(unit_savings)
                        savings_lowest_supplier = tendersure_suppliers[0]["supplier"]

                    try:
                        if item.second_description is None:
                            second_description = "Ac_N/A"
                        else:
                            second_description = item.second_description
                    except:
                        second_description = "Ac_N/A"

                    try:
                        if item.current_supplier is None:
                            current_supplier = "Ac_N/A"
                        else:
                            current_supplier = item.current_supplier.title()
                    except:
                        current_supplier = "Ac_N/A"

                    # if item_current_total != 0:
                    savings_item_row.extend(
                        [
                            "",
                            "Ac_{}".format(item.item_number),
                            item.item_description,
                            "Ac_{}".format(item.quantity),
                            second_description,
                            current_supplier,
                            "T_{}".format(f"{item_current_total:}")
                            if item_current_total > 0
                            else "Ac_N/A",
                            rfq_item_current_total,
                            savings_lowest_supplier,
                            "Tr_{}".format(f"{price_1:}"),
                            item_savings_cell,
                            "T_{}".format(item_savings),
                            "Ac_{}".format(item.percentage_savings(responses)),
                        ]
                    )

                    savings_item_row = tuple(savings_item_row)
                    savings_worksheet.append(savings_item_row)
                    j += 1
                total_percentage_savings = (
                    f"=L{9 + len(rfq_items)+1}/H{9 + len(rfq_items)+1}"
                )
                totals_row = [
                    "",
                    "",
                    "B_Total",
                    "",
                    "",
                    "",
                    f"Tc_=SUM(G10:G{9 + len(rfq_items)})",
                    f"Tc_=SUM(H10:H{9 + len(rfq_items)})",
                    "",
                    f"Tr_=SUM(J10:J{9 + len(rfq_items)})",
                    f"Tr_=SUM(K10:K{9 + len(rfq_items)})",
                    f"Tc_=SUM(L10:L{9 + len(rfq_items)})",
                    "Pc_{}".format(total_percentage_savings),
                ]
                savings_worksheet.append(totals_row)

                savings_worksheet.append(empty_row)
                savings_worksheet.append(["", "hc_#", "Nmrl_3_Notes"])
                savings_worksheet.append(
                    [
                        "",
                        "Ac_1",
                        "Nmrl_3_Where the current prices are lower than the lowest price obtained, current price is retained",
                    ]
                )
                savings_worksheet.append(
                    [
                        "",
                        "Ac_2",
                        "Nmrl_3_N/A means that the current supplier has not been provided",
                    ]
                )
                savings_worksheet.freeze_panes = savings_worksheet["D10"]
            return True

        except Exception as e:
            print(e)
            capture_exception(e)
            return False


@shared_task(bind=True)
def rfq_participation_status(self, job_id=None, category_id=None):
    """
    Get Participation Status for RFQ Invited suppliers
    """
    try:
        if job_id is not None:
            job = Tender.objects.filter(id=job_id).first()
            company = job.company
            data = {}
            time = datetime.datetime.now()
            if job is None:
                return {"message": "Job not an RFQ"}
            else:
                # categories = job.categories.filter(invite_only=True)
                categories = job.categories
                if len(categories) < 1:
                    return {"message": "RFQ job has no closed categories"}
                else:
                    rfq_suppliers = []
                    for cat in categories:
                        rfq = cat
                        if rfq is not None:
                            rfq_participants = rfq.participants
                            rfq_invited_suppliers = rfq.invited_suppliers

                        rfq_data = {
                            "category": cat,
                            "participants": rfq_participants,
                            "invitees": rfq_invited_suppliers,
                        }

                        rfq_suppliers.append(rfq_data)

        # elif category_id is not None:
        #     category = Category.objects.select_related("job").get(id=category_id)
        #     company = category.job.company
        #     job = category.job
        #     data = {}
        #     time = datetime.datetime.now()

        #     if category is None:
        #         return {"message": "RFQ Category does not exist"}
        #     else:
        #         rfq_suppliers = []
        #         rfq = category.rfq
        #         if rfq is not None:
        #             rfq_participants = rfq.participants
        #             rfq_invited_suppliers = rfq.invited_suppliers

        #         rfq_data = {
        #             "category": category,
        #             "participants": list(rfq_participants),
        #             "invitees": rfq_invited_suppliers,
        #         }
        #         rfq_suppliers.append(rfq_data)

        else:
            return {"message": "RFQ job/category does not exist "}

        # Generate Report
        dir_name = Path(
            "media/rfq_participants/%s/%s/%s"
            % (company.company_name, time.year, job.title)
        )  # folder structure
        dir_name.mkdir(parents=True, exist_ok=True)
        time_only = "%d%d_%d%d" % (time.year, time.month, time.day, time.hour)
        filepath = "{}/{}_{}_{}.xlsx".format(
            dir_name, job.unique_reference, job.title, time_only
        )
        # delete similar reports run this month
        match_string = "{}_{}".format(job.id, company.company_name.replace(" ", "_"))
        delete_matching_files_in_directory(dir_name, match_string, ".xlsx")
        # create and save a workbook to later populate
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = f"Summary"

        if company.company_logo_url is not None and company.company_logo_url != "":
            buyer_logo_url = get_file_path(company.company_logo_url)
        else:
            buyer_logo_url = "static/core/img/no-company-image128-128.png"
        qed_logo_anchor = "D1"
        qed_logo_worksheet = worksheet.title
        buyer_logo_anchor = "C1"
        buyer_logo_worksheet = worksheet.title

        worksheet["C5"] = "Nb_{}".format(job.title)
        worksheet["C6"] = "Nb_{}".format(job.unique_reference)
        worksheet["C7"] = "Nb_RFQ Participation Report"
        worksheet["C8"] = "Nb_"

        worksheet.append(
            (
                "",
                "",
                "",
                "hc_Participants",
            )
        )
        worksheet.append(("", "hc_No", "h_Category Name", "h_Invited", "h_Yes", "h_No"))
        worksheet.merge_cells("D9:F9")
        data["Summary_C"] = 62
        data["Summary_D"] = 10
        data["Summary_E"] = 10
        data["Summary_F"] = 10

        total_invited = 0
        total_participants = 0
        total_non_participants = 0
        result = 0
        progress_recorder = ProgressRecorder(self)
        progress_total = len(rfq_suppliers)

        for count, rfq_data in enumerate(rfq_suppliers, start=1):
            result += 1
            progress_recorder.set_progress(result, progress_total)

            category = rfq_data["category"]
            invitees = rfq_data["invitees"]
            parts = rfq_data["participants"]
            participants = parts["participants"]
            participants_count = parts["count"]
            # percentage_participation = "{:.2f}".format(
            #     len(participants) / len(invitees)
            # )
            percentage_participation = "{}%".format(
                round(((participants_count / len(invitees)) * 100))
            )
            non_participants = len(rfq_data["invitees"]) - participants_count
            worksheet.append(
                (
                    "",
                    "Ac_%d" % (count),
                    rfq_data["category"].name,
                    "Ac_{}".format(len(invitees)),
                    "Ac_{}".format(participants_count),
                    non_participants,
                )
            )

            total_invited += len(invitees)
            total_participants += participants_count
            total_non_participants += non_participants

            # create individual sheets
            workbook.create_sheet(category.unique_reference.replace("/", "_"))
            details_worksheet = workbook[category.unique_reference.replace("/", "_")]
            # formating_data = {
            #     "Summary_C": 60,
            # }
            empty_row = [""]
            empty_row = tuple(empty_row)
            details_worksheet.append(empty_row)
            details_worksheet["C2"] = "Nb_{}".format(category.name)
            details_worksheet["C3"] = "Nb_"
            details_worksheet["C4"] = "Number of bidders invited"
            details_worksheet["C5"] = "Number of bidders participated"
            details_worksheet["C6"] = "% Participation"
            details_worksheet["D4"] = "Ac_{}".format(len(invitees))
            details_worksheet["D5"] = "Ac_{}".format(participants_count)
            details_worksheet["D6"] = "Ac_{}".format(percentage_participation)
            details_worksheet["C7"] = "Nb_"

            details_worksheet.append(
                (
                    "",
                    "",
                    "",
                    "",
                    "hc_Participation Status",
                )
            )

            details_worksheet.append(
                (
                    "",
                    "hc_No",
                    "h_Supplier",
                    "h_Phone Number",
                    "h_Yes",
                    "h_No",
                )
            )
            details_worksheet.merge_cells("E8:F8")
            details_worksheet.merge_cells("C2:E2")
            # formating
            data[f'{category.unique_reference.replace("/", "_")}_C'] = 35
            data[f'{category.unique_reference.replace("/", "_")}_D'] = 15
            data[f'{category.unique_reference.replace("/", "_")}_E'] = 10
            data[f'{category.unique_reference.replace("/", "_")}_F'] = 10

            status_yes = ""
            status_no = ""
            for count, supplier in enumerate(invitees, start=1):
                if supplier is not None:
                    if supplier in participants:
                        status_yes = "Yes"
                        status_no = " "
                    else:
                        status_no = "No"
                        status_yes = " "

                    if isinstance(supplier, str):
                        company = supplier
                        phone = "N/A"
                    else:
                        company = supplier.company_name
                        phone = supplier.phone_number

                    details_worksheet.append(
                        (
                            "",
                            "Ac_%d" % (count),
                            company,
                            phone,
                            "Ac_{}".format(status_yes),
                            "Ac_{}".format(status_no),
                        )
                    )

        worksheet.append(
            (
                "",
                "",
                "Total",
                "Ac_{}".format(total_invited),
                "Ac_{}".format(total_participants),
                "Ac_{}".format(total_non_participants),
            )
        )
        workbook.save(filepath)
        format_excel(filepath, data=data)
        insert_image(
            excel_url=filepath,
            worksheet_name=qed_logo_worksheet,
            anchor=qed_logo_anchor,
        )
        insert_image(
            filepath,
            buyer_logo_worksheet,
            buyer_logo_anchor,
            image_url=buyer_logo_url,
        )

        context = {
            "response_message": "Report generated successfully",
            "filepath": filepath.split("/", 1)[1],
        }

        return context

    except Exception as e:
        print(e)
        capture_exception(e)
        context = {
            "messages": "Failed to generate report",
        }
        return context
