import datetime
import os
from pathlib import Path
from zipfile import ZipFile

import pytz
import requests
from celery import shared_task, Task
from celery_progress.backend import ProgressRecorder
from django import apps
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.timezone import make_aware
from django.contrib.contenttypes.models import ContentType
from openpyxl import load_workbook
from sentry_sdk import capture_exception
from apps.core.utils import show, get_file_path
from apps.tender.models import (
    Tender, Category, Section, Question, MarkingScheme, Item, SupplierResponse, 
    QualityAssuranceResponse, SupplierSectionScore, SupplierTechnicalScore, QualityAssurance, ItemResponse, 
    SupplierFinancialTotal, SupplierCategoryScore, QualityAssuranceQuestion, 
    QaTccResponse, QaBusinessPermitResponse, QaCr12Response, QaIncorporationCertificateResponse, QaNationalIdResponse,
    QaNcaaResponse, QaPinCertificateResponse, QaPoisonsBoardResponse
)
from apps.tender.reports import TenderQARankingReport, DueDiligenceRankingReport, TenderInterimReport, \
    ConsolidatedTenderSummaryReport, TenderSummaryReport, FinancialRatiosReport
from backend.celery import app
from backend.storage_backends import PrivateMediaStorage
from apps.ocr.ocr import verify_tcc, verify_cr12, verify_businesspermit, verify_poisonsboardcert, verify_pincert, \
    verify_incorporationcert
from apps.common.utils import get_local_filepath


def job_questions_upload(job_id):
    messages = []
    job = Tender.objects.get(id=job_id)
    categories = Category.objects.filter(tender_id=job_id)
    if categories.count() > 0:

        try:
            # A = PrivateMediaStorage()
            # headers = {"ResponseContentDisposition": f"attachment;"}
            # time = datetime.datetime.now()
            # file_url = A.url(
            #     f"{job.questions_template}", expire=300, parameters=headers, http_method="GET",
            # )
            # dir_name = Path("media/temp/{}/{}/{}".format(time.year, time.month, time.day))
            # dir_name.mkdir(parents=True, exist_ok=True)
            # file_name = os.path.basename(f"{job.questions_template}")
            # filepath = "{}/{}".format(dir_name, file_name)
            # r = requests.get(file_url)
            # with open("{}".format(filepath), "wb") as f:
            #     f.write(r.content)

            workbook = load_workbook(job.question_template, data_only=True)

            worksheets = workbook.sheetnames
            if len(worksheets) < 1:
                messages.append('There are no sections to upload from the excel')
                context = {
                    "messages": messages
                }
                return context

            data = upload_question_data(workbook=workbook, worksheets=worksheets, job=job)

            messages.append(
                "Imports succeeded. {} new sections and {} new questions added to {} categories.".format(
                    data['sections_count'], data['questions_count'], data['categories_count']
                ),
            )
            context = {
                "messages": messages
            }
            return context

        except Exception as e:
            # for category in categories:
            #     for section in category.sections:
            #         section.delete()
            print(e)
            messages.append(f"{str(e)}")
            messages.append("There is a problem with your upload excel. Check instructions carefully, correct, upload.")
            context = {
                "messages": messages
            }
            return context

    else:
        context = {
            "messages": messages
        }
        return context


def upload_question_data(workbook, worksheets, job):
    categories_count = 0
    questions_count = 0
    sections_count = 0
    for sheet in worksheets:
        category_sheet = workbook.get_sheet_by_name(sheet)
        category = Category.objects.filter(tender_id=job.id, unique_reference=sheet).first()

        if category is not None:
            trans_category_name = category_sheet["D2"].value
            category_short_name = category_sheet["C3"].value
            category.category_short_name = category_short_name
            category.trans_name = trans_category_name
            category.save()
            categories_count += 1
            # delete all old questions and sections
            for section in category.sections:
                section.delete()

            for i in range(6, category_sheet.max_row + 1):
                # for each of the rows create a question after checking for existing section
                section_name = category_sheet["A{}".format(i)].value
                trans_name = category_sheet["B{}".format(i)].value
                short_name = category_sheet["C{}".format(i)].value
                old_sections = Section.objects.filter(
                    name=section_name, category_id=category.id
                )
                if old_sections.count() > 0:
                    section = old_sections.first()
                else:
                    section = Section.objects.create(
                        name=section_name, trans_name=trans_name,
                        short_name=short_name, category_id=category.id,
                        description="Questions in {}".format(
                            section_name
                        ),
                    )
                    sections_count += 1

                trans_description = category_sheet["F{}".format(i)].value
                if trans_description == "" or trans_description == "Question short description":
                    question_short_description = category_sheet["D{}".format(i)].value
                    trans_short_description = category_sheet["E{}".format(i)].value
                else:
                    question_short_description = category_sheet["G{}".format(i)].value
                    trans_short_description = category_sheet["H{}".format(i)].value

                q = Question.objects.create(
                    section=section,
                    description=category_sheet["E{}".format(i)].value,
                    trans_description=trans_description,
                    short_description=question_short_description,
                    trans_short_description=trans_short_description,
                    answer_type=category_sheet["I{}".format(i)].value,
                    is_required=category_sheet["J{}".format(i)].value,
                    is_scored=category_sheet["K{}".format(i)].value,
                    max_score=category_sheet["L{}".format(i)].value,
                    is_qa=category_sheet["M{}".format(i)].value,
                    is_dd=category_sheet["N{}".format(i)].value,
                )
                questions_count += 1

                if q.answer_type == Question.TYPE_CHECKBOX or q.answer_type == Question.TYPE_SELECT:
                    MarkingScheme.objects.create(
                        question_id=q.id,
                        options=category_sheet["O{}".format(i)].value,
                        score=category_sheet["P{}".format(i)].value,
                    )

    context = {
        "categories_count": categories_count,
        "questions_count": questions_count,
        "sections_count": sections_count
    }
    return context


def category_questions_upload(job_id, category_id):
    messages = []
    categories_count = 0
    questions_count = 0
    sections_count = 0
    category = Category.objects.filter(id=category_id).first()

    if category is not None:

        try:
            # A = PrivateMediaStorage()
            # headers = {"ResponseContentDisposition": f"attachment;"}
            # time = datetime.datetime.now()
            # file_url = A.url(
            #     f"{category.questions_template}", expire=300, parameters=headers, http_method="GET",
            # )
            # dir_name = Path("media/temp/{}/{}/{}".format(time.year, time.month, time.day))
            # dir_name.mkdir(parents=True, exist_ok=True)
            # file_name = os.path.basename(f"{job.questions_template}")
            # filepath = "{}/{}".format(dir_name, file_name)
            # r = requests.get(file_url)
            # with open("{}".format(filepath), "wb") as f:
            #     f.write(r.content)

            workbook = load_workbook(category.question_template, data_only=True)

            worksheets = workbook.sheetnames[1:]
            if len(worksheets) < 1:
                messages.append('There are no sections to upload from the excel')
                context = {
                    "messages": messages
                }
                return context

            category_sheet = workbook.get_sheet_by_name('Questions')

            trans_category_name = category_sheet["D2"].value
            category_short_name = category_sheet["C3"].value
            category.category_short_name = category_short_name
            category.trans_name = trans_category_name
            category.save()
            categories_count += 1

            # delete all old questions and sections
            sections = Section.objects.filter(category_id=category_id)
            sections.delete()

            for i in range(5, category_sheet.max_row + 1):
                # for each of the rows create a question after checking for existing section
                section_name = category_sheet["A{}".format(i)].value
                trans_name = category_sheet["B{}".format(i)].value
                short_name = category_sheet["C{}".format(i)].value
                old_sections = Section.objects.filter(
                    name=section_name, category=category
                )
                if old_sections.count() > 0:
                    section = old_sections.first()
                else:
                    section = Section.objects.create(
                        name=section_name,
                        trans_name=trans_name,
                        short_name=short_name,
                        description="Questions in {}".format(
                            section_name
                        ),
                        category=category,
                    )
                    sections_count += 1

                trans_description = category_sheet["F{}".format(i)].value
                if trans_description == "" or trans_description == "Question short description":
                    question_short_description = category_sheet["D{}".format(i)].value
                    trans_short_description = category_sheet["E{}".format(i)].value
                else:
                    question_short_description = category_sheet["G{}".format(i)].value
                    trans_short_description = category_sheet["H{}".format(i)].value

                q = Question.objects.create(
                    section=section,
                    description=category_sheet["E{}".format(i)].value,
                    trans_description=trans_description,
                    short_description=question_short_description,
                    trans_short_description=trans_short_description,
                    answer_type=category_sheet["I{}".format(i)].value,
                    is_required=category_sheet["J{}".format(i)].value,
                    is_scored=category_sheet["K{}".format(i)].value,
                    max_score=category_sheet["L{}".format(i)].value,
                    is_qa=category_sheet["M{}".format(i)].value,
                    is_dd=category_sheet["N{}".format(i)].value,
                )
                questions_count += 1

                if q.answer_type == Question.TYPE_CHECKBOX or q.answer_type == Question.TYPE_SELECT:
                    MarkingScheme.objects.create(
                        question_id=q.id,
                        options=category_sheet["O{}".format(i)].value,
                        score=category_sheet["P{}".format(i)].value,
                    )

            messages.append(
                "Imports succeeded. {} new sections and {} new questions added to {} categories.".format(
                    sections_count, questions_count, categories_count
                ),
            )
            context = {
                "messages": messages,
                "categories_count": categories_count,
                "questions_count": questions_count,
                "sections_count": sections_count
            }
            return context

        except Exception as e:
            sections = Section.objects.filter(category_id=category_id)
            sections.delete()

            messages.append(f"{str(e)}")
            messages.append("There is a problem with your upload excel. Check instructions carefully, correct, upload.")
            context = {
                "messages": messages
            }
            print(f"{e}")
            return context

    else:
        context = {
            "messages": messages
        }
        return context


@shared_task(bind=True)
def create_tender_items(self, category_id):
    category = Category.objects.filter(id=category_id).first()
    if category is not None:
        if category.self_evaluate:
            """
            Self Evaluate, evaluated by system
            """
            A = PrivateMediaStorage()
            headers = {"ResponseContentDisposition": f"attachment;"}
            file_url = A.url(
                f"{category.items_template}",
                expire=300,
                parameters=headers,
                http_method="GET",
            )
            time = datetime.datetime.now()
            dir_name = Path(
                "media/temp/{}/{}/{}".format(time.year, time.month, time.day)
            )  # folder structure
            dir_name.mkdir(parents=True, exist_ok=True)
            file_name = os.path.basename(f"{category.items_template}")
            filepath = "{}/{}".format(dir_name, file_name)
            r = requests.get(file_url)
            # filepath = os.path.basename(f"{category.items_template}")

            with open("{}".format(filepath), "wb") as f:
                f.write(r.content)
            # filepath = get_local_filepath(category.items_template.url)

            workbook = load_workbook(filepath, data_only=True)
            try:
                worksheet = workbook.get_sheet_by_name("RFQ ITEMS")

                for i in range(15, worksheet.max_row + 1):
                    try:
                        if len(worksheet.cell(row=i, column=2).value) > 1:
                            if worksheet.cell(row=14, column=3).value != "UoM":
                                unit_of_measure = "N/A"
                            else:
                                unit_of_measure = worksheet.cell(row=i, column=3).value

                            if worksheet.cell(row=14, column=5).value != "Specification 1":
                                specification_1 = "N/A"
                            else:
                                specification_1 = worksheet.cell(row=i, column=5).value

                            if worksheet.cell(row=14, column=6).value != "Specification 2":
                                specification_2 = "N/A"
                            else:
                                specification_2 = worksheet.cell(row=i, column=6).value

                            if worksheet.cell(row=14, column=4).value != "Qty":
                                qty = 1
                            else:
                                qty = worksheet.cell(row=i, column=4).value
                                
                            data = {
                                "category_id": category.id,
                                "description": worksheet.cell(
                                    row=i, column=2
                                ).value,
                                "unit_of_measure": unit_of_measure,
                                "specification_1": specification_1,
                                "specification_2": specification_2,
                                "number": worksheet.cell(row=i, column=1).value,
                                "quantity": qty,
                                "current_price": 0,
                                "price_validity_months": 12,
                            }
                            new_tender_category = Item(**data)
                            new_tender_category.save()
                            print("created got here")

                        else:
                            message = "Column 2 must not be empty"
                            return {"result": "error", "error_message": message}

                    except Exception as e:
                        print(e)
                        message = str(e)
                        return {"result": "error", "error_message": message}

                message = "RFQ Created successfully"
                return {
                    "result": "success",
                    "success_message": message,
                    "category": new_tender_category.id,
                }
            except Exception as e:
                message = "Column 2 must not be empty"
                return {"result": "error", "error_message": message}

        else:
            Item.objects.create(
                category=category,
                description="Test Item",
                item_code="N/A",
                unit_of_measure="N/A",
                current_price=0,
                price_validity_months=12,
            )
    message = "RFQ or Category is None"
    return {"result": "error", "error_message": message}


# @shared_task(bind=True)
def upload_job_current_suppliers(job_id):
    """
    Uploads current suppliers and saves to the Job Model
    """
    # progress_recorder = ProgressRecorder(self)
    result = 0

    messages = []
    job = Tender.objects.get(id=job_id)
    company = job.company
    
    try:
        workbook = load_workbook(job.current_suppliers, data_only=True)
    except Exception as e:
        capture_exception(e)
        context = {
            "messages": messages.append("Excel not found"),
            "response_message": "Error uploading current suppliers"
        }
        return context

    try:
        worksheet = workbook["current_suppliers"]
    except Exception as e:
        capture_exception(e)
        context = {
            "messages": messages.append("Invalid current_suppliers sheet not found"),
            "response_message": "Error uploading current suppliers"
        }
        return context

    try:
        total_for_progress = worksheet.max_row
        count = 1
        for i in range(4, worksheet.max_row + 1):
            result += 1
            # progress_recorder.set_progress(result, total_for_progress)

            qed_code = worksheet["C{}".format(i)].value
            company_name = worksheet["D{}".format(i)].value
            company_email = worksheet["E{}".format(i)].value
            phone = worksheet["F{}".format(i)].value

            category_type = apps.apps.get_model('core', 'CategoryType').objects.filter(
                innitials=qed_code
            ).first()
            
            # modify to pick based on category unique reference to avoid duplicate suppliers in categories
            categories = Category.objects.filter(tender=job, category_type=category_type)

            for category in categories:
                current_supplier = apps.apps.get_model('core', 'CurrentSupplier').objects.filter(
                    supplier_name=company_name, category_id=category.id).first()

                if current_supplier is not None:
                    current_supplier = apps.apps.get_model('core', 'CurrentSupplier').objects.update_or_create(
                        id=current_supplier.id,
                        supplier_name=company_name,
                        category_id=category.id,
                        job_id=job_id,
                        target=ContentType.objects.get_for_model(apps.apps.get_model('prequal', 'Category')),                   
                        defaults={"alternative_email": company_email},
                    )
                else:
                    current_supplier = apps.apps.get_model('core', 'CurrentSupplier').objects.create(
                        company=company,
                        category_id=category.id,
                        job_id=job_id,
                        supplier_name=company_name,
                        supplier_email=company_email,
                        supplier_phone=phone,
                        target=ContentType.objects.get_for_model(apps.apps.get_model('prequal', 'Category'))
                    )

            count += 1

        messages.append(
            "Imports succeeded. {} current suppliers added.".format(count)
        )
        context = {
            "response_message": "Current Suppliers Uploaded Successfully",
            "messages": messages
        }
        return context
    except Exception as e:
        capture_exception(e)
        context = {
            "messages": messages.append(f"{e}"),
            "response_message": "Error uploading current suppliers"
        }
        print(context)
        return context


class EvaluateTender(Task):
    name = "EvaluateTender"
    progress_recorder = None
    result = 0

    def run(self, *args, **kwargs):
        self.progress_recorder = ProgressRecorder(self)
        context = self.calculate_technical_scores(kwargs["category_id"])
        return context

    def calculate_technical_scores(self, category_id):
        category = Category.objects.filter(id=category_id).first()

        sections = Section.objects.filter(category_id=category_id)
        supplier_responses = SupplierResponse.objects.filter(question__section__category_id=category_id)

        suppliers = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
            id__in=supplier_responses.only('supplier_id').values('supplier_id').distinct()
        )
        qa = QualityAssurance.objects.filter(category_id=category_id).first()
        qa_responses = QualityAssuranceResponse.objects.filter(
            quality_assurance_question__quality_assurance__category_id=category_id
        )
        questions = Question.objects.filter(section__category_id=category_id)

        for supplier in suppliers:
            supplier_score = 0
            for section in sections:
                section_questions = questions.filter(section_id=section.id)
                section_score = 0
                for question in section_questions:
                    score = 0
                    if question.is_scored:
                        question_response = supplier_responses.filter(
                            supplier_id=supplier.id, question_id=question.id).first()
                        if (question_response is not None and question_response.options is not None
                                and len(question_response.options) > 0):
                            if question.answer_type == Question.TYPE_UPLOAD:
                                score = float(question.max_score)
                            elif question_response.options in question.options:
                                response_index = question.options.index(question_response.options)
                                score = float(question.scores[response_index])
                            elif question_response.options == "True":
                                score = float(question.max_score)
                    section_score += score

                SupplierSectionScore.objects.update_or_create(
                    section_id=section.id,
                    supplier_id=supplier.id,
                    defaults={"score": section_score}
                )
                supplier_score += section_score

            technically_qualified = False
            if supplier_score >= 50:
                technically_qualified = True

            meets_all_mandatory_requirements = True
            qa_responses = qa_responses.filter(supplier_id=supplier.id)
            if qa_responses.filter(outcome="Fail").count() > 0:
                meets_all_mandatory_requirements = False
            elif qa_responses.count() < 1:
                meets_all_mandatory_requirements = False

            # calculate score after qa
            if qa:
                supplier_responses = supplier_responses.filter(supplier_id=supplier.id)
                score_after_qa = self.calculate_qa_scores(category_id, supplier, supplier_responses, questions)
            else:
                score_after_qa = None

            try:
                score_over_total = supplier_score / (category.total_technical_score)
            except:
                score_over_total = 0

            weighted_score = score_over_total * float(category.pass_score)

            SupplierTechnicalScore.objects.update_or_create(
                category_id=category_id,
                supplier_id=supplier.id,
                defaults={
                    "score": supplier_score,
                    "weighted_score": weighted_score,
                    "score_after_qa": score_after_qa,
                    "meets_all_mandatory_requirements": meets_all_mandatory_requirements,
                    "technically_qualified": technically_qualified
                },
            )

        self.result += 1
        self.progress_recorder.set_progress(self.result, 7)

        # rank participants according to category score
        self.technical_rank_participants(category_id=category_id)

        # rank participants according to qa score
        if qa:
            self.qa_rank_participants(category_id=category_id)
        self.calculate_financial_totals(category_id=category_id)
        self.financial_rank_participants(category_id=category_id)
        self.calculate_total_tender_scores(category_id=category_id)
        context = {
            "response_message": "Score Evaluation Complete"
        }
        return context

    def calculate_qa_scores(self, category_id, supplier, supplier_responses, questions):
        total_score = 0

        for section in Section.objects.filter(category_id=category_id):
            section_score = 0
            questions = questions.filter(section_id=section.id)
            for question in questions:
                if question.is_qa and question.is_scored:
                    score = 0
                    question_response = supplier_responses.filter(
                        supplier_id=supplier.id, question_id=question.id).first()
                    if (question_response is not None and question_response.options is not None
                            and len(question_response.options) > 0):
                        if question.answer_type == Question.TYPE_UPLOAD:
                            score = float(question.max_score)
                        elif question_response.options in question.options:
                            response_index = question.options.index(question_response.options)
                            score = float(question.scores[response_index])
                        elif question_response.options == "True":
                            score = float(question.max_score)
                    total_score += score
                else:
                    question_in_qa = QualityAssuranceResponse.objects.filter(
                        quality_assurance_question__question_id=question.id,
                        supplier_id=supplier.id,
                    ).first()
                    if question_in_qa is not None:
                        if question_in_qa.score_after_qa is not None:
                            total_score += float(question_in_qa.score_after_qa)
                            section_score += float(question_in_qa.score_after_qa)

            SupplierSectionScore.objects.update_or_create(
                section_id=section.id, supplier_id=supplier.id,
                defaults={
                    "score_after_qa": section_score
                }
            )
        return total_score

    def technical_rank_participants(self, category_id):
        scores = SupplierTechnicalScore.objects.filter(category_id=category_id).order_by(
            "-score"
        )
        rank = 0
        first_score = scores.first()
        if first_score is not None:
            first_score.rank = 1
            first_score.save()
            prev_score = first_score
            for score in scores:
                if score.score == prev_score.score:
                    score.rank = prev_score.rank
                else:
                    score.rank = rank + 1
                rank += 1
                score.save()
                prev_score = score
        self.result += 1
        self.progress_recorder.set_progress(self.result, 7)
        context = {
            "response_message": "Done ranking technical part of the tender"
        }
        return context

    def qa_rank_participants(self, category_id):
        qa = QualityAssurance.objects.filter(category_id=category_id).first()
        if qa is not None:
            scores = SupplierTechnicalScore.objects.filter(category_id=category_id).order_by(
                "-score_after_qa"
            )
            rank_after_qa = 0
            if scores.first().score_after_qa is not None:
                first_score = scores.first()
            else:
                first_score = scores[0]

            if first_score is not None:
                first_score.rank_after_qa = 1
                first_score.save()
                prev_score = first_score
                for score in scores:
                    if score.score_after_qa == None:
                        score.rank_after_qa = None
                        rank_after_qa += 0
                        score.save()
                        continue
                    elif score.score_after_qa == prev_score.score_after_qa:
                        score.rank_after_qa = prev_score.rank_after_qa
                    else:
                        score.rank_after_qa = rank_after_qa + 1
                    rank_after_qa += 1
                    score.save()
                    prev_score = score
            self.result += 1
            self.progress_recorder.set_progress(self.result, 7)
            context = {
                "response_message": "Done ranking qa"
            }
            return context
        else:
            self.result += 1
            self.progress_recorder.set_progress(self.result, 7)
            context = {
                "response_message": "No qa instance"
            }
            return context

    def calculate_financial_totals(self, category_id):
        suppliers = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
            id__in=ItemResponse.objects.filter(
                item__category_id=category_id).only('supplier_id').values('supplier_id').distinct()
        )
        item_count = Item.objects.filter(category_id=category_id).count()
        responses = ItemResponse.objects.filter(item__category_id=category_id)

        lowest_price = 100000000000000000

        for supplier in suppliers:
            responses = responses.filter(supplier_id=supplier.id)
            total = 0
            is_outlier = False
            is_blank = False
            for response in responses:
                # print(f"{response.item.description} Outlier {response.item.outlier_score}")
                outlier_score = response.item.outlier_score
                value = show(response.total)
                if value == 0:
                    is_blank = True
                if float(value) < float(outlier_score):
                    outlier = True
                    if outlier is True:
                        is_outlier = True
                total += value
                response.value = value
                response.save()
            if item_count == responses.count():
                if total is not None and is_outlier is not None:
                    obj, created = SupplierFinancialTotal.objects.update_or_create(
                        supplier_id=supplier.id,
                        category_id=category_id,
                        defaults={
                            "total": total,
                            "has_outlier": is_outlier,
                            "has_blank": is_blank,
                        },
                    )
            else:
                obj, created = SupplierFinancialTotal.objects.update_or_create(
                    supplier_id=supplier.id,
                    category_id=category_id,
                    defaults={
                        "total": total,
                        "has_outlier": True,
                        "has_blank": True,
                    },
                )
            if not is_blank and not is_outlier and total < lowest_price:
                lowest_price = total

        # calculate financial score
        self.calculate_financial_score(
            suppliers=suppliers, category_id=category_id, lowest_price=lowest_price)
        # print(f'lowest_cost {lowest_price}')

        self.result += 1
        self.progress_recorder.set_progress(self.result, 7)
        context = {
            "response_message": "Done calculating financial totals"
        }
        return context

    def financial_rank_participants(self, category_id):
        try:
            # do not rank if supplier is not technically qualified or does not meet all mandatory requirements
            scores = SupplierFinancialTotal.objects.filter(category_id=category_id, has_outlier=False).order_by("score")
            if scores:
                rank = 0
                first_score = scores.first()
                first_score.rank = 1
                first_score.save()
                prev_score = first_score
                for rfq_score in scores:
                    if rfq_score.score == prev_score.score:
                        rfq_score.rank = prev_score.rank
                    else:
                        rfq_score.rank = rank + 1

                    if rfq_score.has_outlier:
                        rfq_score.rank = 0
                    rank += 1
                    rfq_score.save()
                    prev_score = rfq_score
            self.result += 1
            self.progress_recorder.set_progress(self.result, 7)
            context = {
                "response_message": "Done ranking financial part of the tender"
            }
            return context
        except Exception as e:
            capture_exception(e)
            print(e)
            self.result += 1
            self.progress_recorder.set_progress(self.result, 7)
            context = {
                "response_message": "An error occurred during ranking of financial participants"
            }
            return context

    def calculate_financial_score(self, suppliers, category_id, lowest_price):
        category = apps.apps.get_model('tender', 'Category').objects.filter(
            id=category_id).first()
        # print('got to rfq score')
        totals = SupplierFinancialTotal.objects.filter(category_id=category_id)
        try:
            for supplier in suppliers:
                pricing_instance = totals.filter(supplier_id=supplier.id).first()
                if not pricing_instance.has_outlier:
                    try:
                        if totals.count() > 1:
                            rfq_score = (float(lowest_price) / float(pricing_instance.total)) * float(
                                category.financial_weight
                            )
                        else:
                            # print('got to rfq score 30')
                            rfq_score = 30
                    except Exception as e:
                        capture_exception(e)
                        print(e)
                        rfq_score = 0

                    pricing_instance.score = rfq_score
                    pricing_instance.save()
                else:
                    pricing_instance.rfq_score = 0
                    pricing_instance.save()
            self.result += 1
            self.progress_recorder.set_progress(self.result, 7)

            context = {
                "response_message": "Done calculating the financial score"
            }
            return context
        except Exception as e:
            capture_exception(e)
            print(e)
            context = {
                "response_message": "An error occurred during tender score calculation"
            }
            self.result += 1
            self.progress_recorder.set_progress(self.result, 7)
            return context

    def calculate_total_tender_scores(self, category_id):
        t_suppliers = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
            id__in=SupplierResponse.objects.filter(
                question__section__category_id=category_id).only('supplier_id').values('supplier_id').distinct()
        ).only('id')

        f_suppliers = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
            id__in=ItemResponse.objects.filter(
                item__category_id=category_id).only('supplier_id').values('supplier_id').distinct()
        ).only('id')

        suppliers = t_suppliers.union(f_suppliers)
        try:
            pricing_instances = apps.apps.get_model('tender', 'SupplierFinancialTotal').objects.filter(
                category_id=category_id
            )
            technical_instances = apps.apps.get_model('tender', 'SupplierTechnicalScore').objects.filter(
                category_id=category_id
            )
            for supplier in suppliers:
                pricing_instance = pricing_instances.filter(supplier_id=supplier.id).first()
                technical_instance = technical_instances.filter(supplier_id=supplier.id).first()

                try:
                    technical_weighted_score = technical_instance.weighted_score
                except:
                    technical_weighted_score = 0

                try:
                    financial_weighted_score = pricing_instance.score
                except:
                    financial_weighted_score = 0

                total = float(technical_weighted_score) + float(financial_weighted_score)
                SupplierCategoryScore.objects.update_or_create(
                    supplier_id=supplier.id,
                    category_id=category_id,
                    defaults={
                        "technical_score": technical_weighted_score,
                        "financial_score": financial_weighted_score,
                        "meets_all_mandatory_requirements": technical_instance.meets_all_mandatory_requirements,
                        "technically_qualified": technical_instance.technically_qualified,
                        "has_outlier": pricing_instance.has_outlier,
                        "has_blank": pricing_instance.has_blank,
                        "score": total
                    }
                )
            self.result += 1
            self.progress_recorder.set_progress(self.result, 7)
            context = {
                "response_message": "Total tender score calculation complete"
            }
            return context
        except Exception as e:
            self.result += 1
            self.progress_recorder.set_progress(self.result, 7)
            capture_exception(e)
            print(e)
            context = {
                "response_message": "Total tender score calculation error"
            }
            return context


@shared_task(bind=True)
def existing_supplier_email_invite(self, supplier_id, category_id):
    # try:
    supplier = apps.apps.get_model('suppliers', 'Supplier').objects.filter(id=supplier_id).first()
    category = apps.apps.get_model('tender', 'Category').objects.filter(id=category_id).first()
    job = category.tender
    company = job.company
    email_config = None
    to_email = supplier.email

    email_subject = job.title.upper() + " INVITATION"
    if email_config is not None:
        message = render_to_string(
            "tender/emails/category_invitation_email.html",
            {
                "supplier": supplier, "category": category,
                "buyer_name": f"{company.company_name} Team",
                "buyer_logo": "company_logo", "company": company,
            },
        )
        email_config.send_email(email_subject, message, bcc=[to_email])
    else:
        if category.invite_only is True:
            body = render_to_string(
                "tender/emails/category_invitation_email.html",
                {
                    "supplier": supplier, "category": category,
                    "buyer_name": "Tendersure Team", "buyer_logo": "tendersure_logo",
                },
            )
            # Attach bidder instructions
            if job.bidding_instructions:
                A = PrivateMediaStorage()
                headers = {"ResponseContentDisposition": f"attachment;"}
                time = datetime.datetime.now()
                file_url = A.url(
                    f"{job.bidding_instructions}",
                    expire=300,
                    parameters=headers,
                    http_method="GET",
                )
                dir_name = Path(
                    "media/temp/{}/{}/{}".format(time.year, time.month, time.day)
                )
                # folder structure
                dir_name.mkdir(parents=True, exist_ok=True)
                file_name = os.path.basename(f"{job.bidding_instructions}")
                filepath = "{}/{}".format(dir_name, file_name)
                r = requests.get(file_url)
                with open("{}".format(filepath), "wb") as f:
                    f.write(r.content)
                email = EmailMultiAlternatives(email_subject, body, to=[to_email])
                email.attach_alternative(body, "text/html")
                email.attach_file(filepath)
                email.send()
            else:
                email = EmailMultiAlternatives(email_subject, body, to=[to_email])
                email.attach_alternative(body, "text/html")
                email.send()

        else:
            message = render_to_string(
                "tender/emails/category_invitation_email.html",
                {
                    "supplier": supplier,
                    "category": category,
                    "buyer_name": "Tendersure Team",
                    "buyer_logo": "tendersure_logo",
                },
            )
            email = EmailMultiAlternatives(email_subject, message, to=[to_email])
            email.attach_alternative(message, "text/html")
            email.send()
    # except Exception as e:
    #     capture_exception(e)
    #     print(e)


@shared_task(bind=True)
def non_existent_invites_email(self, user_email, category):
    try:
        supplier = "Esteemed Vendor"
        job = category.tender
        company = job.company

        email_config = None
        to_email = user_email
        email_subject = "Tender Invitation"

        if category.invite_only is True:
            body = render_to_string(
                "tender/emails/category_invitation_email_non_existent_supplier.html",
                {
                    "supplier": supplier, "category": category, "company": company,
                    "buyer_name": "Tendersure Team", "buyer_logo": "tendersure_logo",
                },
            )
            # Attach bidder instructions
            A = PrivateMediaStorage()
            headers = {"ResponseContentDisposition": f"attachment;"}
            time = datetime.datetime.now()
            file_url = A.url(
                f"{job.bidding_instructions}",
                expire=300,
                parameters=headers,
                http_method="GET",
            )
            dir_name = Path(
                "media/temp/{}/{}/{}".format(time.year, time.month, time.day)
            )
            # folder structure
            dir_name.mkdir(parents=True, exist_ok=True)
            file_name = os.path.basename(f"{job.bidding_instructions}")
            filepath = "{}/{}".format(dir_name, file_name)
            r = requests.get(file_url)
            with open("{}".format(filepath), "wb") as f:
                f.write(r.content)
            email = EmailMultiAlternatives(email_subject, body, to=[to_email])
            email.attach_alternative(body, "text/html")
            email.attach_file(filepath)
            email.send(fail_silently=True)

        else:
            body = render_to_string(
                "tender/emails/category_invitation_email_non_existent_supplier.html",
                {
                    "supplier": supplier, "category": category, "company": company,
                    "buyer_name": "Tendersure Team", "buyer_logo": "tendersure_logo",
                },
            )
            email = EmailMultiAlternatives(email_subject, body, to=[to_email])
            email.attach_alternative(body, "text/html")
            email.send(fail_silently=True)
    except Exception as e:
        capture_exception(e)
        print(e)


class ZipQuestionFiles(Task):
    name = "ZipTenderQuestionFiles"
    progress_recorder = None
    result = 0

    def run(self, *args, **kwargs):
        self.progress_recorder = ProgressRecorder(self)
        context = self.zip_files(kwargs["question_id"])
        return context

    def zip_files(self, question_id):
        progress_recorder = ProgressRecorder(self)
        result = 0
        question = Question.objects.filter(id=question_id).first()
        supplier_responses = (
            SupplierResponse.objects.filter(question_id=question_id).distinct()
        )
        print(supplier_responses)

        total_for_progress = supplier_responses.count()
        print(total_for_progress)
        # dir_name = Path(
        #     f"media/temp/zip/question_documents/{question_id}"
        # )  # folder structure
        dir_name = Path(
            f"media/question_documents/{question_id}"
        )  # folder structure
        dir_name.mkdir(parents=True, exist_ok=True)
        zip_name = f"{dir_name}/{question.short_description.replace('/', '_').replace(' ', '_')}.zip"
        # zip_name = (
        #     f"{question.short_description.replace('/', '_').replace(' ', '_')}.zip"
        # )
        with ZipFile(zip_name, "w") as zipObj2:
            for supplier_response in supplier_responses:
                progress_recorder.set_progress(result, total_for_progress)
                if supplier_response is not None:
                    try:
                        dir_name.mkdir(parents=True, exist_ok=True)
                        file_name = os.path.basename(
                            f"{supplier_response.document_url}"
                        )
                        filepath = f"{dir_name}/{file_name}"
                        if supplier_response.document_url.name != "":
                            r = requests.get(supplier_response.document_url.url)

                            with open("{}".format(filepath), "wb") as f:
                                f.write(r.content)

                            zipObj2.write(f"{filepath}")
                        else:
                            sup_prof = apps.apps.get_model('suppliers', 'SupplierCompanyProfile').objects.filter(
                                supplier__id=supplier_response.supplier.id).order_by("id").last()
                            if sup_prof.registration_cert_url.name != "":
                                n_file_name = os.path.basename(f"{sup_prof.registration_cert_url}")
                                n_filepath = f"{dir_name}/{n_file_name}"
                                r = requests.get(sup_prof.registration_cert_url.url)

                                with open("{}".format(n_filepath), "wb") as f:
                                    f.write(r.content)

                                zipObj2.write(f"{n_filepath}")

                    except Exception as e:
                        capture_exception(e)
                        print(e)
                result += 1
        # shutil.rmtree(
        #     f"media/temp/zip/question_documents/{question_id}", ignore_errors=True
        # )
        context = {
            "filepath": zip_name,
            "response_message": "File generated successfully",
        }

        return context

@shared_task(bind=True)
def submit_tender_rfq(self, category_id, supplier_id, price_template):
    messages = []
    category = Category.objects.filter(id=category_id).first()
    supplier = apps.apps.get_model("suppliers", "Supplier").objects.get(id=supplier_id)

    # A = PrivateMediaStorage()
    # headers = {"ResponseContentDisposition": f"attachment;"}
    # time = datetime.datetime.now()
    # file_url = A.url(
    #     f"{excel_url}", expire=300, parameters=headers, http_method="GET"
    # )
    # dir_name = Path(
    #     "media/temp/{}/{}/{}".format(time.year, time.month, time.day)
    # )  # folder structure
    # dir_name.mkdir(parents=True, exist_ok=True)
    # file_name = os.path.basename(f"{excel_url}")
    # filepath = "{}/{}".format(dir_name, file_name)
    # r = requests.get(file_url)

    # with open("{}".format(filepath), "wb") as f:
    #     f.write(r.content)

    # filepath = get_local_filepath(category.items_template.url)
    filepath = get_local_filepath(price_template)
    print(filepath)

    if category.self_evaluate == False:
        try:
            workbook = load_workbook(filepath, data_only=True)
        except Exception as e:
            workbook = None

        if not workbook:
            context = {
                "tender": category,
                "job": category.tender,
                "company": category.company,
            }
            messages.append(
                "Please make sure the document you are uploading is an excel sheet."
            )

        updated_items = False
        old_response = None
        item = Item.objects.filter(
            category=category, item_description="Test Item", item_number=1
        ).first()
        # check for old response on the item and update or create a new one.
        old_responses = ItemResponse.objects.filter(
            item=item, supplier=supplier
        )
        if old_responses.count() > 0:
            old_response = old_responses.first()
            total = 1
            if total is not None:
                old_response.total = hash(total)
            else:
                old_response.total = hash(0)
            old_response.cell_data = "cell data"
            old_response.column_data = "column data"
            # old_response.excel_url = excel_url
            old_response.save()
            updated_items = True
        else:
            new_total = ""
            total = 1
            if total is not None:
                new_total = hash(total)
            else:
                new_total = hash(0)
            ItemResponse.objects.create(
                supplier=supplier,
                item=item,
                total=new_total,
                cell_data="cell data",
                column_data="column data",
                item_number=1,
                # excel_url=excel_url,
            )
    else:
        workbook = load_workbook(filepath, data_only=True)
        if len(workbook.worksheets) > 2:
            messages.append(
                "Please make sure the document you are uploading has the same format as the downloaded template"
            )
            context = {
                "result": "error",
                "response_message": messages,
            }
            return context

        if not "RFQ ITEMS" in workbook.sheetnames:
            messages.append(
                'Please make sure the document you are uploading has a worksheet with the name "RFQ ITEMS"'
            )
            context = {
                "result": "error",
                "response_message": messages,
            }
            return context

        try:
            worksheet = workbook["RFQ ITEMS"]
            old_response = None
            updated_items = False
            # capture submissions for the rest of the items
            for i in range(5, worksheet.max_row + 1):
                if (
                    worksheet.cell(row=i, column=2).value != None
                    and worksheet.cell(row=i, column=1).value != "No"
                ):
                    column_data = ""
                    cell_data = ""
                    for k in range(5, worksheet.max_column):
                        column_data += "%s," % str(
                            worksheet.cell(row=4, column=k).value
                        )
                        cell_data += "%s," % str(worksheet.cell(row=i, column=k).value)

                    item = Item.objects.filter(
                        category=category,
                        item_description=worksheet.cell(row=i, column=2).value,
                        item_number=worksheet.cell(row=i, column=1).value,
                    ).first()
                    # if rfq_item is None:
                    #     raise ValueError("Item mismatch")
                    if item is not None:
                        total = worksheet.cell(row=i, column=worksheet.max_column).value
                        item_number = worksheet.cell(row=i, column=1).value

                        # check for old response on the item and update or create a new one.
                        old_responses = ItemResponse.objects.filter(
                            item=item, supplier=supplier
                        )
                        if old_responses.count() > 0:
                            old_response = old_responses.first()
                            if total is not None:
                                old_response.total = hash(total)
                            else:
                                old_response.total = hash(0)
                            old_response.cell_data = cell_data
                            old_response.column_data = column_data
                            old_response.save()
                            updated_items = True
                        else:
                            new_total = ""
                            if total is not None:
                                new_total = hash(total)
                            else:
                                new_total = hash(0)
                            item_response = ItemResponse.objects.filter(
                                item=item, supplier=supplier
                            ).first()
                            if item_response is not None:
                                item_response.total = new_total
                                item_response.cell_data = cell_data
                                item_response.column_data = column_data
                                item_response.save()
                            else:
                                ItemResponse.objects.create(
                                    supplier=supplier,
                                    item=item,
                                    total=new_total,
                                    cell_data=cell_data,
                                    column_data=column_data,
                                    item_number=item_number,
                                )
            # create supplier response
            messages.append(
                "RFQ Item Responses created successfuly"
            )
            context = {
                "result": "success",
                "response_message": messages,
            }
            return context
            # supplier.send_financial_responses(rfq.id, type)
        except Exception as e:
            print(e)
            # todo delete that excel from the server
            # "Rfq / 2019 / 08 / 24 / RFQ_Template_for_Ouma_Company_Buyers.xlsx"

            # dir_name = "media/%s" % excel_url.rsplit('/', 1)[0]
            # match_string = excel_url.rsplit('/', 1)[-1]
            # delete_matching_files_in_directory(dir_name, match_string, ".xlsx")

            messages.append(
                "There is a problem with your uploaded Excel. Verify and retry."
            )
            context = {
                "result": "error",
                "response_message": messages,
            }
            return context
        

class ConductOcr(Task):
    name = "ConductOcr"
    messages = []
    responses = []

    def run(self, *args, **kwargs):
        qa_instance = QualityAssurance.objects.filter(category_id=kwargs["category_id"]).first()
        self.ocr_tcc(qa_instance, kwargs['category_id'])
        self.ocr_cr12(qa_instance, kwargs['category_id'])
        self.ocr_business_permit(qa_instance, kwargs['category_id'])
        self.ocr_poisons_board_license(qa_instance, kwargs['category_id'])
        self.ocr_tax_pin_cert(qa_instance, kwargs['category_id'])
        self.ocr_incorporation_certificate(qa_instance, kwargs['category_id'])
        self.ocr_nca_certificate(qa_instance, kwargs['category_id'])

        context = {
            "response_message": "OCR process complete",
            "messages": self.messages,
            "responses": self.responses,
        }
        print(context)
        return context

    def ocr_tcc(self, qa_instance, category_id):
        tcc = QualityAssuranceQuestion.objects.filter(
            question__short_description__icontains='Tax compliance',
            quality_assurance_id=qa_instance.id
        ).first()
        if tcc:
            try:
                supplier_responses = SupplierResponse.objects.filter(
                    question_id=tcc.question_id
                )
                q = Question.objects.filter(
                    section__category_id=category_id, short_description="PIN number").first()
                for supplier_response in supplier_responses:
                    r = verify_tcc(document=f"{supplier_response.document_url}")
                    self.responses.append(r)
                    # if r['kra_pin'] and r['expiry_date']:
                    date = make_aware(r['expiry_date'], timezone=pytz.timezone("Africa/Nairobi")) if r[
                        'expiry_date'] else None
                    instance, created = QualityAssuranceResponse.objects.update_or_create(
                        supplier_id=supplier_response.supplier_id,
                        quality_assurance_question_id=tcc.id,
                        defaults={
                            "number": r['kra_pin'], "date": date, "ocr_response": r
                        }
                    )
                    pin_number_outcome = "Fail"
                    if q is not None:
                        s_r = SupplierResponse.objects.filter(
                            question_id=q.id, supplier_id=supplier_response.supplier_id).first()
                        if s_r is not None:
                            if s_r.options:
                                if s_r.options == r['kra_pin']:
                                    pin_number_outcome = "Pass"

                    expiry_date_outcome = "Fail"
                    if date is not None:
                        if datetime.datetime.now().date() <= date.date():
                            expiry_date_outcome = "Pass"

                    QaTccResponse.objects.update_or_create(
                        qa_response_id=instance.id,
                        defaults={
                            # company name
                            "pin_number": r['kra_pin'], "pin_number_outcome": pin_number_outcome,
                            "expiry_date": date, "expiry_date_outcome": expiry_date_outcome
                        }
                    )
                self.messages.append(f'Processing done for TCC {supplier_responses.count()}')
            except Exception as e:
                # capture_exception(e)
                print(f"TCC {e}")
        else:
            self.messages.append('NO TCC IN QA')

    def ocr_cr12(self, qa_instance, category_id):
        cr12 = QualityAssuranceQuestion.objects.filter(
            question__short_description__icontains='CR12 Certificate or equivalent',
            quality_assurance_id=qa_instance.id
        ).first()
        if cr12:
            try:
                supplier_responses = SupplierResponse.objects.filter(
                    question_id=cr12.question_id
                )
                q = Question.objects.filter(
                    section__category_id=category_id,
                    short_description="Company registration certificate number").first()
                for supplier_response in supplier_responses:
                    r = verify_cr12(document=f"{supplier_response.document_url}")
                    self.responses.append(f'{r}')
                    # if r['cr12_ref_no']:
                    instance, created = QualityAssuranceResponse.objects.update_or_create(
                        supplier_id=supplier_response.supplier_id,
                        quality_assurance_question_id=cr12.id,
                        defaults={
                            "number": r['cr12_ref_no'],
                            "date": make_aware(
                                datetime.datetime.now(), timezone=pytz.timezone("Africa/Nairobi")),
                            "ocr_response": r
                        }
                    )

                    company_number_outcome = "Fail"
                    if q is not None:
                        s_r = SupplierResponse.objects.filter(
                            question_id=q.id, supplier_id=supplier_response.supplier_id).first()
                        if s_r is not None:
                            if s_r.options:
                                if s_r.options == r['cr12_ref_no']:
                                    company_number_outcome = "Pass"

                    QaCr12Response.objects.update_or_create(
                        qa_response_id=instance.id,
                        defaults={
                            "company_number": r["cr12_ref_no"],
                            # "document_date": data["date"],
                            "company_number_outcome": company_number_outcome,
                        }
                    )
                self.messages.append(f'Processing done for CR12 {supplier_responses.count()}')
            except Exception as e:
                # capture_exception(e)
                print(f"CR12 {e}")
        else:
            self.messages.append('NO CR12 IN QA')

    def ocr_business_permit(self, qa_instance, category_id):
        business_permit = QualityAssuranceQuestion.objects.filter(
            question__short_description__icontains='Business permit',
            quality_assurance_id=qa_instance.id
        ).first()
        if business_permit:
            try:
                supplier_responses = SupplierResponse.objects.filter(
                    question_id=business_permit.question_id,
                )
                for supplier_response in supplier_responses:
                    r = verify_businesspermit(document=f"{supplier_response.document_url}")
                    self.responses.append(r)
                    # if r['business_id']:
                    instance, created = QualityAssuranceResponse.objects.update_or_create(
                        supplier_id=supplier_response.supplier_id,
                        quality_assurance_question_id=business_permit.id,
                        defaults={
                            "number": r['business_id'],
                            "date": make_aware(
                                datetime.datetime.now(), timezone=pytz.timezone("Africa/Nairobi")),
                            "ocr_response": r
                        }
                    )

                    business_name_outcome = "Fail"
                    supplier = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                        id=supplier_response.supplier_id).first()
                    if supplier is not None:
                        if supplier.company_name.lower() == r['business_name']:
                            business_name_outcome = "Pass"

                    QaBusinessPermitResponse.objects.update_or_create(
                        qa_response_id=instance.id,
                        defaults={
                            "business_id": r['business_id'], "date": r["year"],
                            "business_name": r["business_name"],
                            "business_name_outcome": business_name_outcome
                        }
                    )
                    # QualityAssuranceResponse.objects.update_or_create()
                self.messages.append(f'Processing done for Business Permit {supplier_responses.count()}')
            except Exception as e:
                # capture_exception(e)
                print(f"Permit {e}")
        else:
            self.messages.append('NO BUSINESS PERMIT IN QA')

    def ocr_poisons_board_license(self, qa_instance, category_id):
        poisons_boardcert = QualityAssuranceQuestion.objects.filter(
            question__short_description__icontains='Pharmacy and poisons board License',
            quality_assurance_id=qa_instance.id
        ).first()
        if poisons_boardcert:
            try:
                supplier_responses = SupplierResponse.objects.filter(
                    question_id=poisons_boardcert.question_id
                )
                for supplier_response in supplier_responses:
                    r = verify_poisonsboardcert(document=f"{supplier_response.document_url}")
                    self.responses.append(r)
                    # if r['business_id'] and r['expiry_date']:
                    date = make_aware(r['expiry_date'], timezone=pytz.timezone("Africa/Nairobi")) if r[
                        'expiry_date'] else None
                    instance, created = QualityAssuranceResponse.objects.update_or_create(
                        supplier_id=supplier_response.supplier_id,
                        quality_assurance_question_id=poisons_boardcert.id,
                        defaults={
                            "number": r['company_name'],
                            "date": date,
                            "ocr_response": r
                        }
                    )

                    company_name_outcome = "Fail"
                    supplier = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                        id=supplier_response.supplier_id).first()
                    if supplier is not None:
                        if supplier.company_name.lower() == r['business_name']:
                            company_name_outcome = "Pass"

                    expiry_date_outcome = "Fail"
                    if datetime.datetime.now().date() <= date.date():
                        expiry_date_outcome = "Pass"

                    QaPoisonsBoardResponse.objects.update_or_create(
                        qa_response_id=instance.id,
                        defaults={
                            "expiry_date": date,
                            "company_name": r["company_name"],
                            "company_name_outcome": company_name_outcome,
                            "expiry_date_outcome": expiry_date_outcome
                        }
                    )
                    # QualityAssuranceResponse.objects.update_or_create()
                self.messages.append(f'Processing done for POISONS {supplier_responses.count()}')
            except Exception as e:
                # capture_exception(e)
                print(f"Poisons {e}")
        else:
            self.messages.append('NO POISIONS BOARD QUESTION IN QA')

    def ocr_tax_pin_cert(self, qa_instance, category_id):
        pin_cert = QualityAssuranceQuestion.objects.filter(
            question__short_description__icontains='PIN certificate',
            quality_assurance_id=qa_instance.id
        ).first()
        if pin_cert:
            try:
                supplier_responses = SupplierResponse.objects.filter(
                    question_id=pin_cert.question_id
                )
                q = Question.objects.filter(
                    section__category_id=category_id, short_description="PIN number").first()
                for supplier_response in supplier_responses:
                    r = verify_pincert(document=f"{supplier_response.document_url}")
                    self.responses.append(r)
                    # if r['pin']:
                    instance, created = QualityAssuranceResponse.objects.update_or_create(
                        supplier_id=supplier_response.supplier_id,
                        quality_assurance_question_id=pin_cert.id,
                        defaults={
                            "number": r['pin'],
                            "date": make_aware(
                                datetime.datetime.now(), timezone=pytz.timezone("Africa/Nairobi")),
                            "ocr_response": r
                        }
                    )

                    tax_pin_outcome = "Fail"
                    if q is not None:
                        s_r = SupplierResponse.objects.filter(
                            question_id=q.id, supplier_id=supplier_response.supplier_id).first()
                        if s_r is not None:
                            if s_r.options:
                                if s_r.options == r['pin']:
                                    tax_pin_outcome = "Pass"

                    QaPinCertificateResponse.objects.update_or_create(
                        qa_response_id=instance.id,
                        defaults={"pin_number": r['pin'], "tax_pin_outcome": tax_pin_outcome}
                    )
                    # QualityAssuranceResponse.objects.update_or_create()
                self.messages.append(f'Processing done for PIN Cert {supplier_responses.count()}')
            except Exception as e:
                # capture_exception(e)
                print(f"Pin cert {e}")
        else:
            self.messages.append("NO PIN CERT QUESTION IN QA")

    def ocr_incorporation_certificate(self, qa_instance, category_id):
        incorporation_cert = QualityAssuranceQuestion.objects.filter(
            question__short_description__icontains='Registration certificate',
            quality_assurance_id=qa_instance.id
        ).first()
        if incorporation_cert:
            try:
                supplier_responses = SupplierResponse.objects.filter(
                    question_id=incorporation_cert.question_id
                )
                q = Question.objects.filter(
                    section__category_id=category_id,
                    short_description="Company registration certificate number").first()

                for supplier_response in supplier_responses:
                    r = verify_incorporationcert(document=f"{supplier_response.document_url}")
                    self.responses.append(r)
                    # if r['number'] and r['date']:
                    date = date = make_aware(r['date'], timezone=pytz.timezone("Africa/Nairobi")) if r['date'] else None
                    instance, created = QualityAssuranceResponse.objects.update_or_create(
                        supplier_id=supplier_response.supplier_id,
                        quality_assurance_question_id=incorporation_cert.id,
                        defaults={
                            "number": r['number'],
                            "date": date,
                            "ocr_response": r
                        }
                    )

                    company_name_outcome = "Fail"
                    supplier = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                        id=self.context['participant_id']).first()
                    if supplier is not None:
                        if supplier.company_name.lower() == r['name'].lower():
                            company_name_outcome = "Pass"

                    company_number_outcome = "Fail"
                    if q is not None:
                        s_r = SupplierResponse.objects.filter(
                            question_id=q.id, supplier_id=self.context['participant_id']).first()
                        if s_r is not None:
                            if s_r.options:
                                if s_r.options == r['number']:
                                    company_number_outcome = "Pass"

                    QaIncorporationCertificateResponse.objects.update_or_create(
                        qa_response_id=instance.id,
                        defaults={
                            "company_name": r["name"],
                            "company_name_outcome": company_name_outcome,
                            "company_number": r["number"],
                            "company_number_outcome": company_number_outcome
                        }
                    )
                self.messages.append(f'Processing done for Incorporation Cert {supplier_responses.count()}')
            except Exception as e:
                print(f"Incorporation cert {e}")
        else:
            self.messages.append('NO INCORPORATION QUESTION IN QA')

    def ocr_nca_certificate(self, qa_instance, category_id):
        nca = QualityAssuranceQuestion.objects.filter(
            question__short_description__icontains='NCA certificate',
            quality_assurance_id=qa_instance.id
        ).first()
        return


@shared_task(bind=True)
def import_category_suppliers(self, job_id):
    messages = []

    job = Tender.objects.filter(id=job_id).first()
    filepath = get_file_path(job.category_suppliers)

    workbook = load_workbook(filepath, data_only=True)

    worksheets = workbook.sheetnames[1:]
    if len(worksheets) < 1:
        messages.append("There are no sheets to upload from the excel.")

        context = {"messages": messages}
        return context

    categories_count = 0
    suppliers = 0

    progress_recorder = ProgressRecorder(self)
    result = 0

    for sheet in worksheets:
        result += 1
        progress_recorder.set_progress(result, len(worksheets))

        sheet_name_list = sheet.split("_")
        category_sheet = workbook[sheet]
        CategoryType = apps.apps.get_model('core', 'CategoryType')
        category_type = CategoryType.objects.filter(innitials=sheet_name_list[-1])

        if category_type.count() == 1:
            categories_count += 1
            category_type = category_type.first()
            for i in range(7, category_sheet.max_row + 1):
                company_name = category_sheet["C{}".format(i)].value
                primary_email = category_sheet["E{}".format(i)].value
                alternative_email = category_sheet["F{}".format(i)].value
                primary_phone = category_sheet["G{}".format(i)].value
                alternative_phone = category_sheet["H{}".format(i)].value
                country = category_sheet["I{}".format(i)].value

                supplier = None
                supplier1 = None
                supplier2 = None

                if primary_email != None:
                    supplier1 = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                        username=primary_email).first()
                elif alternative_email != None:
                    supplier2 = apps.apps.get_model('suppliers', 'Supplier').objects.filter(
                        username=alternative_email
                    ).first()

                if supplier1:
                    supplier = supplier1
                elif supplier2:
                    supplier = supplier2

                if supplier is None:
                    suppliers += 1
                    try:
                        apps.apps.get_model('core', 'CategoryTypeSupplier').objects.update_or_create(
                            primary_email=primary_email, alternative_email=alternative_email,
                            category_type=category_type,
                            defaults={
                                "primary_email": primary_email, "alternative_email": alternative_email,
                                "primary_phone": primary_phone, "alternative_phone": alternative_phone,
                                "category_type": category_type, "company_name": company_name,
                                "country": country,
                            },
                        )
                    except Exception as e:
                        capture_exception(e)
                        print(e)
                else:
                    suppliers += 1
                    try:

                        apps.apps.get_model('core', 'CategoryTypeSupplier').objects.update_or_create(
                            supplier=supplier, category_type=category_type,
                            defaults={
                                "supplier": supplier, "primary_email": primary_email,
                                "alternative_email": alternative_email, "primary_phone": primary_phone,
                                "alternative_phone": alternative_phone, "category_type": category_type,
                                "company_name": company_name, "country": country,
                            },
                        )
                    except Exception as e:
                        capture_exception(e)
                        print(e)
    os.remove(filepath)
    messages.append(
        "Imports succeeded. {} suppliers added to {} categories.".format(
            suppliers, categories_count
        )
    )
    context = {"messages": messages}
    return context


app.register_task(FinancialRatiosReport)
app.register_task(TenderSummaryReport)
app.register_task(ZipQuestionFiles)
app.register_task(ConsolidatedTenderSummaryReport)
app.register_task(EvaluateTender)
app.register_task(TenderQARankingReport())
app.register_task(DueDiligenceRankingReport())
app.register_task(TenderInterimReport())